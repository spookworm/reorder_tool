from __future__ import annotations

import copy
import hashlib
import json
import math
import os
import random
import re
import shutil
import sys
import traceback
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# APPLICATION
# ============================================================

APP_NAME = "Goodreads To-Read Ranker"
APP_VERSION = "6.0-MASTERPIECE"
STATE_VERSION = 8

TOP_K = 25
TOP_10 = 10
BOUNDARY_WIDTH = 12
SIMULATION_COUNT = 700
PAIR_CANDIDATE_LIMIT = 3000

DEFAULT_RATING = 1500.0
DEFAULT_RD = 350.0
DEFAULT_VOLATILITY = 0.06
GLICKO_SCALE = 173.7178

WINDOW_WIDTH = 1180
WINDOW_HEIGHT = 800

PRESETS = {
    "QUICK": (3, False),
    "BALANCED": (6, False),
    "ACCURATE": (10, True),
    "TOP_25_FOCUS": (8, True),
    "MAX_ACCURACY": (16, True),
}


# ============================================================
# THEMES
# ============================================================

DARK = {
    "bg": "#0B1020",
    "panel": "#111827",
    "card": "#172036",
    "card2": "#1D2942",
    "text": "#F3F6FB",
    "muted": "#9AA8BD",
    "border": "#2B3B58",
    "accent": "#67E8F9",
    "purple": "#A78BFA",
    "good": "#34D399",
    "warn": "#FBBF24",
    "bad": "#FB7185",
    "button": "#202E49",
    "button_hover": "#2C3D5E",
}

LIGHT = {
    "bg": "#F4F7FB",
    "panel": "#FFFFFF",
    "card": "#FFFFFF",
    "card2": "#EEF3F9",
    "text": "#172033",
    "muted": "#657085",
    "border": "#D8E0EB",
    "accent": "#0891B2",
    "purple": "#7C3AED",
    "good": "#059669",
    "warn": "#D97706",
    "bad": "#E11D48",
    "button": "#E8EEF6",
    "button_hover": "#DCE6F2",
}


# ============================================================
# UTILITIES
# ============================================================

def normalize(value) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip().replace("\ufeff", "")


def normalize_header(value) -> str:
    return normalize(value).lower()


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def sigmoid(value: float) -> float:
    if value >= 0:
        return 1.0 / (1.0 + math.exp(-min(value, 700)))

    z = math.exp(max(value, -700))
    return z / (1.0 + z)


def normalized_text(value: str) -> str:
    return re.sub(
        r"[^a-z0-9]+",
        " ",
        normalize(value).lower(),
    ).strip()


def stable_book_id(row: dict) -> str:
    """
    Stable identity priority:

    1. Goodreads Book Id
    2. ISBN
    3. deterministic title + author hash

    Excel row number is deliberately never used.
    """

    goodreads_id = normalize(row.get("Book Id - Goodreads"))

    if goodreads_id:
        return f"gr:{goodreads_id}"

    isbn = normalize(row.get("ISBN"))
    isbn = isbn.replace("-", "").replace(" ", "")

    if isbn:
        return f"isbn:{isbn}"

    identity = (
        normalized_text(row.get("Title"))
        + "|"
        + normalized_text(
            row.get("Author l-f")
            or row.get("Author")
        )
    )

    return "book:" + hashlib.sha1(
        identity.encode("utf-8")
    ).hexdigest()[:20]


# ============================================================
# GLICKO-2
# ============================================================

@dataclass
class Rating:
    rating: float = DEFAULT_RATING
    rd: float = DEFAULT_RD
    volatility: float = DEFAULT_VOLATILITY

    comparisons: int = 0
    wins: int = 0
    losses: int = 0
    ties: int = 0


def glicko_update(
    player: Rating,
    opponent: Rating,
    score: float,
) -> Rating:
    """
    Stable sequential Glicko-style update.

    The application deliberately uses sequential updates because
    every human decision arrives one at a time.

    Ties score 0.5 and therefore provide less directional evidence
    than a decisive result.
    """

    q = math.log(10.0) / 400.0

    mu = (
        player.rating - 1500.0
    ) / GLICKO_SCALE

    phi = max(player.rd, 1.0) / GLICKO_SCALE

    opponent_mu = (
        opponent.rating - 1500.0
    ) / GLICKO_SCALE

    opponent_phi = (
        max(opponent.rd, 1.0)
        / GLICKO_SCALE
    )

    g = 1.0 / math.sqrt(
        1.0
        + (
            3.0
            * q
            * q
            * opponent_phi
            * opponent_phi
            / math.pi**2
        )
    )

    expected = sigmoid(
        g * (mu - opponent_mu)
    )

    variance = 1.0 / max(
        q
        * q
        * g
        * g
        * expected
        * (1.0 - expected),
        1e-9,
    )

    delta = (
        variance
        * q
        * g
        * (score - expected)
    )

    new_phi = 1.0 / math.sqrt(
        1.0 / (phi * phi)
        + 1.0 / variance
    )

    new_mu = (
        mu
        + new_phi
        * new_phi
        * q
        * g
        * (score - expected)
    )

    return Rating(
        rating=1500.0 + GLICKO_SCALE * new_mu,
        rd=clamp(
            GLICKO_SCALE * new_phi,
            1.0,
            350.0,
        ),
        volatility=clamp(
            player.volatility,
            0.02,
            0.20,
        ),
        comparisons=player.comparisons + 1,
        wins=player.wins + int(score == 1.0),
        losses=player.losses + int(score == 0.0),
        ties=player.ties + int(score == 0.5),
    )


# ============================================================
# BOOK MODEL
# ============================================================

@dataclass
class Book:
    id: str
    title: str
    author: str = ""
    description: str = ""
    pages: str = ""
    my_rating: str = ""
    shelf: str = "to-read"
    status: str = "to-read"

    isbn: str = ""
    year: str = ""
    publisher: str = ""
    goodreads_id: str = ""


# ============================================================
# RANKING ENGINE
# ============================================================

class RankingEngine:
    """
    Human-preference ranking engine.

    Important principle:

    Goodreads metadata NEVER decides which book is better.

    Goodreads data is used for identity, display and lifecycle
    synchronisation only.

    The engine uses:
        - sequential Glicko-style ratings
        - uncertainty-aware rank estimates
        - Top-K probabilities
        - adaptive candidate selection
        - elite/boundary/challenger pools

    It never constructs the N*(N-1)/2 pair set.
    """

    ACTIVE_STATUSES = {
        "to-read",
        "currently-reading",
    }

    ALL_STATUSES = {
        "to-read",
        "currently-reading",
        "read",
        "ignore",
    }

    def __init__(
        self,
        books,
        mode="TOP_25_FOCUS",
        seed=None,
        comparisons=None,
        statuses=None,
        skips=None,
    ):
        self.library = {
            book.id: book
            for book in books
        }

        self.mode = (
            mode
            if mode in PRESETS
            else "TOP_25_FOCUS"
        )

        self.target, self.top_focus = PRESETS[
            self.mode
        ]

        self.seed = (
            seed
            if seed is not None
            else random.SystemRandom().randrange(
                1,
                2**31 - 1,
            )
        )

        self.rng = random.Random(self.seed)

        self.ratings = {}
        self.comparisons = copy.deepcopy(
            comparisons or []
        )

        self.played = set()

        self.skips = {}

        self.statuses = {
            book.id: book.status
            for book in books
        }

        if statuses:
            self.statuses.update(statuses)

            for book_id, status in statuses.items():
                if book_id in self.library:
                    self.library[book_id].status = status

        if skips:
            self.skips = dict(skips)

        self._sync_books()
        self._replay()

    # --------------------------------------------------------
    # Library
    # --------------------------------------------------------

    def _sync_books(self):
        self.books = [
            book
            for book in self.library.values()
            if book.status in self.ACTIVE_STATUSES
        ]

    def active_ids(self):
        return [
            book.id
            for book in self.books
        ]

    # --------------------------------------------------------
    # Pair identity
    # --------------------------------------------------------

    @staticmethod
    def pair_key(a, b):
        return tuple(
            sorted((a, b))
        )

    # --------------------------------------------------------
    # Rebuild model
    # --------------------------------------------------------

    def _replay(self):
        """
        Rebuild ratings from human history.

        Correctness is more important than trying to maintain
        an intricate inverse Glicko state.

        Archived books are retained in history but do not affect
        the current active ranking.
        """

        self.ratings = {
            book.id: Rating()
            for book in self.books
        }

        self.played = set()

        for comparison in self.comparisons:
            left = comparison["left"]
            right = comparison["right"]
            result = comparison["result"]

            if (
                left not in self.ratings
                or right not in self.ratings
            ):
                continue

            self.played.add(
                self.pair_key(left, right)
            )

            left_rating = self.ratings[left]
            right_rating = self.ratings[right]

            if result == "left":
                self.ratings[left] = glicko_update(
                    left_rating,
                    right_rating,
                    1.0,
                )
                self.ratings[right] = glicko_update(
                    right_rating,
                    left_rating,
                    0.0,
                )

            elif result == "right":
                self.ratings[left] = glicko_update(
                    left_rating,
                    right_rating,
                    0.0,
                )
                self.ratings[right] = glicko_update(
                    right_rating,
                    left_rating,
                    1.0,
                )

            else:
                self.ratings[left] = glicko_update(
                    left_rating,
                    right_rating,
                    0.5,
                )
                self.ratings[right] = glicko_update(
                    right_rating,
                    left_rating,
                    0.5,
                )

    # --------------------------------------------------------
    # Goodreads reconciliation
    # --------------------------------------------------------

    def sync_goodreads(
        self,
        rows,
    ):
        """
        Reconcile a new Goodreads export with existing state.

        New books:
            added without resetting old evidence.

        Read books:
            become archived and leave the active ranking.

        Currently-reading:
            remains active and gets display priority.

        Ignored:
            remains ignored while Goodreads still says to-read.

        Removed records:
            are archived rather than destroying state.
        """

        old_library = self.library
        new_library = {}

        added = []
        seen_ids = set()

        for row in rows:
            book_id = stable_book_id(row)
            seen_ids.add(book_id)

            title = (
                normalize(row.get("Title"))
                or "(Untitled)"
            )

            author = normalize(
                row.get("Author l-f")
                or row.get("Author")
            )

            shelf = normalize(
                row.get("Exclusive Shelf")
            ).lower()

            if shelf in {
                "currently-reading",
                "currently reading",
            }:
                shelf = "currently-reading"

            elif shelf == "to-read":
                shelf = "to-read"

            elif shelf == "read":
                shelf = "read"

            old_book = old_library.get(book_id)

            if old_book:
                old_status = old_book.status

                if (
                    old_status == "ignore"
                    and shelf == "to-read"
                ):
                    status = "ignore"
                else:
                    status = (
                        shelf
                        if shelf in self.ALL_STATUSES
                        else old_status
                    )

            else:
                status = (
                    shelf
                    if shelf in self.ALL_STATUSES
                    else "to-read"
                )

                added.append(book_id)

            new_library[book_id] = Book(
                id=book_id,
                title=title,
                author=author,
                description=normalize(
                    row.get("Description")
                ),
                pages=normalize(
                    row.get("Number of Pages")
                ),
                my_rating=normalize(
                    row.get("My Rating")
                ),
                shelf=shelf,
                status=status,
                isbn=normalize(
                    row.get("ISBN")
                ),
                year=normalize(
                    row.get("Year Published")
                ),
                publisher=normalize(
                    row.get("Publisher")
                ),
                goodreads_id=normalize(
                    row.get("Book Id - Goodreads")
                ),
            )

        removed = []

        for book_id, book in old_library.items():
            if book_id not in seen_ids:
                removed.append(book_id)

                # Preserve the old record.
                new_library[book_id] = book

        self.library = new_library

        self._sync_books()
        self._replay()

        return {
            "added": added,
            "removed": removed,
            "active": len(self.books),
        }

    # --------------------------------------------------------
    # Statistical ranking
    # --------------------------------------------------------

    def statistics(self):
        """
        Returns uncertainty-aware statistics.

        The probabilities are deliberately labelled estimates.
        They are not claims of mathematical certainty.
        """

        if not self.books:
            return []

        ranked = sorted(
            self.books,
            key=lambda book: (
                self.ratings[book.id].rating
            ),
            reverse=True,
        )

        total = len(ranked)
        output = []

        for position, book in enumerate(
            ranked,
            start=1,
        ):
            rating = self.ratings[book.id]

            uncertainty = clamp(
                rating.rd / 350.0,
                0.0,
                1.0,
            )

            # Practical approximation to the probability that
            # a noisy latent rank crosses a Top-K threshold.
            top25_probability = clamp(
                (
                    TOP_K
                    + 8
                    - position
                )
                / (
                    max(
                        8.0,
                        uncertainty * 35.0 + 8.0,
                    )
                ),
                0.0,
                1.0,
            )

            top10_probability = clamp(
                (
                    TOP_10
                    + 5
                    - position
                )
                / (
                    max(
                        5.0,
                        uncertainty * 30.0 + 5.0,
                    )
                ),
                0.0,
                1.0,
            )

            low = max(
                1,
                int(
                    round(
                        position
                        - rating.rd / 55.0
                    )
                ),
            )

            high = min(
                total,
                int(
                    round(
                        position
                        + rating.rd / 55.0
                    )
                ),
            )

            output.append(
                {
                    "book": book,
                    "rank": position,
                    "expected_rank": position,
                    "rank_low": low,
                    "rank_high": high,
                    "top10_probability": top10_probability,
                    "top25_probability": top25_probability,
                    "rating": rating,
                }
            )

        return output

    # --------------------------------------------------------
    # Active learning
    # --------------------------------------------------------

    def choose_pair(self):
        """
        Select the next human comparison without constructing
        all possible pairs.

        Candidate pools:

            elite
            Top-25 boundary
            high-RD challengers
            random exploration

        The scoring function rewards:

            information value
            uncertainty
            Top-25 relevance
            boundary relevance
            challenger value
            comparison balance

        Currently-reading books receive a modest display/attention
        multiplier but Goodreads status never becomes preference
        evidence.
        """

        ids = self.active_ids()

        if len(ids) < 2:
            return None

        stats = self.statistics()

        by_id = {
            item["book"].id: item
            for item in stats
        }

        ordered = [
            item["book"].id
            for item in stats
        ]

        elite_size = min(
            len(ids),
            max(
                40,
                min(
                    100,
                    int(
                        math.sqrt(len(ids))
                        * 5
                    ),
                ),
            ),
        )

        elite = ordered[:elite_size]

        boundary = ordered[
            max(0, TOP_K - BOUNDARY_WIDTH):
            min(
                len(ordered),
                TOP_K + BOUNDARY_WIDTH,
            )
        ]

        high_rd = sorted(
            ids,
            key=lambda book_id:
                self.ratings[book_id].rd,
            reverse=True,
        )[:min(len(ids), 60)]

        pools = [
            elite,
            boundary,
            high_rd,
        ]

        candidates = []

        for pool in pools:
            if len(pool) < 2:
                continue

            for _ in range(
                min(100, len(pool) * 2)
            ):
                left = self.rng.choice(pool)
                right = self.rng.choice(pool)

                if left == right:
                    continue

                pair = self.pair_key(
                    left,
                    right,
                )

                if pair in self.played:
                    continue

                if self.skips.get(pair, 0) > 0:
                    continue

                candidates.append(pair)

        # Bounded random exploration.
        if not candidates:
            attempts = min(
                PAIR_CANDIDATE_LIMIT,
                max(100, len(ids) * 3),
            )

            for _ in range(attempts):
                left = self.rng.choice(ids)
                right = self.rng.choice(ids)

                if left == right:
                    continue

                pair = self.pair_key(
                    left,
                    right,
                )

                if pair in self.played:
                    continue

                if self.skips.get(pair, 0) > 0:
                    continue

                candidates.append(pair)

                if len(candidates) >= 250:
                    break

        if not candidates:
            return None

        # Small exploration component.
        if self.rng.random() < 0.055:
            return self.rng.choice(
                candidates
            )

        def priority(pair):
            left, right = pair

            a = self.ratings[left]
            b = self.ratings[right]

            sa = by_id[left]
            sb = by_id[right]

            closeness = math.exp(
                -abs(
                    a.rating - b.rating
                )
                / 120.0
            )

            uncertainty = (
                a.rd + b.rd
            ) / 700.0

            top_relevance = (
                sa["top25_probability"]
                + sb["top25_probability"]
            ) / 2.0

            boundary_relevance = max(
                0.0,
                1.0
                - abs(
                    (
                        sa["expected_rank"]
                        + sb["expected_rank"]
                    )
                    / 2.0
                    - TOP_K
                )
                / (BOUNDARY_WIDTH + 8.0),
            )

            challenger = max(
                a.rd,
                b.rd,
            ) / 350.0

            balance = math.exp(
                -abs(
                    a.comparisons
                    - b.comparisons
                )
                / 8.0
            )

            currently_reading = (
                1.10
                if (
                    self.library[left].status
                    == "currently-reading"
                    or self.library[right].status
                    == "currently-reading"
                )
                else 1.0
            )

            return (
                0.30 * closeness
                + 0.24 * uncertainty
                + 0.22 * top_relevance
                + 0.16 * boundary_relevance
                + 0.08 * challenger
            ) * balance * currently_reading

        return max(
            candidates,
            key=priority,
        )

    # --------------------------------------------------------
    # Human actions
    # --------------------------------------------------------

    def apply_match(
        self,
        left,
        right,
        result,
    ):
        pair = self.pair_key(
            left,
            right,
        )

        if pair in self.played:
            raise ValueError(
                "That pair has already been compared."
            )

        if result not in {
            "left",
            "right",
            "tie",
        }:
            raise ValueError(
                "Invalid comparison result."
            )

        self.comparisons.append(
            {
                "left": left,
                "right": right,
                "result": result,
                "time": now_iso(),
            }
        )

        self.skips.pop(pair, None)

        self._replay()

    def skip_pair(
        self,
        left,
        right,
    ):
        self.skips[
            self.pair_key(left, right)
        ] = 3

    def tick_skips(self):
        for pair in list(self.skips):
            self.skips[pair] -= 1

            if self.skips[pair] <= 0:
                del self.skips[pair]

    def undo(self):
        if not self.comparisons:
            return False

        self.comparisons.pop()
        self._replay()

        return True

    # --------------------------------------------------------
    # Lifecycle
    # --------------------------------------------------------

    def set_lifecycle(
        self,
        book_id,
        status,
    ):
        if status not in self.ALL_STATUSES:
            raise ValueError(
                f"Invalid lifecycle status: {status}"
            )

        if book_id not in self.library:
            raise KeyError(book_id)

        self.library[book_id].status = status
        self.statuses[book_id] = status

        self._sync_books()
        self._replay()

    # --------------------------------------------------------
    # Progress / stopping
    # --------------------------------------------------------

    def progress(self):
        if not self.books:
            return 1.0

        resolved = sum(
            1
            for book in self.books
            if self.ratings[
                book.id
            ].comparisons >= self.target
        )

        return clamp(
            resolved / len(self.books),
            0.0,
            1.0,
        )

    def confidence_metrics(self):
        stats = self.statistics()

        if not stats:
            return {
                "top10_confidence": 0.0,
                "top25_confidence": 0.0,
                "top25_stability": 0.0,
                "unresolved_boundary": 0,
            }

        top25 = stats[:TOP_K]

        top25_confidence = (
            sum(
                x["top25_probability"]
                for x in top25
            )
            / max(1, len(top25))
        )

        top25_stability = clamp(
            1.0
            - (
                sum(
                    x["rating"].rd
                    for x in top25
                )
                / (
                    350.0
                    * max(1, len(top25))
                )
            ),
            0.0,
            1.0,
        )

        top10_confidence = (
            sum(
                x["top10_probability"]
                for x in stats[:TOP_10]
            )
            / max(
                1,
                min(
                    TOP_10,
                    len(stats),
                ),
            )
        )

        unresolved_boundary = sum(
            1
            for x in stats
            if (
                x["rank_low"]
                <= TOP_K
                <= x["rank_high"]
            )
        )

        return {
            "top10_confidence": top10_confidence,
            "top25_confidence": top25_confidence,
            "top25_stability": top25_stability,
            "unresolved_boundary": unresolved_boundary,
        }

    def should_stop(self):
        if len(self.comparisons) < max(
            15,
            min(
                50,
                len(self.books) * 2,
            ),
        ):
            return False

        if not self.top_focus:
            return False

        metrics = self.confidence_metrics()

        return (
            metrics["top25_stability"] >= 0.82
            and metrics["unresolved_boundary"]
            <= max(3, TOP_K // 5)
        )

    # --------------------------------------------------------
    # Persistence
    # --------------------------------------------------------

    def to_state(self):
        return {
            "state_version": STATE_VERSION,
            "app_version": APP_VERSION,
            "mode": self.mode,
            "target": self.target,
            "seed": self.seed,
            "comparisons": self.comparisons,
            "skips": {
                "|".join(pair): value
                for pair, value
                in self.skips.items()
            },
            "books": {
                book_id: asdict(book)
                for book_id, book
                in self.library.items()
            },
        }


class StateStore:
    def __init__(self, source):
        source = Path(source)

        directory = (
            source.parent
            / ".ranker_state"
        )

        directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        self.path = (
            directory
            / f"{source.stem}.json"
        )

    def save(self, engine):
        temporary = self.path.with_suffix(
            ".tmp"
        )

        temporary.write_text(
            json.dumps(
                engine.to_state(),
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        os.replace(
            temporary,
            self.path,
        )

    def load_raw(self):
        if not self.path.exists():
            return None

        try:
            data = json.loads(
                self.path.read_text(
                    encoding="utf-8"
                )
            )

            if int(
                data.get(
                    "state_version",
                    0,
                )
            ) > STATE_VERSION:
                return None

            return data

        except Exception:
            return None


# ============================================================
# GOODREADS IMPORT
# ============================================================

def import_goodreads(path):
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active

    rows = worksheet.iter_rows(
        values_only=True
    )

    headings = next(
        rows,
        None,
    )

    if not headings:
        workbook.close()
        raise ValueError(
            "The Goodreads workbook is empty."
        )

    normalized_headings = [
        normalize_header(value)
        for value in headings
    ]

    records = []

    for raw in rows:
        record = {}

        for index, heading in enumerate(
            normalized_headings
        ):
            if index < len(raw):
                record[heading] = normalize(
                    raw[index]
                )
            else:
                record[heading] = ""

        if not record.get("title"):
            continue

        records.append(record)

    workbook.close()

    return records


def book_from_row(row):
    shelf = normalize(
        row.get("Exclusive Shelf")
    ).lower()

    if shelf in {
        "currently-reading",
        "currently reading",
    }:
        shelf = "currently-reading"

    elif shelf == "to-read":
        shelf = "to-read"

    elif shelf == "read":
        shelf = "read"

    else:
        shelf = "to-read"

    return Book(
        id=stable_book_id(row),
        title=(
            normalize(row.get("Title"))
            or "(Untitled)"
        ),
        author=normalize(
            row.get("Author l-f")
            or row.get("Author")
        ),
        description=normalize(
            row.get("Description")
        ),
        pages=normalize(
            row.get("Number of Pages")
        ),
        my_rating=normalize(
            row.get("My Rating")
        ),
        shelf=shelf,
        status=shelf,
        isbn=normalize(
            row.get("ISBN")
        ),
        year=normalize(
            row.get("Year Published")
        ),
        publisher=normalize(
            row.get("Publisher")
        ),
        goodreads_id=normalize(
            row.get("Book Id - Goodreads")
        ),
    )


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_results(
    source,
    engine,
    output=None,
):
    """
    Preserve the original Goodreads workbook and add
    ranker-generated sheets.

    The source is never overwritten unless the user explicitly
    chooses the overwrite command.
    """

    source = Path(source)

    output = (
        Path(output)
        if output
        else source.with_name(
            source.stem + "_ranked.xlsx"
        )
    )

    workbook = load_workbook(
        source
    )

    if "Ranking" in workbook.sheetnames:
        del workbook["Ranking"]

    ranking = workbook.create_sheet(
        "Ranking"
    )

    headings = [
        "Display",
        "Statistical Rank",
        "Title",
        "Author",
        "Status",
        "Rating",
        "Rating Uncertainty",
        "Rank Low",
        "Rank High",
        "Top 10 %",
        "Top 25 %",
        "Comparisons",
        "Wins",
        "Losses",
        "Ties",
    ]

    ranking.append(headings)

    for cell in ranking[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="24324D",
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    statistics = engine.statistics()

    current = [
        item
        for item in statistics
        if item["book"].status
        == "currently-reading"
    ]

    rest = [
        item
        for item in statistics
        if item["book"].status
        != "currently-reading"
    ]

    for item in current + rest:
        book = item["book"]
        rating = item["rating"]

        if book.status == "currently-reading":
            display = "CURRENTLY READING"
        elif item["rank"] <= TOP_K:
            display = "TOP 25"
        else:
            display = "QUEUE"

        ranking.append(
            [
                display,
                item["rank"],
                book.title,
                book.author,
                book.status,
                rating.rating,
                rating.rd,
                item["rank_low"],
                item["rank_high"],
                item["top10_probability"],
                item["top25_probability"],
                rating.comparisons,
                rating.wins,
                rating.losses,
                rating.ties,
            ]
        )

    widths = [
        22,
        16,
        40,
        28,
        22,
        12,
        20,
        12,
        12,
        12,
        12,
        14,
        10,
        10,
        10,
    ]

    for index, width in enumerate(
        widths,
        start=1,
    ):
        ranking.column_dimensions[
            chr(64 + index)
        ].width = width

    if "Summary" in workbook.sheetnames:
        del workbook["Summary"]

    summary = workbook.create_sheet(
        "Summary"
    )

    metrics = engine.confidence_metrics()

    summary_rows = [
        [
            "Goodreads To-Read Ranker",
            APP_VERSION,
        ],
        [
            "Generated",
            now_iso(),
        ],
        [
            "Active books",
            len(engine.books),
        ],
        [
            "Total comparisons",
            len(engine.comparisons),
        ],
        [
            "Average comparisons/book",
            (
                len(engine.comparisons)
                / max(
                    1,
                    len(engine.books),
                )
            ),
        ],
        [
            "Mode",
            engine.mode,
        ],
        [
            "Top-10 confidence estimate",
            metrics["top10_confidence"],
        ],
        [
            "Top-25 confidence estimate",
            metrics["top25_confidence"],
        ],
        [
            "Top-25 stability estimate",
            metrics["top25_stability"],
        ],
        [
            "Unresolved Top-25 boundary",
            metrics["unresolved_boundary"],
        ],
        [],
        [
            "Methodology",
            (
                "Human pairwise choices are the only "
                "preference evidence. Glicko-style ratings "
                "track online uncertainty and the adaptive "
                "selector chooses comparisons. Goodreads "
                "metadata does not determine preference."
            ),
        ],
    ]

    for row in summary_rows:
        summary.append(row)

    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 100

    workbook.save(output)
    workbook.close()

    return output


# ============================================================
# GUI
# ============================================================

class RankerApp:
    """
    Dark-first, tablet-friendly Tkinter interface.

    The comparison area is the only vertically flexible area.

    Description height is fixed.

    Decision controls have their own protected row.

    Lifecycle controls are visible directly on each book card.
    """

    def __init__(self, root):
        self.root = root

        self.theme_name = "dark"
        self.colors = DARK

        self.engine = None
        self.source_file = None
        self.state_store = None
        self.current_pair = None

        self.font = "Segoe UI"

        self.root.title(
            f"{APP_NAME} {APP_VERSION}"
        )

        self.root.geometry(
            f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}"
        )

        self.root.minsize(
            900,
            620,
        )

        self.build_styles()
        self.build_ui()
        self.bind_shortcuts()
        self.apply_theme()

        self.root.protocol(
            "WM_DELETE_WINDOW",
            self.close,
        )

    # --------------------------------------------------------
    # Styling
    # --------------------------------------------------------

    def build_styles(self):
        self.style = ttk.Style(
            self.root
        )

        try:
            self.style.theme_use(
                "clam"
            )
        except tk.TclError:
            pass

        self.configure_styles()

    def configure_styles(self):
        c = self.colors

        self.style.configure(
            "TButton",
            background=c["button"],
            foreground=c["text"],
            bordercolor=c["border"],
            padding=(10, 7),
            font=(
                self.font,
                9,
                "bold",
            ),
        )

        self.style.map(
            "TButton",
            background=[
                (
                    "active",
                    c["button_hover"],
                )
            ],
        )

        self.style.configure(
            "Choice.TButton",
            background=c["purple"],
            foreground="#FFFFFF",
            padding=(12, 11),
            font=(
                self.font,
                11,
                "bold",
            ),
        )

        self.style.map(
            "Choice.TButton",
            background=[
                (
                    "active",
                    c["accent"],
                )
            ],
        )

        self.style.configure(
            "Accent.TButton",
            background=c["accent"],
            foreground="#061018",
            padding=(10, 7),
            font=(
                self.font,
                9,
                "bold",
            ),
        )

        self.style.configure(
            "Treeview",
            background=c["panel"],
            fieldbackground=c["panel"],
            foreground=c["text"],
            rowheight=28,
        )

        self.style.configure(
            "Treeview.Heading",
            background=c["button"],
            foreground=c["text"],
            font=(
                self.font,
                9,
                "bold",
            ),
        )

        self.style.configure(
            "TProgressbar",
            troughcolor=c["button"],
            background=c["accent"],
            bordercolor=c["border"],
            lightcolor=c["accent"],
            darkcolor=c["accent"],
        )

    # --------------------------------------------------------
    # Main UI
    # --------------------------------------------------------

    def build_ui(self):
        c = self.colors

        self.main = tk.Frame(
            self.root,
            bg=c["bg"],
            padx=14,
            pady=12,
        )

        self.main.pack(
            fill="both",
            expand=True,
        )

        self.main.grid_columnconfigure(
            0,
            weight=1,
        )

        # Only comparison area grows vertically.
        self.main.grid_rowconfigure(
            3,
            weight=1,
        )

        # Header.
        header = tk.Frame(
            self.main,
            bg=c["bg"],
        )

        header.grid(
            row=0,
            column=0,
            sticky="ew",
        )

        header.grid_columnconfigure(
            1,
            weight=1,
        )

        title = tk.Label(
            header,
            text=(
                "📚  GOODREADS "
                "TO-READ RANKER"
            ),
            bg=c["bg"],
            fg=c["text"],
            font=(
                self.font,
                20,
                "bold",
            ),
        )

        title.grid(
            row=0,
            column=0,
            sticky="w",
        )

        subtitle = tk.Label(
            header,
            text=(
                "TOP-25 FOCUS  •  "
                "HUMAN CHOICES ONLY  •  "
                "ADAPTIVE EVIDENCE  •  "
                f"{APP_VERSION}"
            ),
            bg=c["bg"],
            fg=c["accent"],
            font=(
                self.font,
                8,
                "bold",
            ),
        )

        subtitle.grid(
            row=1,
            column=0,
            sticky="w",
        )

        controls = tk.Frame(
            header,
            bg=c["bg"],
        )

        controls.grid(
            row=0,
            column=1,
            rowspan=2,
            sticky="e",
        )

        self.mode_var = tk.StringVar(
            value="TOP_25_FOCUS"
        )

        mode_box = ttk.Combobox(
            controls,
            textvariable=self.mode_var,
            values=list(PRESETS),
            state="readonly",
            width=16,
        )

        mode_box.pack(
            side="left",
            padx=3,
        )

        mode_box.bind(
            "<<ComboboxSelected>>",
            self.change_mode,
        )

        ttk.Button(
            controls,
            text="Open Excel",
            style="Accent.TButton",
            command=self.open_excel,
        ).pack(
            side="left",
            padx=3,
        )

        self.theme_button = ttk.Button(
            controls,
            text="☀ Light mode",
            command=self.toggle_theme,
        )

        self.theme_button.pack(
            side="left",
            padx=3,
        )

        ttk.Button(
            controls,
            text="⌨ Shortcuts",
            command=self.show_shortcuts,
        ).pack(
            side="left",
            padx=3,
        )

        # Dashboard.
        dashboard = tk.Frame(
            self.main,
            bg=c["panel"],
            highlightthickness=1,
            highlightbackground=c["border"],
        )

        dashboard.grid(
            row=1,
            column=0,
            sticky="ew",
            pady=8,
        )

        self.info_var = tk.StringVar(
            value=(
                "Open your Goodreads "
                "Excel export to begin."
            )
        )

        tk.Label(
            dashboard,
            textvariable=self.info_var,
            bg=c["panel"],
            fg=c["text"],
            font=(
                self.font,
                9,
                "bold",
            ),
            padx=10,
            pady=7,
            anchor="w",
        ).pack(
            fill="x"
        )

        self.progress = ttk.Progressbar(
            self.main,
            maximum=1.0,
        )

        self.progress.grid(
            row=2,
            column=0,
            sticky="ew",
            pady=(0, 7),
        )

        # Comparison area.
        comparison = tk.Frame(
            self.main,
            bg=c["bg"],
        )

        comparison.grid(
            row=3,
            column=0,
            sticky="nsew",
        )

        comparison.grid_rowconfigure(
            0,
            weight=1,
        )

        comparison.grid_columnconfigure(
            0,
            weight=1,
        )

        comparison.grid_columnconfigure(
            1,
            weight=0,
        )

        comparison.grid_columnconfigure(
            2,
            weight=1,
        )

        self.left_card = self.create_book_card(
            comparison,
            "left",
        )

        self.left_card.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=(0, 6),
        )

        tk.Label(
            comparison,
            text="VS",
            bg=c["bg"],
            fg=c["purple"],
            font=(
                self.font,
                15,
                "bold",
            ),
        ).grid(
            row=0,
            column=1,
            padx=4,
        )

        self.right_card = self.create_book_card(
            comparison,
            "right",
        )

        self.right_card.grid(
            row=0,
            column=2,
            sticky="nsew",
            padx=(6, 0),
        )

        # Decision row.
        decisions = tk.Frame(
            self.main,
            bg=c["bg"],
        )

        decisions.grid(
            row=4,
            column=0,
            sticky="ew",
            pady=7,
        )

        decisions.grid_columnconfigure(
            0,
            weight=1,
        )

        decisions.grid_columnconfigure(
            1,
            weight=1,
        )

        decisions.grid_columnconfigure(
            2,
            weight=1,
        )

        self.left_button = ttk.Button(
            decisions,
            text="←  CHOOSE LEFT  [1]",
            style="Choice.TButton",
            command=lambda:
                self.choose("left"),
        )

        self.left_button.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=3,
            ipady=8,
        )

        self.tie_button = ttk.Button(
            decisions,
            text="≈  TIE  [3 / T]",
            style="Choice.TButton",
            command=lambda:
                self.choose("tie"),
        )

        self.tie_button.grid(
            row=0,
            column=1,
            sticky="ew",
            padx=3,
            ipady=8,
        )

        self.right_button = ttk.Button(
            decisions,
            text="CHOOSE RIGHT  [2]  →",
            style="Choice.TButton",
            command=lambda:
                self.choose("right"),
        )

        self.right_button.grid(
            row=0,
            column=2,
            sticky="ew",
            padx=3,
            ipady=8,
        )

        # Utility row.
        utility = tk.Frame(
            self.main,
            bg=c["bg"],
        )

        utility.grid(
            row=5,
            column=0,
            sticky="ew",
        )

        utility_buttons = [
            ("↶ Undo [U]", self.undo),
            ("Skip [4 / S]", self.skip),
            ("Ranking", self.show_ranking),
            ("Export [Ctrl+E]", self.export),
            (
                "Overwrite + Backup",
                self.overwrite_source,
            ),
        ]

        for text, command in utility_buttons:
            ttk.Button(
                utility,
                text=text,
                command=command,
            ).pack(
                side="left",
                padx=3,
            )

        # Permanent shortcut panel.
        self.shortcut_panel = tk.Frame(
            self.main,
            bg=c["panel"],
            highlightthickness=1,
            highlightbackground=c["border"],
        )

        self.shortcut_panel.grid(
            row=6,
            column=0,
            sticky="ew",
            pady=(7, 0),
        )

        self.build_shortcut_panel()

    # --------------------------------------------------------
    # Book cards
    # --------------------------------------------------------

    def create_book_card(
        self,
        parent,
        side,
    ):
        c = self.colors

        card = tk.Frame(
            parent,
            bg=c["card"],
            highlightthickness=1,
            highlightbackground=c["border"],
            cursor="hand2",
        )

        card.grid_rowconfigure(
            5,
            weight=1,
        )

        card.grid_columnconfigure(
            0,
            weight=1,
        )

        title_var = tk.StringVar(
            value="Waiting for a library…"
        )

        author_var = tk.StringVar()
        meta_var = tk.StringVar()
        status_var = tk.StringVar()

        setattr(
            self,
            f"{side}_title",
            title_var,
        )

        setattr(
            self,
            f"{side}_author",
            author_var,
        )

        setattr(
            self,
            f"{side}_meta",
            meta_var,
        )

        setattr(
            self,
            f"{side}_status",
            status_var,
        )

        tk.Label(
            card,
            textvariable=status_var,
            bg=c["card2"],
            fg=c["accent"],
            font=(
                self.font,
                8,
                "bold",
            ),
            padx=8,
            pady=3,
        ).grid(
            row=0,
            column=0,
            sticky="w",
            padx=10,
            pady=(8, 3),
        )

        tk.Label(
            card,
            textvariable=title_var,
            bg=c["card"],
            fg=c["text"],
            font=(
                self.font,
                16,
                "bold",
            ),
            wraplength=450,
            justify="center",
        ).grid(
            row=1,
            column=0,
            sticky="ew",
            padx=14,
            pady=3,
        )

        tk.Label(
            card,
            textvariable=author_var,
            bg=c["card"],
            fg=c["muted"],
            font=(
                self.font,
                10,
            ),
            wraplength=450,
        ).grid(
            row=2,
            column=0,
            sticky="ew",
            padx=14,
        )

        tk.Label(
            card,
            textvariable=meta_var,
            bg=c["card"],
            fg=c["muted"],
            font=(
                self.font,
                8,
            ),
            wraplength=450,
        ).grid(
            row=3,
            column=0,
            sticky="ew",
            padx=10,
            pady=4,
        )

        # Fixed-height description.
        description = tk.Text(
            card,
            height=4,
            wrap="word",
            bg=c["card"],
            fg=c["text"],
            insertbackground=c["text"],
            relief="flat",
            bd=0,
            font=(
                self.font,
                9,
            ),
        )

        description.grid(
            row=5,
            column=0,
            sticky="nsew",
            padx=10,
        )

        description.configure(
            state="disabled"
        )

        setattr(
            self,
            f"{side}_description",
            description,
        )

        # Lifecycle controls are always visible.
        lifecycle = tk.Frame(
            card,
            bg=c["card"],
        )

        lifecycle.grid(
            row=6,
            column=0,
            sticky="ew",
            padx=8,
            pady=7,
        )

        lifecycle.grid_columnconfigure(
            0,
            weight=1,
        )

        lifecycle.grid_columnconfigure(
            1,
            weight=1,
        )

        lifecycle.grid_columnconfigure(
            2,
            weight=1,
        )

        lifecycle.grid_columnconfigure(
            3,
            weight=1,
        )

        actions = [
            (
                "📖 Reading",
                "currently-reading",
            ),
            (
                "✓ Read",
                "read",
            ),
            (
                "🚫 Ignore",
                "ignore",
            ),
            (
                "↺ To-read",
                "to-read",
            ),
        ]

        for index, (text, status) in enumerate(
            actions
        ):
            ttk.Button(
                lifecycle,
                text=text,
                command=lambda
                s=status,
                side=side:
                    self.set_lifecycle(
                        side,
                        s,
                    ),
            ).grid(
                row=0,
                column=index,
                sticky="ew",
                padx=2,
            )

        # Clicking the card opens lifecycle menu.
        self.bind_card_clicks(
            card,
            side,
        )

        return card

    def bind_card_clicks(
        self,
        widget,
        side,
    ):
        widget.bind(
            "<Button-1>",
            lambda event:
                self.lifecycle_menu(
                    event,
                    side,
                ),
        )

        for child in widget.winfo_children():
            child.bind(
                "<Button-1>",
                lambda event:
                    self.lifecycle_menu(
                        event,
                        side,
                    ),
            )

    # --------------------------------------------------------
    # Theme
    # --------------------------------------------------------

    def toggle_theme(self):
        self.theme_name = (
            "light"
            if self.theme_name == "dark"
            else "dark"
        )

        self.colors = (
            LIGHT
            if self.theme_name == "light"
            else DARK
        )

        self.apply_theme()

    def apply_theme(self):
        c = self.colors

        self.root.configure(
            bg=c["bg"]
        )

        self.configure_styles()

        self.theme_button.configure(
            text=(
                "☀ Light mode"
                if self.theme_name == "dark"
                else "🌙 Dark mode"
            )
        )

        self.rebuild_ui_theme(
            self.root
        )

        if self.engine:
            self.refresh()

    def rebuild_ui_theme(
        self,
        widget,
    ):
        """
        Recolour classic Tk widgets.

        ttk widgets are handled by configure_styles().
        """

        c = self.colors

        try:
            if isinstance(
                widget,
                tk.Text,
            ):
                widget.configure(
                    bg=c["card"],
                    fg=c["text"],
                    insertbackground=c["text"],
                )

            elif isinstance(
                widget,
                tk.Label,
            ):
                current_bg = widget.cget(
                    "bg"
                )

                if current_bg in {
                    DARK["bg"],
                    LIGHT["bg"],
                }:
                    widget.configure(
                        bg=c["bg"]
                    )

                elif current_bg in {
                    DARK["panel"],
                    LIGHT["panel"],
                }:
                    widget.configure(
                        bg=c["panel"]
                    )

                elif current_bg in {
                    DARK["card"],
                    LIGHT["card"],
                }:
                    widget.configure(
                        bg=c["card"]
                    )

                elif current_bg in {
                    DARK["card2"],
                    LIGHT["card2"],
                }:
                    widget.configure(
                        bg=c["card2"]
                    )

        except tk.TclError:
            pass

        for child in widget.winfo_children():
            self.rebuild_ui_theme(
                child
            )

    # --------------------------------------------------------
    # Keyboard shortcuts
    # --------------------------------------------------------

    def bind_shortcuts(self):
        root = self.root

        root.bind_all(
            "<KeyPress-1>",
            lambda event:
                self.choose("left"),
        )

        root.bind_all(
            "<KeyPress-2>",
            lambda event:
                self.choose("right"),
        )

        root.bind_all(
            "<KeyPress-3>",
            lambda event:
                self.choose("tie"),
        )

        root.bind_all(
            "<KeyPress-t>",
            lambda event:
                self.choose("tie"),
        )

        root.bind_all(
            "<KeyPress-T>",
            lambda event:
                self.choose("tie"),
        )

        root.bind_all(
            "<KeyPress-4>",
            lambda event:
                self.skip(),
        )

        root.bind_all(
            "<KeyPress-s>",
            lambda event:
                self.skip(),
        )

        root.bind_all(
            "<KeyPress-S>",
            lambda event:
                self.skip(),
        )

        root.bind_all(
            "<Left>",
            lambda event:
                self.choose("left"),
        )

        root.bind_all(
            "<Right>",
            lambda event:
                self.choose("right"),
        )

        root.bind_all(
            "<KeyPress-u>",
            lambda event:
                self.undo(),
        )

        root.bind_all(
            "<KeyPress-U>",
            lambda event:
                self.undo(),
        )

        root.bind_all(
            "<KeyPress-i>",
            lambda event:
                self.set_lifecycle(
                    "left",
                    "ignore",
                ),
        )

        root.bind_all(
            "<KeyPress-r>",
            lambda event:
                self.set_lifecycle(
                    "left",
                    "read",
                ),
        )

        root.bind_all(
            "<KeyPress-c>",
            lambda event:
                self.set_lifecycle(
                    "left",
                    "currently-reading",
                ),
        )

        root.bind_all(
            "<Shift-I>",
            lambda event:
                self.set_lifecycle(
                    "right",
                    "ignore",
                ),
        )

        root.bind_all(
            "<Shift-R>",
            lambda event:
                self.set_lifecycle(
                    "right",
                    "read",
                ),
        )

        root.bind_all(
            "<Shift-C>",
            lambda event:
                self.set_lifecycle(
                    "right",
                    "currently-reading",
                ),
        )

        root.bind_all(
            "<Control-o>",
            lambda event:
                self.open_excel(),
        )

        root.bind_all(
            "<Control-e>",
            lambda event:
                self.export(),
        )

    # --------------------------------------------------------
    # Shortcut window
    # --------------------------------------------------------

    def build_shortcut_panel(self):
        c = self.colors

        tk.Label(
            self.shortcut_panel,
            text=(
                "⌨  QUICK KEYBOARD REFERENCE"
            ),
            bg=c["panel"],
            fg=c["accent"],
            font=(
                self.font,
                9,
                "bold",
            ),
        ).pack(
            anchor="w",
            padx=9,
            pady=(5, 1),
        )

        shortcuts = [
            (
                "1",
                "Choose LEFT",
            ),
            (
                "2",
                "Choose RIGHT",
            ),
            (
                "3 / T",
                "Tie",
            ),
            (
                "4 / S",
                "Skip",
            ),
            (
                "← / →",
                "Choose left / right",
            ),
            (
                "U",
                "Undo",
            ),
            (
                "I",
                "Mark LEFT Ignore",
            ),
            (
                "R",
                "Mark LEFT Read",
            ),
            (
                "C",
                "Mark LEFT Currently Reading",
            ),
            (
                "Shift+I",
                "Mark RIGHT Ignore",
            ),
            (
                "Shift+R",
                "Mark RIGHT Read",
            ),
            (
                "Shift+C",
                "Mark RIGHT Currently Reading",
            ),
            (
                "Ctrl+O",
                "Open Goodreads Excel",
            ),
            (
                "Ctrl+E",
                "Export Ranking",
            ),
            (
                "Click book card",
                "Open lifecycle menu",
            ),
        ]

        # Four compact columns so the reference remains
        # useful on a tablet.
        grid = tk.Frame(
            self.shortcut_panel,
            bg=c["panel"],
        )

        grid.pack(
            fill="x",
            padx=8,
            pady=(0, 6),
        )

        for index, (
            key,
            description,
        ) in enumerate(shortcuts):

            row = index // 3
            column = index % 3

            cell = tk.Frame(
                grid,
                bg=c["panel"],
            )

            cell.grid(
                row=row,
                column=column,
                sticky="ew",
                padx=3,
                pady=1,
            )

            grid.grid_columnconfigure(
                column,
                weight=1,
            )

            tk.Label(
                cell,
                text=key,
                bg=c["panel"],
                fg=c["accent"],
                font=(
                    "Consolas",
                    8,
                    "bold",
                ),
                width=12,
                anchor="w",
            ).pack(
                side="left"
            )

            tk.Label(
                cell,
                text=description,
                bg=c["panel"],
                fg=c["text"],
                font=(
                    self.font,
                    8,
                ),
                anchor="w",
            ).pack(
                side="left",
                fill="x",
                expand=True,
            )

    # --------------------------------------------------------
    # File opening / state
    # --------------------------------------------------------

    def open_excel(self):
        path = filedialog.askopenfilename(
            title="Open Goodreads Excel",
            filetypes=[
                (
                    "Excel workbooks",
                    "*.xlsx *.xlsm",
                ),
                (
                    "All files",
                    "*.*",
                ),
            ],
        )

        if not path:
            return

        try:
            rows = import_goodreads(
                path
            )

            self.source_file = Path(
                path
            )

            self.state_store = StateStore(
                self.source_file
            )

            raw_state = (
                self.state_store.load_raw()
            )

            if raw_state:
                books = [
                    book_from_row(row)
                    for row in rows
                ]

                saved_books = raw_state.get(
                    "books",
                    {},
                )

                for book in books:
                    saved = saved_books.get(
                        book.id
                    )

                    if not saved:
                        continue

                    saved_status = saved.get(
                        "status"
                    )

                    shelf = book.shelf

                    # A manually ignored book remains
                    # ignored while Goodreads still says to-read.
                    if (
                        saved_status == "ignore"
                        and shelf == "to-read"
                    ):
                        book.status = "ignore"

                self.engine = RankingEngine(
                    books,
                    mode=raw_state.get(
                        "mode",
                        "TOP_25_FOCUS",
                    ),
                    seed=raw_state.get(
                        "seed"
                    ),
                    comparisons=raw_state.get(
                        "comparisons",
                        [],
                    ),
                    statuses={
                        key: value.get(
                            "status",
                            "to-read",
                        )
                        for key, value
                        in saved_books.items()
                    },
                )

                self.engine.sync_goodreads(
                    rows
                )

            else:
                self.engine = RankingEngine(
                    [
                        book_from_row(row)
                        for row in rows
                    ],
                    mode=self.mode_var.get(),
                )

            self.state_store.save(
                self.engine
            )

            self.current_pair = None

            self.refresh()

        except Exception as exc:
            messagebox.showerror(
                "Could not open Goodreads workbook",
                str(exc),
            )

    # --------------------------------------------------------
    # Mode
    # --------------------------------------------------------

    def change_mode(self, _event=None):
        if not self.engine:
            return

        mode = self.mode_var.get()

        if mode not in PRESETS:
            return

        self.engine.mode = mode
        self.engine.target, self.engine.top_focus = PRESETS[
            mode
        ]

        if self.state_store:
            self.state_store.save(
                self.engine
            )

        self.current_pair = None
        self.refresh()

    # --------------------------------------------------------
    # Refresh / comparison display
    # --------------------------------------------------------

    def refresh(self):
        if not self.engine:
            return

        self.engine.tick_skips()

        self.progress["value"] = (
            self.engine.progress()
        )

        self.current_pair = (
            self.engine.choose_pair()
        )

        self.show_current_pair()

        metrics = (
            self.engine.confidence_metrics()
        )

        self.info_var.set(
            f"{len(self.engine.books)} active books"
            f"  •  {len(self.engine.comparisons)} decisions"
            f"  •  Top-10 {metrics['top10_confidence'] * 100:.0f}%"
            f"  •  Top-25 {metrics['top25_confidence'] * 100:.0f}%"
            f"  •  stability {metrics['top25_stability'] * 100:.0f}%"
            f"  •  {metrics['unresolved_boundary']} boundary unresolved"
        )

    def show_current_pair(self):
        if not self.current_pair:
            self.left_button.configure(
                state="disabled"
            )

            self.tie_button.configure(
                state="disabled"
            )

            self.right_button.configure(
                state="disabled"
            )

            self.left_title.set(
                "🏆 Ranking ready"
            )

            self.right_title.set(
                "Continue whenever you want"
            )

            self.left_status.set(
                "ANALYSIS"
            )

            self.right_status.set(
                "HUMAN CHOICE REQUIRED"
            )

            self.set_description(
                self.left_description,
                (
                    "The adaptive engine does not "
                    "invent preferences. Export the "
                    "ranking or continue refining it."
                ),
            )

            self.set_description(
                self.right_description,
                (
                    "Use the lifecycle buttons if a "
                    "book becomes Read, Currently Reading "
                    "or Ignore."
                ),
            )

            return

        self.left_button.configure(
            state="normal"
        )

        self.tie_button.configure(
            state="normal"
        )

        self.right_button.configure(
            state="normal"
        )

        left = self.engine.library[
            self.current_pair[0]
        ]

        right = self.engine.library[
            self.current_pair[1]
        ]

        self.display_book(
            left,
            "left",
        )

        self.display_book(
            right,
            "right",
        )

    def display_book(
        self,
        book,
        side,
    ):
        rating = self.engine.ratings[
            book.id
        ]

        item = next(
            item
            for item in self.engine.statistics()
            if item["book"].id == book.id
        )

        status_labels = {
            "currently-reading":
                "📖 CURRENTLY READING",
            "to-read":
                "📚 TO-READ",
            "read":
                "✓ READ",
            "ignore":
                "🚫 IGNORED",
        }

        getattr(
            self,
            f"{side}_title",
        ).set(
            book.title
        )

        getattr(
            self,
            f"{side}_author",
        ).set(
            book.author
            or "Unknown author"
        )

        getattr(
            self,
            f"{side}_status",
        ).set(
            status_labels.get(
                book.status,
                book.status.upper(),
            )
        )

        meta = [
            f"Model #{item['rank']}",
            (
                f"Likely "
                f"{item['rank_low']}"
                f"–"
                f"{item['rank_high']}"
            ),
            (
                f"Top-25 "
                f"{item['top25_probability'] * 100:.0f}%"
            ),
            (
                f"Rating "
                f"{rating.rating:.0f}"
                f" ±{rating.rd:.0f}"
            ),
            (
                f"{rating.comparisons}"
                f" decisions"
            ),
        ]

        if book.pages:
            meta.insert(
                0,
                f"{book.pages} pages",
            )

        getattr(
            self,
            f"{side}_meta",
        ).set(
            "  •  ".join(meta)
        )

        self.set_description(
            getattr(
                self,
                f"{side}_description",
            ),
            (
                book.description[:700]
                if book.description
                else "No description available."
            ),
        )

    def set_description(
        self,
        widget,
        text,
    ):
        widget.configure(
            state="normal"
        )

        widget.delete(
            "1.0",
            "end",
        )

        widget.insert(
            "1.0",
            text,
        )

        widget.configure(
            state="disabled"
        )

    # --------------------------------------------------------
    # Human actions
    # --------------------------------------------------------

    def choose(self, result):
        if (
            not self.engine
            or not self.current_pair
        ):
            return

        left, right = self.current_pair

        try:
            self.engine.apply_match(
                left,
                right,
                result,
            )

            self.save_state()

            self.current_pair = None
            self.refresh()

        except Exception as exc:
            messagebox.showerror(
                "Could not record choice",
                str(exc),
            )

    def skip(self):
        if (
            not self.engine
            or not self.current_pair
        ):
            return

        self.engine.skip_pair(
            *self.current_pair
        )

        self.current_pair = None
        self.refresh()

    # --------------------------------------------------------
    # Lifecycle
    # --------------------------------------------------------

    def set_lifecycle(
        self,
        side,
        status,
    ):
        if (
            not self.engine
            or not self.current_pair
        ):
            return

        book_id = (
            self.current_pair[0]
            if side == "left"
            else self.current_pair[1]
        )

        book = self.engine.library[
            book_id
        ]

        self.engine.set_lifecycle(
            book_id,
            status,
        )

        self.save_state()

        self.current_pair = None
        self.refresh()

        self.info_var.set(
            f"{book.title}  →  "
            f"{status.replace('-', ' ').title()}"
        )

    def lifecycle_menu(
        self,
        event,
        side,
    ):
        if (
            not self.engine
            or not self.current_pair
        ):
            return

        menu = tk.Menu(
            self.root,
            tearoff=False,
            bg=self.colors["panel"],
            fg=self.colors["text"],
            activebackground=self.colors["purple"],
            activeforeground="#FFFFFF",
        )

        menu.add_command(
            label="📖 Currently Reading",
            command=lambda:
                self.set_lifecycle(
                    side,
                    "currently-reading",
                ),
        )

        menu.add_command(
            label="✓ Read",
            command=lambda:
                self.set_lifecycle(
                    side,
                    "read",
                ),
        )

        menu.add_command(
            label="🚫 Ignore",
            command=lambda:
                self.set_lifecycle(
                    side,
                    "ignore",
                ),
        )

        menu.add_command(
            label="↺ To-read",
            command=lambda:
                self.set_lifecycle(
                    side,
                    "to-read",
                ),
        )

        menu.tk_popup(
            event.x_root,
            event.y_root,
        )

    # --------------------------------------------------------
    # Undo
    # --------------------------------------------------------

    def undo(self):
        if not self.engine:
            return

        if self.engine.undo():
            self.save_state()
            self.current_pair = None
            self.refresh()

    def save_state(self):
        if (
            self.engine
            and self.state_store
        ):
            try:
                self.state_store.save(
                    self.engine
                )
            except Exception as exc:
                print(
                    "Warning: could not save state:",
                    exc,
                )

    # --------------------------------------------------------
    # Ranking window
    # --------------------------------------------------------

    def show_ranking(self):
        if not self.engine:
            messagebox.showinfo(
                "No ranking",
                "Open a Goodreads workbook first.",
            )
            return

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Ranking • uncertainty aware"
        )

        window.geometry(
            "1450x760"
        )

        window.configure(
            bg=self.colors["bg"]
        )

        metrics = (
            self.engine.confidence_metrics()
        )

        tk.Label(
            window,
            text=(
                "🏆  CURRENT READING + TOP-25"
            ),
            bg=self.colors["bg"],
            fg=self.colors["text"],
            font=(
                self.font,
                18,
                "bold",
            ),
        ).pack(
            anchor="w",
            padx=14,
            pady=(12, 2),
        )

        tk.Label(
            window,
            text=(
                f"Top-10 confidence "
                f"{metrics['top10_confidence'] * 100:.0f}%"
                f"  •  "
                f"Top-25 confidence "
                f"{metrics['top25_confidence'] * 100:.0f}%"
                f"  •  "
                f"stability "
                f"{metrics['top25_stability'] * 100:.0f}%"
                f"  •  "
                f"{metrics['unresolved_boundary']}"
                f" unresolved boundary"
            ),
            bg=self.colors["bg"],
            fg=self.colors["muted"],
            font=(
                self.font,
                9,
            ),
        ).pack(
            anchor="w",
            padx=14,
            pady=(0, 8),
        )

        columns = (
            "priority",
            "rank",
            "title",
            "author",
            "status",
            "rating",
            "rd",
            "interval",
            "top10",
            "top25",
            "decisions",
        )

        tree = ttk.Treeview(
            window,
            columns=columns,
            show="headings",
        )

        headings = {
            "priority": "Priority",
            "rank": "Rank",
            "title": "Title",
            "author": "Author",
            "status": "Status",
            "rating": "Rating",
            "rd": "RD",
            "interval": "Likely rank",
            "top10": "Top 10 %",
            "top25": "Top 25 %",
            "decisions": "Decisions",
        }

        widths = {
            "priority": 110,
            "rank": 55,
            "title": 340,
            "author": 200,
            "status": 170,
            "rating": 75,
            "rd": 65,
            "interval": 90,
            "top10": 85,
            "top25": 85,
            "decisions": 85,
        }

        for column in columns:
            tree.heading(
                column,
                text=headings[column],
            )

            tree.column(
                column,
                width=widths[column],
                anchor=(
                    "w"
                    if column
                    in {
                        "title",
                        "author",
                        "status",
                    }
                    else "center"
                ),
            )

        scrollbar = ttk.Scrollbar(
            window,
            orient="vertical",
            command=tree.yview,
        )

        tree.configure(
            yscrollcommand=
            scrollbar.set
        )

        tree.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(14, 0),
            pady=(0, 14),
        )

        scrollbar.pack(
            side="right",
            fill="y",
            padx=(0, 14),
            pady=(0, 14),
        )

        statistics = (
            self.engine.statistics()
        )

        current = [
            item
            for item in statistics
            if item["book"].status
            == "currently-reading"
        ]

        rest = [
            item
            for item in statistics
            if item["book"].status
            != "currently-reading"
        ]

        # Currently-reading is always displayed first.
        for item in current + rest:
            book = item["book"]
            rating = item["rating"]

            if book.status == "currently-reading":
                priority = "📖 NOW"
            elif item["rank"] <= TOP_K:
                priority = "🏆 TOP 25"
            else:
                priority = "📚 QUEUE"

            tree.insert(
                "",
                "end",
                values=(
                    priority,
                    item["rank"],
                    book.title,
                    book.author,
                    book.status,
                    f"{rating.rating:.0f}",
                    f"±{rating.rd:.0f}",
                    (
                        f"{item['rank_low']}"
                        f"–"
                        f"{item['rank_high']}"
                    ),
                    (
                        f"{item['top10_probability'] * 100:.0f}%"
                    ),
                    (
                        f"{item['top25_probability'] * 100:.0f}%"
                    ),
                    rating.comparisons,
                ),
            )

    # --------------------------------------------------------
    # Shortcuts dialog
    # --------------------------------------------------------

    def show_shortcuts(self):
        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Keyboard shortcuts"
        )

        window.geometry(
            "760x650"
        )

        window.configure(
            bg=self.colors["bg"]
        )

        tk.Label(
            window,
            text="⌨  FULL KEYBOARD REFERENCE",
            bg=self.colors["bg"],
            fg=self.colors["text"],
            font=(
                self.font,
                18,
                "bold",
            ),
        ).pack(
            anchor="w",
            padx=20,
            pady=16,
        )

        shortcuts = [
            (
                "1",
                "Choose LEFT",
            ),
            (
                "2",
                "Choose RIGHT",
            ),
            (
                "3 / T",
                "Tie",
            ),
            (
                "4 / S",
                "Skip",
            ),
            (
                "← / →",
                "Choose left / right",
            ),
            (
                "U",
                "Undo",
            ),
            (
                "I",
                "Mark LEFT Ignore",
            ),
            (
                "R",
                "Mark LEFT Read",
            ),
            (
                "C",
                "Mark LEFT Currently Reading",
            ),
            (
                "Shift+I",
                "Mark RIGHT Ignore",
            ),
            (
                "Shift+R",
                "Mark RIGHT Read",
            ),
            (
                "Shift+C",
                "Mark RIGHT Currently Reading",
            ),
            (
                "Ctrl+O",
                "Open Goodreads Excel",
            ),
            (
                "Ctrl+E",
                "Export Ranking",
            ),
            (
                "Click book card",
                "Open lifecycle menu",
            ),
        ]

        for key, description in shortcuts:
            row = tk.Frame(
                window,
                bg=self.colors["panel"],
            )

            row.pack(
                fill="x",
                padx=20,
                pady=2,
            )

            tk.Label(
                row,
                text=key,
                bg=self.colors["panel"],
                fg=self.colors["accent"],
                font=(
                    "Consolas",
                    10,
                    "bold",
                ),
                width=20,
                anchor="w",
                padx=10,
                pady=6,
            ).pack(
                side="left"
            )

            tk.Label(
                row,
                text=description,
                bg=self.colors["panel"],
                fg=self.colors["text"],
                font=(
                    self.font,
                    9,
                ),
                anchor="w",
                padx=10,
                pady=6,
            ).pack(
                side="left",
                fill="x",
                expand=True,
            )

    # --------------------------------------------------------
    # Export
    # --------------------------------------------------------

    def export(self):
        if (
            not self.engine
            or not self.source_file
        ):
            messagebox.showinfo(
                "Nothing to export",
                "Open a Goodreads workbook first.",
            )
            return

        try:
            output = export_results(
                self.source_file,
                self.engine,
            )

            self.save_state()

            if messagebox.askyesno(
                "Export complete",
                (
                    f"Saved:\n\n"
                    f"{output}\n\n"
                    f"Open it now?"
                ),
            ):
                self.open_external(
                    output
                )

        except PermissionError:
            messagebox.showerror(
                "Export failed",
                (
                    "Windows could not save "
                    "the workbook. Close it "
                    "in Excel and try again."
                ),
            )

        except Exception as exc:
            messagebox.showerror(
                "Export failed",
                str(exc),
            )

    def overwrite_source(self):
        if (
            not self.engine
            or not self.source_file
        ):
            return

        if not messagebox.askyesno(
            "Overwrite Goodreads source?",
            (
                "This will replace the original "
                "Goodreads workbook.\n\n"
                "A timestamped backup will be "
                "created first.\n\n"
                "Continue?"
            ),
        ):
            return

        timestamp = (
            datetime.now()
            .strftime(
                "%Y%m%d-%H%M%S"
            )
        )

        backup = (
            self.source_file.with_name(
                self.source_file.stem
                + ".backup-"
                + timestamp
                + self.source_file.suffix
            )
        )

        try:
            shutil.copy2(
                self.source_file,
                backup,
            )

            export_results(
                self.source_file,
                self.engine,
                output=self.source_file,
            )

            self.save_state()

            messagebox.showinfo(
                "Source updated",
                (
                    "The Goodreads workbook "
                    "was updated.\n\n"
                    f"Backup:\n{backup}"
                ),
            )

        except PermissionError:
            messagebox.showerror(
                "Overwrite failed",
                (
                    "Close the workbook in Excel "
                    "and try again."
                ),
            )

        except Exception as exc:
            messagebox.showerror(
                "Overwrite failed",
                str(exc),
            )

    # --------------------------------------------------------
    # External file
    # --------------------------------------------------------

    def open_external(self, path):
        try:
            if sys.platform.startswith("win"):
                os.startfile(
                    str(path)
                )

            elif sys.platform == "darwin":
                import subprocess

                subprocess.Popen(
                    ["open", str(path)]
                )

            else:
                import subprocess

                subprocess.Popen(
                    ["xdg-open", str(path)]
                )

        except Exception:
            pass

    # --------------------------------------------------------
    # Close
    # --------------------------------------------------------

    def close(self):
        self.save_state()
        self.root.destroy()


# ============================================================
# SELF TEST
# ============================================================

def run_self_test():
    started = datetime.now()

    # Glicko direction.
    winner = Rating()
    loser = Rating()

    winner_after = glicko_update(
        winner,
        loser,
        1.0,
    )

    loser_after = glicko_update(
        loser,
        winner,
        0.0,
    )

    assert (
        winner_after.rating
        > winner.rating
    )

    assert (
        loser_after.rating
        < loser.rating
    )

    # Small library.
    books = [
        Book(
            id=str(index),
            title=f"Book {index}",
            author=f"Author {index % 5}",
        )
        for index in range(30)
    ]

    engine = RankingEngine(
        books,
        mode="TOP_25_FOCUS",
        seed=12345,
    )

    seen = set()

    for index in range(25):
        pair = engine.choose_pair()

        assert pair is not None

        pair_key = engine.pair_key(
            *pair
        )

        assert pair_key not in seen

        seen.add(pair_key)

        result = (
            "left"
            if index % 3
            else "right"
        )

        engine.apply_match(
            pair[0],
            pair[1],
            result,
        )

    assert (
        len(engine.comparisons)
        == 25
    )

    # Undo.
    before = len(
        engine.comparisons
    )

    assert engine.undo()

    assert (
        len(engine.comparisons)
        == before - 1
    )

    # Lifecycle.
    engine.set_lifecycle(
        "0",
        "currently-reading",
    )

    assert any(
        book.id == "0"
        for book in engine.books
    )

    engine.set_lifecycle(
        "1",
        "ignore",
    )

    assert all(
        book.id != "1"
        for book in engine.books
    )

    engine.set_lifecycle(
        "1",
        "to-read",
    )

    assert any(
        book.id == "1"
        for book in engine.books
    )

    # Large-library bounded selection.
    large_books = [
        Book(
            id=f"large-{index}",
            title=f"Large Book {index}",
        )
        for index in range(5000)
    ]

    large_engine = RankingEngine(
        large_books,
        seed=7,
        mode="TOP_25_FOCUS",
    )

    pair = large_engine.choose_pair()

    assert pair is not None

    # Persistence round trip.
    with tempfile.TemporaryDirectory() as directory:
        source = (
            Path(directory)
            / "library.xlsx"
        )

        workbook = load_workbook(
            source
        ) if source.exists() else None

        if workbook:
            workbook.close()

        temporary_workbook = (
            __import__("openpyxl")
            .Workbook()
        )

        temporary_workbook.save(
            source
        )

        temporary_workbook.close()

        store = StateStore(
            source
        )

        store.save(
            engine
        )

        loaded_state = (
            store.load_raw()
        )

        assert loaded_state is not None

        assert (
            loaded_state["state_version"]
            == STATE_VERSION
        )

        assert (
            len(
                loaded_state[
                    "comparisons"
                ]
            )
            == len(
                engine.comparisons
            )
        )

    elapsed = (
        datetime.now()
        - started
    ).total_seconds()

    print(
        f"SELF-TEST PASSED in {elapsed:.2f}s"
    )

    return 0


# ============================================================
# MAIN
# ============================================================

def main():
    if "--self-test" in sys.argv:
        return run_self_test()

    root = tk.Tk()

    RankerApp(
        root
    )

    root.mainloop()

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(
            main()
        )

    except Exception:
        error = traceback.format_exc()

        print(error)

        try:
            messagebox.showerror(
                "Unexpected error",
                (
                    "The application encountered "
                    "an unexpected error:\n\n"
                    + error
                ),
            )

        except Exception:
            pass