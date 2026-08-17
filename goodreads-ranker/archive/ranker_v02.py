from __future__ import annotations

import copy
import json
import math
import os
import random
import re
import subprocess
import sys
import traceback
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

APP_NAME = "Goodreads To-Read Ranker"
APP_VERSION = "3.0"

# ------------------------------------------------------------
# Glicko defaults
# ------------------------------------------------------------

DEFAULT_RATING = 1500.0
DEFAULT_RD = 350.0
DEFAULT_VOLATILITY = 0.06

GLICKO_SCALE = 173.7178
TAU = 0.5

# ------------------------------------------------------------
# Comparison budgets
#
# These are MAXIMUM average-ish targets, not rigid quotas.
# The adaptive engine can finish earlier when the ranking is
# sufficiently stable.
# ------------------------------------------------------------

MIN_TARGET_COMPARISONS = 3
MAX_TARGET_COMPARISONS = 20

QUICK_COMPARISONS = 3
BALANCED_COMPARISONS = 6
ACCURATE_COMPARISONS = 12

DEFAULT_TARGET_COMPARISONS = BALANCED_COMPARISONS

# ------------------------------------------------------------
# Adaptive ranking
# ------------------------------------------------------------

MIN_INITIAL_COMPARISONS = 2

# Refit Bradley-Terry every N decisions.
BT_REFRESH_INTERVAL = 12

# Recalculate stability every N decisions.
STABILITY_INTERVAL = 12

# Minimum total decisions before adaptive early stopping.
MIN_ADAPTIVE_DECISIONS = 50

# How stable the ranking should be before allowing an
# automatic early finish.
DEFAULT_STABILITY_THRESHOLD = 0.985

# More stringent requirement for the very top of the ranking.
TOP_STABILITY_THRESHOLD = 0.975

# Don't stop until at least this many books have enough evidence.
MIN_COVERAGE_FOR_STOP = 0.92

# Maximum number of candidate pairs examined when selecting
# the next pair.
PAIR_CANDIDATE_LIMIT = 5000

# Small exploration probability prevents the active learner
# from becoming too deterministic.
EXPLORATION_RATE = 0.055

# Every so often, force a long-range comparison to keep the
# comparison graph connected.
LONG_RANGE_RATE = 0.025

# ------------------------------------------------------------
# Bradley-Terry regularization
# ------------------------------------------------------------

BT_MAX_ITERATIONS = 80
BT_TOLERANCE = 1e-5

# L2 penalty on latent preference strength.
# Larger = more conservative / less extreme.
BT_REGULARIZATION = 0.035

# Convert BT latent strength into a familiar rating scale.
BT_RATING_SCALE = 173.7178

# ------------------------------------------------------------
# Goodreads star prior
# ------------------------------------------------------------

USE_GOODREADS_PRIOR = True

# Strength of prior from My Rating.
#
# This is deliberately modest. Pairwise preferences should
# dominate the final ranking.
GOODREADS_PRIOR_STRENGTH = 0.15

# ------------------------------------------------------------
# State
# ------------------------------------------------------------

STATE_DIRECTORY_NAME = ".ranker_state"
STATE_VERSION = 5

# ------------------------------------------------------------
# UI
# ------------------------------------------------------------

WINDOW_WIDTH = 1280
WINDOW_HEIGHT = 850


# ============================================================
# GENERAL UTILITIES
# ============================================================

def normalize(value) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def normalize_header(value) -> str:
    return normalize(value).replace("\ufeff", "").strip().lower()


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def safe_filename(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*]', "_", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value or "ranking"


def truncate(text: str, length: int) -> str:
    text = normalize(text)

    if len(text) <= length:
        return text

    return text[:length - 1].rstrip() + "…"


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def safe_exp(value: float) -> float:
    if value > 700:
        return math.exp(700)

    if value < -700:
        return math.exp(-700)

    return math.exp(value)


def sigmoid(value: float) -> float:
    if value >= 0:
        z = math.exp(-min(value, 700))
        return 1.0 / (1.0 + z)

    z = math.exp(max(value, -700))
    return z / (1.0 + z)


def parse_goodreads_rating(value: str) -> Optional[float]:
    """
    Goodreads "My Rating" is normally an integer from 0 to 5.
    Return None when unavailable/invalid.
    """
    value = normalize(value)

    if not value:
        return None

    try:
        rating = float(value)
    except ValueError:
        return None

    if rating <= 0:
        return None

    return clamp(rating, 1.0, 5.0)


# ============================================================
# GLICKO-2
# ============================================================

@dataclass
class Rating:
    rating: float = DEFAULT_RATING
    rd: float = DEFAULT_RD
    volatility: float = DEFAULT_VOLATILITY

    comparisons: int = 0
    wins: int = 0
    losses: int = 0
    ties: int = 0


def to_glicko(rating: Rating):
    return (
        (rating.rating - 1500.0) / GLICKO_SCALE,
        rating.rd / GLICKO_SCALE,
    )


def from_glicko(
    mu: float,
    phi: float,
    volatility: float,
) -> Rating:
    return Rating(
        rating=1500.0 + GLICKO_SCALE * mu,
        rd=max(1.0, GLICKO_SCALE * phi),
        volatility=volatility,
    )


def g_function(phi: float) -> float:
    return 1.0 / math.sqrt(
        1.0
        + 3.0 * phi * phi / (math.pi * math.pi)
    )


def expected_score(
    mu: float,
    opponent_mu: float,
    opponent_phi: float,
) -> float:
    exponent = -g_function(opponent_phi) * (
        mu - opponent_mu
    )

    if exponent > 700:
        return 0.0

    if exponent < -700:
        return 1.0

    return 1.0 / (1.0 + math.exp(exponent))


def volatility_function(
    x: float,
    delta: float,
    phi: float,
    variance: float,
    a: float,
) -> float:
    exp_x = safe_exp(x)

    base = (
        phi * phi
        + variance
        + exp_x
    )

    denominator = 2.0 * base * base

    numerator = exp_x * (
        delta * delta
        - phi * phi
        - variance
        - exp_x
    )

    return (
        numerator / denominator
        - (x - a) / (TAU * TAU)
    )


def calculate_volatility(
    phi: float,
    volatility: float,
    delta: float,
    variance: float,
) -> float:
    a = math.log(
        max(volatility * volatility, 1e-12)
    )

    A = a

    if delta * delta > phi * phi + variance:
        B = math.log(
            max(
                delta * delta
                - phi * phi
                - variance,
                1e-12,
            )
        )

    else:
        k = 1

        while True:
            B = a - k * TAU

            value = volatility_function(
                B,
                delta,
                phi,
                variance,
                a,
            )

            if value >= 0:
                break

            k += 1

            if k > 100:
                B = a - 100 * TAU
                break

    f_a = volatility_function(
        A,
        delta,
        phi,
        variance,
        a,
    )

    f_b = volatility_function(
        B,
        delta,
        phi,
        variance,
        a,
    )

    for _ in range(100):
        if abs(B - A) < 1e-8:
            break

        denominator = f_b - f_a

        if abs(denominator) < 1e-15:
            break

        C = A + (
            (A - B) * f_a / denominator
        )

        f_c = volatility_function(
            C,
            delta,
            phi,
            variance,
            a,
        )

        if f_c * f_b < 0:
            A = B
            f_a = f_b

        else:
            f_a /= 2.0

        B = C
        f_b = f_c

    result = math.exp(A / 2.0)

    return clamp(
        result,
        0.01,
        1.0,
    )


def glicko_update(
    player: Rating,
    opponents: list[Rating],
    scores: list[float],
) -> Rating:
    if not opponents:
        result = copy.deepcopy(player)

        phi = player.rd / GLICKO_SCALE

        phi_star = math.sqrt(
            phi * phi
            + player.volatility
            * player.volatility
        )

        result.rd = min(
            GLICKO_SCALE * phi_star,
            350.0,
        )

        return result

    mu, phi = to_glicko(player)

    variance_inverse = 0.0
    score_sum = 0.0

    for opponent, score in zip(
        opponents,
        scores,
    ):
        opponent_mu, opponent_phi = (
            to_glicko(opponent)
        )

        g_value = g_function(
            opponent_phi
        )

        expected = expected_score(
            mu,
            opponent_mu,
            opponent_phi,
        )

        variance_inverse += (
            g_value
            * g_value
            * expected
            * (1.0 - expected)
        )

        score_sum += (
            g_value
            * (score - expected)
        )

    if variance_inverse <= 0:
        return copy.deepcopy(player)

    variance = 1.0 / variance_inverse

    delta = variance * score_sum

    new_volatility = calculate_volatility(
        phi,
        player.volatility,
        delta,
        variance,
    )

    phi_star = math.sqrt(
        phi * phi
        + new_volatility
        * new_volatility
    )

    new_phi = 1.0 / math.sqrt(
        1.0 / (phi_star * phi_star)
        + 1.0 / variance
    )

    new_mu = (
        mu
        + new_phi * new_phi * score_sum
    )

    return from_glicko(
        new_mu,
        new_phi,
        new_volatility,
    )


# ============================================================
# BOOK MODEL
# ============================================================

@dataclass
class Book:
    id: str
    title: str
    author: str

    pages: str = ""
    year: str = ""
    publication_year: str = ""
    description: str = ""
    my_rating: str = ""
    isbn: str = ""
    book_id: str = ""

    original_row: int = 0
    original: dict = None

    def __post_init__(self):
        if self.original is None:
            self.original = {}


# ============================================================
# GOODREADS IMPORTER
# ============================================================

def find_header(headers: list, wanted: str):
    wanted = normalize_header(wanted)

    for header in headers:
        if normalize_header(header) == wanted:
            return header

    return None


def find_first_header(
    headers: list,
    candidates: list[str],
):
    for candidate in candidates:
        found = find_header(
            headers,
            candidate,
        )

        if found is not None:
            return found

    return None


def load_goodreads(path: Path):
    if not path.exists():
        raise FileNotFoundError(
            f"File not found:\n{path}"
        )

    if path.suffix.lower() != ".xlsx":
        raise ValueError(
            "Please select a Goodreads .xlsx export."
        )

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook.active

        rows = worksheet.iter_rows(
            values_only=True
        )

        try:
            headers = list(next(rows))

        except StopIteration:
            raise ValueError(
                "The Excel file is empty."
            )

        shelf_header = find_header(
            headers,
            "Exclusive Shelf",
        )

        title_header = find_header(
            headers,
            "Title",
        )

        if shelf_header is None:
            raise ValueError(
                'Could not find the "Exclusive Shelf" column.'
            )

        if title_header is None:
            raise ValueError(
                'Could not find the "Title" column.'
            )

        author_header = find_first_header(
            headers,
            [
                "Author l-f",
                "Author",
                "Authors",
                "Author Name",
            ],
        )

        header_index = {
            header: index
            for index, header in enumerate(headers)
            if header is not None
        }

        def get_value(values, header):
            if header is None:
                return ""

            index = header_index.get(header)

            if index is None:
                return ""

            if index >= len(values):
                return ""

            return normalize(
                values[index]
            )

        pages_header = find_header(
            headers,
            "Pages",
        )

        year_header = find_header(
            headers,
            "Year Published",
        )

        original_year_header = find_header(
            headers,
            "Original Publication Year",
        )

        description_header = find_header(
            headers,
            "Description",
        )

        my_rating_header = find_header(
            headers,
            "My Rating",
        )

        isbn_header = find_header(
            headers,
            "ISBN",
        )

        book_id_header = find_header(
            headers,
            "Book Id - Goodreads",
        )

        first_name_header = find_header(
            headers,
            "Author First Name",
        )

        last_name_header = find_header(
            headers,
            "Author Last Name",
        )

        books = []

        excel_row_number = 2

        for raw_values in rows:
            values = list(raw_values)

            while len(values) < len(headers):
                values.append(None)

            shelf = get_value(
                values,
                shelf_header,
            ).lower()

            if shelf != "to-read":
                excel_row_number += 1
                continue

            title = get_value(
                values,
                title_header,
            )

            if not title:
                excel_row_number += 1
                continue

            author = ""

            if author_header is not None:
                author = get_value(
                    values,
                    author_header,
                )

            if not author:
                first_name = get_value(
                    values,
                    first_name_header,
                )

                last_name = get_value(
                    values,
                    last_name_header,
                )

                author = " ".join(
                    part
                    for part in (
                        first_name,
                        last_name,
                    )
                    if part
                )

            goodreads_id = get_value(
                values,
                book_id_header,
            )

            if goodreads_id:
                book_id = (
                    "goodreads:"
                    + goodreads_id
                )

            else:
                book_id = (
                    "row:"
                    + str(excel_row_number)
                )

            original = {}

            for index, header in enumerate(
                headers
            ):
                if header is None:
                    continue

                value = (
                    values[index]
                    if index < len(values)
                    else None
                )

                if isinstance(
                    value,
                    datetime,
                ):
                    value = value.isoformat()

                original[str(header)] = value

            books.append(
                Book(
                    id=book_id,
                    title=title,
                    author=author,
                    pages=get_value(
                        values,
                        pages_header,
                    ),
                    year=get_value(
                        values,
                        year_header,
                    ),
                    publication_year=get_value(
                        values,
                        original_year_header,
                    ),
                    description=get_value(
                        values,
                        description_header,
                    ),
                    my_rating=get_value(
                        values,
                        my_rating_header,
                    ),
                    isbn=get_value(
                        values,
                        isbn_header,
                    ),
                    book_id=goodreads_id,
                    original_row=excel_row_number,
                    original=original,
                )
            )

            excel_row_number += 1

        return headers, books

    finally:
        workbook.close()


# ============================================================
# RANKING ENGINE
# ============================================================

class RankingEngine:
    """
    Hybrid adaptive ranking engine.

    Architecture:

        1. Glicko-2
           Used for fast online uncertainty updates.

        2. Swiss-style exploration
           Gives every book initial evidence without wasting
           comparisons on random pairs.

        3. Active pair selection
           Focuses on uncertain and rank-sensitive pairs.

        4. Regularized Bradley-Terry
           Periodically fits the entire comparison history
           into one global preference model.

        5. Adaptive stopping
           Stops when the ranking has become stable instead
           of blindly forcing every book to receive N matches.

    This is deliberately designed for large Goodreads shelves.
    """

    def __init__(
        self,
        books: list[Book],
        target_comparisons: int = DEFAULT_TARGET_COMPARISONS,
        random_seed: Optional[int] = None,
    ):
        self.books = books

        self.target_comparisons = max(
            MIN_TARGET_COMPARISONS,
            min(
                MAX_TARGET_COMPARISONS,
                int(target_comparisons),
            ),
        )

        self.ratings = {
            book.id: Rating()
            for book in books
        }

        self.comparisons = []

        # ----------------------------------------------------
        # Cached state
        # ----------------------------------------------------

        self.counts = {
            book.id: 0
            for book in books
        }

        self.played = set()

        self.wins = {
            book.id: 0.0
            for book in books
        }

        # ----------------------------------------------------
        # Model cache
        # ----------------------------------------------------

        self.bt_strengths = {
            book.id: 1.0
            for book in books
        }

        self.bt_ratings = {
            book.id: DEFAULT_RATING
            for book in books
        }

        self.bt_dirty = True
        self.last_bt_comparison_count = 0

        # ----------------------------------------------------
        # Ranking stability
        # ----------------------------------------------------

        self.last_stable_order = []
        self.stability = 0.0
        self.top_stability = 0.0

        self.last_stability_check = 0

        # ----------------------------------------------------
        # Active-learning state
        # ----------------------------------------------------

        self.random = random.Random(
            random_seed
        )

        self.seed = random_seed

        self.selection_counter = 0

        self.phase = "exploration"

        # ----------------------------------------------------
        # Fast undo
        # ----------------------------------------------------

        self.undo_snapshots = []

        # ----------------------------------------------------
        # Statistics
        # ----------------------------------------------------

        self.pair_selection_count = 0

        self.rejected_selection_count = 0

        self.rebuild_indexes()

    # ========================================================
    # BASIC INDEXES
    # ========================================================

    def book_map(self):
        return {
            book.id: book
            for book in self.books
        }

    @staticmethod
    def pair_key(
        first: str,
        second: str,
    ) -> str:
        return "|".join(
            sorted([first, second])
        )

    def rebuild_indexes(self):
        self.counts = {
            book.id: 0
            for book in self.books
        }

        self.played = set()

        self.wins = {
            book.id: 0.0
            for book in self.books
        }

        for match in self.comparisons:
            left = match.get("left")
            right = match.get("right")
            result = match.get("result")

            if (
                left not in self.ratings
                or right not in self.ratings
            ):
                continue

            self.counts[left] += 1
            self.counts[right] += 1

            self.played.add(
                self.pair_key(
                    left,
                    right,
                )
            )

            if result == "left":
                self.wins[left] += 1.0

            elif result == "right":
                self.wins[right] += 1.0

            elif result == "tie":
                self.wins[left] += 0.5
                self.wins[right] += 0.5

        self.bt_dirty = True

    def played_pairs(self):
        return self.played

    def comparison_counts(self):
        return self.counts

    # ========================================================
    # SNAPSHOTS / UNDO
    # ========================================================

    def make_snapshot(self):
        return {
            "ratings": copy.deepcopy(
                self.ratings
            ),
            "comparisons": copy.deepcopy(
                self.comparisons
            ),
            "counts": copy.deepcopy(
                self.counts
            ),
            "played": copy.deepcopy(
                self.played
            ),
            "wins": copy.deepcopy(
                self.wins
            ),
            "bt_strengths": copy.deepcopy(
                self.bt_strengths
            ),
            "bt_ratings": copy.deepcopy(
                self.bt_ratings
            ),
            "bt_dirty": self.bt_dirty,
            "last_bt_comparison_count": (
                self.last_bt_comparison_count
            ),
            "stability": self.stability,
            "top_stability": self.top_stability,
            "last_stable_order": copy.deepcopy(
                self.last_stable_order
            ),
            "phase": self.phase,
            "selection_counter": (
                self.selection_counter
            ),
        }

    def restore_snapshot(self, snapshot):
        self.ratings = copy.deepcopy(
            snapshot["ratings"]
        )

        self.comparisons = copy.deepcopy(
            snapshot["comparisons"]
        )

        self.counts = copy.deepcopy(
            snapshot["counts"]
        )

        self.played = copy.deepcopy(
            snapshot["played"]
        )

        self.wins = copy.deepcopy(
            snapshot["wins"]
        )

        self.bt_strengths = copy.deepcopy(
            snapshot["bt_strengths"]
        )

        self.bt_ratings = copy.deepcopy(
            snapshot["bt_ratings"]
        )

        self.bt_dirty = snapshot["bt_dirty"]

        self.last_bt_comparison_count = (
            snapshot[
                "last_bt_comparison_count"
            ]
        )

        self.stability = snapshot[
            "stability"
        ]

        self.top_stability = snapshot[
            "top_stability"
        ]

        self.last_stable_order = copy.deepcopy(
            snapshot[
                "last_stable_order"
            ]
        )

        self.phase = snapshot["phase"]

        self.selection_counter = (
            snapshot[
                "selection_counter"
            ]
        )

    def undo(self) -> bool:
        if not self.undo_snapshots:
            return False

        snapshot = self.undo_snapshots.pop()

        self.restore_snapshot(
            snapshot
        )

        return True

    # ========================================================
    # GOODREADS PRIOR
    # ========================================================

    def goodreads_prior_beta(
        self,
        book: Book,
    ) -> float:
        if not USE_GOODREADS_PRIOR:
            return 0.0

        stars = parse_goodreads_rating(
            book.my_rating
        )

        if stars is None:
            return 0.0

        # Convert 1-5 stars to approximately
        # -2 .. +2 latent preference strength.
        centered = stars - 3.0

        return (
            centered
            * GOODREADS_PRIOR_STRENGTH
        )

    # ========================================================
    # MATCH APPLICATION
    # ========================================================

    def apply_match(
        self,
        left_id: str,
        right_id: str,
        result: str,
    ):
        if left_id not in self.ratings:
            raise ValueError(
                "Unknown left book."
            )

        if right_id not in self.ratings:
            raise ValueError(
                "Unknown right book."
            )

        if left_id == right_id:
            raise ValueError(
                "A book cannot be compared with itself."
            )

        if result not in (
            "left",
            "right",
            "tie",
        ):
            raise ValueError(
                "Invalid comparison result."
            )

        pair = self.pair_key(
            left_id,
            right_id,
        )

        if pair in self.played:
            raise ValueError(
                "That pair has already been compared."
            )

        # ----------------------------------------------------
        # Snapshot before mutation.
        # ----------------------------------------------------

        self.undo_snapshots.append(
            self.make_snapshot()
        )

        # Keep undo memory bounded.
        if len(self.undo_snapshots) > 100:
            self.undo_snapshots.pop(0)

        old_left = copy.deepcopy(
            self.ratings[left_id]
        )

        old_right = copy.deepcopy(
            self.ratings[right_id]
        )

        if result == "left":
            left_score = 1.0
            right_score = 0.0

        elif result == "right":
            left_score = 0.0
            right_score = 1.0

        else:
            left_score = 0.5
            right_score = 0.5

        new_left = glicko_update(
            old_left,
            [old_right],
            [left_score],
        )

        new_right = glicko_update(
            old_right,
            [old_left],
            [right_score],
        )

        new_left.comparisons = (
            old_left.comparisons + 1
        )

        new_right.comparisons = (
            old_right.comparisons + 1
        )

        new_left.wins = old_left.wins
        new_left.losses = old_left.losses
        new_left.ties = old_left.ties

        new_right.wins = old_right.wins
        new_right.losses = old_right.losses
        new_right.ties = old_right.ties

        if result == "left":
            new_left.wins += 1
            new_right.losses += 1

        elif result == "right":
            new_right.wins += 1
            new_left.losses += 1

        else:
            new_left.ties += 1
            new_right.ties += 1

        self.ratings[left_id] = new_left
        self.ratings[right_id] = new_right

        self.counts[left_id] += 1
        self.counts[right_id] += 1

        self.played.add(pair)

        if result == "left":
            self.wins[left_id] += 1.0

        elif result == "right":
            self.wins[right_id] += 1.0

        else:
            self.wins[left_id] += 0.5
            self.wins[right_id] += 0.5

        self.comparisons.append(
            {
                "left": left_id,
                "right": right_id,
                "result": result,
                "timestamp": now_iso(),
            }
        )

        self.selection_counter += 1

        self.bt_dirty = True

        self.update_phase()

        # Refit global model periodically.
        if (
            len(self.comparisons)
            - self.last_bt_comparison_count
            >= BT_REFRESH_INTERVAL
        ):
            self.fit_bradley_terry()

        # Recalculate stability periodically.
        if (
            len(self.comparisons)
            - self.last_stability_check
            >= STABILITY_INTERVAL
        ):
            self.calculate_stability()

    # ========================================================
    # HISTORY REBUILD
    # ========================================================

    def rebuild_from_history(self):
        history = copy.deepcopy(
            self.comparisons
        )

        self.ratings = {
            book.id: Rating()
            for book in self.books
        }

        self.comparisons = []

        self.rebuild_indexes()

        self.undo_snapshots = []

        for match in history:
            try:
                self.apply_match(
                    match["left"],
                    match["right"],
                    match["result"],
                )

            except Exception:
                continue

        self.undo_snapshots = []

        self.fit_bradley_terry()
        self.calculate_stability()

    # ========================================================
    # PHASE MANAGEMENT
    # ========================================================

    def update_phase(self):
        if not self.books:
            self.phase = "complete"
            return

        if self.minimum_comparisons() < MIN_INITIAL_COMPARISONS:
            self.phase = "exploration"
            return

        if len(self.comparisons) < max(
            MIN_ADAPTIVE_DECISIONS,
            len(self.books),
        ):
            self.phase = "calibration"
            return

        self.phase = "active"

    def minimum_comparisons(self) -> int:
        if not self.books:
            return 0

        return min(
            self.counts[book.id]
            for book in self.books
        )

    def average_comparisons(self) -> float:
        if not self.books:
            return 0.0

        return (
            sum(
                self.counts[book.id]
                for book in self.books
            )
            / len(self.books)
        )

    def coverage(self) -> float:
        if not self.books:
            return 0.0

        covered = sum(
            1
            for book in self.books
            if self.counts[book.id]
            >= MIN_INITIAL_COMPARISONS
        )

        return covered / len(self.books)

    # ========================================================
    # SWISS-STYLE EXPLORATION
    # ========================================================

    def choose_exploration_pair(self):
        """
        During early ranking, prioritize books with little/no
        evidence and pair them with similarly experienced books.

        This is effectively a lightweight Swiss-style design:
        coverage first, then progressively stronger local
        comparisons.
        """

        undercovered = [
            book
            for book in self.books
            if self.counts[book.id]
            < MIN_INITIAL_COMPARISONS
        ]

        if len(undercovered) < 2:
            return None

        minimum = min(
            self.counts[book.id]
            for book in undercovered
        )

        pool = [
            book
            for book in undercovered
            if self.counts[book.id]
            <= minimum + 1
        ]

        self.random.shuffle(pool)

        # Prefer rating proximity once enough evidence exists.
        primary = pool[0]

        candidates = []

        for opponent in pool[1:]:
            pair = self.pair_key(
                primary.id,
                opponent.id,
            )

            if pair in self.played:
                continue

            distance = abs(
                self.ratings[
                    primary.id
                ].rating
                - self.ratings[
                    opponent.id
                ].rating
            )

            score = (
                -distance
                + self.random.random() * 80.0
            )

            candidates.append(
                (
                    score,
                    opponent,
                )
            )

        if not candidates:
            return None

        candidates.sort(
            key=lambda item: item[0],
            reverse=True,
        )

        return (
            primary.id,
            candidates[0][1].id,
        )

    # ========================================================
    # ACTIVE LEARNING
    # ========================================================

    def probability_left_wins(
        self,
        left: Rating,
        right: Rating,
    ) -> float:
        """
        Glicko-inspired pair probability.

        RD is folded into the effective uncertainty so that
        very uncertain books are more likely to be selected
        for useful comparisons.
        """

        rating_difference = (
            left.rating - right.rating
        )

        uncertainty = math.sqrt(
            1.0
            + (
                left.rd * left.rd
                + right.rd * right.rd
            )
            / (
                350.0 * 350.0
            )
        )

        effective_difference = (
            rating_difference
            / uncertainty
        )

        return sigmoid(
            effective_difference
            / 173.7178
        )

    def pair_information_score(
        self,
        left: Book,
        right: Book,
        ranks: dict[str, int],
    ) -> float:
        """
        Acquisition function.

        High score means:

        * outcome is uncertain
        * both books are uncertain
        * pair is close in ranking
        * pair could affect the ordering
        * neither book is already over-sampled

        This is intentionally a practical approximation of
        expected-information-gain style active sampling.
        """

        left_rating = self.ratings[
            left.id
        ]

        right_rating = self.ratings[
            right.id
        ]

        probability = (
            self.probability_left_wins(
                left_rating,
                right_rating,
            )
        )

        # Binary entropy, 0..1.
        if (
            probability <= 0.0
            or probability >= 1.0
        ):
            entropy = 0.0

        else:
            entropy = -(
                probability
                * math.log2(probability)
                + (
                    1.0 - probability
                )
                * math.log2(
                    1.0 - probability
                )
            )

        combined_rd = (
            left_rating.rd
            + right_rating.rd
        ) / 2.0

        uncertainty_factor = clamp(
            combined_rd / 250.0,
            0.25,
            2.0,
        )

        left_rank = ranks.get(
            left.id,
            len(self.books) // 2,
        )

        right_rank = ranks.get(
            right.id,
            len(self.books) // 2,
        )

        rank_gap = abs(
            left_rank - right_rank
        )

        # Nearby ranks are much more consequential.
        boundary_factor = (
            1.0
            / (
                1.0
                + rank_gap / 12.0
            )
        )

        # Extra emphasis near the top.
        top_factor = 1.0

        best_rank = min(
            left_rank,
            right_rank,
        )

        if best_rank <= 10:
            top_factor = 1.65

        elif best_rank <= 25:
            top_factor = 1.35

        elif best_rank <= 100:
            top_factor = 1.10

        # Avoid repeatedly selecting books that already have
        # lots of evidence.
        average_count = (
            self.counts[left.id]
            + self.counts[right.id]
        ) / 2.0

        evidence_penalty = 1.0 / (
            1.0
            + max(
                0.0,
                average_count
                - self.average_comparisons(),
            )
            * 0.12
        )

        # Mild preference for books with large RD.
        rd_bonus = (
            1.0
            + combined_rd / 500.0
        )

        score = (
            entropy
            * uncertainty_factor
            * boundary_factor
            * top_factor
            * evidence_penalty
            * rd_bonus
        )

        return score

    def choose_active_pair(self):
        if len(self.books) < 2:
            return None

        ranking = self.ranking()

        ranks = {
            item["book"].id: index + 1
            for index, item in enumerate(
                ranking
            )
        }

        # ----------------------------------------------------
        # Strategic long-range exploration.
        # ----------------------------------------------------

        if (
            self.random.random()
            < LONG_RANGE_RATE
            and len(self.books) >= 4
        ):
            sorted_books = [
                item["book"]
                for item in ranking
            ]

            left_index = self.random.randrange(
                0,
                max(1, len(sorted_books) // 4),
            )

            right_index = self.random.randrange(
                max(1, len(sorted_books) * 3 // 4),
                len(sorted_books),
            )

            left = sorted_books[
                left_index
            ]

            right = sorted_books[
                right_index
            ]

            if (
                left.id != right.id
                and self.pair_key(
                    left.id,
                    right.id,
                )
                not in self.played
            ):
                return (
                    left.id,
                    right.id,
                )

        # ----------------------------------------------------
        # Candidate pool.
        #
        # Don't inspect every possible pair. For large shelves,
        # candidate generation is deliberately local.
        # ----------------------------------------------------

        ranked_books = [
            item["book"]
            for item in ranking
        ]

        candidates = []

        n = len(ranked_books)

        # Window size grows slowly with library size.
        window = max(
            4,
            min(
                20,
                int(
                    math.sqrt(n) * 2.5
                ),
            ),
        )

        # Generate nearby ranking pairs.
        for index, left in enumerate(
            ranked_books
        ):
            upper = min(
                n,
                index + window + 1,
            )

            for j in range(
                index + 1,
                upper,
            ):
                right = ranked_books[j]

                if (
                    self.pair_key(
                        left.id,
                        right.id,
                    )
                    in self.played
                ):
                    continue

                score = (
                    self.pair_information_score(
                        left,
                        right,
                        ranks,
                    )
                )

                candidates.append(
                    (
                        score,
                        left,
                        right,
                    )
                )

        # ----------------------------------------------------
        # Add high-RD exploratory pairs.
        # ----------------------------------------------------

        uncertain_books = sorted(
            self.books,
            key=lambda book: (
                self.ratings[
                    book.id
                ].rd
            ),
            reverse=True,
        )

        uncertain_books = uncertain_books[
            :min(40, len(uncertain_books))
        ]

        for left in uncertain_books:
            possible = []

            for right in ranked_books:
                if left.id == right.id:
                    continue

                if (
                    self.pair_key(
                        left.id,
                        right.id,
                    )
                    in self.played
                ):
                    continue

                possible.append(right)

                if len(possible) >= 12:
                    break

            for right in possible:
                score = (
                    self.pair_information_score(
                        left,
                        right,
                        ranks,
                    )
                )

                candidates.append(
                    (
                        score * 1.12,
                        left,
                        right,
                    )
                )

        if not candidates:
            return self.choose_any_unplayed_pair()

        candidates.sort(
            key=lambda item: item[0],
            reverse=True,
        )

        # Explore occasionally among the strongest candidates.
        top_count = min(
            12,
            len(candidates),
        )

        if (
            self.random.random()
            < EXPLORATION_RATE
        ):
            selected = self.random.choice(
                candidates[:top_count]
            )

        else:
            # Weighted choice gives the very best pair a strong
            # preference without becoming completely deterministic.
            shortlist = candidates[
                :min(20, len(candidates))
            ]

            weights = [
                max(
                    0.001,
                    item[0],
                )
                for item in shortlist
            ]

            selected = self.random.choices(
                shortlist,
                weights=weights,
                k=1,
            )[0]

        self.pair_selection_count += 1

        return (
            selected[1].id,
            selected[2].id,
        )

    def choose_any_unplayed_pair(self):
        """
        Fast fallback.

        Uses rating-sorted local search rather than the old
        O(N²) full-pair fallback.
        """

        ranking = self.ranking()

        ranked_books = [
            item["book"]
            for item in ranking
        ]

        n = len(ranked_books)

        window = min(
            30,
            max(
                5,
                int(math.sqrt(n) * 3),
            ),
        )

        for index, left in enumerate(
            ranked_books
        ):
            for offset in range(
                1,
                window + 1,
            ):
                j = index + offset

                if j >= n:
                    break

                right = ranked_books[j]

                if (
                    self.pair_key(
                        left.id,
                        right.id,
                    )
                    not in self.played
                ):
                    return (
                        left.id,
                        right.id,
                    )

        # Last resort. This should almost never be reached.
        for left in ranked_books:
            for right in ranked_books:
                if left.id == right.id:
                    continue

                if (
                    self.pair_key(
                        left.id,
                        right.id,
                    )
                    not in self.played
                ):
                    return (
                        left.id,
                        right.id,
                    )

        return None

    def choose_pair(self):
        if len(self.books) < 2:
            return None

        if self.is_finished():
            return None

        self.update_phase()

        if self.phase == "exploration":
            pair = (
                self.choose_exploration_pair()
            )

            if pair is not None:
                return pair

        return self.choose_active_pair()

    # ========================================================
    # BRADLEY-TERRY
    # ========================================================

    def fit_bradley_terry(self):
        """
        Fit a regularized Bradley-Terry-style logistic model.

        beta_i represents latent preference strength:

            P(i beats j)
                = sigmoid(beta_i - beta_j)

        A small L2 penalty prevents pathological infinite
        scores when the comparison graph is sparse or one-sided.

        The implementation is intentionally dependency-free:
        no numpy/scipy/pandas required.
        """

        if not self.books:
            return

        ids = [
            book.id
            for book in self.books
        ]

        index = {
            book_id: i
            for i, book_id in enumerate(ids)
        }

        n = len(ids)

        # Start from current Glicko ratings.
        beta = [
            (
                self.ratings[
                    book_id
                ].rating
                - 1500.0
            )
            / BT_RATING_SCALE
            for book_id in ids
        ]

        # Blend in Goodreads priors very gently.
        for i, book in enumerate(
            self.books
        ):
            beta[i] += (
                self.goodreads_prior_beta(
                    book
                )
                * 0.35
            )

        # Sparse comparison records.
        records = []

        for match in self.comparisons:
            left_id = match.get("left")
            right_id = match.get("right")
            result = match.get("result")

            if (
                left_id not in index
                or right_id not in index
            ):
                continue

            if result == "left":
                score = 1.0

            elif result == "right":
                score = 0.0

            elif result == "tie":
                score = 0.5

            else:
                continue

            records.append(
                (
                    index[left_id],
                    index[right_id],
                    score,
                )
            )

        if not records:
            self.bt_strengths = {
                book.id: 1.0
                for book in self.books
            }

            self.bt_ratings = {
                book.id: self.ratings[
                    book.id
                ].rating
                for book in self.books
            }

            self.bt_dirty = False
            self.last_bt_comparison_count = (
                len(self.comparisons)
            )

            return

        # ----------------------------------------------------
        # Diagonal-Newton logistic fitting.
        #
        # This is O(number of comparisons) per iteration and
        # therefore scales much better than dense N x N
        # pairwise matrices.
        # ----------------------------------------------------

        for _ in range(
            BT_MAX_ITERATIONS
        ):
            gradient = [
                0.0
            ] * n

            hessian = [
                BT_REGULARIZATION
            ] * n

            for i, j, score in records:
                difference = (
                    beta[i]
                    - beta[j]
                )

                probability = sigmoid(
                    difference
                )

                error = (
                    score
                    - probability
                )

                gradient[i] += error
                gradient[j] -= error

                information = (
                    probability
                    * (1.0 - probability)
                )

                hessian[i] += information
                hessian[j] += information

            # ------------------------------------------------
            # Goodreads prior as a weak pull toward its
            # initial latent value.
            # ------------------------------------------------

            if USE_GOODREADS_PRIOR:
                for i, book in enumerate(
                    self.books
                ):
                    prior = (
                        self.goodreads_prior_beta(
                            book
                        )
                    )

                    if prior == 0:
                        continue

                    prior_weight = (
                        GOODREADS_PRIOR_STRENGTH
                        * 0.25
                    )

                    gradient[i] += (
                        prior_weight
                        * (prior - beta[i])
                    )

                    hessian[i] += (
                        prior_weight
                    )

            max_change = 0.0

            for i in range(n):
                step = (
                    gradient[i]
                    / max(
                        hessian[i],
                        1e-8,
                    )
                )

                # Damp very large jumps.
                step = clamp(
                    step,
                    -0.35,
                    0.35,
                )

                beta[i] += step

                max_change = max(
                    max_change,
                    abs(step),
                )

            # Center latent strengths.
            mean_beta = (
                sum(beta) / n
            )

            beta = [
                value - mean_beta
                for value in beta
            ]

            if (
                max_change
                < BT_TOLERANCE
            ):
                break

        # ----------------------------------------------------
        # Convert latent strengths to positive strengths.
        # ----------------------------------------------------

        strengths = {}

        ratings = {}

        for i, book_id in enumerate(
            ids
        ):
            safe_beta = clamp(
                beta[i],
                -8.0,
                8.0,
            )

            strength = safe_exp(
                safe_beta
            )

            strengths[
                book_id
            ] = strength

            ratings[
                book_id
            ] = (
                1500.0
                + BT_RATING_SCALE
                * safe_beta
            )

        self.bt_strengths = strengths
        self.bt_ratings = ratings

        self.bt_dirty = False

        self.last_bt_comparison_count = (
            len(self.comparisons)
        )

    # ========================================================
    # RANKING
    # ========================================================

    def ensure_model(self):
        if self.bt_dirty:
            self.fit_bradley_terry()

    def ranking(self):
        self.ensure_model()

        result = []

        for book in self.books:
            rating = self.ratings[
                book.id
            ]

            final_rating = (
                self.bt_ratings.get(
                    book.id,
                    rating.rating,
                )
            )

            result.append(
                {
                    "book": book,
                    "rating": rating,
                    "final_rating": final_rating,
                    "strength": (
                        self.bt_strengths.get(
                            book.id,
                            1.0,
                        )
                    ),
                }
            )

        result.sort(
            key=lambda item: (
                item["final_rating"],
                item["rating"].rating,
                -item["rating"].rd,
                item["book"].title.lower(),
            ),
            reverse=True,
        )

        return result

    # ========================================================
    # STABILITY
    # ========================================================

    def calculate_stability(self):
        if not self.books:
            return 1.0

        ranking = self.ranking()

        current_order = [
            item["book"].id
            for item in ranking
        ]

        if not self.last_stable_order:
            self.last_stable_order = (
                current_order
            )

            self.stability = 0.0
            self.top_stability = 0.0

            self.last_stability_check = (
                len(self.comparisons)
            )

            return 0.0

        previous_positions = {
            book_id: index
            for index, book_id in enumerate(
                self.last_stable_order
            )
        }

        total_displacement = 0.0

        for current_index, book_id in enumerate(
            current_order
        ):
            old_index = previous_positions.get(
                book_id,
                current_index,
            )

            total_displacement += abs(
                current_index
                - old_index
            )

        n = len(current_order)

        # Maximum useful average displacement is roughly n.
        average_displacement = (
            total_displacement / n
        )

        normalized = (
            average_displacement
            / max(
                1.0,
                n * 0.20,
            )
        )

        self.stability = clamp(
            1.0 - normalized,
            0.0,
            1.0,
        )

        # Top-K stability is more important for a reading
        # recommender.
        top_k = min(
            25,
            n,
        )

        current_top = current_order[
            :top_k
        ]

        previous_top_positions = {
            book_id: index
            for index, book_id in enumerate(
                self.last_stable_order[
                    :top_k
                ]
            )
        }

        top_displacement = 0.0

        for index, book_id in enumerate(
            current_top
        ):
            old_index = (
                previous_top_positions.get(
                    book_id,
                    top_k,
                )
            )

            top_displacement += abs(
                index - old_index
            )

        top_average = (
            top_displacement
            / max(1, top_k)
        )

        top_normalized = (
            top_average
            / max(
                1.0,
                top_k * 0.20,
            )
        )

        self.top_stability = clamp(
            1.0 - top_normalized,
            0.0,
            1.0,
        )

        self.last_stable_order = (
            current_order
        )

        self.last_stability_check = (
            len(self.comparisons)
        )

        return self.stability

    # ========================================================
    # PROGRESS
    # ========================================================

    def progress(self):
        """
        Progress is based on evidence coverage, but the
        adaptive model gets additional credit for stability.
        """

        if not self.books:
            return 0.0

        average = self.average_comparisons()

        evidence_progress = clamp(
            average
            / self.target_comparisons,
            0.0,
            1.0,
        )

        stability_bonus = (
            self.stability
            * 0.20
        )

        coverage_bonus = (
            self.coverage()
            * 0.15
        )

        return clamp(
            (
                evidence_progress * 0.65
                + stability_bonus
                + coverage_bonus
            ),
            0.0,
            1.0,
        )

    # ========================================================
    # ADAPTIVE STOPPING
    # ========================================================

    def all_at_target(self):
        return all(
            self.counts[book.id]
            >= self.target_comparisons
            for book in self.books
        )

    def is_adaptively_finished(self):
        if not self.books:
            return True

        total = len(
            self.comparisons
        )

        if total < max(
            MIN_ADAPTIVE_DECISIONS,
            len(self.books),
        ):
            return False

        if self.coverage() < (
            MIN_COVERAGE_FOR_STOP
        ):
            return False

        if self.minimum_comparisons() < (
            MIN_INITIAL_COMPARISONS
        ):
            return False

        if self.stability < (
            DEFAULT_STABILITY_THRESHOLD
        ):
            return False

        if self.top_stability < (
            TOP_STABILITY_THRESHOLD
        ):
            return False

        return True

    def is_finished(self):
        if self.all_at_target():
            return True

        return self.is_adaptively_finished()

    # ========================================================
    # DIAGNOSTICS
    # ========================================================

    def top_confidence(self, count=10):
        ranking = self.ranking()

        result = []

        for index, item in enumerate(
            ranking[:count],
            start=1,
        ):
            rating = item["rating"]

            confidence = clamp(
                1.0
                - (
                    rating.rd
                    / 350.0
                ),
                0.0,
                1.0,
            )

            result.append(
                {
                    "rank": index,
                    "book": item["book"],
                    "rating": item[
                        "final_rating"
                    ],
                    "confidence": confidence,
                    "rd": rating.rd,
                }
            )

        return result

    def phase_label(self):
        return {
            "exploration": "Exploring",
            "calibration": "Calibrating",
            "active": "Actively refining",
            "complete": "Complete",
        }.get(
            self.phase,
            "Ranking",
        )


# ============================================================
# SAVED STATE
# ============================================================

class StateStore:
    def __init__(
        self,
        source_file: Path,
    ):
        self.source_file = source_file

        self.directory = (
            source_file.parent
            / STATE_DIRECTORY_NAME
        )

        self.directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        self.path = (
            self.directory
            / (
                safe_filename(
                    source_file.stem
                )
                + ".json"
            )
        )

    def save(
        self,
        engine: RankingEngine,
    ):
        data = {
            "version": STATE_VERSION,
            "source_file": str(
                self.source_file.resolve()
            ),
            "target_comparisons": (
                engine.target_comparisons
            ),
            "seed": engine.seed,
            "books": [
                asdict(book)
                for book in engine.books
            ],
            "comparisons": (
                engine.comparisons
            ),
            "saved_at": now_iso(),
        }

        temporary = self.path.with_suffix(
            ".tmp"
        )

        with temporary.open(
            "w",
            encoding="utf-8",
        ) as file:
            json.dump(
                data,
                file,
                indent=2,
                ensure_ascii=False,
            )

        temporary.replace(
            self.path
        )

    def load(
        self,
        books: list[Book],
    ) -> Optional[RankingEngine]:
        if not self.path.exists():
            return None

        try:
            with self.path.open(
                "r",
                encoding="utf-8",
            ) as file:
                data = json.load(file)

            saved_ids = {
                book.get("id")
                for book in data.get(
                    "books",
                    [],
                )
            }

            current_ids = {
                book.id
                for book in books
            }

            if saved_ids != current_ids:
                return None

            target = int(
                data.get(
                    "target_comparisons",
                    DEFAULT_TARGET_COMPARISONS,
                )
            )

            seed = data.get(
                "seed"
            )

            engine = RankingEngine(
                books,
                target,
                seed,
            )

            engine.comparisons = (
                copy.deepcopy(
                    data.get(
                        "comparisons",
                        [],
                    )
                )
            )

            # Old state files are accepted.
            engine.rebuild_from_history()

            return engine

        except Exception:
            return None


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_results(
    source_file: Path,
    engine: RankingEngine,
):
    workbook = load_workbook(
        filename=source_file
    )

    try:
        if "Ranking" in workbook.sheetnames:
            del workbook["Ranking"]

        ranking_sheet = workbook.create_sheet(
            "Ranking",
            0,
        )

        headers = [
            "Rank",
            "Title",
            "Author",
            "Rating",
            "Uncertainty",
            "Comparisons",
            "Wins",
            "Losses",
            "Ties",
            "Confidence",
            "Goodreads Book Id",
        ]

        ranking_sheet.append(
            headers
        )

        ranked_items = engine.ranking()

        for rank, item in enumerate(
            ranked_items,
            start=1,
        ):
            book = item["book"]
            rating = item["rating"]

            confidence = clamp(
                1.0
                - rating.rd / 350.0,
                0.0,
                1.0,
            )

            ranking_sheet.append(
                [
                    rank,
                    book.title,
                    book.author,
                    round(
                        item["final_rating"],
                        2,
                    ),
                    round(
                        rating.rd,
                        2,
                    ),
                    rating.comparisons,
                    rating.wins,
                    rating.losses,
                    rating.ties,
                    round(
                        confidence * 100,
                        1,
                    ),
                    book.book_id,
                ]
            )

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="2F5597",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        for cell in ranking_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                vertical="center"
            )

        ranking_sheet.freeze_panes = "A2"

        ranking_sheet.auto_filter.ref = (
            ranking_sheet.dimensions
        )

        widths = [
            8,
            45,
            30,
            14,
            16,
            15,
            10,
            10,
            10,
            14,
            22,
        ]

        for index, width in enumerate(
            widths,
            start=1,
        ):
            ranking_sheet.column_dimensions[
                get_column_letter(index)
            ].width = width

        # ----------------------------------------------------
        # Diagnostics sheet
        # ----------------------------------------------------

        if "Diagnostics" in workbook.sheetnames:
            del workbook["Diagnostics"]

        diagnostics_sheet = (
            workbook.create_sheet(
                "Diagnostics"
            )
        )

        diagnostics_sheet.append(
            [
                "Metric",
                "Value",
            ]
        )

        diagnostics = [
            (
                "Books",
                len(engine.books),
            ),
            (
                "Comparisons",
                len(engine.comparisons),
            ),
            (
                "Average comparisons/book",
                round(
                    engine.average_comparisons(),
                    2,
                ),
            ),
            (
                "Minimum comparisons/book",
                engine.minimum_comparisons(),
            ),
            (
                "Evidence coverage",
                round(
                    engine.coverage()
                    * 100,
                    1,
                ),
            ),
            (
                "Ranking stability",
                round(
                    engine.stability
                    * 100,
                    1,
                ),
            ),
            (
                "Top-25 stability",
                round(
                    engine.top_stability
                    * 100,
                    1,
                ),
            ),
            (
                "Phase",
                engine.phase_label(),
            ),
            (
                "Maximum comparisons/book",
                engine.target_comparisons,
            ),
            (
                "Adaptive stopping",
                (
                    "Complete"
                    if engine.is_finished()
                    else "Still refining"
                ),
            ),
        ]

        for row in diagnostics:
            diagnostics_sheet.append(
                row
            )

        for cell in diagnostics_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        diagnostics_sheet.column_dimensions[
            "A"
        ].width = 35

        diagnostics_sheet.column_dimensions[
            "B"
        ].width = 30

        # ----------------------------------------------------
        # Comparisons sheet
        # ----------------------------------------------------

        if "Comparisons" in workbook.sheetnames:
            del workbook["Comparisons"]

        comparison_sheet = workbook.create_sheet(
            "Comparisons"
        )

        comparison_sheet.append(
            [
                "Left Book",
                "Left Author",
                "Right Book",
                "Right Author",
                "Result",
                "Timestamp",
            ]
        )

        book_map = engine.book_map()

        for match in engine.comparisons:
            left = book_map.get(
                match["left"]
            )

            right = book_map.get(
                match["right"]
            )

            if left is None or right is None:
                continue

            if match["result"] == "left":
                result = left.title

            elif match["result"] == "right":
                result = right.title

            else:
                result = "Tie"

            comparison_sheet.append(
                [
                    left.title,
                    left.author,
                    right.title,
                    right.author,
                    result,
                    match["timestamp"],
                ]
            )

        for cell in comparison_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        comparison_sheet.freeze_panes = "A2"

        comparison_widths = {
            "A": 45,
            "B": 30,
            "C": 45,
            "D": 30,
            "E": 45,
            "F": 22,
        }

        for column, width in (
            comparison_widths.items()
        ):
            comparison_sheet.column_dimensions[
                column
            ].width = width

        # ----------------------------------------------------
        # Original sheet enrichment
        # ----------------------------------------------------

        original_sheet = None

        for sheet in workbook.worksheets:
            if sheet.title not in (
                "Ranking",
                "Comparisons",
                "Diagnostics",
            ):
                original_sheet = sheet
                break

        if original_sheet is None:
            original_sheet = workbook.active

        existing_headers = [
            cell.value
            for cell in original_sheet[1]
        ]

        extra_headers = [
            "Rank",
            "Rank Rating",
            "Rating Uncertainty",
            "Pairwise Comparisons",
            "Wins",
            "Losses",
            "Ties",
            "Confidence",
        ]

        start_column = (
            len(existing_headers)
            + 1
        )

        for offset, header in enumerate(
            extra_headers
        ):
            cell = original_sheet.cell(
                row=1,
                column=(
                    start_column
                    + offset
                ),
            )

            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        ranked = {}

        for rank, item in enumerate(
            ranked_items,
            start=1,
        ):
            ranked[
                item["book"].id
            ] = (
                rank,
                item["final_rating"],
                item["rating"],
            )

        book_id_column = None

        for index, header in enumerate(
            existing_headers,
            start=1,
        ):
            if (
                normalize_header(header)
                == normalize_header(
                    "Book Id - Goodreads"
                )
            ):
                book_id_column = index
                break

        books_by_row = {
            book.original_row: book
            for book in engine.books
        }

        books_by_id = {
            book.id: book
            for book in engine.books
        }

        for row in range(
            2,
            original_sheet.max_row + 1,
        ):
            book = None

            if book_id_column is not None:
                goodreads_id = normalize(
                    original_sheet.cell(
                        row=row,
                        column=book_id_column,
                    ).value
                )

                if goodreads_id:
                    candidate_id = (
                        "goodreads:"
                        + goodreads_id
                    )

                    book = books_by_id.get(
                        candidate_id
                    )

            if book is None:
                book = books_by_row.get(
                    row
                )

            if book is None:
                continue

            ranking_data = ranked.get(
                book.id
            )

            if ranking_data is None:
                continue

            (
                rank,
                final_rating,
                rating,
            ) = ranking_data

            confidence = clamp(
                1.0
                - rating.rd / 350.0,
                0.0,
                1.0,
            )

            values = [
                rank,
                round(
                    final_rating,
                    2,
                ),
                round(
                    rating.rd,
                    2,
                ),
                rating.comparisons,
                rating.wins,
                rating.losses,
                rating.ties,
                round(
                    confidence * 100,
                    1,
                ),
            ]

            for offset, value in enumerate(
                values
            ):
                original_sheet.cell(
                    row=row,
                    column=(
                        start_column
                        + offset
                    ),
                    value=value,
                )

        output_path = (
            source_file.parent
            / (
                safe_filename(
                    source_file.stem
                )
                + "_ranked.xlsx"
            )
        )

        workbook.save(
            output_path
        )

        return output_path

    finally:
        workbook.close()


# ============================================================
# GUI APPLICATION
# ============================================================

class RankerApp:
    def __init__(
        self,
        root: tk.Tk,
    ):
        self.root = root

        self.root.title(
            f"{APP_NAME} {APP_VERSION}"
        )

        self.root.geometry(
            f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}"
        )

        self.root.minsize(
            1050,
            720,
        )

        self.engine = None
        self.source_file = None
        self.state_store = None
        self.current_pair = None

        self.target_var = tk.IntVar(
            value=DEFAULT_TARGET_COMPARISONS
        )

        self.auto_stop_var = tk.BooleanVar(
            value=True
        )

        self.status_var = tk.StringVar(
            value=(
                "Open your Goodreads Excel export "
                "to begin."
            )
        )

        self.progress_var = tk.DoubleVar(
            value=0.0
        )

        self.stats_var = tk.StringVar(
            value=""
        )

        self.left_title_var = tk.StringVar()
        self.left_author_var = tk.StringVar()
        self.left_meta_var = tk.StringVar()

        self.right_title_var = tk.StringVar()
        self.right_author_var = tk.StringVar()
        self.right_meta_var = tk.StringVar()

        self.build_styles()
        self.build_menu()
        self.build_ui()
        self.bind_keys()

        self.root.protocol(
            "WM_DELETE_WINDOW",
            self.on_close,
        )

    # ========================================================
    # STYLES
    # ========================================================

    def build_styles(self):
        style = ttk.Style()

        try:
            style.theme_use(
                "vista"
            )
        except tk.TclError:
            pass

        style.configure(
            "Title.TLabel",
            font=(
                "Segoe UI",
                23,
                "bold",
            ),
        )

        style.configure(
            "BookTitle.TLabel",
            font=(
                "Segoe UI",
                18,
                "bold",
            ),
        )

        style.configure(
            "BookAuthor.TLabel",
            font=(
                "Segoe UI",
                12,
            ),
        )

        style.configure(
            "BookMeta.TLabel",
            font=(
                "Segoe UI",
                10,
            ),
            foreground="#666666",
        )

        style.configure(
            "Choice.TButton",
            font=(
                "Segoe UI",
                12,
                "bold",
            ),
            padding=14,
        )

        style.configure(
            "Preset.TButton",
            font=(
                "Segoe UI",
                9,
                "bold",
            ),
            padding=(8, 5),
        )

        style.configure(
            "Secondary.TButton",
            font=(
                "Segoe UI",
                10,
            ),
            padding=8,
        )

    # ========================================================
    # MENU
    # ========================================================

    def build_menu(self):
        menu = tk.Menu(self.root)

        file_menu = tk.Menu(
            menu,
            tearoff=False,
        )

        file_menu.add_command(
            label="Open Goodreads Excel…",
            command=self.open_file,
            accelerator="Ctrl+O",
        )

        file_menu.add_command(
            label="Export Ranking",
            command=self.export,
            accelerator="Ctrl+E",
        )

        file_menu.add_separator()

        file_menu.add_command(
            label="Exit",
            command=self.on_close,
        )

        menu.add_cascade(
            label="File",
            menu=file_menu,
        )

        ranking_menu = tk.Menu(
            menu,
            tearoff=False,
        )

        ranking_menu.add_command(
            label="Undo Last Choice",
            command=self.undo,
            accelerator="U",
        )

        ranking_menu.add_command(
            label="Finish Now",
            command=self.finish_now,
        )

        ranking_menu.add_command(
            label="View Ranking",
            command=self.show_ranking,
        )

        ranking_menu.add_command(
            label="View Diagnostics",
            command=self.show_diagnostics,
        )

        menu.add_cascade(
            label="Ranking",
            menu=ranking_menu,
        )

        help_menu = tk.Menu(
            menu,
            tearoff=False,
        )

        help_menu.add_command(
            label="Keyboard Shortcuts",
            command=self.show_shortcuts,
        )

        help_menu.add_command(
            label="About",
            command=self.show_about,
        )

        menu.add_cascade(
            label="Help",
            menu=help_menu,
        )

        self.root.config(
            menu=menu
        )

    # ========================================================
    # MAIN UI
    # ========================================================

    def build_ui(self):
        outer = ttk.Frame(
            self.root,
            padding=22,
        )

        outer.pack(
            fill="both",
            expand=True,
        )

        # ----------------------------------------------------
        # Header
        # ----------------------------------------------------

        header = ttk.Frame(
            outer
        )

        header.pack(
            fill="x",
            pady=(0, 10),
        )

        ttk.Label(
            header,
            text=(
                "📚 Goodreads To-Read Ranker"
            ),
            style="Title.TLabel",
        ).pack(
            side="left"
        )

        controls = ttk.Frame(
            header
        )

        controls.pack(
            side="right"
        )

        ttk.Label(
            controls,
            text="Max comparisons/book:",
        ).pack(
            side="left",
            padx=(0, 6),
        )

        self.spinbox = ttk.Spinbox(
            controls,
            from_=MIN_TARGET_COMPARISONS,
            to=MAX_TARGET_COMPARISONS,
            textvariable=self.target_var,
            width=5,
        )

        self.spinbox.pack(
            side="left"
        )

        self.spinbox.bind(
            "<Return>",
            self.target_changed,
        )

        ttk.Checkbutton(
            controls,
            text="Stop when stable",
            variable=self.auto_stop_var,
        ).pack(
            side="left",
            padx=(12, 0),
        )

        ttk.Button(
            controls,
            text="Open Excel",
            command=self.open_file,
        ).pack(
            side="left",
            padx=(12, 0),
        )

        # ----------------------------------------------------
        # Speed presets
        # ----------------------------------------------------

        preset_frame = ttk.Frame(
            outer
        )

        preset_frame.pack(
            fill="x",
            pady=(0, 12),
        )

        ttk.Label(
            preset_frame,
            text="Strategy:",
        ).pack(
            side="left",
            padx=(0, 7),
        )

        ttk.Button(
            preset_frame,
            text="⚡ Turbo · 3",
            style="Preset.TButton",
            command=lambda: self.set_target(
                QUICK_COMPARISONS
            ),
        ).pack(
            side="left",
            padx=2,
        )

        ttk.Button(
            preset_frame,
            text="Quick · 3",
            style="Preset.TButton",
            command=lambda: self.set_target(
                QUICK_COMPARISONS
            ),
        ).pack(
            side="left",
            padx=2,
        )

        ttk.Button(
            preset_frame,
            text="Balanced · 6",
            style="Preset.TButton",
            command=lambda: self.set_target(
                BALANCED_COMPARISONS
            ),
        ).pack(
            side="left",
            padx=2,
        )

        ttk.Button(
            preset_frame,
            text="Accurate · 12",
            style="Preset.TButton",
            command=lambda: self.set_target(
                ACCURATE_COMPARISONS
            ),
        ).pack(
            side="left",
            padx=2,
        )

        ttk.Label(
            preset_frame,
            text=(
                "Adaptive engine prioritises uncertain "
                "ranking boundaries."
            ),
            foreground="#666666",
        ).pack(
            side="left",
            padx=(10, 0),
        )

        # ----------------------------------------------------
        # Status
        # ----------------------------------------------------

        status = ttk.Frame(
            outer
        )

        status.pack(
            fill="x",
            pady=(0, 12),
        )

        ttk.Label(
            status,
            textvariable=self.status_var,
        ).pack(
            anchor="w"
        )

        ttk.Progressbar(
            status,
            variable=self.progress_var,
            maximum=100,
        ).pack(
            fill="x",
            pady=(7, 0),
        )

        ttk.Label(
            status,
            textvariable=self.stats_var,
        ).pack(
            anchor="w",
            pady=(4, 0),
        )

        # ----------------------------------------------------
        # Cards
        # ----------------------------------------------------

        choices = ttk.Frame(
            outer
        )

        choices.pack(
            fill="both",
            expand=True,
        )

        self.left_card = (
            self.make_book_card(
                choices,
                "left",
            )
        )

        self.left_card.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(0, 8),
        )

        middle = ttk.Frame(
            choices,
            width=70,
        )

        middle.pack(
            side="left",
            fill="y",
        )

        middle.pack_propagate(
            False
        )

        ttk.Label(
            middle,
            text="VS",
            font=(
                "Segoe UI",
                16,
                "bold",
            ),
        ).pack(
            pady=(155, 10)
        )

        self.right_card = (
            self.make_book_card(
                choices,
                "right",
            )
        )

        self.right_card.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(8, 0),
        )

        # ----------------------------------------------------
        # Main buttons
        # ----------------------------------------------------

        buttons = ttk.Frame(
            outer
        )

        buttons.pack(
            fill="x",
            pady=(12, 0),
        )

        self.left_button = ttk.Button(
            buttons,
            text="←  Choose Left  [1]",
            command=lambda: self.choose(
                "left"
            ),
            style="Choice.TButton",
        )

        self.left_button.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 5),
        )

        self.tie_button = ttk.Button(
            buttons,
            text="≈  Tie  [3]",
            command=lambda: self.choose(
                "tie"
            ),
            style="Choice.TButton",
        )

        self.tie_button.pack(
            side="left",
            fill="x",
            expand=True,
            padx=5,
        )

        self.right_button = ttk.Button(
            buttons,
            text="Choose Right  →  [2]",
            command=lambda: self.choose(
                "right"
            ),
            style="Choice.TButton",
        )

        self.right_button.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(5, 0),
        )

        # ----------------------------------------------------
        # Bottom controls
        # ----------------------------------------------------

        bottom = ttk.Frame(
            outer
        )

        bottom.pack(
            fill="x",
            pady=(10, 0),
        )

        ttk.Button(
            bottom,
            text="Undo [U]",
            command=self.undo,
            style="Secondary.TButton",
        ).pack(
            side="left"
        )

        ttk.Button(
            bottom,
            text="Skip [4]",
            command=self.skip,
            style="Secondary.TButton",
        ).pack(
            side="left",
            padx=7,
        )

        ttk.Button(
            bottom,
            text="Finish Now",
            command=self.finish_now,
            style="Secondary.TButton",
        ).pack(
            side="left"
        )

        ttk.Label(
            bottom,
            text=(
                "Keyboard: 1 left · 2 right · "
                "3 tie · 4 skip · U undo"
            ),
            foreground="#666666",
        ).pack(
            side="left",
            padx=12,
        )

        ttk.Button(
            bottom,
            text="View Ranking",
            command=self.show_ranking,
            style="Secondary.TButton",
        ).pack(
            side="right"
        )

        ttk.Button(
            bottom,
            text="Export Excel",
            command=self.export,
            style="Secondary.TButton",
        ).pack(
            side="right",
            padx=7,
        )

    # ========================================================
    # BOOK CARD
    # ========================================================

    def make_book_card(
        self,
        parent,
        side,
    ):
        frame = tk.Frame(
            parent,
            bg="#F6F7F9",
            highlightbackground="#D8DCE2",
            highlightthickness=1,
        )

        content = ttk.Frame(
            frame,
            padding=22,
        )

        content.pack(
            fill="both",
            expand=True,
        )

        if side == "left":
            title_var = (
                self.left_title_var
            )
            author_var = (
                self.left_author_var
            )
            meta_var = (
                self.left_meta_var
            )

        else:
            title_var = (
                self.right_title_var
            )
            author_var = (
                self.right_author_var
            )
            meta_var = (
                self.right_meta_var
            )

        ttk.Label(
            content,
            textvariable=title_var,
            style="BookTitle.TLabel",
            wraplength=410,
            justify="center",
        ).pack(
            pady=(25, 10)
        )

        ttk.Label(
            content,
            textvariable=author_var,
            style="BookAuthor.TLabel",
            wraplength=410,
            justify="center",
        ).pack(
            pady=(0, 18)
        )

        ttk.Label(
            content,
            textvariable=meta_var,
            style="BookMeta.TLabel",
            wraplength=410,
            justify="center",
        ).pack(
            pady=(0, 15)
        )

        description = tk.Text(
            content,
            height=9,
            width=45,
            wrap="word",
            relief="flat",
            borderwidth=0,
            bg="#F6F7F9",
            fg="#444444",
            font=(
                "Segoe UI",
                10,
            ),
            padx=5,
            pady=5,
        )

        description.pack(
            fill="both",
            expand=True,
        )

        description.configure(
            state="disabled"
        )

        if side == "left":
            self.left_description = (
                description
            )

        else:
            self.right_description = (
                description
            )

        return frame

    # ========================================================
    # KEYBOARD
    # ========================================================

    def bind_keys(self):
        self.root.bind(
            "<KeyPress-1>",
            lambda event: self.choose(
                "left"
            ),
        )

        self.root.bind(
            "<KeyPress-2>",
            lambda event: self.choose(
                "right"
            ),
        )

        self.root.bind(
            "<KeyPress-3>",
            lambda event: self.choose(
                "tie"
            ),
        )

        self.root.bind(
            "<KeyPress-4>",
            lambda event: self.skip(),
        )

        self.root.bind(
            "<Left>",
            lambda event: self.choose(
                "left"
            ),
        )

        self.root.bind(
            "<Right>",
            lambda event: self.choose(
                "right"
            ),
        )

        self.root.bind(
            "<KeyPress-t>",
            lambda event: self.choose(
                "tie"
            ),
        )

        self.root.bind(
            "<KeyPress-T>",
            lambda event: self.choose(
                "tie"
            ),
        )

        self.root.bind(
            "<KeyPress-s>",
            lambda event: self.skip(),
        )

        self.root.bind(
            "<KeyPress-S>",
            lambda event: self.skip(),
        )

        self.root.bind(
            "<KeyPress-u>",
            lambda event: self.undo(),
        )

        self.root.bind(
            "<KeyPress-U>",
            lambda event: self.undo(),
        )

        self.root.bind(
            "<Control-o>",
            lambda event: self.open_file(),
        )

        self.root.bind(
            "<Control-e>",
            lambda event: self.export(),
        )

    # ========================================================
    # FILE
    # ========================================================

    def open_file(self):
        path = filedialog.askopenfilename(
            title=(
                "Open Goodreads Excel export"
            ),
            filetypes=[
                (
                    "Excel files",
                    "*.xlsx",
                ),
                (
                    "All files",
                    "*.*",
                ),
            ],
        )

        if not path:
            return

        self.load_file(
            Path(path)
        )

    def load_file(
        self,
        path: Path,
    ):
        try:
            _, books = load_goodreads(
                path
            )

        except Exception as exc:
            messagebox.showerror(
                "Could not open Goodreads file",
                str(exc),
            )
            return

        if len(books) < 2:
            messagebox.showwarning(
                "Not enough books",
                (
                    'I found fewer than two books where '
                    '"Exclusive Shelf" is "to-read".'
                ),
            )
            return

        self.source_file = path

        self.state_store = (
            StateStore(path)
        )

        existing = (
            self.state_store.load(
                books
            )
        )

        if existing is not None:
            resume = messagebox.askyesno(
                "Resume previous ranking?",
                (
                    f"I found a saved ranking for:\n\n"
                    f"{path.name}\n\n"
                    f"{len(existing.comparisons)} "
                    f"comparisons have already been made.\n\n"
                    f"Resume it?"
                ),
            )

            if resume:
                self.engine = existing

            else:
                self.engine = (
                    RankingEngine(
                        books,
                        self.target_var.get(),
                    )
                )

        else:
            self.engine = (
                RankingEngine(
                    books,
                    self.target_var.get(),
                )
            )

        self.target_var.set(
            self.engine.target_comparisons
        )

        self.current_pair = None

        self.status_var.set(
            (
                f"{len(books)} books on your "
                f"to-read shelf · {path.name}"
            )
        )

        self.refresh()

    # ========================================================
    # SPEED PRESETS
    # ========================================================

    def set_target(
        self,
        value: int,
    ):
        self.target_var.set(
            value
        )

        self.target_changed()

    def target_changed(
        self,
        event=None,
    ):
        if self.engine is None:
            return

        try:
            value = int(
                self.target_var.get()
            )

        except ValueError:
            value = (
                DEFAULT_TARGET_COMPARISONS
            )

        value = max(
            MIN_TARGET_COMPARISONS,
            min(
                MAX_TARGET_COMPARISONS,
                value,
            ),
        )

        self.target_var.set(
            value
        )

        self.engine.target_comparisons = (
            value
        )

        self.save_state()

        self.current_pair = None

        self.refresh()

    # ========================================================
    # REFRESH
    # ========================================================

    def refresh(self):
        if self.engine is None:
            return

        # Ensure model is current before selecting the pair.
        self.engine.ensure_model()

        if (
            self.auto_stop_var.get()
            and self.engine.is_finished()
        ):
            self.show_finished()

            return

        if (
            self.current_pair is None
            or not self.current_pair_valid()
        ):
            self.current_pair = (
                self.engine.choose_pair()
            )

        if self.current_pair is None:
            self.show_finished()

            return

        self.enable_choices()

        left_id, right_id = (
            self.current_pair
        )

        book_map = (
            self.engine.book_map()
        )

        left = book_map[
            left_id
        ]

        right = book_map[
            right_id
        ]

        self.show_book(
            left,
            self.engine.ratings[
                left.id
            ],
            self.left_title_var,
            self.left_author_var,
            self.left_meta_var,
            self.left_description,
        )

        self.show_book(
            right,
            self.engine.ratings[
                right.id
            ],
            self.right_title_var,
            self.right_author_var,
            self.right_meta_var,
            self.right_description,
        )

        progress = (
            self.engine.progress()
            * 100.0
        )

        self.progress_var.set(
            progress
        )

        comparisons = len(
            self.engine.comparisons
        )

        average = (
            self.engine.average_comparisons()
        )

        coverage = (
            self.engine.coverage()
            * 100.0
        )

        stability = (
            self.engine.stability
            * 100.0
        )

        top_stability = (
            self.engine.top_stability
            * 100.0
        )

        self.stats_var.set(
            (
                f"{comparisons} decisions · "
                f"{average:.1f} avg/book · "
                f"{coverage:.0f}% covered · "
                f"stability {stability:.0f}% · "
                f"top-25 {top_stability:.0f}%"
            )
        )

        self.status_var.set(
            (
                f"{self.engine.phase_label()} · "
                "Which book would you rather read?"
            )
        )

    def show_finished(self):
        self.disable_choices()

        self.left_title_var.set(
            "🏆 Ranking ready"
        )

        self.left_author_var.set("")
        self.left_meta_var.set("")

        self.right_title_var.set(
            "You can export now"
        )

        self.right_author_var.set("")
        self.right_meta_var.set("")

        self.set_description(
            self.left_description,
            (
                "The adaptive ranking engine has "
                "determined that additional comparisons "
                "are unlikely to materially improve the "
                "current ordering."
            ),
        )

        self.set_description(
            self.right_description,
            (
                "You can continue refining by turning "
                "off 'Stop when stable', or export the "
                "current ranking to Excel."
            ),
        )

        self.progress_var.set(
            max(
                95.0,
                self.engine.progress()
                * 100.0,
            )
        )

        self.stats_var.set(
            (
                f"{len(self.engine.comparisons)} decisions · "
                f"{self.engine.average_comparisons():.1f} "
                f"average/book · "
                f"ranking stability "
                f"{self.engine.stability * 100:.0f}% · "
                f"top-25 stability "
                f"{self.engine.top_stability * 100:.0f}%"
            )
        )

        self.status_var.set(
            "Ranking complete. Export when ready."
        )

    def current_pair_valid(self):
        if (
            self.engine is None
            or self.current_pair is None
        ):
            return False

        left, right = (
            self.current_pair
        )

        if left not in self.engine.ratings:
            return False

        if right not in self.engine.ratings:
            return False

        if left == right:
            return False

        return (
            self.engine.pair_key(
                left,
                right,
            )
            not in self.engine.played
        )

    # ========================================================
    # DISPLAY BOOK
    # ========================================================

    def show_book(
        self,
        book,
        rating,
        title_var,
        author_var,
        meta_var,
        description_widget,
    ):
        title_var.set(
            book.title
        )

        author_var.set(
            book.author
            if book.author
            else "Unknown author"
        )

        meta = []

        if book.pages:
            meta.append(
                f"{book.pages} pages"
            )

        if book.year:
            meta.append(
                f"Published {book.year}"
            )

        meta.append(
            f"Rating {rating.rating:.0f}"
        )

        meta.append(
            f"±{rating.rd:.0f}"
        )

        meta.append(
            f"{rating.comparisons} comparisons"
        )

        if book.my_rating:
            meta.append(
                f"My Goodreads: {book.my_rating}"
            )

        meta_var.set(
            "  ·  ".join(meta)
        )

        description = (
            book.description
            if book.description
            else "No description available."
        )

        self.set_description(
            description_widget,
            description,
        )

    def set_description(
        self,
        widget,
        text,
    ):
        widget.configure(
            state="normal"
        )

        widget.delete(
            "1.0",
            "end",
        )

        widget.insert(
            "1.0",
            truncate(
                text,
                1800,
            ),
        )

        widget.configure(
            state="disabled"
        )

    # ========================================================
    # CHOICES
    # ========================================================

    def choose(
        self,
        result: str,
    ):
        if (
            self.engine is None
            or self.current_pair is None
        ):
            return

        left_id, right_id = (
            self.current_pair
        )

        try:
            self.engine.apply_match(
                left_id,
                right_id,
                result,
            )

            self.save_state()

        except Exception as exc:
            messagebox.showerror(
                "Could not record choice",
                str(exc),
            )

            return

        self.current_pair = None

        self.refresh()

    def skip(self):
        if self.engine is None:
            return

        self.current_pair = None

        self.refresh()

    # ========================================================
    # FINISH
    # ========================================================

    def finish_now(self):
        if self.engine is None:
            return

        self.engine.fit_bradley_terry()
        self.engine.calculate_stability()

        self.save_state()

        self.current_pair = None

        self.show_finished()

    # ========================================================
    # SAVE
    # ========================================================

    def save_state(self):
        if (
            self.engine is None
            or self.state_store is None
        ):
            return

        try:
            self.state_store.save(
                self.engine
            )

        except Exception as exc:
            print(
                "Warning: could not save state:",
                exc,
            )

    # ========================================================
    # UNDO
    # ========================================================

    def undo(self):
        if self.engine is None:
            return

        if self.engine.undo():
            self.save_state()

            self.current_pair = None

            self.refresh()

    # ========================================================
    # BUTTON STATES
    # ========================================================

    def enable_choices(self):
        self.left_button.configure(
            state="normal"
        )

        self.tie_button.configure(
            state="normal"
        )

        self.right_button.configure(
            state="normal"
        )

    def disable_choices(self):
        self.left_button.configure(
            state="disabled"
        )

        self.tie_button.configure(
            state="disabled"
        )

        self.right_button.configure(
            state="disabled"
        )

    # ========================================================
    # RANKING WINDOW
    # ========================================================

    def show_ranking(self):
        if self.engine is None:
            messagebox.showinfo(
                "No ranking",
                "Open a Goodreads file first.",
            )

            return

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Current Ranking"
        )

        window.geometry(
            "1150x700"
        )

        frame = ttk.Frame(
            window,
            padding=15,
        )

        frame.pack(
            fill="both",
            expand=True,
        )

        ttk.Label(
            frame,
            text="Current Ranking",
            style="Title.TLabel",
        ).pack(
            anchor="w",
            pady=(0, 5),
        )

        ttk.Label(
            frame,
            text=(
                "Global ranking is fitted from all "
                "pairwise decisions."
            ),
            foreground="#666666",
        ).pack(
            anchor="w",
            pady=(0, 10),
        )

        columns = (
            "rank",
            "title",
            "author",
            "rating",
            "rd",
            "comparisons",
            "confidence",
            "wins",
            "losses",
            "ties",
        )

        tree = ttk.Treeview(
            frame,
            columns=columns,
            show="headings",
        )

        headings = {
            "rank": "Rank",
            "title": "Title",
            "author": "Author",
            "rating": "Rating",
            "rd": "Uncertainty",
            "comparisons": "Comparisons",
            "confidence": "Confidence",
            "wins": "Wins",
            "losses": "Losses",
            "ties": "Ties",
        }

        widths = {
            "rank": 55,
            "title": 340,
            "author": 220,
            "rating": 80,
            "rd": 95,
            "comparisons": 100,
            "confidence": 90,
            "wins": 65,
            "losses": 65,
            "ties": 55,
        }

        for column in columns:
            tree.heading(
                column,
                text=headings[column],
            )

            tree.column(
                column,
                width=widths[column],
                anchor=(
                    "w"
                    if column in (
                        "title",
                        "author",
                    )
                    else "center"
                ),
            )

        scrollbar = ttk.Scrollbar(
            frame,
            orient="vertical",
            command=tree.yview,
        )

        tree.configure(
            yscrollcommand=scrollbar.set
        )

        tree.pack(
            side="left",
            fill="both",
            expand=True,
        )

        scrollbar.pack(
            side="right",
            fill="y",
        )

        for rank, item in enumerate(
            self.engine.ranking(),
            start=1,
        ):
            book = item["book"]
            rating = item["rating"]

            confidence = clamp(
                1.0
                - rating.rd / 350.0,
                0.0,
                1.0,
            )

            tree.insert(
                "",
                "end",
                values=(
                    rank,
                    book.title,
                    book.author,
                    f"{item['final_rating']:.0f}",
                    f"±{rating.rd:.0f}",
                    rating.comparisons,
                    f"{confidence * 100:.0f}%",
                    rating.wins,
                    rating.losses,
                    rating.ties,
                ),
            )

    # ========================================================
    # DIAGNOSTICS WINDOW
    # ========================================================

    def show_diagnostics(self):
        if self.engine is None:
            messagebox.showinfo(
                "No ranking",
                "Open a Goodreads file first.",
            )

            return

        self.engine.ensure_model()

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Ranking Diagnostics"
        )

        window.geometry(
            "700x650"
        )

        frame = ttk.Frame(
            window,
            padding=20,
        )

        frame.pack(
            fill="both",
            expand=True,
        )

        ttk.Label(
            frame,
            text="Ranking Diagnostics",
            style="Title.TLabel",
        ).pack(
            anchor="w",
            pady=(0, 15),
        )

        rows = [
            (
                "Books",
                len(self.engine.books),
            ),
            (
                "Total comparisons",
                len(
                    self.engine.comparisons
                ),
            ),
            (
                "Average comparisons/book",
                f"{self.engine.average_comparisons():.2f}",
            ),
            (
                "Minimum comparisons/book",
                self.engine.minimum_comparisons(),
            ),
            (
                "Evidence coverage",
                f"{self.engine.coverage() * 100:.1f}%",
            ),
            (
                "Maximum target",
                self.engine.target_comparisons,
            ),
            (
                "Engine phase",
                self.engine.phase_label(),
            ),
            (
                "Ranking stability",
                f"{self.engine.stability * 100:.1f}%",
            ),
            (
                "Top-25 stability",
                f"{self.engine.top_stability * 100:.1f}%",
            ),
            (
                "Adaptive stopping",
                (
                    "READY"
                    if self.engine.is_adaptively_finished()
                    else "NOT YET"
                ),
            ),
        ]

        for label, value in rows:
            row = ttk.Frame(
                frame
            )

            row.pack(
                fill="x",
                pady=4,
            )

            ttk.Label(
                row,
                text=label,
                width=30,
            ).pack(
                side="left"
            )

            ttk.Label(
                row,
                text=str(value),
            ).pack(
                side="left"
            )

        ttk.Separator(
            frame
        ).pack(
            fill="x",
            pady=15,
        )

        ttk.Label(
            frame,
            text=(
                "How the engine works"
            ),
            font=(
                "Segoe UI",
                13,
                "bold",
            ),
        ).pack(
            anchor="w",
            pady=(0, 8),
        )

        explanation = (
            "1. Exploration gives every book initial evidence.\n"
            "2. Glicko-2 tracks live rating and uncertainty.\n"
            "3. Active selection concentrates comparisons around "
            "uncertain ranking boundaries.\n"
            "4. A regularized Bradley-Terry model periodically "
            "refits the entire comparison history.\n"
            "5. Ranking stability determines when additional "
            "questions stop being worthwhile."
        )

        ttk.Label(
            frame,
            text=explanation,
            justify="left",
            wraplength=620,
        ).pack(
            anchor="w"
        )

    # ========================================================
    # EXPORT
    # ========================================================

    def export(self):
        if (
            self.engine is None
            or self.source_file is None
        ):
            messagebox.showinfo(
                "Nothing to export",
                "Open a Goodreads file first.",
            )

            return

        try:
            # Always make sure final model is current.
            self.engine.fit_bradley_terry()
            self.engine.calculate_stability()

            output_path = export_results(
                self.source_file,
                self.engine,
            )

            self.save_state()

        except PermissionError:
            messagebox.showerror(
                "Could not save",
                (
                    "Windows could not save the ranked "
                    "workbook.\n\n"
                    "If the output file is open in Excel, "
                    "close it and try again."
                ),
            )

            return

        except Exception as exc:
            messagebox.showerror(
                "Export failed",
                str(exc),
            )

            return

        open_file = messagebox.askyesno(
            "Ranking exported",
            (
                "Ranking exported successfully.\n\n"
                f"{output_path}\n\n"
                "Open it now?"
            ),
        )

        if open_file:
            self.open_external(
                output_path
            )

    @staticmethod
    def open_external(
        path: Path,
    ):
        try:
            if sys.platform.startswith(
                "win"
            ):
                os.startfile(
                    str(path)
                )

            elif sys.platform == "darwin":
                subprocess.Popen(
                    [
                        "open",
                        str(path),
                    ]
                )

            else:
                subprocess.Popen(
                    [
                        "xdg-open",
                        str(path),
                    ]
                )

        except Exception:
            pass

    # ========================================================
    # HELP
    # ========================================================

    def show_shortcuts(self):
        messagebox.showinfo(
            "Keyboard shortcuts",
            (
                "1   Choose left book\n"
                "2   Choose right book\n"
                "3   Tie / equal\n"
                "4   Skip pair\n"
                "←   Choose left book\n"
                "→   Choose right book\n"
                "T   Tie / equal\n"
                "S   Skip pair\n"
                "U   Undo\n"
                "Ctrl+O   Open Goodreads Excel\n"
                "Ctrl+E   Export ranking"
            ),
        )

    def show_about(self):
        messagebox.showinfo(
            "About",
            (
                f"{APP_NAME}\n"
                f"Version {APP_VERSION}\n\n"
                "A large-library adaptive pairwise "
                "ranking engine for Goodreads shelves.\n\n"
                "ENGINE:\n"
                "• Glicko-2 live uncertainty tracking\n"
                "• Active-learning pair selection\n"
                "• Swiss-style initial exploration\n"
                "• Regularized Bradley-Terry global ranking\n"
                "• Adaptive early stopping\n"
                "• Ranking stability analysis\n"
                "• Top-ranking refinement\n"
                "• Goodreads My Rating as a weak prior\n\n"
                "PERFORMANCE:\n"
                "• Cached comparison indexes\n"
                "• No repeated pair scans\n"
                "• Fast undo snapshots\n"
                "• Sparse comparison fitting\n"
                "• Efficient Excel export lookup\n\n"
                "Only rows where \"Exclusive Shelf\" "
                "equals \"to-read\" are included.\n\n"
                "Your original Goodreads workbook is "
                "never modified."
            ),
        )

    # ========================================================
    # CLOSE
    # ========================================================

    def on_close(self):
        self.save_state()
        self.root.destroy()


# ============================================================
# MAIN
# ============================================================

def main():
    root = tk.Tk()

    RankerApp(
        root
    )

    root.mainloop()


if __name__ == "__main__":
    try:
        main()

    except Exception:
        error = traceback.format_exc()

        print(error)

        try:
            messagebox.showerror(
                "Unexpected error",
                (
                    "The application encountered an "
                    "unexpected error:\n\n"
                    + error
                ),
            )

        except Exception:
            pass