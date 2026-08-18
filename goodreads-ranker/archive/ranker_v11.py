from __future__ import annotations

import copy
import json
import math
import os
import random
import re
import subprocess
import sys
import tempfile
import time
import traceback
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

APP_NAME = "Goodreads To-Read Ranker"
APP_VERSION = "4.0"

# ------------------------------------------------------------
# Glicko defaults
# ------------------------------------------------------------

DEFAULT_RATING = 1500.0
DEFAULT_RD = 350.0
DEFAULT_VOLATILITY = 0.06

GLICKO_SCALE = 173.7178
TAU = 0.5

# ------------------------------------------------------------
# Comparison budgets
#
# These are MAXIMUM average-ish targets, not rigid quotas.
# The adaptive engine can finish earlier when the ranking is
# sufficiently stable.
# ------------------------------------------------------------

MIN_TARGET_COMPARISONS = 3
MAX_TARGET_COMPARISONS = 20

QUICK_COMPARISONS = 3
BALANCED_COMPARISONS = 6
ACCURATE_COMPARISONS = 12

DEFAULT_TARGET_COMPARISONS = BALANCED_COMPARISONS

# ------------------------------------------------------------
# Adaptive ranking
# ------------------------------------------------------------

MIN_INITIAL_COMPARISONS = 2

# Refit Bradley-Terry every N decisions.
BT_REFRESH_INTERVAL = 12

# Recalculate stability every N decisions.
STABILITY_INTERVAL = 12

# Minimum total decisions before adaptive early stopping.
MIN_ADAPTIVE_DECISIONS = 50

# How stable the ranking should be before allowing an
# automatic early finish.
DEFAULT_STABILITY_THRESHOLD = 0.985

# More stringent requirement for the very top of the ranking.
TOP_STABILITY_THRESHOLD = 0.975

# Don't stop until at least this many books have enough evidence.
MIN_COVERAGE_FOR_STOP = 0.92

# Maximum number of candidate pairs examined when selecting
# the next pair.
PAIR_CANDIDATE_LIMIT = 5000

# Small exploration probability prevents the active learner
# from becoming too deterministic.
EXPLORATION_RATE = 0.055

# Every so often, force a long-range comparison to keep the
# comparison graph connected.
LONG_RANGE_RATE = 0.025

# ------------------------------------------------------------
# Bradley-Terry regularization
# ------------------------------------------------------------

BT_MAX_ITERATIONS = 80
BT_TOLERANCE = 1e-5

# L2 penalty on latent preference strength.
# Larger = more conservative / less extreme.
BT_REGULARIZATION = 0.035

# Convert BT latent strength into a familiar rating scale.
BT_RATING_SCALE = 173.7178

# ------------------------------------------------------------
# Goodreads star prior
# ------------------------------------------------------------

USE_GOODREADS_PRIOR = True

# Strength of prior from My Rating.
#
# This is deliberately modest. Pairwise preferences should
# dominate the final ranking.
GOODREADS_PRIOR_STRENGTH = 0.15

# ------------------------------------------------------------
# State
# ------------------------------------------------------------

STATE_DIRECTORY_NAME = ".ranker_state"
STATE_VERSION = 6

# ------------------------------------------------------------
# UI
# ------------------------------------------------------------

WINDOW_WIDTH = 1280
WINDOW_HEIGHT = 850


# ============================================================
# GENERAL UTILITIES
# ============================================================

def normalize(value) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def normalize_header(value) -> str:
    return normalize(value).replace("\ufeff", "").strip().lower()


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def safe_filename(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*]', "_", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value or "ranking"


def truncate(text: str, length: int) -> str:
    text = normalize(text)

    if len(text) <= length:
        return text

    return text[:length - 1].rstrip() + "…"


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def safe_exp(value: float) -> float:
    if value > 700:
        return math.exp(700)

    if value < -700:
        return math.exp(-700)

    return math.exp(value)


def sigmoid(value: float) -> float:
    if value >= 0:
        z = math.exp(-min(value, 700))
        return 1.0 / (1.0 + z)

    z = math.exp(max(value, -700))
    return z / (1.0 + z)


def parse_goodreads_rating(value: str) -> Optional[float]:
    """
    Goodreads "My Rating" is normally an integer from 0 to 5.
    Return None when unavailable/invalid.
    """
    value = normalize(value)

    if not value:
        return None

    try:
        rating = float(value)
    except ValueError:
        return None

    if rating <= 0:
        return None

    return clamp(rating, 1.0, 5.0)


# ============================================================
# GLICKO-2
# ============================================================

@dataclass
class Rating:
    rating: float = DEFAULT_RATING
    rd: float = DEFAULT_RD
    volatility: float = DEFAULT_VOLATILITY

    comparisons: int = 0
    wins: int = 0
    losses: int = 0
    ties: int = 0


def to_glicko(rating: Rating):
    return (
        (rating.rating - 1500.0) / GLICKO_SCALE,
        rating.rd / GLICKO_SCALE,
    )


def from_glicko(
    mu: float,
    phi: float,
    volatility: float,
) -> Rating:
    return Rating(
        rating=1500.0 + GLICKO_SCALE * mu,
        rd=max(1.0, GLICKO_SCALE * phi),
        volatility=volatility,
    )


def g_function(phi: float) -> float:
    return 1.0 / math.sqrt(
        1.0
        + 3.0 * phi * phi / (math.pi * math.pi)
    )


def expected_score(
    mu: float,
    opponent_mu: float,
    opponent_phi: float,
) -> float:
    exponent = -g_function(opponent_phi) * (
        mu - opponent_mu
    )

    if exponent > 700:
        return 0.0

    if exponent < -700:
        return 1.0

    return 1.0 / (1.0 + math.exp(exponent))


def volatility_function(
    x: float,
    delta: float,
    phi: float,
    variance: float,
    a: float,
) -> float:
    exp_x = safe_exp(x)

    base = (
        phi * phi
        + variance
        + exp_x
    )

    denominator = 2.0 * base * base

    numerator = exp_x * (
        delta * delta
        - phi * phi
        - variance
        - exp_x
    )

    return (
        numerator / denominator
        - (x - a) / (TAU * TAU)
    )


def calculate_volatility(
    phi: float,
    volatility: float,
    delta: float,
    variance: float,
) -> float:
    a = math.log(
        max(volatility * volatility, 1e-12)
    )

    A = a

    if delta * delta > phi * phi + variance:
        B = math.log(
            max(
                delta * delta
                - phi * phi
                - variance,
                1e-12,
            )
        )

    else:
        k = 1

        while True:
            B = a - k * TAU

            value = volatility_function(
                B,
                delta,
                phi,
                variance,
                a,
            )

            if value >= 0:
                break

            k += 1

            if k > 100:
                B = a - 100 * TAU
                break

    f_a = volatility_function(
        A,
        delta,
        phi,
        variance,
        a,
    )

    f_b = volatility_function(
        B,
        delta,
        phi,
        variance,
        a,
    )

    for _ in range(100):
        if abs(B - A) < 1e-8:
            break

        denominator = f_b - f_a

        if abs(denominator) < 1e-15:
            break

        C = A + (
            (A - B) * f_a / denominator
        )

        f_c = volatility_function(
            C,
            delta,
            phi,
            variance,
            a,
        )

        if f_c * f_b < 0:
            A = B
            f_a = f_b

        else:
            f_a /= 2.0

        B = C
        f_b = f_c

    result = math.exp(A / 2.0)

    return clamp(
        result,
        0.01,
        1.0,
    )


def glicko_update(
    player: Rating,
    opponents: list[Rating],
    scores: list[float],
) -> Rating:
    if not opponents:
        result = copy.deepcopy(player)

        phi = player.rd / GLICKO_SCALE

        phi_star = math.sqrt(
            phi * phi
            + player.volatility
            * player.volatility
        )

        result.rd = min(
            GLICKO_SCALE * phi_star,
            350.0,
        )

        return result

    mu, phi = to_glicko(player)

    variance_inverse = 0.0
    score_sum = 0.0

    for opponent, score in zip(
        opponents,
        scores,
    ):
        opponent_mu, opponent_phi = (
            to_glicko(opponent)
        )

        g_value = g_function(
            opponent_phi
        )

        expected = expected_score(
            mu,
            opponent_mu,
            opponent_phi,
        )

        variance_inverse += (
            g_value
            * g_value
            * expected
            * (1.0 - expected)
        )

        score_sum += (
            g_value
            * (score - expected)
        )

    if variance_inverse <= 0:
        return copy.deepcopy(player)

    variance = 1.0 / variance_inverse

    delta = variance * score_sum

    new_volatility = calculate_volatility(
        phi,
        player.volatility,
        delta,
        variance,
    )

    phi_star = math.sqrt(
        phi * phi
        + new_volatility
        * new_volatility
    )

    new_phi = 1.0 / math.sqrt(
        1.0 / (phi_star * phi_star)
        + 1.0 / variance
    )

    new_mu = (
        mu
        + new_phi * new_phi * score_sum
    )

    return from_glicko(
        new_mu,
        new_phi,
        new_volatility,
    )


# ============================================================
# BOOK MODEL
# ============================================================

@dataclass
class Book:
    id: str
    title: str
    author: str

    pages: str = ""
    year: str = ""
    publication_year: str = ""
    description: str = ""
    my_rating: str = ""
    isbn: str = ""
    book_id: str = ""

    original_row: int = 0
    original: dict = None

    def __post_init__(self):
        if self.original is None:
            self.original = {}


# ============================================================
# GOODREADS IMPORTER
# ============================================================

def find_header(headers: list, wanted: str):
    wanted = normalize_header(wanted)

    for header in headers:
        if normalize_header(header) == wanted:
            return header

    return None


def find_first_header(
    headers: list,
    candidates: list[str],
):
    for candidate in candidates:
        found = find_header(
            headers,
            candidate,
        )

        if found is not None:
            return found

    return None


def load_goodreads(path: Path):
    if not path.exists():
        raise FileNotFoundError(
            f"File not found:\n{path}"
        )

    if path.suffix.lower() != ".xlsx":
        raise ValueError(
            "Please select a Goodreads .xlsx export."
        )

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook.active

        rows = worksheet.iter_rows(
            values_only=True
        )

        try:
            headers = list(next(rows))

        except StopIteration:
            raise ValueError(
                "The Excel file is empty."
            )

        shelf_header = find_header(
            headers,
            "Exclusive Shelf",
        )

        title_header = find_header(
            headers,
            "Title",
        )

        if shelf_header is None:
            raise ValueError(
                'Could not find the "Exclusive Shelf" column.'
            )

        if title_header is None:
            raise ValueError(
                'Could not find the "Title" column.'
            )

        author_header = find_first_header(
            headers,
            [
                "Author l-f",
                "Author",
                "Authors",
                "Author Name",
            ],
        )

        header_index = {
            header: index
            for index, header in enumerate(headers)
            if header is not None
        }

        def get_value(values, header):
            if header is None:
                return ""

            index = header_index.get(header)

            if index is None:
                return ""

            if index >= len(values):
                return ""

            return normalize(
                values[index]
            )

        pages_header = find_header(
            headers,
            "Pages",
        )

        year_header = find_header(
            headers,
            "Year Published",
        )

        original_year_header = find_header(
            headers,
            "Original Publication Year",
        )

        description_header = find_header(
            headers,
            "Description",
        )

        my_rating_header = find_header(
            headers,
            "My Rating",
        )

        isbn_header = find_header(
            headers,
            "ISBN",
        )

        book_id_header = find_header(
            headers,
            "Book Id - Goodreads",
        )

        first_name_header = find_header(
            headers,
            "Author First Name",
        )

        last_name_header = find_header(
            headers,
            "Author Last Name",
        )

        books = []
        seen_book_ids = set()

        excel_row_number = 2

        for raw_values in rows:
            values = list(raw_values)

            while len(values) < len(headers):
                values.append(None)

            shelf = get_value(
                values,
                shelf_header,
            ).lower()

            if shelf != "to-read":
                excel_row_number += 1
                continue

            title = get_value(
                values,
                title_header,
            )

            if not title:
                excel_row_number += 1
                continue

            author = ""

            if author_header is not None:
                author = get_value(
                    values,
                    author_header,
                )

            if not author:
                first_name = get_value(
                    values,
                    first_name_header,
                )

                last_name = get_value(
                    values,
                    last_name_header,
                )

                author = " ".join(
                    part
                    for part in (
                        first_name,
                        last_name,
                    )
                    if part
                )

            goodreads_id = get_value(
                values,
                book_id_header,
            )

            if goodreads_id:
                candidate_id = "goodreads:" + goodreads_id
                if candidate_id in seen_book_ids:
                    # Goodreads IDs should be unique. Keep both rows rather than
                    # silently merging them; the duplicate receives a stable row ID.
                    book_id = "row:" + str(excel_row_number)
                else:
                    book_id = candidate_id
            else:
                book_id = "row:" + str(excel_row_number)

            while book_id in seen_book_ids:
                book_id = "row:" + str(excel_row_number) + ":dup"
            seen_book_ids.add(book_id)

            original = {}

            for index, header in enumerate(
                headers
            ):
                if header is None:
                    continue

                value = (
                    values[index]
                    if index < len(values)
                    else None
                )

                if isinstance(
                    value,
                    datetime,
                ):
                    value = value.isoformat()

                original[str(header)] = value

            books.append(
                Book(
                    id=book_id,
                    title=title,
                    author=author,
                    pages=get_value(
                        values,
                        pages_header,
                    ),
                    year=get_value(
                        values,
                        year_header,
                    ),
                    publication_year=get_value(
                        values,
                        original_year_header,
                    ),
                    description=get_value(
                        values,
                        description_header,
                    ),
                    my_rating=get_value(
                        values,
                        my_rating_header,
                    ),
                    isbn=get_value(
                        values,
                        isbn_header,
                    ),
                    book_id=goodreads_id,
                    original_row=excel_row_number,
                    original=original,
                )
            )

            excel_row_number += 1

        return headers, books

    finally:
        workbook.close()


# ============================================================
# RANKING ENGINE
# ============================================================

@dataclass
class RankingConfig:
    name: str = "BALANCED"
    top_k: int = 25
    top10_k: int = 10
    top50_k: int = 50
    elite_pool_min: int = 50
    elite_pool_max: int = 100
    boundary_width: int = 10
    top10_boundary: int = 8
    min_comparisons: int = 1
    target_comparisons: int = 6
    global_budget: int = 0
    simulation_count: int = 160
    simulation_refresh: int = 6
    exploration_rate: float = 0.055
    long_range_rate: float = 0.02
    challenger_rate: float = 0.18
    max_pair_candidates: int = 5000
    local_window: int = 10
    high_rd_pool: int = 80
    under_tested_pool: int = 100
    max_undo: int = 100
    top25_membership_threshold: float = 0.90
    top10_membership_threshold: float = 0.88
    top25_stability_threshold: float = 0.90
    top10_stability_threshold: float = 0.90
    rank_swap_probability: float = 0.12
    stop_min_decisions: int = 40
    stop_min_coverage: float = 0.70
    bt_refresh_interval: int = 12
    bt_max_iterations: int = 60
    bt_tolerance: float = 1e-5
    bt_regularization: float = 0.035
    glicko_rating_scale: float = GLICKO_SCALE


PRESETS = {
    "QUICK": RankingConfig(
        name="QUICK", target_comparisons=3, global_budget=0,
        simulation_count=96, simulation_refresh=8,
        elite_pool_min=40, elite_pool_max=60,
        boundary_width=7, local_window=7,
        stop_min_decisions=25, stop_min_coverage=0.45,
    ),
    "BALANCED": RankingConfig(
        name="BALANCED", target_comparisons=6, global_budget=0,
        simulation_count=160, simulation_refresh=6,
        elite_pool_min=50, elite_pool_max=80,
        boundary_width=10, local_window=10,
        stop_min_decisions=40, stop_min_coverage=0.60,
    ),
    "ACCURATE": RankingConfig(
        name="ACCURATE", target_comparisons=12, global_budget=0,
        simulation_count=240, simulation_refresh=5,
        elite_pool_min=70, elite_pool_max=120,
        boundary_width=14, local_window=14,
        stop_min_decisions=70, stop_min_coverage=0.75,
        top25_membership_threshold=0.94,
        top10_membership_threshold=0.92,
    ),
    "TOP_25_FOCUS": RankingConfig(
        name="TOP_25_FOCUS", target_comparisons=8, global_budget=0,
        simulation_count=220, simulation_refresh=4,
        elite_pool_min=70, elite_pool_max=120,
        boundary_width=18, top10_boundary=12,
        local_window=16, high_rd_pool=100, under_tested_pool=120,
        challenger_rate=0.28, stop_min_decisions=55,
        stop_min_coverage=0.45,
        top25_membership_threshold=0.94,
        top10_membership_threshold=0.92,
        top25_stability_threshold=0.94,
        top10_stability_threshold=0.94,
    ),
    "MAX_ACCURACY": RankingConfig(
        name="MAX_ACCURACY", target_comparisons=18, global_budget=0,
        simulation_count=360, simulation_refresh=3,
        elite_pool_min=100, elite_pool_max=160,
        boundary_width=22, top10_boundary=16,
        local_window=20, high_rd_pool=140, under_tested_pool=160,
        challenger_rate=0.35, stop_min_decisions=100,
        stop_min_coverage=0.55,
        top25_membership_threshold=0.97,
        top10_membership_threshold=0.96,
        top25_stability_threshold=0.96,
        top10_stability_threshold=0.96,
    ),
}


def make_config(name: str, target: Optional[int] = None) -> RankingConfig:
    key = str(name or "BALANCED").upper()
    if key not in PRESETS:
        key = "BALANCED"
    config = copy.deepcopy(PRESETS[key])
    if target is not None:
        config.target_comparisons = int(clamp(int(target), MIN_TARGET_COMPARISONS, MAX_TARGET_COMPARISONS))
    return config


class RankingEngine:
    """Large-library, top-K-focused pairwise ranking engine.

    Glicko-2 supplies the online rating/uncertainty state.  A sparse,
    regularised Bradley-Terry fit supplies a globally coherent point ranking.
    Monte-Carlo rank simulation estimates top-K membership and rank intervals.
    Pair selection is active learning: it samples bounded candidate pools and
    spends human decisions where the result can change the important part of
    the ordering.  It deliberately never materialises the O(N²) pair set.
    """

    def __init__(self, books: list[Book], target_comparisons: int = DEFAULT_TARGET_COMPARISONS,
                 random_seed: Optional[int] = None, mode: str = "BALANCED"):
        self.books = books
        self.seed = random_seed if random_seed is not None else random.SystemRandom().randrange(1, 2**31 - 1)
        self.random = random.Random(self.seed)
        self.mode = str(mode or "BALANCED").upper()
        self.config = make_config(self.mode, target_comparisons)
        self.target_comparisons = self.config.target_comparisons
        self.ratings = {book.id: Rating() for book in books}
        self.comparisons: list[dict] = []
        self.counts = {book.id: 0 for book in books}
        self.wins = {book.id: 0.0 for book in books}
        self.played: set[str] = set()
        self.recent_skips: dict[str, int] = {}
        self.undo_snapshots: list[dict] = []
        self.bt_strengths = {book.id: 1.0 for book in books}
        self.bt_ratings = {book.id: DEFAULT_RATING for book in books}
        self.bt_dirty = True
        self.last_bt_comparison_count = 0
        self.selection_counter = 0
        self.phase = "exploration"
        self.stability = 0.0
        self.top_stability = 0.0
        self.top10_stability = 0.0
        self.last_stable_order: list[str] = []
        self.last_top25: set[str] = set()
        self.last_top10: set[str] = set()
        self.last_stability_check = 0
        self.pair_selection_count = 0
        self.rejected_selection_count = 0
        self._book_map = {book.id: book for book in books}
        self._sorted_cache: list[dict] = []
        self._sorted_dirty = True
        self._analysis_cache: Optional[dict] = None
        self._analysis_comparison_count = -1
        self._analysis_model_signature = None
        self._ranking_cache: Optional[list[dict]] = None
        self._rank_map_cache: Optional[dict[str, int]] = None
        self._ranking_comparison_count = -1
        self._skipped_since_pair: set[str] = set()
        self._last_selected_pair: Optional[str] = None
        self.rebuild_indexes()

    # --------------------------- indexes ---------------------------
    def book_map(self):
        return self._book_map

    @staticmethod
    def pair_key(first: str, second: str) -> str:
        return "|".join(sorted((str(first), str(second))))

    def rebuild_indexes(self):
        self.counts = {book.id: 0 for book in self.books}
        self.wins = {book.id: 0.0 for book in self.books}
        self.played = set()
        for match in self.comparisons:
            left, right, result = match.get("left"), match.get("right"), match.get("result")
            if left not in self.ratings or right not in self.ratings or left == right:
                continue
            self.counts[left] += 1
            self.counts[right] += 1
            self.played.add(self.pair_key(left, right))
            if result == "left":
                self.wins[left] += 1.0
            elif result == "right":
                self.wins[right] += 1.0
            elif result == "tie":
                self.wins[left] += 0.5
                self.wins[right] += 0.5
        self._sorted_dirty = True
        self._ranking_cache = None
        self._rank_map_cache = None
        self._analysis_cache = None

    def played_pairs(self):
        return self.played

    def comparison_counts(self):
        return self.counts

    # --------------------------- state / undo ---------------------------
    def make_snapshot(self):
        return {
            "ratings": copy.deepcopy(self.ratings),
            "comparisons": copy.deepcopy(self.comparisons),
            "recent_skips": dict(self.recent_skips),
            "selection_counter": self.selection_counter,
            "mode": self.mode,
            "target_comparisons": self.target_comparisons,
        }

    def restore_snapshot(self, snapshot):
        self.ratings = copy.deepcopy(snapshot["ratings"])
        self.comparisons = copy.deepcopy(snapshot["comparisons"])
        self.recent_skips = dict(snapshot.get("recent_skips", {}))
        self.selection_counter = int(snapshot.get("selection_counter", len(self.comparisons)))
        self.mode = snapshot.get("mode", self.mode)
        self.config = make_config(self.mode, int(snapshot.get("target_comparisons", self.target_comparisons)))
        self.target_comparisons = self.config.target_comparisons
        self.bt_dirty = True
        self.last_bt_comparison_count = 0
        self.rebuild_indexes()
        self.fit_bradley_terry(force=True)
        self.calculate_stability(force=True)

    def undo(self) -> bool:
        if not self.undo_snapshots:
            return False
        self.restore_snapshot(self.undo_snapshots.pop())
        return True

    # --------------------------- model ---------------------------
    def goodreads_prior_beta(self, book: Book) -> float:
        if not USE_GOODREADS_PRIOR:
            return 0.0
        stars = parse_goodreads_rating(book.my_rating)
        if stars is None:
            return 0.0
        return (stars - 3.0) * GOODREADS_PRIOR_STRENGTH

    def apply_match(self, left_id: str, right_id: str, result: str):
        if left_id not in self.ratings or right_id not in self.ratings:
            raise ValueError("Unknown book in comparison.")
        if left_id == right_id:
            raise ValueError("A book cannot be compared with itself.")
        if result not in {"left", "right", "tie"}:
            raise ValueError("Invalid comparison result.")
        pair = self.pair_key(left_id, right_id)
        if pair in self.played:
            raise ValueError("That pair has already been compared.")

        self.undo_snapshots.append(self.make_snapshot())
        if len(self.undo_snapshots) > self.config.max_undo:
            del self.undo_snapshots[:-self.config.max_undo]

        old_left = copy.deepcopy(self.ratings[left_id])
        old_right = copy.deepcopy(self.ratings[right_id])
        scores = (1.0, 0.0) if result == "left" else ((0.0, 1.0) if result == "right" else (0.5, 0.5))
        self.ratings[left_id] = glicko_update(old_left, [old_right], [scores[0]])
        self.ratings[right_id] = glicko_update(old_right, [old_left], [scores[1]])
        self.ratings[left_id].comparisons = old_left.comparisons + 1
        self.ratings[right_id].comparisons = old_right.comparisons + 1
        self.ratings[left_id].wins, self.ratings[left_id].losses, self.ratings[left_id].ties = old_left.wins, old_left.losses, old_left.ties
        self.ratings[right_id].wins, self.ratings[right_id].losses, self.ratings[right_id].ties = old_right.wins, old_right.losses, old_right.ties
        if result == "left":
            self.ratings[left_id].wins += 1; self.ratings[right_id].losses += 1
            self.wins[left_id] += 1.0
        elif result == "right":
            self.ratings[right_id].wins += 1; self.ratings[left_id].losses += 1
            self.wins[right_id] += 1.0
        else:
            self.ratings[left_id].ties += 1; self.ratings[right_id].ties += 1
            self.wins[left_id] += 0.5; self.wins[right_id] += 0.5
        self.counts[left_id] += 1
        self.counts[right_id] += 1
        self.played.add(pair)
        self.comparisons.append({"left": left_id, "right": right_id, "result": result, "timestamp": now_iso()})
        self.selection_counter += 1
        self.recent_skips.pop(pair, None)
        self.bt_dirty = True
        self._sorted_dirty = True
        self._ranking_cache = None
        # Keep the previous simulation cache for a few decisions; active learning
        # can safely reuse it until the configured refresh interval.
        self.update_phase()
        if len(self.comparisons) - self.last_bt_comparison_count >= self.config.bt_refresh_interval:
            self.fit_bradley_terry()
        if len(self.comparisons) - self.last_stability_check >= STABILITY_INTERVAL:
            self.calculate_stability()

    def rebuild_from_history(self):
        history = copy.deepcopy(self.comparisons)
        self.ratings = {book.id: Rating() for book in self.books}
        self.comparisons = []
        self.recent_skips = {}
        self.rebuild_indexes()
        self.undo_snapshots = []
        for match in history:
            try:
                self.apply_match(match["left"], match["right"], match["result"])
            except Exception:
                continue
        self.undo_snapshots = []
        self.fit_bradley_terry(force=True)
        self.calculate_stability(force=True)

    def fit_bradley_terry(self, force: bool = False):
        if not self.books:
            return
        if not force and not self.bt_dirty and self.last_bt_comparison_count == len(self.comparisons):
            return
        n = len(self.books)
        ids = [book.id for book in self.books]
        index = {book_id: i for i, book_id in enumerate(ids)}
        beta = [(self.ratings[b].rating - DEFAULT_RATING) / BT_RATING_SCALE for b in ids]
        for i, book in enumerate(self.books):
            beta[i] += self.goodreads_prior_beta(book) * 0.35
        records = []
        for match in self.comparisons:
            li, ri, result = match.get("left"), match.get("right"), match.get("result")
            if li not in index or ri not in index or result not in {"left", "right", "tie"}:
                continue
            score = 1.0 if result == "left" else (0.0 if result == "right" else 0.5)
            records.append((index[li], index[ri], score))
        if not records:
            self.bt_strengths = {book.id: 1.0 for book in self.books}
            self.bt_ratings = {book.id: self.ratings[book.id].rating for book in self.books}
            self.bt_dirty = False
            self.last_bt_comparison_count = 0
            self._ranking_cache = None
            return

        reg = max(self.config.bt_regularization, 1e-5)
        for _ in range(self.config.bt_max_iterations):
            gradient = [-reg * b for b in beta]
            hessian = [reg for _ in range(n)]
            for i, j, score in records:
                p = sigmoid(beta[i] - beta[j])
                # A tie contributes half a win to each side but is deliberately
                # weaker evidence than a decisive result.
                weight = 0.55 if score == 0.5 else 1.0
                error = weight * (score - p)
                curvature = weight * max(p * (1.0 - p), 1e-5)
                gradient[i] += error
                gradient[j] -= error
                hessian[i] += curvature
                hessian[j] += curvature
            max_change = 0.0
            for i in range(n):
                step = clamp(gradient[i] / max(hessian[i], 1e-8), -0.30, 0.30)
                beta[i] += step
                max_change = max(max_change, abs(step))
            mean_beta = sum(beta) / n
            beta = [b - mean_beta for b in beta]
            if max_change < self.config.bt_tolerance:
                break
        strengths = {}
        ratings = {}
        for i, book_id in enumerate(ids):
            b = clamp(beta[i], -8.0, 8.0)
            strengths[book_id] = safe_exp(b)
            ratings[book_id] = DEFAULT_RATING + BT_RATING_SCALE * b
        self.bt_strengths = strengths
        self.bt_ratings = ratings
        self.bt_dirty = False
        self.last_bt_comparison_count = len(self.comparisons)
        self._ranking_cache = None
        self._rank_map_cache = None
        self._analysis_cache = None

    def ensure_model(self):
        if self.bt_dirty:
            self.fit_bradley_terry()

    # --------------------------- ranking ---------------------------
    def ranking(self):
        self.ensure_model()
        if self._ranking_cache is not None and self._ranking_comparison_count == len(self.comparisons):
            return self._ranking_cache
        result = []
        for book in self.books:
            rating = self.ratings[book.id]
            final_rating = self.bt_ratings.get(book.id, rating.rating)
            result.append({"book": book, "rating": rating, "final_rating": final_rating,
                           "strength": self.bt_strengths.get(book.id, 1.0)})
        result.sort(key=lambda item: (item["final_rating"], item["rating"].rating, -item["rating"].rd,
                                      item["book"].title.casefold()), reverse=True)
        self._ranking_cache = result
        self._rank_map_cache = {item["book"].id: i + 1 for i, item in enumerate(result)}
        self._ranking_comparison_count = len(self.comparisons)
        return result

    # --------------------------- uncertainty simulation ---------------------------
    def _model_signature(self):
        if not self.books:
            return 0
        return (len(self.comparisons), round(sum(self.ratings[b.id].rating for b in self.books), 3),
                round(sum(self.ratings[b.id].rd for b in self.books), 3), self.mode)

    def _candidate_ids_for_analysis(self):
        ranked = self.ranking()
        n = len(ranked)
        elite_size = min(n, max(self.config.elite_pool_min, min(self.config.elite_pool_max, int(math.sqrt(max(n, 1)) * 2.0) + self.config.top_k)))
        ids = {item["book"].id for item in ranked[:elite_size]}
        ids.update(item["book"].id for item in ranked[:min(n, self.config.top_k + self.config.boundary_width)])
        return ids

    def _simulate_topk(self):
        if not self.books:
            return {"top10": {}, "top25": {}, "top50": {}, "rank_samples": {}, "stability": 1.0}
        ranked = self.ranking()
        n = len(ranked)
        sim_count = max(32, int(self.config.simulation_count))
        top10 = min(self.config.top10_k, n)
        top25 = min(self.config.top_k, n)
        top50 = min(self.config.top50_k, n)
        candidates = self._candidate_ids_for_analysis()
        # Include a small safety band around the current boundary.
        candidate_list = [item["book"].id for item in ranked[:min(n, max(top50 + self.config.boundary_width * 2, len(candidates)))]]
        candidate_set = set(candidate_list) | candidates
        membership10 = {book_id: 0 for book_id in candidate_set}
        membership25 = {book_id: 0 for book_id in candidate_set}
        membership50 = {book_id: 0 for book_id in candidate_set}
        rank_samples = {book_id: [] for book_id in candidate_set}
        current_top25 = [item["book"].id for item in ranked[:top25]]
        current_top10 = [item["book"].id for item in ranked[:top10]]
        current_top25_set = set(current_top25)
        current_top10_set = set(current_top10)
        # Use a private deterministic stream derived from the user seed and history length.
        rng = random.Random((self.seed * 1000003 + len(self.comparisons) * 9176 + sum((i + 1) * ord(ch) for i, ch in enumerate(self.mode))) & 0xFFFFFFFF)
        samples = []
        for book in self.books:
            r = self.ratings[book.id]
            # Glicko RD is a one-sigma-ish uncertainty proxy in rating points.
            sigma = clamp(r.rd * 0.72, 20.0, 350.0)
            center = self.bt_ratings.get(book.id, r.rating)
            samples.append((book.id, center, sigma))
        # Heap keeps only the top-50 simulated books, avoiding full O(N log N) sorting per draw.
        import heapq
        top_heap_limit = max(top50, min(80, top25 + self.config.boundary_width * 2))
        stable_overlap_sum = 0.0
        for _ in range(sim_count):
            heap = []
            for book_id, center, sigma in samples:
                value = center + rng.gauss(0.0, sigma)
                entry = (value, book_id)
                if len(heap) < top_heap_limit:
                    heapq.heappush(heap, entry)
                elif value > heap[0][0]:
                    heapq.heapreplace(heap, entry)
            heap.sort(reverse=True)
            order = [book_id for _, book_id in heap]
            positions = {book_id: i + 1 for i, book_id in enumerate(order)}
            sim_top10 = set(order[:top10])
            sim_top25 = set(order[:top25])
            sim_top50 = set(order[:top50])
            stable_overlap_sum += len(sim_top25 & current_top25_set) / max(1, top25)
            for book_id in candidate_set:
                pos = positions.get(book_id)
                if pos is None:
                    # Outside the retained top heap. We only need a conservative
                    # lower-information rank estimate for these candidates.
                    continue
                rank_samples[book_id].append(pos)
                if book_id in sim_top10:
                    membership10[book_id] += 1
                if book_id in sim_top25:
                    membership25[book_id] += 1
                if book_id in sim_top50:
                    membership50[book_id] += 1
        top10_probs = {k: v / sim_count for k, v in membership10.items()}
        top25_probs = {k: v / sim_count for k, v in membership25.items()}
        top50_probs = {k: v / sim_count for k, v in membership50.items()}
        intervals = {}
        for book_id, values in rank_samples.items():
            if not values:
                current = next((i + 1 for i, x in enumerate(ranked) if x["book"].id == book_id), n)
                intervals[book_id] = (current, current, current)
                continue
            values.sort()
            lo = values[max(0, int(len(values) * 0.10) - 1)]
            med = values[int(len(values) * 0.50)]
            hi = values[min(len(values) - 1, int(len(values) * 0.90))]
            intervals[book_id] = (lo, med, hi)
        # Top-K stability is membership stability, not false certainty about exact order.
        membership_stability = stable_overlap_sum / sim_count
        return {"top10": top10_probs, "top25": top25_probs, "top50": top50_probs,
                "rank_samples": intervals, "stability": membership_stability,
                "current_top25": current_top25, "current_top10": current_top10,
                "simulations": sim_count}

    def analysis(self, force: bool = False):
        self.ensure_model()
        sig = self._model_signature()
        if not force and self._analysis_cache is not None and self._analysis_comparison_count == len(self.comparisons) and self._analysis_model_signature == sig:
            return self._analysis_cache
        # During active interaction, avoid rerunning simulations after every click.
        if not force and self._analysis_cache is not None and len(self.comparisons) - self._analysis_comparison_count < self.config.simulation_refresh:
            return self._analysis_cache
        self._analysis_cache = self._simulate_topk()
        self._analysis_comparison_count = len(self.comparisons)
        self._analysis_model_signature = sig
        return self._analysis_cache

    def book_stats(self, book_id: str):
        ranking = self.ranking()
        if self._rank_map_cache is None:
            self._rank_map_cache = {item["book"].id: i + 1 for i, item in enumerate(ranking)}
        rank_map = self._rank_map_cache
        analysis = self.analysis()
        rank = rank_map.get(book_id, len(ranking))
        interval = analysis.get("rank_samples", {}).get(book_id, (rank, rank, rank))
        return {
            "rank": rank,
            "expected_rank": interval[1],
            "rank_low": interval[0],
            "rank_high": interval[2],
            "top10_probability": analysis.get("top10", {}).get(book_id, 0.0),
            "top25_probability": analysis.get("top25", {}).get(book_id, 0.0),
            "top50_probability": analysis.get("top50", {}).get(book_id, 0.0),
        }

    # --------------------------- active learning ---------------------------
    def probability_left_wins(self, left: Rating, right: Rating) -> float:
        # Bradley-Terry/Glicko hybrid: point difference plus both RDs.
        scale = math.sqrt(BT_RATING_SCALE ** 2 + 0.25 * (left.rd ** 2 + right.rd ** 2))
        return sigmoid((left.rating - right.rating) / max(scale, 1.0))

    def _normal_entropy(self, p: float) -> float:
        p = clamp(p, 1e-9, 1 - 1e-9)
        return -(p * math.log2(p) + (1 - p) * math.log2(1 - p))

    def _pair_candidate_score(self, left: Book, right: Book, ranks: dict[str, int], analysis: dict) -> float:
        lr, rr = self.ratings[left.id], self.ratings[right.id]
        p = self.probability_left_wins(lr, rr)
        entropy = self._normal_entropy(p)
        combined_rd = math.sqrt(lr.rd * lr.rd + rr.rd * rr.rd) / math.sqrt(2.0)
        uncertainty = 0.35 + clamp(combined_rd / 220.0, 0.0, 1.75)
        lrank, rrank = ranks.get(left.id, len(self.books)), ranks.get(right.id, len(self.books))
        best, worst = min(lrank, rrank), max(lrank, rrank)
        gap = abs(lrank - rrank)
        boundary = 1.0 + 2.8 * math.exp(-((best - self.config.top_k) / max(3.0, self.config.boundary_width)) ** 2)
        if best <= self.config.top10_k:
            boundary += 1.8 * math.exp(-(gap / max(3.0, self.config.top10_boundary)) ** 2)
        elif best <= self.config.top_k:
            boundary += 1.1
        top_prob = max(analysis.get("top25", {}).get(left.id, 0.0), analysis.get("top25", {}).get(right.id, 0.0))
        top_relevance = 0.65 + 2.0 * top_prob
        challenger = 1.0
        if lrank > self.config.top_k or rrank > self.config.top_k:
            challenger = 1.0 + 1.8 * max(analysis.get("top25", {}).get(left.id, 0.0), analysis.get("top25", {}).get(right.id, 0.0))
        coverage = 1.0 + 0.55 / math.sqrt(1.0 + min(lr.comparisons, rr.comparisons))
        closeness = 1.0 / (1.0 + gap / max(6.0, self.config.boundary_width * 1.5))
        rank_sensitivity = 1.0 + 1.5 * closeness
        count_balance = 1.0 + min(lr.comparisons, rr.comparisons) / max(1.0, max(lr.comparisons, rr.comparisons, 1.0)) * 0.35
        pair_key = self.pair_key(left.id, right.id)
        recent_skip = pair_key in self.recent_skips and (self.selection_counter - self.recent_skips[pair_key]) <= 24
        repeat_penalty = 0.72 if recent_skip else 1.0
        return entropy * uncertainty * boundary * top_relevance * challenger * coverage * rank_sensitivity * count_balance * repeat_penalty

    def _candidate_pool(self):
        ranked = self.ranking()
        n = len(ranked)
        if n < 2:
            return []
        analysis = self.analysis()
        elite_size = min(n, max(self.config.elite_pool_min, min(self.config.elite_pool_max, self.config.top_k + int(math.sqrt(n) * 3))))
        ids = []
        seen = set()
        def add(item):
            bid = item["book"].id
            if bid not in seen:
                seen.add(bid); ids.append(item["book"])
        for item in ranked[:elite_size]: add(item)
        for item in ranked[max(0, self.config.top_k - self.config.boundary_width):min(n, self.config.top_k + self.config.boundary_width + 1)]: add(item)
        if analysis.get("top25"):
            for bid, prob in sorted(analysis["top25"].items(), key=lambda x: x[1], reverse=True)[:self.config.elite_pool_max]:
                book = self._book_map.get(bid)
                if book is not None and prob >= 0.05: add({"book": book})
        uncertain = sorted(self.books, key=lambda b: self.ratings[b.id].rd, reverse=True)[:min(self.config.high_rd_pool, n)]
        for book in uncertain: add({"book": book})
        undertested = sorted(self.books, key=lambda b: (self.counts[b.id], -self.ratings[b.id].rd, b.title.casefold()))[:min(self.config.under_tested_pool, n)]
        for book in undertested: add({"book": book})
        return ids

    def choose_exploration_pair(self):
        """Swiss-style bootstrap with bounded coverage.

        The objective is not to give every book an equal quota. We sample
        enough of the library to create a useful scaffold, while retaining a
        rotating pool of unseen books so a strong challenger can still enter
        the active phase later.
        """
        n = len(self.books)
        if n < 2:
            return None
        ranked = self.ranking()
        # Deterministically rotate through the library so every region has a
        # chance to enter the scaffold, without materialising all pairs.
        cursor = (self.selection_counter * 2) % n
        probe_count = min(n, max(60, self.config.elite_pool_max * 2))
        probe = [ranked[(cursor + i) % n]["book"] for i in range(probe_count)]
        # Prefer books with no evidence, then books with only one result.
        probe.sort(key=lambda b: (self.counts[b.id], -self.ratings[b.id].rd, b.title.casefold()))
        primary = probe[0]
        candidates = [b for b in probe[1:] if self.pair_key(primary.id, b.id) not in self.played]
        if not candidates:
            for primary in probe:
                candidates = [b for b in probe if b.id != primary.id and self.pair_key(primary.id, b.id) not in self.played]
                if candidates:
                    break
        if not candidates:
            return self.choose_any_unplayed_pair()
        # Similar current ratings make the first few comparisons more informative
        # without using any Goodreads metadata as preference evidence.
        candidates.sort(key=lambda b: (abs(self.ratings[primary.id].rating - self.ratings[b.id].rating), self.counts[b.id]))
        return primary.id, candidates[0].id

    def choose_active_pair(self):
        if len(self.books) < 2:
            return None
        ranked = self.ranking()
        ranks = {item["book"].id: i + 1 for i, item in enumerate(ranked)}
        analysis = self.analysis()
        pool = self._candidate_pool()
        if len(pool) < 2:
            return self.choose_any_unplayed_pair()
        candidates = []
        n = len(ranked)
        # Local neighbourhoods in the current ordering are the main candidate generator.
        pool_set = {b.id for b in pool}
        rank_to_book = {ranks[b.id]: b for b in pool}
        sorted_pool = sorted(pool, key=lambda b: ranks[b.id])
        for i, left in enumerate(sorted_pool):
            for j in range(i + 1, min(len(sorted_pool), i + self.config.local_window + 1)):
                right = sorted_pool[j]
                pair = self.pair_key(left.id, right.id)
                if pair in self.played:
                    continue
                score = self._pair_candidate_score(left, right, ranks, analysis)
                candidates.append((score, left, right))
        # Explicit boundary crossovers: a borderline book vs several established elite books.
        boundary_books = [b for b in sorted_pool if abs(ranks[b.id] - self.config.top_k) <= self.config.boundary_width or ranks[b.id] <= self.config.top10_k]
        elite_books = sorted_pool[:min(len(sorted_pool), self.config.top_k + self.config.boundary_width)]
        for left in boundary_books[:self.config.boundary_width * 3 + 20]:
            for right in elite_books[:min(35, len(elite_books))]:
                if left.id == right.id:
                    continue
                pair = self.pair_key(left.id, right.id)
                if pair in self.played:
                    continue
                score = self._pair_candidate_score(left, right, ranks, analysis) * 1.20
                candidates.append((score, left, right))
        # Challengers: high top-25 probability outside the current top 25.
        top_probs = analysis.get("top25", {})
        challengers = [self._book_map[bid] for bid, p in sorted(top_probs.items(), key=lambda x: x[1], reverse=True)
                       if bid in self._book_map and ranks.get(bid, n + 1) > self.config.top_k and p >= 0.08]
        for challenger in challengers[:min(25, len(challengers))]:
            for opponent in elite_books[:min(20, len(elite_books))]:
                if challenger.id == opponent.id:
                    continue
                pair = self.pair_key(challenger.id, opponent.id)
                if pair in self.played:
                    continue
                candidates.append((self._pair_candidate_score(challenger, opponent, ranks, analysis) * 1.45,
                                   challenger, opponent))
        # High-RD books get a few targeted tests, but only against a bounded elite sample.
        high_rd = sorted(self.books, key=lambda b: self.ratings[b.id].rd, reverse=True)[:min(self.config.high_rd_pool, n)]
        for uncertain in high_rd[:40]:
            if ranks.get(uncertain.id, n + 1) > self.config.top_k + self.config.boundary_width and top_probs.get(uncertain.id, 0.0) < 0.05:
                continue
            for opponent in elite_books[:min(12, len(elite_books))]:
                if uncertain.id == opponent.id:
                    continue
                pair = self.pair_key(uncertain.id, opponent.id)
                if pair in self.played:
                    continue
                candidates.append((self._pair_candidate_score(uncertain, opponent, ranks, analysis) * 1.12,
                                   uncertain, opponent))
        if not candidates:
            return self.choose_any_unplayed_pair()
        # Keep memory and sorting bounded.
        if len(candidates) > self.config.max_pair_candidates:
            candidates = sorted(candidates, key=lambda x: x[0], reverse=True)[:self.config.max_pair_candidates]
        candidates.sort(key=lambda x: x[0], reverse=True)
        shortlist = candidates[:min(24, len(candidates))]
        # Deliberate exploration is among high-value pairs, never the entire library.
        if self.random.random() < self.config.exploration_rate:
            selected = self.random.choice(shortlist[:min(12, len(shortlist))])
        else:
            weights = [max(0.001, c[0]) for c in shortlist]
            selected = self.random.choices(shortlist, weights=weights, k=1)[0]
        self.pair_selection_count += 1
        self._last_selected_pair = self.pair_key(selected[1].id, selected[2].id)
        return selected[1].id, selected[2].id

    def choose_any_unplayed_pair(self):
        ranked = self.ranking()
        n = len(ranked)
        # Bounded local windows, then a deterministic pseudo-random sample. No O(N²) fallback.
        books = [x["book"] for x in ranked]
        window = min(40, max(6, int(math.sqrt(max(n, 1)) * 2.5)))
        for i, left in enumerate(books):
            for j in range(i + 1, min(n, i + window + 1)):
                right = books[j]
                if self.pair_key(left.id, right.id) not in self.played:
                    return left.id, right.id
        sample = books if n <= 300 else self.random.sample(books, min(300, n))
        for _ in range(min(1000, len(sample) * 4)):
            left, right = self.random.sample(sample, 2)
            if self.pair_key(left.id, right.id) not in self.played:
                return left.id, right.id
        return None

    def skip_pair(self, left_id: str, right_id: str):
        if left_id not in self.ratings or right_id not in self.ratings or left_id == right_id:
            return
        pair = self.pair_key(left_id, right_id)
        # Skip is never evidence. It is only a short-lived anti-loop signal.
        self.recent_skips[pair] = self.selection_counter
        cutoff = self.selection_counter - 24
        self.recent_skips = {k: v for k, v in self.recent_skips.items() if v >= cutoff}
        self.selection_counter += 1

    def choose_pair(self):
        if len(self.books) < 2 or self.is_finished():
            return None
        self.update_phase()
        if self.phase == "exploration":
            pair = self.choose_exploration_pair()
            if pair is not None:
                return pair
        # Occasionally force a bounded long-range challenge to avoid a disconnected graph.
        if self.random.random() < self.config.long_range_rate and len(self.books) >= 8:
            ranked = [x["book"] for x in self.ranking()]
            elite = ranked[:min(len(ranked), self.config.elite_pool_max)]
            tail_start = max(len(elite), int(len(ranked) * 0.55))
            tail = ranked[tail_start: min(len(ranked), tail_start + 300)]
            if tail:
                for _ in range(30):
                    left = self.random.choice(elite); right = self.random.choice(tail)
                    if left.id != right.id and self.pair_key(left.id, right.id) not in self.played:
                        return left.id, right.id
        return self.choose_active_pair()

    # --------------------------- stability / stopping ---------------------------
    def minimum_comparisons(self) -> int:
        return min((self.counts[b.id] for b in self.books), default=0)

    def average_comparisons(self) -> float:
        return sum(self.counts.values()) / max(1, len(self.books))

    def coverage(self) -> float:
        if not self.books:
            return 0.0
        return sum(1 for b in self.books if self.counts[b.id] >= self.config.min_comparisons) / len(self.books)

    def calculate_stability(self, force: bool = False):
        if not self.books:
            self.stability = self.top_stability = self.top10_stability = 1.0
            return 1.0
        ranking = self.ranking()
        order = [x["book"].id for x in ranking]
        top25 = set(order[:min(self.config.top_k, len(order))])
        top10 = set(order[:min(self.config.top10_k, len(order))])
        if not self.last_stable_order:
            self.last_stable_order = order
            self.last_top25 = top25
            self.last_top10 = top10
            self.stability = 0.0
            self.top_stability = 0.0
            self.top10_stability = 0.0
        else:
            pos = {bid: i for i, bid in enumerate(self.last_stable_order)}
            avg_disp = sum(abs(i - pos.get(bid, i)) for i, bid in enumerate(order)) / max(1, len(order))
            self.stability = clamp(1.0 - avg_disp / max(1.0, len(order) * 0.15), 0.0, 1.0)
            self.top_stability = len(top25 & self.last_top25) / max(1, len(top25))
            self.top10_stability = len(top10 & self.last_top10) / max(1, len(top10))
            self.last_stable_order = order
            self.last_top25 = top25
            self.last_top10 = top10
        self.last_stability_check = len(self.comparisons)
        return self.stability

    def _top25_confidence(self):
        analysis = self.analysis()
        probs = analysis.get("top25", {})
        ranked = self.ranking()
        current = [x["book"].id for x in ranked[:min(self.config.top_k, len(ranked))]]
        if not current:
            return 1.0
        return min(probs.get(bid, 0.0) for bid in current) if probs else 0.0

    def _top10_confidence(self):
        analysis = self.analysis()
        probs = analysis.get("top10", {})
        ranked = self.ranking()
        current = [x["book"].id for x in ranked[:min(self.config.top10_k, len(ranked))]]
        return min((probs.get(bid, 0.0) for bid in current), default=0.0)

    def top25_stability_score(self):
        analysis = self.analysis()
        return clamp(float(analysis.get("stability", self.top_stability)), 0.0, 1.0)

    def unresolved_top25_swaps(self) -> int:
        analysis = self.analysis()
        ranked = self.ranking()
        ids = [x["book"].id for x in ranked]
        count = 0
        boundary = min(len(ids), self.config.top_k + self.config.boundary_width)
        for i in range(min(self.config.top_k, len(ids))):
            bid = ids[i]
            p = analysis.get("top25", {}).get(bid, 0.0)
            if p < self.config.top25_membership_threshold:
                count += 1
        for i in range(self.config.top_k, boundary):
            bid = ids[i]
            if analysis.get("top25", {}).get(bid, 0.0) >= self.config.rank_swap_probability:
                count += 1
        return count

    def estimated_additional_comparisons(self) -> int:
        unresolved = self.unresolved_top25_swaps()
        confidence = self._top25_confidence()
        base = max(0.0, (1.0 - confidence) * 3.5 + unresolved * 1.6)
        if self.phase == "exploration":
            base += max(0, len(self.books) - sum(1 for b in self.books if self.counts[b.id] >= 1)) * 0.35
        return int(math.ceil(base))

    def progress(self):
        if not self.books:
            return 0.0
        analysis = self.analysis()
        top_conf = self._top25_confidence()
        # In Top-25 mode, stability dominates progress; long-tail coverage is deliberately secondary.
        coverage = self.coverage()
        stability = self.top25_stability_score()
        if self.mode == "TOP_25_FOCUS":
            return clamp(0.15 * min(1.0, self.average_comparisons() / max(1, self.target_comparisons)) +
                         0.45 * stability + 0.40 * top_conf, 0.0, 1.0)
        return clamp(0.25 * min(1.0, self.average_comparisons() / max(1, self.target_comparisons)) +
                     0.25 * coverage + 0.25 * stability + 0.25 * top_conf, 0.0, 1.0)

    def all_at_target(self):
        # Compatibility method: target is a ceiling, not a quota in adaptive modes.
        return all(self.counts[b.id] >= self.target_comparisons for b in self.books)

    def is_adaptively_finished(self):
        n = len(self.books)
        if n < 2:
            return True
        if len(self.comparisons) < min(self.config.stop_min_decisions, max(10, n // 2)):
            return False
        if self.coverage() < self.config.stop_min_coverage and n <= 500:
            return False
        top25_conf = self._top25_confidence()
        top10_conf = self._top10_confidence()
        top25_stability = self.top25_stability_score()
        # Also require a reasonably resolved boundary; this is the main guard against premature stopping.
        boundary_ok = self.unresolved_top25_swaps() <= max(2, int(self.config.boundary_width * 0.20))
        return (top25_conf >= self.config.top25_membership_threshold and
                top10_conf >= self.config.top10_membership_threshold and
                top25_stability >= self.config.top25_stability_threshold and
                self.top10_stability >= self.config.top10_stability_threshold and boundary_ok)

    def is_finished(self):
        # Global budget, if explicitly configured, is a hard cap.
        if self.config.global_budget and len(self.comparisons) >= self.config.global_budget:
            return True
        return self.is_adaptively_finished()

    def exploration_target_decisions(self) -> int:
        n = len(self.books)
        if n <= 100:
            return max(20, n)
        # Broad evidence is useful, but Top-25 mode intentionally stops broad
        # coverage much earlier and lets active learning do the expensive work.
        if self.mode == "TOP_25_FOCUS":
            return min(n // 2, max(80, int(0.08 * n) + self.config.elite_pool_max))
        if self.mode == "MAX_ACCURACY":
            return min(n, max(100, int(0.18 * n)))
        return min(n, max(80, int(0.12 * n) + self.config.elite_pool_min))

    def update_phase(self):
        if len(self.books) < 2:
            self.phase = "complete"
            return
        if len(self.comparisons) < self.exploration_target_decisions():
            self.phase = "exploration"
        elif len(self.comparisons) < max(self.config.stop_min_decisions, min(2 * self.config.top_k, self.exploration_target_decisions() + 20)):
            self.phase = "calibration"
        else:
            self.phase = "active"

    # --------------------------- diagnostics ---------------------------
    def top_confidence(self, count=10):
        ranking = self.ranking()
        analysis = self.analysis()
        result = []
        for index, item in enumerate(ranking[:count], start=1):
            stats = self.book_stats(item["book"].id)
            result.append({"rank": index, "book": item["book"], "rating": item["final_rating"],
                           "confidence": stats["top10_probability"] if index <= 10 else stats["top25_probability"],
                           "rd": item["rating"].rd, **stats})
        return result

    def phase_label(self):
        return {"exploration": "Exploring", "calibration": "Calibrating", "active": "Top-K refinement",
                "complete": "Top-25 stable"}.get(self.phase, "Ranking")

    def diagnostics(self):
        return {
            "books": len(self.books), "comparisons": len(self.comparisons),
            "average_comparisons": self.average_comparisons(), "minimum_comparisons": self.minimum_comparisons(),
            "coverage": self.coverage(), "stability": self.stability,
            "top25_stability": self.top25_stability_score(), "top10_stability": self.top10_stability,
            "top25_confidence": self._top25_confidence(), "top10_confidence": self._top10_confidence(),
            "unresolved_boundary": self.unresolved_top25_swaps(),
            "estimated_additional": self.estimated_additional_comparisons(),
            "mode": self.mode, "target_comparisons": self.target_comparisons,
            "seed": self.seed, "phase": self.phase_label(),
        }


# ============================================================
# SAVED STATE
# ============================================================

class StateStore:
    def __init__(self, source_file: Path):
        self.source_file = source_file
        self.directory = source_file.parent / STATE_DIRECTORY_NAME
        self.directory.mkdir(parents=True, exist_ok=True)
        self.path = self.directory / (safe_filename(source_file.stem) + ".json")

    def save(self, engine: RankingEngine):
        data = {
            "version": STATE_VERSION,
            "model_version": APP_VERSION,
            "source_file": str(self.source_file.resolve()),
            "mode": engine.mode,
            "target_comparisons": engine.target_comparisons,
            "seed": engine.seed,
            "books": [asdict(book) for book in engine.books],
            "comparisons": engine.comparisons,
            "recent_skips": engine.recent_skips,
            "saved_at": now_iso(),
        }
        temporary = self.path.with_suffix(".tmp")
        with temporary.open("w", encoding="utf-8") as file:
            json.dump(data, file, indent=2, ensure_ascii=False, separators=(",", ":"))
            file.flush()
            os.fsync(file.fileno())
        temporary.replace(self.path)

    def load(self, books: list[Book]) -> Optional[RankingEngine]:
        if not self.path.exists():
            return None
        try:
            with self.path.open("r", encoding="utf-8") as file:
                data = json.load(file)
            version = int(data.get("version", 0))
            if version > STATE_VERSION:
                return None
            saved_ids = {str(book.get("id")) for book in data.get("books", []) if book.get("id") is not None}
            current_ids = {book.id for book in books}
            if saved_ids != current_ids:
                return None
            target = int(data.get("target_comparisons", DEFAULT_TARGET_COMPARISONS))
            seed = data.get("seed")
            mode = str(data.get("mode", "BALANCED")).upper()
            if mode not in PRESETS:
                mode = "BALANCED"
            engine = RankingEngine(books, target, seed, mode)
            raw_comparisons = data.get("comparisons", [])
            if not isinstance(raw_comparisons, list):
                return None
            valid = []
            seen = set()
            for match in raw_comparisons:
                if not isinstance(match, dict):
                    continue
                left, right, result = match.get("left"), match.get("right"), match.get("result")
                if left not in current_ids or right not in current_ids or left == right or result not in {"left", "right", "tie"}:
                    continue
                key = engine.pair_key(left, right)
                if key in seen:
                    continue
                seen.add(key)
                valid.append({"left": left, "right": right, "result": result, "timestamp": str(match.get("timestamp", ""))})
            engine.comparisons = valid
            saved_skips = {str(k): int(v) for k, v in data.get("recent_skips", {}).items() if isinstance(k, str)}
            engine.rebuild_from_history()
            engine.recent_skips = saved_skips
            return engine
        except Exception:
            return None


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_results(source_file: Path, engine: RankingEngine):
    workbook = load_workbook(filename=source_file)
    try:
        engine.fit_bradley_terry(force=True)
        engine.calculate_stability(force=True)
        ranked_items = engine.ranking()
        analysis = engine.analysis(force=True)

        for sheet_name in ("Ranking", "Summary", "Diagnostics", "Comparisons"):
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

        header_fill = PatternFill(fill_type="solid", fgColor="2F5597")
        header_font = Font(bold=True, color="FFFFFF")

        ranking_sheet = workbook.create_sheet("Ranking", 0)
        ranking_headers = [
            "Rank", "Title", "Author", "Rating", "Rating Uncertainty", "Estimated Rank",
            "Rank Low", "Rank High", "Top 10 Probability", "Top 25 Probability", "Top 50 Probability",
            "Comparisons", "Wins", "Losses", "Ties", "Status", "Goodreads Book Id",
        ]
        ranking_sheet.append(ranking_headers)
        book_stats_cache = {}
        for rank, item in enumerate(ranked_items, start=1):
            book = item["book"]
            rating = item["rating"]
            stats = engine.book_stats(book.id)
            book_stats_cache[book.id] = stats
            p10, p25, p50 = stats["top10_probability"], stats["top25_probability"], stats["top50_probability"]
            status = "Top 10" if rank <= 10 else ("Top 25" if rank <= 25 else ("Borderline" if p25 >= 0.20 else "Long tail"))
            if rating.rd >= 220 or stats["rank_high"] - stats["rank_low"] >= max(8, engine.config.boundary_width):
                status += " · Uncertain"
            ranking_sheet.append([
                rank, book.title, book.author, round(item["final_rating"], 2), round(rating.rd, 2),
                stats["expected_rank"], stats["rank_low"], stats["rank_high"], round(p10 * 100, 1),
                round(p25 * 100, 1), round(p50 * 100, 1), rating.comparisons, rating.wins, rating.losses,
                rating.ties, status, book.book_id,
            ])
        for cell in ranking_sheet[1]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(vertical="center")
        ranking_sheet.freeze_panes = "A2"
        ranking_sheet.auto_filter.ref = ranking_sheet.dimensions
        widths = [8, 45, 30, 14, 18, 15, 11, 11, 18, 18, 18, 15, 10, 10, 10, 24, 20]
        for i, width in enumerate(widths, 1):
            ranking_sheet.column_dimensions[get_column_letter(i)].width = width

        summary = workbook.create_sheet("Summary", 1)
        summary.append(["Goodreads To-Read Ranker", APP_VERSION])
        summary.append([])
        d = engine.diagnostics()
        summary_rows = [
            ("Generated", now_iso()), ("Books", len(engine.books)), ("Comparisons", len(engine.comparisons)),
            ("Mode", engine.mode), ("Target comparisons/book", engine.target_comparisons),
            ("Average comparisons/book", round(engine.average_comparisons(), 2)),
            ("Evidence coverage", round(engine.coverage() * 100, 1)),
            ("Top-10 membership confidence", round(d["top10_confidence"] * 100, 1)),
            ("Top-25 membership confidence", round(d["top25_confidence"] * 100, 1)),
            ("Top-10 stability", round(engine.top10_stability * 100, 1)),
            ("Top-25 stability", round(engine.top25_stability_score() * 100, 1)),
            ("Unresolved Top-25 boundary items", engine.unresolved_top25_swaps()),
            ("Estimated additional comparisons", engine.estimated_additional_comparisons()),
            ("Phase", engine.phase_label()), ("Random seed", engine.seed),
        ]
        for row in summary_rows:
            summary.append(row)
        summary.append([])
        summary.append(["Methodology", "Human pairwise choices are the only preference evidence. Goodreads fields are not used to infer preference; they are retained for display/export. Glicko-2 tracks online uncertainty, Bradley-Terry provides a global point model, Monte Carlo estimates top-K uncertainty, and active learning chooses the next comparison."])
        for cell in summary[1]: cell.fill = header_fill; cell.font = header_font
        summary.column_dimensions["A"].width = 38; summary.column_dimensions["B"].width = 110
        summary.freeze_panes = "A3"

        diagnostics_sheet = workbook.create_sheet("Diagnostics")
        diagnostics_sheet.append(["Metric", "Value"])
        for key, value in d.items():
            if isinstance(value, float): value = round(value, 4)
            diagnostics_sheet.append([key, value])
        diagnostics_sheet.append(["Simulation count", analysis.get("simulations", 0)])
        diagnostics_sheet.append(["Top-25 simulation stability", round(analysis.get("stability", 0.0), 4)])
        for cell in diagnostics_sheet[1]: cell.fill = header_fill; cell.font = header_font
        diagnostics_sheet.column_dimensions["A"].width = 40; diagnostics_sheet.column_dimensions["B"].width = 35

        comparison_sheet = workbook.create_sheet("Comparisons")
        comparison_sheet.append(["Left Book", "Left Author", "Right Book", "Right Author", "Result", "Timestamp"])
        book_map = engine.book_map()
        for match in engine.comparisons:
            left, right = book_map.get(match["left"]), book_map.get(match["right"])
            if left is None or right is None: continue
            result = left.title if match["result"] == "left" else (right.title if match["result"] == "right" else "Tie")
            comparison_sheet.append([left.title, left.author, right.title, right.author, result, match.get("timestamp", "")])
        for cell in comparison_sheet[1]: cell.fill = header_fill; cell.font = header_font
        comparison_sheet.freeze_panes = "A2"
        for col, width in {"A":45,"B":30,"C":45,"D":30,"E":45,"F":22}.items(): comparison_sheet.column_dimensions[col].width = width

        # Preserve the original Goodreads worksheet and append analysis columns to it.
        original_sheet = None
        for sheet in workbook.worksheets:
            if sheet.title not in {"Ranking", "Summary", "Diagnostics", "Comparisons"}:
                original_sheet = sheet; break
        if original_sheet is not None:
            existing_headers = [cell.value for cell in original_sheet[1]]
            extra_headers = ["Rank", "Rank Rating", "Rating Uncertainty", "Estimated Rank", "Rank Low", "Rank High",
                             "Top 10 Probability", "Top 25 Probability", "Top 50 Probability", "Pairwise Comparisons",
                             "Wins", "Losses", "Ties", "Status"]
            start_column = len(existing_headers) + 1
            for offset, header in enumerate(extra_headers):
                cell = original_sheet.cell(row=1, column=start_column + offset, value=header)
                cell.fill = header_fill; cell.font = header_font
            book_id_column = None
            for index, header in enumerate(existing_headers, start=1):
                if normalize_header(header) == normalize_header("Book Id - Goodreads"):
                    book_id_column = index; break
            books_by_row = {book.original_row: book for book in engine.books}
            books_by_id = {book.id: book for book in engine.books}
            for row in range(2, original_sheet.max_row + 1):
                book = None
                if book_id_column is not None:
                    gid = normalize(original_sheet.cell(row=row, column=book_id_column).value)
                    if gid: book = books_by_id.get("goodreads:" + gid)
                if book is None: book = books_by_row.get(row)
                if book is None: continue
                rank_data = next((item for item in ranked_items if item["book"].id == book.id), None)
                if rank_data is None: continue
                rating = rank_data["rating"]; stats = book_stats_cache[book.id]
                rank = stats["rank"]
                status = "Top 10" if rank <= 10 else ("Top 25" if rank <= 25 else ("Borderline" if stats["top25_probability"] >= .20 else "Long tail"))
                values = [rank, round(rank_data["final_rating"],2), round(rating.rd,2), stats["expected_rank"], stats["rank_low"], stats["rank_high"],
                          round(stats["top10_probability"]*100,1), round(stats["top25_probability"]*100,1), round(stats["top50_probability"]*100,1),
                          rating.comparisons, rating.wins, rating.losses, rating.ties, status]
                for offset, value in enumerate(values): original_sheet.cell(row=row, column=start_column + offset, value=value)

        output_path = source_file.parent / (safe_filename(source_file.stem) + "_ranked.xlsx")
        workbook.save(output_path)
        return output_path
    finally:
        workbook.close()


# ============================================================
# GUI APPLICATION
# ============================================================

class RankerApp:
    def __init__(
        self,
        root: tk.Tk,
    ):
        self.root = root

        self.root.title(
            f"{APP_NAME} {APP_VERSION}"
        )

        self.root.geometry(
            f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}"
        )

        self.root.minsize(
            1050,
            720,
        )

        self.engine = None
        self.source_file = None
        self.state_store = None
        self.current_pair = None

        self.target_var = tk.IntVar(
            value=DEFAULT_TARGET_COMPARISONS
        )

        self.mode_var = tk.StringVar(
            value="BALANCED"
        )

        self.auto_stop_var = tk.BooleanVar(
            value=True
        )

        self.status_var = tk.StringVar(
            value=(
                "Open your Goodreads Excel export "
                "to begin."
            )
        )

        self.progress_var = tk.DoubleVar(
            value=0.0
        )

        self.stats_var = tk.StringVar(
            value=""
        )

        self.left_title_var = tk.StringVar()
        self.left_author_var = tk.StringVar()
        self.left_meta_var = tk.StringVar()

        self.right_title_var = tk.StringVar()
        self.right_author_var = tk.StringVar()
        self.right_meta_var = tk.StringVar()

        self.build_styles()
        self.build_menu()
        self.build_ui()
        self.bind_keys()

        self.root.protocol(
            "WM_DELETE_WINDOW",
            self.on_close,
        )

    # ========================================================
    # STYLES
    # ========================================================

    def build_styles(self):
        style = ttk.Style()

        try:
            style.theme_use(
                "vista"
            )
        except tk.TclError:
            pass

        style.configure(
            "Title.TLabel",
            font=(
                "Segoe UI",
                23,
                "bold",
            ),
        )

        style.configure(
            "BookTitle.TLabel",
            font=(
                "Segoe UI",
                18,
                "bold",
            ),
        )

        style.configure(
            "BookAuthor.TLabel",
            font=(
                "Segoe UI",
                12,
            ),
        )

        style.configure(
            "BookMeta.TLabel",
            font=(
                "Segoe UI",
                10,
            ),
            foreground="#666666",
        )

        style.configure(
            "Choice.TButton",
            font=(
                "Segoe UI",
                12,
                "bold",
            ),
            padding=14,
        )

        style.configure(
            "Preset.TButton",
            font=(
                "Segoe UI",
                9,
                "bold",
            ),
            padding=(8, 5),
        )

        style.configure(
            "Secondary.TButton",
            font=(
                "Segoe UI",
                10,
            ),
            padding=8,
        )

    # ========================================================
    # MENU
    # ========================================================

    def build_menu(self):
        menu = tk.Menu(self.root)

        file_menu = tk.Menu(
            menu,
            tearoff=False,
        )

        file_menu.add_command(
            label="Open Goodreads Excel…",
            command=self.open_file,
            accelerator="Ctrl+O",
        )

        file_menu.add_command(
            label="Export Ranking",
            command=self.export,
            accelerator="Ctrl+E",
        )

        file_menu.add_separator()

        file_menu.add_command(
            label="Exit",
            command=self.on_close,
        )

        menu.add_cascade(
            label="File",
            menu=file_menu,
        )

        ranking_menu = tk.Menu(
            menu,
            tearoff=False,
        )

        ranking_menu.add_command(
            label="Undo Last Choice",
            command=self.undo,
            accelerator="U",
        )

        ranking_menu.add_command(
            label="Finish Now",
            command=self.finish_now,
        )

        ranking_menu.add_command(
            label="View Ranking",
            command=self.show_ranking,
        )

        ranking_menu.add_command(
            label="View Diagnostics",
            command=self.show_diagnostics,
        )

        menu.add_cascade(
            label="Ranking",
            menu=ranking_menu,
        )

        help_menu = tk.Menu(
            menu,
            tearoff=False,
        )

        help_menu.add_command(
            label="Keyboard Shortcuts",
            command=self.show_shortcuts,
        )

        help_menu.add_command(
            label="About",
            command=self.show_about,
        )

        menu.add_cascade(
            label="Help",
            menu=help_menu,
        )

        self.root.config(
            menu=menu
        )

    # ========================================================
    # MAIN UI
    # ========================================================

    def build_ui(self):
        """Build a responsive, tablet-first comparison interface.

        The central comparison area is the only vertically elastic region.
        Descriptions and shortcut/help areas have deliberately bounded
        geometry so long Goodreads text can never push the decision controls
        below the visible viewport.
        """
        self.root.configure(bg=self.colors["bg"])

        outer = ttk.Frame(self.root, style="App.TFrame", padding=12)
        outer.pack(fill="both", expand=True)
        outer.columnconfigure(0, weight=1)
        outer.columnconfigure(1, weight=1)
        # Only the comparison cards expand vertically.
        outer.rowconfigure(4, weight=1)

        # ---------- Header ----------
        header = ttk.Frame(outer, style="App.TFrame")
        header.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 7))
        header.columnconfigure(0, weight=1)

        ttk.Label(
            header,
            text="📚  GOODREADS  ·  TO-READ RANKER",
            style="Title.TLabel",
        ).grid(row=0, column=0, sticky="w")

        self.theme_button = ttk.Button(
            header,
            text=("☀ Light Mode" if self.dark_mode else "🌙 Dark Mode"),
            command=self.toggle_theme,
            style="Secondary.TButton",
        )
        self.theme_button.grid(row=0, column=1, padx=(5, 4))

        ttk.Button(
            header,
            text="📂 Open Excel",
            command=self.open_file,
            style="Secondary.TButton",
        ).grid(row=0, column=2, padx=(0, 0))

        # ---------- Strategy ----------
        strategy = ttk.Frame(outer, style="App.TFrame")
        strategy.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 6))
        strategy.columnconfigure(7, weight=1)

        ttk.Label(strategy, text="Strategy").grid(
            row=0, column=0, padx=(0, 5), sticky="w"
        )

        modes = [
            ("⚡ Quick", "QUICK"),
            ("Balanced", "BALANCED"),
            ("Accurate", "ACCURATE"),
            ("🏆 Top 25 Focus", "TOP_25_FOCUS"),
            ("Maximum", "MAX_ACCURACY"),
        ]
        for i, (label, mode) in enumerate(modes, start=1):
            ttk.Button(
                strategy,
                text=label,
                style="Preset.TButton",
                command=lambda m=mode: self.set_mode(m),
            ).grid(row=0, column=i, padx=2, sticky="w")

        ttk.Label(
            strategy,
            textvariable=self.mode_var,
            style="Help.TLabel",
        ).grid(row=0, column=6, padx=(8, 10), sticky="w")

        ttk.Checkbutton(
            strategy,
            text="Stop when stable",
            variable=self.auto_stop_var,
        ).grid(row=0, column=7, padx=5, sticky="e")

        ttk.Label(strategy, text="Target").grid(
            row=0, column=8, padx=(8, 3), sticky="e"
        )

        self.spinbox = ttk.Spinbox(
            strategy,
            from_=MIN_TARGET_COMPARISONS,
            to=MAX_TARGET_COMPARISONS,
            textvariable=self.target_var,
            width=5,
        )
        self.spinbox.grid(row=0, column=9, sticky="e")
        self.spinbox.bind("<Return>", self.target_changed)

        # ---------- Status / dashboard ----------
        status = ttk.Frame(outer, style="App.TFrame")
        status.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(0, 7))
        status.columnconfigure(0, weight=1)

        ttk.Label(
            status,
            textvariable=self.status_var,
            style="Status.TLabel",
        ).grid(row=0, column=0, sticky="w")

        ttk.Label(
            status,
            textvariable=self.stats_var,
            style="Help.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(2, 0))

        ttk.Progressbar(
            status,
            variable=self.progress_var,
            maximum=100,
        ).grid(row=2, column=0, sticky="ew", pady=(4, 0))

        # ---------- Comparison cards ----------
        choices = ttk.Frame(outer, style="App.TFrame")
        choices.grid(
            row=4,
            column=0,
            columnspan=2,
            sticky="nsew",
            pady=(1, 5),
        )
        choices.columnconfigure(0, weight=1, uniform="book")
        choices.columnconfigure(1, weight=1, uniform="book")
        choices.rowconfigure(0, weight=1)

        self.left_card = self.make_book_card(choices, "left")
        self.left_card.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=(0, 5),
        )

        self.right_card = self.make_book_card(choices, "right")
        self.right_card.grid(
            row=0,
            column=1,
            sticky="nsew",
            padx=(5, 0),
        )

        # ---------- Primary decision controls ----------
        buttons = ttk.Frame(outer, style="App.TFrame")
        buttons.grid(
            row=5,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(1, 5),
        )
        buttons.columnconfigure(0, weight=1, uniform="decision")
        buttons.columnconfigure(1, weight=1, uniform="decision")
        buttons.columnconfigure(2, weight=1, uniform="decision")

        self.left_button = ttk.Button(
            buttons,
            text="←  CHOOSE LEFT   [1]",
            command=lambda: self.choose("left"),
            style="Choice.TButton",
        )
        self.left_button.grid(row=0, column=0, sticky="ew", padx=(0, 4))

        self.tie_button = ttk.Button(
            buttons,
            text="≈  TIE   [3 / T]",
            command=lambda: self.choose("tie"),
            style="Choice.TButton",
        )
        self.tie_button.grid(row=0, column=1, sticky="ew", padx=4)

        self.right_button = ttk.Button(
            buttons,
            text="CHOOSE RIGHT   [2]  →",
            command=lambda: self.choose("right"),
            style="Choice.TButton",
        )
        self.right_button.grid(row=0, column=2, sticky="ew", padx=(4, 0))

        # ---------- Secondary actions ----------
        actions = ttk.Frame(outer, style="App.TFrame")
        actions.grid(
            row=6,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(0, 4),
        )

        ttk.Button(
            actions,
            text="↶ Undo [U]",
            command=self.undo,
            style="Secondary.TButton",
        ).pack(side="left", padx=(0, 3))

        ttk.Button(
            actions,
            text="Skip [4 / S]",
            command=self.skip,
            style="Secondary.TButton",
        ).pack(side="left", padx=3)

        ttk.Button(
            actions,
            text="🏁 Finish",
            command=self.finish_now,
            style="Secondary.TButton",
        ).pack(side="left", padx=3)

        ttk.Button(
            actions,
            text="⌨ Shortcuts",
            command=self.show_shortcuts,
            style="Secondary.TButton",
        ).pack(side="left", padx=3)

        ttk.Button(
            actions,
            text="Ranking",
            command=self.show_ranking,
            style="Secondary.TButton",
        ).pack(side="right", padx=3)

        ttk.Button(
            actions,
            text="Export",
            command=self.export,
            style="Secondary.TButton",
        ).pack(side="right", padx=3)

        ttk.Button(
            actions,
            text="Overwrite + Backup",
            command=self.export_overwrite,
            style="Secondary.TButton",
        ).pack(side="right", padx=3)

        # ---------- Always-visible shortcut reference ----------
        # This is intentionally a real panel, not a tiny one-line label.
        # It remains visible at the bottom of the main comparison screen.
        shortcut_panel = tk.Frame(
            outer,
            bg=self.colors["panel"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
            padx=9,
            pady=6,
        )
        shortcut_panel.grid(
            row=7,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(0, 0),
        )
        shortcut_panel.columnconfigure(0, weight=1)

        tk.Label(
            shortcut_panel,
            text="⌨  QUICK KEYBOARD REFERENCE",
            bg=self.colors["panel"],
            fg=self.colors["accent"],
            font=("Segoe UI", 9, "bold"),
        ).grid(row=0, column=0, sticky="w")

        shortcut_row_1 = (
            "1  Choose LEFT    •    2  Choose RIGHT    •    "
            "3 / T  Tie    •    4 / S  Skip    •    "
            "←  LEFT    •    →  RIGHT    •    U  Undo"
        )
        shortcut_row_2 = (
            "I  LEFT → Ignore    •    R  LEFT → Read    •    "
            "C  LEFT → Currently Reading    •    "
            "Shift+I  RIGHT → Ignore    •    Shift+R  RIGHT → Read    •    "
            "Shift+C  RIGHT → Currently Reading"
        )
        shortcut_row_3 = (
            "Ctrl+O  Open Goodreads Excel    •    "
            "Ctrl+E  Export Ranking    •    "
            "Click a book card  Lifecycle menu"
        )

        for row, text in enumerate(
            (shortcut_row_1, shortcut_row_2, shortcut_row_3),
            start=1,
        ):
            tk.Label(
                shortcut_panel,
                text=text,
                bg=self.colors["panel"],
                fg=self.colors["text"] if row < 3 else self.colors["muted"],
                font=("Segoe UI", 8 if row == 3 else 8.5),
                anchor="w",
                justify="left",
            ).grid(
                row=row,
                column=0,
                sticky="ew",
                pady=(1 if row == 1 else 0, 0),
            )

    def toggle_theme(self):
        """Switch the complete application between dark and light themes."""
        self.dark_mode = not self.dark_mode

        # Preserve the current window geometry while rebuilding widgets.
        geometry = self.root.geometry()

        for widget in list(self.root.winfo_children()):
            widget.destroy()

        self.build_styles()
        self.build_menu()
        self.build_ui()
        self.bind_keys()

        self.root.geometry(geometry)

        if self.engine is not None:
            self.refresh()

    # ========================================================
    # BOOK CARD
    # ========================================================

    def make_book_card(
        self,
        parent,
        side,
    ):
        frame = tk.Frame(
            parent,
            bg="#F6F7F9",
            highlightbackground="#D8DCE2",
            highlightthickness=1,
        )

        content = ttk.Frame(
            frame,
            padding=22,
        )

        content.pack(
            fill="both",
            expand=True,
        )

        if side == "left":
            title_var = (
                self.left_title_var
            )
            author_var = (
                self.left_author_var
            )
            meta_var = (
                self.left_meta_var
            )

        else:
            title_var = (
                self.right_title_var
            )
            author_var = (
                self.right_author_var
            )
            meta_var = (
                self.right_meta_var
            )

        ttk.Label(
            content,
            textvariable=title_var,
            style="BookTitle.TLabel",
            wraplength=410,
            justify="center",
        ).pack(
            pady=(25, 10)
        )

        ttk.Label(
            content,
            textvariable=author_var,
            style="BookAuthor.TLabel",
            wraplength=410,
            justify="center",
        ).pack(
            pady=(0, 18)
        )

        ttk.Label(
            content,
            textvariable=meta_var,
            style="BookMeta.TLabel",
            wraplength=410,
            justify="center",
        ).pack(
            pady=(0, 15)
        )

        description = tk.Text(
            content,
            height=9,
            width=45,
            wrap="word",
            relief="flat",
            borderwidth=0,
            bg="#F6F7F9",
            fg="#444444",
            font=(
                "Segoe UI",
                10,
            ),
            padx=5,
            pady=5,
        )

        description.pack(
            fill="both",
            expand=True,
        )

        description.configure(
            state="disabled"
        )

        if side == "left":
            self.left_description = (
                description
            )

        else:
            self.right_description = (
                description
            )

        return frame

    # ========================================================
    # KEYBOARD
    # ========================================================

    def bind_keys(self):
        """Install keyboard controls on the whole application.

        bind_all is used deliberately: focus may be on a button, spinbox,
        card, or another child widget. The ranking workflow should never
        require the user to click the comparison area before a shortcut works.
        """

        # Remove/reinstall only our application-level bindings.
        bindings = [
            ("<KeyPress-1>", lambda e: self.choose("left")),
            ("<KeyPress-2>", lambda e: self.choose("right")),
            ("<KeyPress-3>", lambda e: self.choose("tie")),
            ("<KeyPress-4>", lambda e: self.skip()),
            ("<Left>", lambda e: self.choose("left")),
            ("<Right>", lambda e: self.choose("right")),
            ("<KeyPress-t>", lambda e: self.choose("tie")),
            ("<KeyPress-T>", lambda e: self.choose("tie")),
            ("<KeyPress-s>", lambda e: self.skip()),
            ("<KeyPress-S>", lambda e: self.skip()),
            ("<KeyPress-u>", lambda e: self.undo()),
            ("<KeyPress-U>", lambda e: self.undo()),

            # Left-card lifecycle shortcuts.
            ("<KeyPress-i>", lambda e: self.change_card_status("left", "ignore")),
            ("<KeyPress-I>", lambda e: self.change_card_status("left", "ignore")),
            ("<KeyPress-r>", lambda e: self.change_card_status("left", "read")),
            ("<KeyPress-R>", lambda e: self.change_card_status("left", "read")),
            ("<KeyPress-c>", lambda e: self.change_card_status("left", "currently-reading")),
            ("<KeyPress-C>", lambda e: self.change_card_status("left", "currently-reading")),

            # Right-card lifecycle shortcuts. Explicit Shift bindings are
            # required; plain uppercase KeyPress bindings are not equivalent
            # on every Tk/Windows keyboard configuration.
            ("<Shift-KeyPress-I>", lambda e: self.change_card_status("right", "ignore")),
            ("<Shift-KeyPress-R>", lambda e: self.change_card_status("right", "read")),
            ("<Shift-KeyPress-C>", lambda e: self.change_card_status("right", "currently-reading")),

            # File shortcuts.
            ("<Control-KeyPress-o>", lambda e: self.open_file()),
            ("<Control-KeyPress-e>", lambda e: self.export()),
        ]

        for sequence, callback in bindings:
            self.root.bind_all(sequence, callback, add="+")
    # ========================================================
    # FILE
    # ========================================================

    def open_file(self):
        path = filedialog.askopenfilename(
            title=(
                "Open Goodreads Excel export"
            ),
            filetypes=[
                (
                    "Excel files",
                    "*.xlsx",
                ),
                (
                    "All files",
                    "*.*",
                ),
            ],
        )

        if not path:
            return

        self.load_file(
            Path(path)
        )

    def load_file(
        self,
        path: Path,
    ):
        try:
            _, books = load_goodreads(
                path
            )

        except Exception as exc:
            messagebox.showerror(
                "Could not open Goodreads file",
                str(exc),
            )
            return

        if len(books) < 2:
            messagebox.showwarning(
                "Not enough books",
                (
                    'I found fewer than two books where '
                    '"Exclusive Shelf" is "to-read".'
                ),
            )
            return

        self.source_file = path

        self.state_store = (
            StateStore(path)
        )

        existing = (
            self.state_store.load(
                books
            )
        )

        if existing is not None:
            resume = messagebox.askyesno(
                "Resume previous ranking?",
                (
                    f"I found a saved ranking for:\n\n"
                    f"{path.name}\n\n"
                    f"{len(existing.comparisons)} "
                    f"comparisons have already been made.\n\n"
                    f"Resume it?"
                ),
            )

            if resume:
                self.engine = existing

            else:
                self.engine = (
                    RankingEngine(
                        books,
                        self.target_var.get(),
                    )
                )

        else:
            self.engine = (
                RankingEngine(
                    books,
                    self.target_var.get(),
                )
            )

        self.target_var.set(
            self.engine.target_comparisons
        )
        self.mode_var.set(getattr(self.engine, "mode", "BALANCED"))

        self.current_pair = None

        self.status_var.set(
            (
                f"{len(books)} books on your "
                f"to-read shelf · {path.name}"
            )
        )

        self.refresh()

    # ========================================================
    # SPEED PRESETS / STRATEGIES
    # ========================================================

    def set_mode(self, mode: str):
        if self.engine is None:
            self.mode_var.set(mode)
            return
        mode = str(mode).upper()
        if mode not in PRESETS:
            mode = "BALANCED"
        self.mode_var.set(mode)
        current_target = self.engine.target_comparisons
        self.engine.mode = mode
        self.engine.config = make_config(mode, current_target)
        self.engine.target_comparisons = self.engine.config.target_comparisons
        self.engine._analysis_cache = None
        self.engine._ranking_cache = None
        self.engine.current_mode_changed = True
        self.current_pair = None
        self.save_state()
        self.refresh()

    def set_target(
        self,
        value: int,
    ):
        self.target_var.set(
            value
        )

        self.target_changed()

    def target_changed(
        self,
        event=None,
    ):
        if self.engine is None:
            return

        try:
            value = int(
                self.target_var.get()
            )

        except ValueError:
            value = (
                DEFAULT_TARGET_COMPARISONS
            )

        value = max(
            MIN_TARGET_COMPARISONS,
            min(
                MAX_TARGET_COMPARISONS,
                value,
            ),
        )

        self.target_var.set(
            value
        )

        self.engine.target_comparisons = value
        self.engine.config = make_config(self.engine.mode, value)
        self.engine._analysis_cache = None
        self.engine._ranking_cache = None

        self.save_state()

        self.current_pair = None

        self.refresh()

    # ========================================================
    # REFRESH
    # ========================================================

    def refresh(self):
        if self.engine is None:
            return

        # Ensure model is current before selecting the pair.
        self.engine.ensure_model()

        if (
            self.auto_stop_var.get()
            and self.engine.is_finished()
        ):
            self.show_finished()

            return

        if (
            self.current_pair is None
            or not self.current_pair_valid()
        ):
            self.current_pair = (
                self.engine.choose_pair()
            )

        if self.current_pair is None:
            self.show_finished()

            return

        self.enable_choices()

        left_id, right_id = (
            self.current_pair
        )

        book_map = (
            self.engine.book_map()
        )

        left = book_map[
            left_id
        ]

        right = book_map[
            right_id
        ]

        self.show_book(
            left,
            self.engine.ratings[
                left.id
            ],
            self.left_title_var,
            self.left_author_var,
            self.left_meta_var,
            self.left_description,
        )

        self.show_book(
            right,
            self.engine.ratings[
                right.id
            ],
            self.right_title_var,
            self.right_author_var,
            self.right_meta_var,
            self.right_description,
        )

        progress = (
            self.engine.progress()
            * 100.0
        )

        self.progress_var.set(
            progress
        )

        comparisons = len(
            self.engine.comparisons
        )

        average = (
            self.engine.average_comparisons()
        )

        coverage = (
            self.engine.coverage()
            * 100.0
        )

        stability = (
            self.engine.stability
            * 100.0
        )

        top_stability = (
            self.engine.top_stability
            * 100.0
        )

        diagnostics = self.engine.diagnostics()
        self.stats_var.set(
            (
                f"{comparisons} decisions · {average:.1f} avg/book · "
                f"coverage {coverage:.0f}% · Top-25 confidence {diagnostics['top25_confidence'] * 100:.0f}% · "
                f"Top-25 stability {diagnostics['top25_stability'] * 100:.0f}% · "
                f"~{diagnostics['estimated_additional']} more"
            )
        )

        self.status_var.set(
            (
                f"{self.engine.phase_label()} · "
                "Which book would you rather read?"
            )
        )

    def show_finished(self):
        self.disable_choices()

        self.left_title_var.set(
            "🏆 Ranking ready"
        )

        self.left_author_var.set("")
        self.left_meta_var.set("")

        self.right_title_var.set(
            "You can export now"
        )

        self.right_author_var.set("")
        self.right_meta_var.set("")

        self.set_description(
            self.left_description,
            (
                "The adaptive ranking engine has "
                "determined that additional comparisons "
                "are unlikely to materially improve the "
                "current ordering."
            ),
        )

        self.set_description(
            self.right_description,
            (
                "You can continue refining by turning "
                "off 'Stop when stable', or export the "
                "current ranking to Excel."
            ),
        )

        self.progress_var.set(
            max(
                95.0,
                self.engine.progress()
                * 100.0,
            )
        )

        diagnostics = self.engine.diagnostics()
        self.stats_var.set(
            (
                f"{len(self.engine.comparisons)} decisions · "
                f"Top-10 {diagnostics['top10_confidence'] * 100:.0f}% · "
                f"Top-25 {diagnostics['top25_confidence'] * 100:.0f}% · "
                f"boundary unresolved {diagnostics['unresolved_boundary']} · "
                f"{diagnostics['phase']}"
            )
        )

        self.status_var.set(
            "Ranking complete. Export when ready."
        )

    def current_pair_valid(self):
        if (
            self.engine is None
            or self.current_pair is None
        ):
            return False

        left, right = (
            self.current_pair
        )

        if left not in self.engine.ratings:
            return False

        if right not in self.engine.ratings:
            return False

        if left == right:
            return False

        return (
            self.engine.pair_key(
                left,
                right,
            )
            not in self.engine.played
        )

    # ========================================================
    # DISPLAY BOOK
    # ========================================================

    def show_book(
        self,
        book,
        rating,
        title_var,
        author_var,
        meta_var,
        description_widget,
    ):
        title_var.set(
            book.title
        )

        author_var.set(
            book.author
            if book.author
            else "Unknown author"
        )

        meta = []

        if book.pages:
            meta.append(
                f"{book.pages} pages"
            )

        if book.year:
            meta.append(
                f"Published {book.year}"
            )

        stats = self.engine.book_stats(book.id) if self.engine is not None else None
        if stats is not None:
            meta.append(f"Current #{stats['rank']}")
            meta.append(f"Top 25 {stats['top25_probability'] * 100:.0f}%")
            meta.append(f"Likely {stats['rank_low']}–{stats['rank_high']}")

        meta.append(f"Rating {rating.rating:.0f}")
        meta.append(f"±{rating.rd:.0f}")
        meta.append(f"{rating.comparisons} comparisons")

        if book.my_rating:
            meta.append(
                f"My Goodreads: {book.my_rating}"
            )

        meta_var.set(
            "  ·  ".join(meta)
        )

        description = (
            book.description
            if book.description
            else "No description available."
        )

        self.set_description(
            description_widget,
            description,
        )

    def set_description(
        self,
        widget,
        text,
    ):
        widget.configure(
            state="normal"
        )

        widget.delete(
            "1.0",
            "end",
        )

        widget.insert(
            "1.0",
            truncate(
                text,
                1800,
            ),
        )

        widget.configure(
            state="disabled"
        )

    # ========================================================
    # CHOICES
    # ========================================================

    def choose(
        self,
        result: str,
    ):
        if (
            self.engine is None
            or self.current_pair is None
        ):
            return

        left_id, right_id = (
            self.current_pair
        )

        try:
            self.engine.apply_match(
                left_id,
                right_id,
                result,
            )

            self.save_state()

        except Exception as exc:
            messagebox.showerror(
                "Could not record choice",
                str(exc),
            )

            return

        self.current_pair = None

        self.refresh()

    def skip(self):
        if self.engine is None or self.current_pair is None:
            return
        left, right = self.current_pair
        self.engine.skip_pair(left, right)
        self.current_pair = None
        self.refresh()

    # ========================================================
    # FINISH
    # ========================================================

    def finish_now(self):
        if self.engine is None:
            return

        self.engine.fit_bradley_terry()
        self.engine.calculate_stability()

        self.save_state()

        self.current_pair = None

        self.show_finished()

    # ========================================================
    # SAVE
    # ========================================================

    def save_state(self):
        if (
            self.engine is None
            or self.state_store is None
        ):
            return

        try:
            self.state_store.save(
                self.engine
            )

        except Exception as exc:
            print(
                "Warning: could not save state:",
                exc,
            )

    # ========================================================
    # UNDO
    # ========================================================

    def undo(self):
        if self.engine is None:
            return

        if self.engine.undo():
            self.save_state()

            self.current_pair = None

            self.refresh()

    # ========================================================
    # BUTTON STATES
    # ========================================================

    def enable_choices(self):
        self.left_button.configure(
            state="normal"
        )

        self.tie_button.configure(
            state="normal"
        )

        self.right_button.configure(
            state="normal"
        )

    def disable_choices(self):
        self.left_button.configure(
            state="disabled"
        )

        self.tie_button.configure(
            state="disabled"
        )

        self.right_button.configure(
            state="disabled"
        )

    # ========================================================
    # RANKING WINDOW
    # ========================================================

    def show_ranking(self):
        if self.engine is None:
            messagebox.showinfo("No ranking", "Open a Goodreads file first.")
            return
        self.engine.ensure_model()
        window = tk.Toplevel(self.root)
        window.title("Current Ranking · Top-K uncertainty")
        window.geometry("1450x760")
        frame = ttk.Frame(window, padding=12)
        frame.pack(fill="both", expand=True)
        d = self.engine.diagnostics()
        ttk.Label(frame, text="Current Ranking", style="Title.TLabel").pack(anchor="w")
        ttk.Label(frame, text=(
            f"{self.engine.mode} · Top-25 confidence {d['top25_confidence']*100:.0f}% · "
            f"Top-25 stability {d['top25_stability']*100:.0f}% · "
            f"{d['unresolved_boundary']} unresolved boundary items · "
            f"~{d['estimated_additional']} additional decisions estimated"
        ), foreground="#666666").pack(anchor="w", pady=(0, 8))
        columns = ("rank","title","author","rating","rd","expected","interval","top10","top25","comparisons","record","status")
        tree = ttk.Treeview(frame, columns=columns, show="headings")
        headings = {
            "rank":"Rank","title":"Title","author":"Author","rating":"Model","rd":"RD",
            "expected":"Likely rank","interval":"10–90% rank","top10":"Top 10 %","top25":"Top 25 %",
            "comparisons":"Decisions","record":"W/L/T","status":"Status"
        }
        widths = {"rank":55,"title":320,"author":200,"rating":80,"rd":65,"expected":85,"interval":105,"top10":85,"top25":85,"comparisons":75,"record":75,"status":150}
        for c in columns:
            tree.heading(c, text=headings[c]); tree.column(c, width=widths[c], anchor="w" if c in {"title","author","status"} else "center")
        scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        tree.pack(side="left", fill="both", expand=True); scroll.pack(side="right", fill="y")
        for rank, item in enumerate(self.engine.ranking(), start=1):
            book = item["book"]; rating = item["rating"]; st = self.engine.book_stats(book.id)
            status = "TOP 10" if rank <= 10 else ("TOP 25" if rank <= 25 else ("BORDERLINE" if st["top25_probability"] >= .20 else "Long tail"))
            if rating.rd >= 220 or st["rank_high"] - st["rank_low"] >= self.engine.config.boundary_width:
                status += " · uncertain"
            record = f"{rating.wins}/{rating.losses}/{rating.ties}"
            tree.insert("", "end", values=(rank, book.title, book.author, f"{item['final_rating']:.0f}", f"±{rating.rd:.0f}",
                                              st["expected_rank"], f"{st['rank_low']}–{st['rank_high']}",
                                              f"{st['top10_probability']*100:.0f}%", f"{st['top25_probability']*100:.0f}%",
                                              rating.comparisons, record, status))

    # ========================================================
    # DIAGNOSTICS WINDOW
    # ========================================================

    def show_diagnostics(self):
        if self.engine is None:
            messagebox.showinfo("No ranking", "Open a Goodreads file first.")
            return
        self.engine.ensure_model()
        d = self.engine.diagnostics()
        window = tk.Toplevel(self.root)
        window.title("Top-K Ranking Diagnostics")
        window.geometry("780x720")
        frame = ttk.Frame(window, padding=20); frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Ranking Diagnostics", style="Title.TLabel").pack(anchor="w", pady=(0, 12))
        rows = [
            ("Books", d["books"]), ("Human decisions", d["comparisons"]),
            ("Average decisions/book", f"{d['average_comparisons']:.2f}"),
            ("Evidence coverage", f"{d['coverage']*100:.1f}%"),
            ("Mode", d["mode"]), ("Phase", d["phase"]),
            ("Top-10 membership confidence", f"{d['top10_confidence']*100:.1f}%"),
            ("Top-25 membership confidence", f"{d['top25_confidence']*100:.1f}%"),
            ("Top-10 stability", f"{d['top10_stability']*100:.1f}%"),
            ("Top-25 stability", f"{d['top25_stability']*100:.1f}%"),
            ("Unresolved Top-25 boundary", d["unresolved_boundary"]),
            ("Estimated additional decisions", d["estimated_additional"]),
            ("Simulation count", self.engine.analysis().get("simulations", 0)),
            ("Seed", d["seed"]),
        ]
        for label, value in rows:
            row=ttk.Frame(frame); row.pack(fill="x", pady=3)
            ttk.Label(row, text=label, width=34).pack(side="left")
            ttk.Label(row, text=str(value)).pack(side="left")
        ttk.Separator(frame).pack(fill="x", pady=14)
        ttk.Label(frame, text="How this version allocates evidence", font=("Segoe UI",13,"bold")).pack(anchor="w", pady=(0,8))
        explanation=(
            "The system does not try to perfectly rank the long tail. It first establishes broad evidence, "
            "then builds a bounded elite pool and concentrates comparisons on the Top 10, Top 25 boundary, "
            "high-uncertainty challengers, and pairs whose outcome can materially change the important ordering.\n\n"
            "Top-K probabilities and rank intervals are Monte-Carlo estimates from the current rating and uncertainty model. "
            "They are deliberately labelled estimates, not guarantees. Goodreads ratings and metadata are never used to decide which book you prefer.\n\n"
            "A ranking is considered complete only when Top-25 membership, Top-10 membership, simulation stability, and the boundary are sufficiently resolved."
        )
        ttk.Label(frame, text=explanation, justify="left", wraplength=720).pack(anchor="w")

    # ========================================================
    # EXPORT
    # ========================================================

    def export(self):
        if (
            self.engine is None
            or self.source_file is None
        ):
            messagebox.showinfo(
                "Nothing to export",
                "Open a Goodreads file first.",
            )

            return

        try:
            # Always make sure final model is current.
            self.engine.fit_bradley_terry()
            self.engine.calculate_stability()

            output_path = export_results(
                self.source_file,
                self.engine,
            )

            self.save_state()

        except PermissionError:
            messagebox.showerror(
                "Could not save",
                (
                    "Windows could not save the ranked "
                    "workbook.\n\n"
                    "If the output file is open in Excel, "
                    "close it and try again."
                ),
            )

            return

        except Exception as exc:
            messagebox.showerror(
                "Export failed",
                str(exc),
            )

            return

        open_file = messagebox.askyesno(
            "Ranking exported",
            (
                "Ranking exported successfully.\n\n"
                f"{output_path}\n\n"
                "Open it now?"
            ),
        )

        if open_file:
            self.open_external(
                output_path
            )

    @staticmethod
    def open_external(
        path: Path,
    ):
        try:
            if sys.platform.startswith(
                "win"
            ):
                os.startfile(
                    str(path)
                )

            elif sys.platform == "darwin":
                subprocess.Popen(
                    [
                        "open",
                        str(path),
                    ]
                )

            else:
                subprocess.Popen(
                    [
                        "xdg-open",
                        str(path),
                    ]
                )

        except Exception:
            pass

    # ========================================================
    # HELP
    # ========================================================

    def show_shortcuts(self):
        messagebox.showinfo(
            "Keyboard shortcuts",
            (
                "1   Choose left book\n"
                "2   Choose right book\n"
                "3   Tie / equal\n"
                "4   Skip pair\n"
                "←   Choose left book\n"
                "→   Choose right book\n"
                "T   Tie / equal\n"
                "S   Skip pair\n"
                "U   Undo\n"
                "Ctrl+O   Open Goodreads Excel\n"
                "Ctrl+E   Export ranking"
            ),
        )

    def show_about(self):
        messagebox.showinfo(
            "About",
            (
                f"{APP_NAME}\n"
                f"Version {APP_VERSION}\n\n"
                "A large-library adaptive pairwise "
                "ranking engine for Goodreads shelves.\n\n"
                "ENGINE:\n"
                "• Glicko-2 live uncertainty tracking\n"
                "• Active-learning pair selection\n"
                "• Swiss-style initial exploration\n"
                "• Regularized Bradley-Terry global ranking\n"
                "• Adaptive early stopping\n"
                "• Ranking stability analysis\n"
                "• Top-ranking refinement\n"
                "• Goodreads My Rating as a weak prior\n\n"
                "PERFORMANCE:\n"
                "• Cached comparison indexes\n"
                "• No repeated pair scans\n"
                "• Fast undo snapshots\n"
                "• Sparse comparison fitting\n"
                "• Efficient Excel export lookup\n\n"
                "Only rows where \"Exclusive Shelf\" "
                "equals \"to-read\" are included.\n\n"
                "Your original Goodreads workbook is "
                "never modified."
            ),
        )

    # ========================================================
    # CLOSE
    # ========================================================

    def on_close(self):
        self.save_state()
        self.root.destroy()




def run_self_test():
    """Run dependency-free sanity checks without starting Tkinter."""
    started = time.perf_counter()

    # Glicko sanity: a decisive win should move the winner up and loser down.
    a, b = Rating(), Rating()
    a2 = glicko_update(a, [b], [1.0])
    b2 = glicko_update(b, [a], [0.0])
    assert a2.rating > a.rating and b2.rating < b.rating
    assert 1.0 <= a2.rd <= 350.0 and 1.0 <= b2.rd <= 350.0

    books = [Book(id=f"b{i}", title=f"Book {i}", author=f"Author {i % 7}") for i in range(30)]
    engine = RankingEngine(books, target_comparisons=6, random_seed=12345, mode="TOP_25_FOCUS")
    seen = set()
    for _ in range(20):
        pair = engine.choose_pair()
        if pair is None:
            break
        key = engine.pair_key(*pair)
        assert key not in seen
        seen.add(key)
        engine.apply_match(pair[0], pair[1], "left" if len(seen) % 3 else "tie")
    assert len(engine.played) == len(engine.comparisons)
    assert len(engine.ranking()) == len(books)
    stats = engine.book_stats(books[0].id)
    assert 1 <= stats["rank_low"] <= len(books)
    assert 0.0 <= stats["top25_probability"] <= 1.0

    # Undo must remove exactly the latest decision without duplicating history.
    before = len(engine.comparisons)
    assert engine.undo()
    assert len(engine.comparisons) == before - 1
    pair = engine.choose_pair()
    assert pair is not None

    # Large-library candidate selection sanity: bounded candidate generation and no quadratic pair list.
    large_books = [Book(id=f"L{i}", title=f"Large Book {i}", author="Author") for i in range(1000)]
    large = RankingEngine(large_books, target_comparisons=8, random_seed=7, mode="TOP_25_FOCUS")
    t0 = time.perf_counter()
    pair = large.choose_pair()
    elapsed = time.perf_counter() - t0
    assert pair is not None and elapsed < 10.0, f"Large-library pair selection too slow: {elapsed:.2f}s"
    assert len(large.played) == 0

    # Save/load round trip using a temporary source path.
    with tempfile.TemporaryDirectory() as td:
        source = Path(td) / "books.xlsx"
        wb = load_workbook(filename=source) if source.exists() else None
        if wb is not None:
            wb.close()
        from openpyxl import Workbook
        wb = Workbook(); wb.save(source); wb.close()
        store = StateStore(source)
        store.save(engine)
        loaded = store.load(books)
        assert loaded is not None
        assert len(loaded.comparisons) == len(engine.comparisons)
        assert loaded.seed == engine.seed
        assert loaded.mode == engine.mode

    print(f"SELF-TEST PASSED in {time.perf_counter() - started:.2f}s")
    return 0

# ============================================================
# MAIN
# ============================================================

def main():
    if "--self-test" in sys.argv:
        return run_self_test()
    root = tk.Tk()
    RankerApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())

    except Exception:
        error = traceback.format_exc()

        print(error)

        try:
            messagebox.showerror(
                "Unexpected error",
                (
                    "The application encountered an "
                    "unexpected error:\n\n"
                    + error
                ),
            )

        except Exception:
            pass