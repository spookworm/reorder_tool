from __future__ import annotations

import copy
import hashlib
import json
import math
import os
import random
import re
import shutil
import subprocess
import sys
import tempfile
import traceback
import webbrowser
import urllib.parse
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# APPLICATION
# ============================================================

APP_NAME = "Goodreads To-Read Ranker"
APP_VERSION = "6.1.0-SHELF-SEGMENTED"
STATE_VERSION = 9

TOP_K = 25
TOP_10 = 10
BOUNDARY_WIDTH = 12

DEFAULT_RATING = 1500.0
DEFAULT_RD = 350.0
DEFAULT_VOLATILITY = 0.06
GLICKO_SCALE = 173.7178

PAIR_CANDIDATE_LIMIT = 3000

WINDOW_WIDTH = 1180
WINDOW_HEIGHT = 800

PRESETS = {
    "QUICK": (3, False),
    "BALANCED": (6, False),
    "ACCURATE": (10, True),
    "TOP_25_FOCUS": (8, True),
    "MAX_ACCURACY": (16, True),
}


# ============================================================
# THEMES
# ============================================================

DARK = {
    "bg": "#0B1020",
    "panel": "#111827",
    "card": "#172036",
    "card2": "#1D2942",
    "text": "#F3F6FB",
    "muted": "#9AA8BD",
    "border": "#2B3B58",
    "accent": "#67E8F9",
    "purple": "#A78BFA",
    "good": "#34D399",
    "warn": "#FBBF24",
    "bad": "#FB7185",
    "button": "#202E49",
    "button_hover": "#2C3D5E",
}

LIGHT = {
    "bg": "#F4F7FB",
    "panel": "#FFFFFF",
    "card": "#FFFFFF",
    "card2": "#EEF3F9",
    "text": "#172033",
    "muted": "#657085",
    "border": "#D8E0EB",
    "accent": "#0891B2",
    "purple": "#7C3AED",
    "good": "#059669",
    "warn": "#D97706",
    "bad": "#E11D48",
    "button": "#E8EEF6",
    "button_hover": "#DCE6F2",
}


# ============================================================
# UTILITIES
# ============================================================

def normalize(value) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip().replace("\ufeff", "")


def normalize_header(value) -> str:
    return normalize(value).lower()


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def sigmoid(value: float) -> float:
    if value >= 0:
        return 1.0 / (1.0 + math.exp(-min(value, 700)))

    z = math.exp(max(value, -700))
    return z / (1.0 + z)


def normalized_text(value) -> str:
    return re.sub(
        r"[^a-z0-9]+",
        " ",
        normalize(value).lower(),
    ).strip()


def clean_isbn(value) -> str:
    return (
        normalize(value)
        .replace("-", "")
        .replace(" ", "")
    )


def stable_book_id(row: dict) -> str:
    """
    Stable identity priority:

    1. Goodreads Book Id
    2. ISBN
    3. deterministic title + author hash

    IMPORTANT:
    import_goodreads() normalizes all headers to lowercase,
    so this function deliberately uses lowercase keys.
    """

    if not isinstance(row, dict):
        raise TypeError(
            f"Expected Goodreads row to be a dict, "
            f"got {type(row).__name__}"
        )

    goodreads_id = normalize(
        row.get("book id - goodreads")
    )

    if goodreads_id:
        return f"gr:{goodreads_id}"

    isbn = clean_isbn(
        row.get("isbn")
    )

    if isbn:
        return f"isbn:{isbn}"

    title = normalized_text(
        row.get("title")
    )

    author = normalized_text(
        row.get("author l-f")
        or row.get("author")
    )

    identity = f"{title}|{author}"

    return "book:" + hashlib.sha1(
        identity.encode("utf-8")
    ).hexdigest()[:20]


def safe_float(value, default=0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


# ============================================================
# GLICKO-STYLE RATING
# ============================================================

@dataclass
class Rating:
    rating: float = DEFAULT_RATING
    rd: float = DEFAULT_RD
    volatility: float = DEFAULT_VOLATILITY

    comparisons: int = 0
    wins: int = 0
    losses: int = 0
    ties: int = 0


def glicko_update(
    player: Rating,
    opponent: Rating,
    score: float,
) -> Rating:
    """
    Stable sequential Glicko-style update.

    Every human decision arrives one at a time, so the model
    rebuilds its state by replaying the human comparison history.
    """

    q = math.log(10.0) / 400.0

    mu = (
        player.rating - 1500.0
    ) / GLICKO_SCALE

    phi = max(
        player.rd,
        1.0,
    ) / GLICKO_SCALE

    opponent_mu = (
        opponent.rating - 1500.0
    ) / GLICKO_SCALE

    opponent_phi = (
        max(opponent.rd, 1.0)
        / GLICKO_SCALE
    )

    g = 1.0 / math.sqrt(
        1.0
        + (
            3.0
            * q
            * q
            * opponent_phi
            * opponent_phi
            / math.pi**2
        )
    )

    expected = sigmoid(
        g * (mu - opponent_mu)
    )

    variance = 1.0 / max(
        q
        * q
        * g
        * g
        * expected
        * (1.0 - expected),
        1e-9,
    )

    delta = (
        variance
        * q
        * g
        * (score - expected)
    )

    new_phi = 1.0 / math.sqrt(
        1.0 / (phi * phi)
        + 1.0 / variance
    )

    new_mu = (
        mu
        + new_phi
        * new_phi
        * q
        * g
        * (score - expected)
    )

    return Rating(
        rating=1500.0 + GLICKO_SCALE * new_mu,
        rd=clamp(
            GLICKO_SCALE * new_phi,
            1.0,
            350.0,
        ),
        volatility=clamp(
            player.volatility,
            0.02,
            0.20,
        ),
        comparisons=player.comparisons + 1,
        wins=player.wins + int(score == 1.0),
        losses=player.losses + int(score == 0.0),
        ties=player.ties + int(score == 0.5),
    )


# ============================================================
# SERIES HELPERS
# ============================================================

def series_name(book):
    fields = getattr(book, "goodreads_fields", {}) or {}
    value = fields.get("series", "")
    return normalize(value).strip()


def series_number(book):
    fields = getattr(book, "goodreads_fields", {}) or {}
    value = normalize(fields.get("series-#", "")).strip()

    if not value:
        return None

    try:
        number = int(float(value))
    except (TypeError, ValueError):
        return None

    return number


def series_key(book):
    name = series_name(book)
    number = series_number(book)

    if not name or number is None:
        return None

    return (name.casefold(), number)


# ============================================================
# BOOK MODEL
# ============================================================

@dataclass
class Book:
    id: str
    title: str
    author: str = ""
    description: str = ""
    pages: str = ""
    my_rating: str = ""
    shelf: str = "to-read"
    status: str = "to-read"

    isbn: str = ""
    year: str = ""
    publisher: str = ""
    goodreads_id: str = ""

    # Every normalized Goodreads column is retained here so the
    # book cards can display the complete source record, including
    # columns that are not otherwise used by the ranking engine.
    goodreads_fields: dict = field(default_factory=dict)


# ============================================================
# GOODREADS IMPORT
# ============================================================

def import_goodreads(path):
    """
    Import a Goodreads Excel export.

    ALL dictionary keys returned by this function are lowercase.

    This is important because Goodreads headers such as:

        Title
        Author l-f
        Book Id - Goodreads

    become:

        title
        author l-f
        book id - goodreads
    """

    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(
            f"File does not exist:\n{path}"
        )

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook.active

        rows = worksheet.iter_rows(
            values_only=True
        )

        headings = next(
            rows,
            None,
        )

        if not headings:
            raise ValueError(
                "The Goodreads workbook is empty."
            )

        normalized_headings = [
            normalize_header(value)
            for value in headings
        ]

        if "title" not in normalized_headings:
            raise ValueError(
                "This does not appear to be a Goodreads "
                "export because the 'Title' column was not found."
            )

        records = []

        for raw in rows:
            record = {}

            for index, heading in enumerate(
                normalized_headings
            ):
                if not heading:
                    continue

                if index < len(raw):
                    record[heading] = normalize(
                        raw[index]
                    )
                else:
                    record[heading] = ""

            # Ignore completely empty rows.
            if not any(record.values()):
                continue

            if not record.get("title"):
                continue

            records.append(record)

        return records

    finally:
        workbook.close()


def book_from_row(row):
    """
    Convert one normalized Goodreads row into a Book.

    This function expects ONE dict, not the complete list returned
    by import_goodreads().
    """

    if not isinstance(row, dict):
        raise TypeError(
            f"book_from_row expected a dict, "
            f"got {type(row).__name__}"
        )

    shelf = normalize(
        row.get("exclusive shelf")
    ).lower()

    if shelf in {
        "currently-reading",
        "currently reading",
    }:
        shelf = "currently-reading"

    elif shelf == "to-read":
        shelf = "to-read"

    elif shelf == "read":
        shelf = "read"

    elif shelf == "ignore":
        shelf = "ignore"

    else:
        shelf = "to-read"

    return Book(
        id=stable_book_id(row),

        title=(
            normalize(
                row.get("title")
            )
            or "(Untitled)"
        ),

        author=normalize(
            row.get("author l-f")
            or row.get("author")
        ),

        description=normalize(
            row.get("description")
        ),

        pages=normalize(
            row.get("number of pages")
        ),

        my_rating=normalize(
            row.get("my rating")
        ),

        shelf=shelf,
        status=shelf,

        isbn=clean_isbn(
            row.get("isbn")
        ),

        year=normalize(
            row.get("year published")
        ),

        publisher=normalize(
            row.get("publisher")
        ),

        goodreads_id=normalize(
            row.get("book id - goodreads")
        ),
        goodreads_fields={
            str(key): normalize(value)
            for key, value in row.items()
        },
    )


# ============================================================
# RANKING ENGINE
# ============================================================

class RankingEngine:
    """
    Human-preference ranking engine.

    Goodreads metadata NEVER decides which book is better.

    Goodreads data is used for:
        - identity
        - display
        - lifecycle synchronization

    Human pairwise choices are the preference evidence.

    The engine uses:
        - sequential Glicko-style ratings
        - uncertainty-aware rank estimates
        - Top-K probability estimates
        - adaptive candidate selection
        - elite/boundary/challenger pools

    It never constructs the full N*(N-1)/2 pair set.
    """

    ACTIVE_STATUSES = {
        "to-read",
        "currently-reading",
    }

    ALL_STATUSES = {
        "to-read",
        "currently-reading",
        "read",
        "ignore",
    }

    def __init__(
        self,
        books,
        mode="TOP_25_FOCUS",
        seed=None,
        comparisons=None,
        statuses=None,
        skips=None,
    ):
        if books is None:
            books = []

        self.library = {
            book.id: book
            for book in books
            if isinstance(book, Book)
        }

        self.mode = (
            mode
            if mode in PRESETS
            else "TOP_25_FOCUS"
        )

        self.target, self.top_focus = PRESETS[
            self.mode
        ]

        self.seed = (
            seed
            if seed is not None
            else random.SystemRandom().randrange(
                1,
                2**31 - 1,
            )
        )

        self.rng = random.Random(
            self.seed
        )

        self.ratings = {}

        self.comparisons = (
            copy.deepcopy(comparisons)
            if isinstance(comparisons, list)
            else []
        )

        self.played = set()
        self.skips = {}

        self.statuses = {
            book.id: book.status
            for book in self.library.values()
        }

        if isinstance(statuses, dict):
            self.statuses.update(statuses)

            for book_id, status in statuses.items():
                if (
                    book_id in self.library
                    and status in self.ALL_STATUSES
                ):
                    self.library[
                        book_id
                    ].status = status

        if isinstance(skips, dict):
            self.skips = self._normalize_skips(
                skips
            )

        self._sync_books()
        self._replay()

    # --------------------------------------------------------
    # State compatibility
    # --------------------------------------------------------

    @staticmethod
    def _normalize_skips(skips):
        result = {}

        for key, value in skips.items():
            try:
                count = int(value)
            except (TypeError, ValueError):
                continue

            if isinstance(key, str):
                parts = key.split("|", 1)

                if len(parts) == 2:
                    result[
                        RankingEngine.pair_key(
                            parts[0],
                            parts[1],
                        )
                    ] = count

            elif (
                isinstance(key, (list, tuple))
                and len(key) == 2
            ):
                result[
                    RankingEngine.pair_key(
                        str(key[0]),
                        str(key[1]),
                    )
                ] = count

        return result

    # --------------------------------------------------------
    # Library
    # --------------------------------------------------------

    def _sync_books(self):
        self.books = [
            book
            for book in self.library.values()
            if book.status in self.ACTIVE_STATUSES
        ]

    def active_ids(self):
        return [
            book.id
            for book in self.books
        ]

    # --------------------------------------------------------
    # Pair identity
    # --------------------------------------------------------

    @staticmethod
    def pair_key(a, b):
        return tuple(
            sorted(
                (
                    str(a),
                    str(b),
                )
            )
        )

    # --------------------------------------------------------
    # Rebuild model
    # --------------------------------------------------------

    def _replay(self):
        """
        Rebuild ratings from human history.

        Archived books remain in the comparison history but do not
        affect the current active ranking.
        """

        self.ratings = {
            book.id: Rating()
            for book in self.books
        }

        self.played = set()

        for comparison in self.comparisons:
            if not isinstance(comparison, dict):
                continue

            left = comparison.get("left")
            right = comparison.get("right")
            result = comparison.get("result")

            if not left or not right:
                continue

            if (
                left not in self.ratings
                or right not in self.ratings
            ):
                continue

            if result not in {
                "left",
                "right",
                "tie",
            }:
                continue

            self.played.add(
                self.pair_key(
                    left,
                    right,
                )
            )

            left_rating = self.ratings[
                left
            ]

            right_rating = self.ratings[
                right
            ]

            if result == "left":
                self.ratings[left] = glicko_update(
                    left_rating,
                    right_rating,
                    1.0,
                )

                self.ratings[right] = glicko_update(
                    right_rating,
                    left_rating,
                    0.0,
                )

            elif result == "right":
                self.ratings[left] = glicko_update(
                    left_rating,
                    right_rating,
                    0.0,
                )

                self.ratings[right] = glicko_update(
                    right_rating,
                    left_rating,
                    1.0,
                )

            else:
                self.ratings[left] = glicko_update(
                    left_rating,
                    right_rating,
                    0.5,
                )

                self.ratings[right] = glicko_update(
                    right_rating,
                    left_rating,
                    0.5,
                )

    # --------------------------------------------------------
    # Goodreads reconciliation
    # --------------------------------------------------------

    def sync_goodreads(self, rows):
        """
        Reconcile a new Goodreads export with existing state.

        IMPORTANT:
        rows must be a list of dictionaries.

        New books are added without resetting existing evidence.

        Read books leave the active ranking.

        Currently-reading remains active and is a separate comparison
        cohort: it is only compared with other currently-reading books.

        To-read books are compared only with other to-read books.

        Ignored books remain ignored even if Goodreads says to-read.

        Imported series order is recorded as real ranking evidence: for
        #1, #2, #3 the engine records #1 > #2 and #2 > #3.

        Removed records are retained in the internal library so
        previous identity/history is not silently destroyed.
        """

        if not isinstance(rows, list):
            raise TypeError(
                "sync_goodreads expected a list of Goodreads "
                "row dictionaries."
            )

        old_library = self.library
        new_library = {}

        added = []
        seen_ids = set()

        for row in rows:
            if not isinstance(row, dict):
                continue

            book_id = stable_book_id(row)

            seen_ids.add(book_id)

            title = (
                normalize(
                    row.get("title")
                )
                or "(Untitled)"
            )

            author = normalize(
                row.get("author l-f")
                or row.get("author")
            )

            shelf = normalize(
                row.get("exclusive shelf")
            ).lower()

            if shelf in {
                "currently-reading",
                "currently reading",
            }:
                shelf = "currently-reading"

            elif shelf == "to-read":
                shelf = "to-read"

            elif shelf == "read":
                shelf = "read"

            old_book = old_library.get(
                book_id
            )

            if old_book:
                old_status = old_book.status

                # Manual lifecycle decisions take precedence over
                # the Goodreads source shelf.
                if old_status == "ignore":
                    status = "ignore"

                else:
                    status = (
                        shelf
                        if shelf in self.ALL_STATUSES
                        else old_status
                    )

            else:
                status = (
                    shelf
                    if shelf in self.ALL_STATUSES
                    else "to-read"
                )

                added.append(book_id)

            # The application lifecycle is authoritative.
            # Keep the shelf representation synchronized with it.
            if status in self.ALL_STATUSES:
                shelf = status

            goodreads_fields = {
                str(key): normalize(value)
                for key, value in row.items()
            }

            # Keep the normalized source record consistent with the
            # application's authoritative lifecycle. This matters for
            # Ignore and Currently Reading across later imports.
            goodreads_fields[
                "exclusive shelf"
            ] = shelf

            new_library[book_id] = Book(
                id=book_id,
                title=title,
                author=author,
                description=normalize(
                    row.get("description")
                ),
                pages=normalize(
                    row.get("number of pages")
                ),
                my_rating=normalize(
                    row.get("my rating")
                ),
                shelf=shelf,
                status=status,
                isbn=clean_isbn(
                    row.get("isbn")
                ),
                year=normalize(
                    row.get("year published")
                ),
                publisher=normalize(
                    row.get("publisher")
                ),
                goodreads_id=normalize(
                    row.get("book id - goodreads")
                ),
                goodreads_fields=goodreads_fields,
            )

        removed = []

        for book_id, book in old_library.items():
            if book_id not in seen_ids:
                removed.append(book_id)

                # Preserve old records.
                new_library[book_id] = book

        self.library = new_library

        self._sync_books()
        self._replay()

        # A series is real ranking evidence. When books arrive through an
        # import, establish the ordered relationships as actual comparisons
        # so they are replayed, persisted, counted, and undone exactly like
        # human comparisons. Existing relationships are never duplicated.
        series_comparisons_added = (
            self._apply_series_order_comparisons()
        )

        if series_comparisons_added:
            self._replay()

        return {
            "added": added,
            "removed": removed,
            "active": len(self.books),
            "series_comparisons_added": series_comparisons_added,
        }

    # --------------------------------------------------------
    # Statistical ranking
    # --------------------------------------------------------

    def statistics(self):
        if not self.books:
            return []

        def ranking_key(book):
            rating = self.ratings[book.id].rating
            name = series_name(book)
            number = series_number(book)

            # Series order is a hard ordering constraint within a single
            # series: a smaller valid Series-# always appears above a
            # larger valid Series-# for that same Series. Outside the same
            # series, human preference/rating remains the ranking signal.
            return (
                rating,
                name.casefold() if name else "",
                -number if name and number is not None else float("-inf"),
            )

        # First sort by human evidence, then repair ordering inside each
        # series so that the lowest valid Series-# is always first.
        ranked = sorted(
            self.books,
            key=lambda book: self.ratings[book.id].rating,
            reverse=True,
        )

        groups = {}
        for book in ranked:
            name = series_name(book)
            if name and series_number(book) is not None:
                groups.setdefault(name.casefold(), []).append(book)

        for group_name, members in groups.items():
            ordered_members = sorted(
                members,
                key=lambda book: (
                    series_number(book),
                    -self.ratings[book.id].rating,
                ),
            )
            member_positions = {book.id: index for index, book in enumerate(ordered_members)}
            ranked_positions = {book.id: index for index, book in enumerate(ranked) if book.id in member_positions}

            # Move members into the positions occupied by the series members,
            # preserving the rest of the global ranking.
            positions = sorted(ranked_positions.values())
            for position, member in zip(positions, ordered_members):
                ranked[position] = member

        total = len(ranked)
        output = []

        for position, book in enumerate(
            ranked,
            start=1,
        ):
            rating = self.ratings[
                book.id
            ]

            uncertainty = clamp(
                rating.rd / 350.0,
                0.0,
                1.0,
            )

            top25_probability = clamp(
                (
                    TOP_K
                    + 8
                    - position
                )
                / (
                    max(
                        8.0,
                        uncertainty * 35.0 + 8.0,
                    )
                ),
                0.0,
                1.0,
            )

            top10_probability = clamp(
                (
                    TOP_10
                    + 5
                    - position
                )
                / (
                    max(
                        5.0,
                        uncertainty * 30.0 + 5.0,
                    )
                ),
                0.0,
                1.0,
            )

            low = max(
                1,
                int(
                    round(
                        position
                        - rating.rd / 55.0
                    )
                ),
            )

            high = min(
                total,
                int(
                    round(
                        position
                        + rating.rd / 55.0
                    )
                ),
            )

            output.append(
                {
                    "book": book,
                    "rank": position,
                    "expected_rank": position,
                    "rank_low": low,
                    "rank_high": high,
                    "top10_probability":
                        top10_probability,
                    "top25_probability":
                        top25_probability,
                    "rating": rating,
                }
            )

        return output

    # --------------------------------------------------------
    # Active learning
    # --------------------------------------------------------

    def _pairing_pool(self):
        """Return the active shelf cohort eligible for the next comparison."""
        currently_reading = [
            book.id
            for book in self.books
            if book.status == "currently-reading"
        ]

        # Currently-reading is a hard comparison boundary. If there are
        # two or more currently-reading books, the next comparison MUST
        # come from that cohort and may not use any to-read book.
        if len(currently_reading) >= 2:
            return currently_reading

        to_read = [
            book.id
            for book in self.books
            if book.status == "to-read"
        ]

        # A single currently-reading book cannot be compared to a to-read
        # book, so allow the to-read cohort to continue independently.
        if len(to_read) >= 2:
            return to_read

        return []

    def _same_pairing_cohort(self, left, right):
        """Return True only when both books belong to the same active shelf."""
        if left not in self.library or right not in self.library:
            return False

        left_status = self.library[left].status
        right_status = self.library[right].status

        return (
            left_status == right_status
            and left_status in self.ACTIVE_STATUSES
        )

    def choose_pair(self):
        """Choose the next pair using shelf-local active learning.

        Priority order:
        1. Books with zero comparisons are introduced first.
        2. If two zero-comparison books are available, compare them together.
        3. Otherwise compare a zero-comparison book against the strongest
           available candidate in the same Exclusive Shelf cohort.
        4. Once every book in the cohort has at least one comparison, use the
           normal adaptive Glicko/top-25/boundary selector.

        Currently-reading and to-read books are never mixed in a comparison.
        """
        ids = self._pairing_pool()

        if len(ids) < 2:
            return None

        stats = self.statistics()
        by_id = {
            item["book"].id: item
            for item in stats
            if item["book"].id in set(ids)
        }

        allowed_ids = set(ids)

        # ----------------------------------------------------
        # SERIES ORDER PRIORITY
        # ----------------------------------------------------
        # When a cohort contains multiple books from the same series,
        # always work through the lowest available Series-# first. The
        # comparison itself is still a human decision; series metadata only
        # determines which comparison gets presented first.
        series_groups = {}
        for book_id in ids:
            book = self.library[book_id]
            name = series_name(book)
            number = series_number(book)
            if name and number is not None:
                series_groups.setdefault(name.casefold(), []).append(book)

        series_candidates = []

        for members in series_groups.values():
            members.sort(
                key=lambda book: (
                    series_number(book),
                    -self.ratings[book.id].rating,
                )
            )

            # Compare the earliest numbered book with the next numbered
            # instance. If there is a gap in numbering, the next available
            # larger number is used. Duplicate Series-# values are resolved
            # by existing rating/uncertainty logic.
            for index in range(len(members) - 1):
                left_book = members[index]
                right_book = members[index + 1]

                if series_number(left_book) == series_number(right_book):
                    continue

                left = left_book.id
                right = right_book.id
                pair = self.pair_key(left, right)

                if pair in self.played or self.skips.get(pair, 0) > 0:
                    continue

                if not self._same_pairing_cohort(left, right):
                    continue

                series_candidates.append(pair)

        if series_candidates:
            def series_priority(pair):
                left, right = pair
                left_book = self.library[left]
                right_book = self.library[right]
                left_number = series_number(left_book)
                right_number = series_number(right_book)

                return (
                    -min(left_number, right_number),
                    self.ratings[left].comparisons
                    + self.ratings[right].comparisons,
                    self.ratings[left].rd
                    + self.ratings[right].rd,
                )

            selected = max(
                series_candidates,
                key=series_priority,
            )

            # Always present the lower-numbered series item on the left so
            # the UI has a stable, readable series progression.
            left, right = selected
            if series_number(self.library[left]) > series_number(self.library[right]):
                return (right, left)
            return selected

        # ----------------------------------------------------
        # NEW-BOOK PRIORITY
        # ----------------------------------------------------
        # Rating.comparisons is rebuilt from the human comparison history,
        # so zero means the book has genuinely never been compared.
        unseen = [
            book_id
            for book_id in ids
            if self.ratings[book_id].comparisons == 0
        ]

        if unseen:
            # Prefer two never-compared books. This gives two books evidence
            # with a single human decision and quickly establishes an initial
            # ordering for the cohort.
            if len(unseen) >= 2:
                unseen_pair_candidates = []

                for i, left in enumerate(unseen):
                    for right in unseen[i + 1:]:
                        pair = self.pair_key(left, right)

                        if pair in self.played:
                            continue

                        if self.skips.get(pair, 0) > 0:
                            continue

                        if not self._same_pairing_cohort(left, right):
                            continue

                        unseen_pair_candidates.append(pair)

                if unseen_pair_candidates:
                    def unseen_pair_priority(pair):
                        left, right = pair
                        a = self.ratings[left]
                        b = self.ratings[right]

                        # Prefer the pair with the greatest combined
                        # uncertainty while keeping the initial sample
                        # broad and unbiased.
                        return (
                            a.rd + b.rd,
                            abs(a.rd - b.rd),
                        )

                    return max(
                        unseen_pair_candidates,
                        key=unseen_pair_priority,
                    )

            # If only one unseen book remains, pair it with the best
            # available book from the SAME shelf cohort. This guarantees the
            # unseen book gets its first piece of evidence before the ranker
            # returns to ordinary adaptive comparisons.
            priority_candidates = []

            for left in unseen:
                for right in ids:
                    if left == right:
                        continue

                    pair = self.pair_key(left, right)

                    if pair in self.played:
                        continue

                    if self.skips.get(pair, 0) > 0:
                        continue

                    if not self._same_pairing_cohort(left, right):
                        continue

                    priority_candidates.append(pair)

            if priority_candidates:
                def unseen_book_priority(pair):
                    left, right = pair
                    a = self.ratings[left]
                    b = self.ratings[right]
                    left_stats = by_id[left]
                    right_stats = by_id[right]

                    # Always prefer pairs that contain an unseen book.
                    unseen_count = int(a.comparisons == 0) + int(
                        b.comparisons == 0
                    )

                    top_relevance = (
                        left_stats["top25_probability"]
                        + right_stats["top25_probability"]
                    ) / 2.0

                    boundary_relevance = max(
                        0.0,
                        1.0
                        - abs(
                            (
                                left_stats["expected_rank"]
                                + right_stats["expected_rank"]
                            )
                            / 2.0
                            - TOP_K
                        )
                        / (BOUNDARY_WIDTH + 8.0),
                    )

                    return (
                        unseen_count,
                        max(a.rd, b.rd),
                        top_relevance,
                        boundary_relevance,
                    )

                return max(
                    priority_candidates,
                    key=unseen_book_priority,
                )

        # ----------------------------------------------------
        # NORMAL ADAPTIVE SELECTOR
        # ----------------------------------------------------
        ordered = [
            item["book"].id
            for item in stats
            if item["book"].id in allowed_ids
        ]

        elite_size = min(
            len(ids),
            max(
                40,
                min(
                    100,
                    int(
                        math.sqrt(len(ids)) * 5
                    ),
                ),
            ),
        )

        elite = ordered[:elite_size]

        boundary = ordered[
            max(0, TOP_K - BOUNDARY_WIDTH):
            min(len(ordered), TOP_K + BOUNDARY_WIDTH)
        ]

        high_rd = sorted(
            ids,
            key=lambda book_id: self.ratings[book_id].rd,
            reverse=True,
        )[:min(len(ids), 60)]

        pools = [elite, boundary, high_rd]
        candidates = []

        for pool in pools:
            if len(pool) < 2:
                continue

            for _ in range(min(100, len(pool) * 2)):
                left = self.rng.choice(pool)
                right = self.rng.choice(pool)

                if left == right:
                    continue

                pair = self.pair_key(left, right)

                if pair in self.played:
                    continue

                if self.skips.get(pair, 0) > 0:
                    continue

                if not self._same_pairing_cohort(left, right):
                    continue

                candidates.append(pair)

        # Bounded random exploration.
        if not candidates:
            attempts = min(
                PAIR_CANDIDATE_LIMIT,
                max(100, len(ids) * 3),
            )

            for _ in range(attempts):
                left = self.rng.choice(ids)
                right = self.rng.choice(ids)

                if left == right:
                    continue

                pair = self.pair_key(left, right)

                if pair in self.played:
                    continue

                if self.skips.get(pair, 0) > 0:
                    continue

                if not self._same_pairing_cohort(left, right):
                    continue

                candidates.append(pair)

                if len(candidates) >= 250:
                    break

        if not candidates:
            return None

        if self.rng.random() < 0.055:
            return self.rng.choice(candidates)

        def priority(pair):
            left, right = pair

            a = self.ratings[left]
            b = self.ratings[right]

            sa = by_id[left]
            sb = by_id[right]

            closeness = math.exp(
                -abs(a.rating - b.rating) / 120.0
            )

            uncertainty = (a.rd + b.rd) / 700.0

            top_relevance = (
                sa["top25_probability"]
                + sb["top25_probability"]
            ) / 2.0

            boundary_relevance = max(
                0.0,
                1.0
                - abs(
                    (
                        sa["expected_rank"]
                        + sb["expected_rank"]
                    )
                    / 2.0
                    - TOP_K
                )
                / (BOUNDARY_WIDTH + 8.0),
            )

            challenger = max(a.rd, b.rd) / 350.0

            balance = math.exp(
                -abs(a.comparisons - b.comparisons) / 8.0
            )

            return (
                0.30 * closeness
                + 0.24 * uncertainty
                + 0.22 * top_relevance
                + 0.16 * boundary_relevance
                + 0.08 * challenger
            ) * balance

        return max(candidates, key=priority)

    # --------------------------------------------------------
    # Automatic series-order evidence
    # --------------------------------------------------------

    def _apply_series_order_comparisons(self):
        """
        Record real ranking comparisons implied by series order.

        For a series containing #1, #2, #3, this records:

            #1 > #2
            #2 > #3

        These entries use the same comparison/replay machinery as human
        comparisons and are persisted in ``self.comparisons``. Existing
        pairs are never duplicated.

        Series comparisons respect the same active-shelf cohort rule as
        human comparisons: currently-reading books only compare with
        currently-reading books, and to-read only with to-read.
        """
        groups = {}

        for book in self.library.values():
            if book.status not in self.ACTIVE_STATUSES:
                continue

            name = series_name(book)
            number = series_number(book)

            if not name or number is None:
                continue

            key = (book.status, name.casefold())
            groups.setdefault(key, []).append(book)

        existing = {
            self.pair_key(
                item.get("left"),
                item.get("right"),
            )
            for item in self.comparisons
            if isinstance(item, dict)
            and item.get("left")
            and item.get("right")
        }

        created = 0

        for members in groups.values():
            members.sort(
                key=lambda book: (
                    series_number(book),
                    book.id,
                )
            )

            for left_book, right_book in zip(
                members,
                members[1:],
            ):
                left = left_book.id
                right = right_book.id
                pair = self.pair_key(left, right)

                if left_book.id == right_book.id:
                    continue

                if series_number(left_book) == series_number(right_book):
                    continue

                if pair in existing:
                    continue

                # The lower-numbered book is the winner. This is a genuine
                # comparison, not a rating override.
                self.apply_match(
                    left,
                    right,
                    "left",
                    source="series-order",
                )

                existing.add(pair)
                created += 1

        return created

    # --------------------------------------------------------
    # Human actions
    # --------------------------------------------------------

    def apply_match(
        self,
        left,
        right,
        result,
        source="human",
    ):
        if (
            left not in self.library
            or right not in self.library
        ):
            raise ValueError(
                "One of the selected books no longer exists."
            )

        if left == right:
            raise ValueError(
                "A book cannot be compared with itself."
            )

        if not self._same_pairing_cohort(left, right):
            raise ValueError(
                "Books can only be compared within the same active "
                "Exclusive Shelf cohort."
            )

        pair = self.pair_key(
            left,
            right,
        )

        if pair in self.played:
            raise ValueError(
                "That pair has already been compared."
            )

        if result not in {
            "left",
            "right",
            "tie",
        }:
            raise ValueError(
                "Invalid comparison result."
            )

        if source not in {
            "human",
            "series-order",
        }:
            raise ValueError(
                "Invalid comparison source."
            )

        self.comparisons.append(
            {
                "left": left,
                "right": right,
                "result": result,
                "source": source,
                "time": now_iso(),
            }
        )

        self.skips.pop(
            pair,
            None,
        )

        self._replay()

    def skip_pair(
        self,
        left,
        right,
    ):
        pair = self.pair_key(
            left,
            right,
        )

        self.skips[pair] = 3

    def tick_skips(self):
        for pair in list(
            self.skips
        ):
            self.skips[pair] -= 1

            if self.skips[pair] <= 0:
                del self.skips[pair]

    def undo(self):
        if not self.comparisons:
            return False

        self.comparisons.pop()

        self._replay()

        return True

    # --------------------------------------------------------
    # Lifecycle
    # --------------------------------------------------------

    def set_lifecycle(
        self,
        book_id,
        status,
    ):
        if status not in self.ALL_STATUSES:
            raise ValueError(
                f"Invalid lifecycle status: {status}"
            )

        if book_id not in self.library:
            raise KeyError(book_id)

        book = self.library[book_id]

        # Always update the internal lifecycle status.
        book.status = status
        self.statuses[book_id] = status

        # Keep the Book shelf and Goodreads source data synchronized.
        if status == "ignore":
            book.shelf = "ignore"

            if not isinstance(book.goodreads_fields, dict):
                book.goodreads_fields = {}

            book.goodreads_fields["exclusive shelf"] = "ignore"

        elif status in {
            "to-read",
            "currently-reading",
            "read",
        }:
            book.shelf = status

            if not isinstance(book.goodreads_fields, dict):
                book.goodreads_fields = {}

            book.goodreads_fields["exclusive shelf"] = status

        self._sync_books()
        self._replay()

    # --------------------------------------------------------
    # Progress
    # --------------------------------------------------------

    def progress(self):
        if not self.books:
            return 1.0

        resolved = sum(
            1
            for book in self.books
            if self.ratings[
                book.id
            ].comparisons >= self.target
        )

        return clamp(
            resolved / len(self.books),
            0.0,
            1.0,
        )

    # --------------------------------------------------------
    # Confidence
    # --------------------------------------------------------

    def confidence_metrics(self):
        stats = self.statistics()

        if not stats:
            return {
                "top10_confidence": 0.0,
                "top25_confidence": 0.0,
                "top25_stability": 0.0,
                "unresolved_boundary": 0,
            }

        top25 = stats[:TOP_K]

        top25_confidence = (
            sum(
                x["top25_probability"]
                for x in top25
            )
            / max(
                1,
                len(top25),
            )
        )

        top25_stability = clamp(
            1.0
            - (
                sum(
                    x["rating"].rd
                    for x in top25
                )
                / (
                    350.0
                    * max(
                        1,
                        len(top25),
                    )
                )
            ),
            0.0,
            1.0,
        )

        top10_confidence = (
            sum(
                x["top10_probability"]
                for x in stats[:TOP_10]
            )
            / max(
                1,
                min(
                    TOP_10,
                    len(stats),
                ),
            )
        )

        unresolved_boundary = sum(
            1
            for x in stats
            if (
                x["rank_low"]
                <= TOP_K
                <= x["rank_high"]
            )
        )

        return {
            "top10_confidence":
                top10_confidence,
            "top25_confidence":
                top25_confidence,
            "top25_stability":
                top25_stability,
            "unresolved_boundary":
                unresolved_boundary,
        }

    def should_stop(self):
        if len(self.comparisons) < max(
            15,
            min(
                50,
                len(self.books) * 2,
            ),
        ):
            return False

        if not self.top_focus:
            return False

        metrics = self.confidence_metrics()

        return (
            metrics[
                "top25_stability"
            ] >= 0.82
            and metrics[
                "unresolved_boundary"
            ]
            <= max(
                3,
                TOP_K // 5,
            )
        )

    # --------------------------------------------------------
    # Persistence
    # --------------------------------------------------------

    def to_state(self):
        return {
            "state_version":
                STATE_VERSION,

            "app_version":
                APP_VERSION,

            "mode":
                self.mode,

            "target":
                self.target,

            "seed":
                self.seed,

            "comparisons":
                self.comparisons,

            "skips": {
                "|".join(pair): value
                for pair, value
                in self.skips.items()
            },

            "books": {
                book_id: asdict(book)
                for book_id, book
                in self.library.items()
            },
        }


# ============================================================
# STATE STORE
# ============================================================

class StateStore:
    def __init__(self, source):
        source = Path(source)

        directory = (
            source.parent
            / ".ranker_state"
        )

        directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        self.path = (
            directory
            / f"{source.stem}.json"
        )

    def save(self, engine):
        temporary = self.path.with_suffix(
            ".tmp"
        )

        data = engine.to_state()

        temporary.write_text(
            json.dumps(
                data,
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        os.replace(
            temporary,
            self.path,
        )

    def load_raw(self):
        if not self.path.exists():
            return None

        try:
            data = json.loads(
                self.path.read_text(
                    encoding="utf-8"
                )
            )

            if not isinstance(data, dict):
                return None

            version = int(
                data.get(
                    "state_version",
                    0,
                )
            )

            if version > STATE_VERSION:
                return None

            # Old/malformed state files should never be allowed
            # to crash the application.
            if not isinstance(
                data.get(
                    "comparisons",
                    [],
                ),
                list,
            ):
                data["comparisons"] = []

            if not isinstance(
                data.get(
                    "books",
                    {},
                ),
                dict,
            ):
                data["books"] = {}

            if not isinstance(
                data.get(
                    "skips",
                    {},
                ),
                dict,
            ):
                data["skips"] = {}

            return data

        except Exception:
            return None


# ============================================================
# EXCEL EXPORT
# ============================================================

def excel_column_name(number):
    result = ""

    while number:
        number, remainder = divmod(
            number - 1,
            26,
        )

        result = (
            chr(
                65 + remainder
            )
            + result
        )

    return result

def update_goodreads_source_sheet(
    worksheet,
    engine,
):
    """
    Write the current Book.goodreads_fields values back into
    the original Goodreads worksheet.

    Rows are matched primarily by Goodreads Book Id.
    Existing columns are preserved.
    """

    if worksheet.max_row < 1:
        return

    headers = {}

    for column in range(
        1,
        worksheet.max_column + 1,
    ):
        value = worksheet.cell(
            row=1,
            column=column,
        ).value

        key = normalize_header(value)

        if key:
            headers[key] = column

    id_column = headers.get(
        "book id - goodreads"
    )

    if not id_column:
        raise ValueError(
            "The Goodreads worksheet does not contain "
            "'Book Id - Goodreads', so edited books cannot "
            "be safely matched to their original rows."
        )

    source_rows = {}

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        value = normalize(
            worksheet.cell(
                row=row_number,
                column=id_column,
            ).value
        )

        if value:
            source_rows[value] = row_number

    updated = 0

    for book in engine.library.values():

        goodreads_id = normalize(
            book.goodreads_id
        )

        if not goodreads_id:
            continue

        row_number = source_rows.get(
            goodreads_id
        )

        if not row_number:
            continue

        fields = dict(
            getattr(
                book,
                "goodreads_fields",
                {},
            )
            or {}
        )

        for key, value in fields.items():

            column = headers.get(
                normalize_header(key)
            )

            if not column:
                # Do not invent columns in the middle of
                # the original Goodreads export.
                continue

            # Never overwrite the identity key from the editor.
            if normalize_header(key) == (
                "book id - goodreads"
            ):
                continue

            worksheet.cell(
                row=row_number,
                column=column,
            ).value = value

        updated += 1

    return updated
    
def export_results(
    source,
    engine,
    output=None,
):
    """
    Preserve the original Goodreads workbook and add:

        Ranking
        Summary

    The source is not overwritten unless output is explicitly
    the source path.
    """

    source = Path(source)

    output = (
        Path(output)
        if output
        else source.with_name(
            source.stem
            + "_ranked.xlsx"
        )
    )

    workbook = load_workbook(
        source
    )

    try:
        # --------------------------------------------------------
        # IMPORTANT:
        # Write edits from the cards back into the original
        # Goodreads worksheet before creating Ranking/Summary.
        # --------------------------------------------------------

        source_sheet = workbook.active

        update_goodreads_source_sheet(
            source_sheet,
            engine,
        )

        if "Ranking" in workbook.sheetnames:
            del workbook["Ranking"]

        ranking = workbook.create_sheet(
            "Ranking"
        )

        headings = [
            "Display",
            "Statistical Rank",
            "Title",
            "Author",
            "Status",
            "Rating",
            "Rating Uncertainty",
            "Rank Low",
            "Rank High",
            "Top 10 %",
            "Top 25 %",
            "Comparisons",
            "Wins",
            "Losses",
            "Ties",
        ]

        ranking.append(
            headings
        )

        for cell in ranking[1]:
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = PatternFill(
                "solid",
                fgColor="24324D",
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        statistics = engine.statistics()

        current = [
            item
            for item in statistics
            if item["book"].status
            == "currently-reading"
        ]

        rest = [
            item
            for item in statistics
            if item["book"].status
            != "currently-reading"
        ]

        for item in current + rest:
            book = item["book"]
            rating = item["rating"]

            if (
                book.status
                == "currently-reading"
            ):
                display = "CURRENTLY READING"

            elif item["rank"] <= TOP_K:
                display = "TOP 25"

            else:
                display = "QUEUE"

            ranking.append(
                [
                    display,
                    item["rank"],
                    book.title,
                    book.author,
                    book.status,
                    rating.rating,
                    rating.rd,
                    item["rank_low"],
                    item["rank_high"],
                    item[
                        "top10_probability"
                    ],
                    item[
                        "top25_probability"
                    ],
                    rating.comparisons,
                    rating.wins,
                    rating.losses,
                    rating.ties,
                ]
            )

        widths = [
            22,
            16,
            40,
            28,
            22,
            12,
            20,
            12,
            12,
            12,
            12,
            14,
            10,
            10,
            10,
        ]

        for index, width in enumerate(
            widths,
            start=1,
        ):
            ranking.column_dimensions[
                excel_column_name(index)
            ].width = width

        ranking.freeze_panes = "A2"
        ranking.auto_filter.ref = (
            ranking.dimensions
        )

        if "Summary" in workbook.sheetnames:
            del workbook["Summary"]

        summary = workbook.create_sheet(
            "Summary"
        )

        metrics = (
            engine.confidence_metrics()
        )

        average_comparisons = (
            len(engine.comparisons)
            / max(
                1,
                len(engine.books),
            )
        )

        summary_rows = [
            [
                "Goodreads To-Read Ranker",
                APP_VERSION,
            ],
            [
                "Generated",
                now_iso(),
            ],
            [
                "Active books",
                len(engine.books),
            ],
            [
                "Total comparisons",
                len(engine.comparisons),
            ],
            [
                "Average comparisons/book",
                average_comparisons,
            ],
            [
                "Mode",
                engine.mode,
            ],
            [
                "Top-10 confidence estimate",
                metrics[
                    "top10_confidence"
                ],
            ],
            [
                "Top-25 confidence estimate",
                metrics[
                    "top25_confidence"
                ],
            ],
            [
                "Top-25 stability estimate",
                metrics[
                    "top25_stability"
                ],
            ],
            [
                "Unresolved Top-25 boundary",
                metrics[
                    "unresolved_boundary"
                ],
            ],
            [],
            [
                "Methodology",
                (
                    "Human pairwise choices are the only "
                    "preference evidence. Glicko-style "
                    "ratings track online uncertainty and "
                    "the adaptive selector chooses useful "
                    "comparisons. Goodreads metadata does "
                    "not determine preference."
                ),
            ],
        ]

        for row in summary_rows:
            summary.append(row)

        summary.column_dimensions[
            "A"
        ].width = 34

        summary.column_dimensions[
            "B"
        ].width = 100

        summary["B12"].alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        workbook.save(
            output
        )

    finally:
        workbook.close()

    return output


# ============================================================
# GUI
# ============================================================

class RankerApp:
    def __init__(self, root):
        self.root = root

        self.theme_name = "dark"
        self.colors = DARK

        self.engine = None
        self.source_file = None
        self.state_store = None
        self.current_pair = None

        self.font = "Segoe UI"

        self.root.title(
            f"{APP_NAME} {APP_VERSION}"
        )

        self.root.geometry(
            f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}"
        )

        self.root.minsize(
            900,
            620,
        )

        self.build_styles()
        self.build_ui()
        self.bind_shortcuts()
        self.apply_theme()

        self.root.protocol(
            "WM_DELETE_WINDOW",
            self.close,
        )

    # --------------------------------------------------------
    # Styling
    # --------------------------------------------------------

    def build_styles(self):
        self.style = ttk.Style(
            self.root
        )

        try:
            self.style.theme_use(
                "clam"
            )
        except tk.TclError:
            pass

        self.configure_styles()

    def configure_styles(self):
        c = self.colors

        self.style.configure(
            "TButton",
            background=c["button"],
            foreground=c["text"],
            bordercolor=c["border"],
            padding=(10, 7),
            font=(
                self.font,
                9,
                "bold",
            ),
        )

        self.style.map(
            "TButton",
            background=[
                (
                    "active",
                    c["button_hover"],
                )
            ],
        )

        self.style.configure(
            "Choice.TButton",
            background=c["purple"],
            foreground="#FFFFFF",
            padding=(12, 11),
            font=(
                self.font,
                11,
                "bold",
            ),
        )

        self.style.map(
            "Choice.TButton",
            background=[
                (
                    "active",
                    c["accent"],
                )
            ],
        )

        self.style.configure(
            "Accent.TButton",
            background=c["accent"],
            foreground="#061018",
            padding=(10, 7),
            font=(
                self.font,
                9,
                "bold",
            ),
        )

        self.style.configure(
            "Treeview",
            background=c["panel"],
            fieldbackground=c["panel"],
            foreground=c["text"],
            rowheight=28,
        )

        self.style.configure(
            "Treeview.Heading",
            background=c["button"],
            foreground=c["text"],
            font=(
                self.font,
                9,
                "bold",
            ),
        )

        self.style.configure(
            "TProgressbar",
            troughcolor=c["button"],
            background=c["accent"],
            bordercolor=c["border"],
            lightcolor=c["accent"],
            darkcolor=c["accent"],
        )

    # --------------------------------------------------------
    # Main UI
    # --------------------------------------------------------

    def build_ui(self):
        c = self.colors

        self.main = tk.Frame(
            self.root,
            bg=c["bg"],
            padx=14,
            pady=12,
        )

        self.main.pack(
            fill="both",
            expand=True,
        )

        self.main.grid_columnconfigure(
            0,
            weight=1,
        )

        self.main.grid_rowconfigure(
            3,
            weight=1,
        )

        # Header.
        header = tk.Frame(
            self.main,
            bg=c["bg"],
        )

        header.grid(
            row=0,
            column=0,
            sticky="ew",
        )

        header.grid_columnconfigure(
            1,
            weight=1,
        )

        title = tk.Label(
            header,
            text=(
                "📚  GOODREADS "
                "TO-READ RANKER"
            ),
            bg=c["bg"],
            fg=c["text"],
            font=(
                self.font,
                20,
                "bold",
            ),
        )

        title.grid(
            row=0,
            column=0,
            sticky="w",
        )

        subtitle = tk.Label(
            header,
            text=(
                "TOP-25 FOCUS  •  "
                "HUMAN CHOICES ONLY  •  "
                "ADAPTIVE EVIDENCE  •  "
                f"{APP_VERSION}"
            ),
            bg=c["bg"],
            fg=c["accent"],
            font=(
                self.font,
                8,
                "bold",
            ),
        )

        subtitle.grid(
            row=1,
            column=0,
            sticky="w",
        )

        controls = tk.Frame(
            header,
            bg=c["bg"],
        )

        controls.grid(
            row=0,
            column=1,
            rowspan=2,
            sticky="e",
        )

        self.mode_var = tk.StringVar(
            value="TOP_25_FOCUS"
        )

        mode_box = ttk.Combobox(
            controls,
            textvariable=self.mode_var,
            values=list(PRESETS),
            state="readonly",
            width=16,
        )

        mode_box.pack(
            side="left",
            padx=3,
        )

        mode_box.bind(
            "<<ComboboxSelected>>",
            self.change_mode,
        )

        ttk.Button(
            controls,
            text="Open Excel",
            style="Accent.TButton",
            command=self.open_excel,
        ).pack(
            side="left",
            padx=3,
        )

        self.theme_button = ttk.Button(
            controls,
            text="☀ Light mode",
            command=self.toggle_theme,
        )

        self.theme_button.pack(
            side="left",
            padx=3,
        )

        ttk.Button(
            controls,
            text="⌨ Shortcuts",
            command=self.show_shortcuts,
        ).pack(
            side="left",
            padx=3,
        )

        # Dashboard.
        dashboard = tk.Frame(
            self.main,
            bg=c["panel"],
            highlightthickness=1,
            highlightbackground=c["border"],
        )

        dashboard.grid(
            row=1,
            column=0,
            sticky="ew",
            pady=8,
        )

        self.info_var = tk.StringVar(
            value=(
                "Open your Goodreads "
                "Excel export to begin."
            )
        )

        tk.Label(
            dashboard,
            textvariable=self.info_var,
            bg=c["panel"],
            fg=c["text"],
            font=(
                self.font,
                9,
                "bold",
            ),
            padx=10,
            pady=7,
            anchor="w",
        ).pack(
            fill="x"
        )

        self.progress = ttk.Progressbar(
            self.main,
            maximum=1.0,
        )

        self.progress.grid(
            row=2,
            column=0,
            sticky="ew",
            pady=(0, 7),
        )

        # Comparison area.
        comparison = tk.Frame(
            self.main,
            bg=c["bg"],
        )

        comparison.grid(
            row=3,
            column=0,
            sticky="nsew",
        )

        comparison.grid_rowconfigure(
            0,
            weight=1,
        )

        comparison.grid_columnconfigure(
            0,
            weight=1,
        )

        comparison.grid_columnconfigure(
            1,
            weight=0,
        )

        comparison.grid_columnconfigure(
            2,
            weight=1,
        )

        self.left_card = self.create_book_card(
            comparison,
            "left",
        )

        self.left_card.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=(0, 6),
        )

        tk.Label(
            comparison,
            text="VS",
            bg=c["bg"],
            fg=c["purple"],
            font=(
                self.font,
                15,
                "bold",
            ),
        ).grid(
            row=0,
            column=1,
            padx=4,
        )

        self.right_card = self.create_book_card(
            comparison,
            "right",
        )

        self.right_card.grid(
            row=0,
            column=2,
            sticky="nsew",
            padx=(6, 0),
        )

        # Decision row.
        decisions = tk.Frame(
            self.main,
            bg=c["bg"],
        )

        decisions.grid(
            row=4,
            column=0,
            sticky="ew",
            pady=7,
        )

        for column in range(3):
            decisions.grid_columnconfigure(
                column,
                weight=1,
            )

        self.left_button = ttk.Button(
            decisions,
            text="←  CHOOSE LEFT  [1]",
            style="Choice.TButton",
            command=lambda:
                self.choose("left"),
        )

        self.left_button.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=3,
            ipady=8,
        )

        self.tie_button = ttk.Button(
            decisions,
            text="≈  TIE  [3 / T]",
            style="Choice.TButton",
            command=lambda:
                self.choose("tie"),
        )

        self.tie_button.grid(
            row=0,
            column=1,
            sticky="ew",
            padx=3,
            ipady=8,
        )

        self.right_button = ttk.Button(
            decisions,
            text="CHOOSE RIGHT  [2]  →",
            style="Choice.TButton",
            command=lambda:
                self.choose("right"),
        )

        self.right_button.grid(
            row=0,
            column=2,
            sticky="ew",
            padx=3,
            ipady=8,
        )

        # Utility row.
        utility = tk.Frame(
            self.main,
            bg=c["bg"],
        )

        utility.grid(
            row=5,
            column=0,
            sticky="ew",
        )

        utility_buttons = [
            (
                "↶ Undo [U]",
                self.undo,
            ),
            (
                "Skip [4 / S]",
                self.skip,
            ),
            (
                "Ranking",
                self.show_ranking,
            ),
            (
                "Export [Ctrl+E]",
                self.export,
            ),
            (
                "Overwrite + Backup",
                self.overwrite_source,
            ),
        ]

        for text, command in utility_buttons:
            ttk.Button(
                utility,
                text=text,
                command=command,
            ).pack(
                side="left",
                padx=3,
            )

        # Shortcut panel.
        self.shortcut_panel = tk.Frame(
            self.main,
            bg=c["panel"],
            highlightthickness=1,
            highlightbackground=c["border"],
        )

        self.shortcut_panel.grid(
            row=6,
            column=0,
            sticky="ew",
            pady=(7, 0),
        )

        self.build_shortcut_panel()

    # --------------------------------------------------------
    # Book cards
    # --------------------------------------------------------

    def create_book_card(
        self,
        parent,
        side,
    ):
        c = self.colors

        card = tk.Frame(
            parent,
            bg=c["card"],
            highlightthickness=1,
            highlightbackground=c["border"],
            cursor="hand2",
        )

        card.grid_rowconfigure(
            5,
            weight=1,
        )

        card.grid_columnconfigure(
            0,
            weight=1,
        )

        card.grid_columnconfigure(
            1,
            weight=0,
        )

        title_var = tk.StringVar(
            value="Waiting for a library…"
        )

        author_var = tk.StringVar()
        meta_var = tk.StringVar()
        status_var = tk.StringVar()

        setattr(
            self,
            f"{side}_title",
            title_var,
        )

        setattr(
            self,
            f"{side}_author",
            author_var,
        )

        setattr(
            self,
            f"{side}_meta",
            meta_var,
        )

        setattr(
            self,
            f"{side}_status",
            status_var,
        )

        tk.Label(
            card,
            textvariable=status_var,
            bg=c["card2"],
            fg=c["accent"],
            font=(
                self.font,
                8,
                "bold",
            ),
            padx=8,
            pady=3,
        ).grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="w",
            padx=10,
            pady=(8, 3),
        )

        title_label = tk.Label(
            card,
            textvariable=title_var,
            bg=c["card"],
            fg=c["text"],
            font=(
                self.font,
                16,
                "bold",
            ),
            wraplength=450,
            justify="center",
        )

        title_label.grid(
            row=1,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=14,
            pady=(6, 4),
        )

        search_button = tk.Button(
            card,
            text="🔎  Search this book",
            bg=c["button"],
            fg=c["text"],
            activebackground=c["button_hover"],
            activeforeground=c["accent"],
            font=(
                self.font,
                9,
                "bold",
            ),
            relief="flat",
            bd=0,
            cursor="hand2",
            padx=12,
            pady=6,
            command=lambda current_side=side:
                self.search_current_book(current_side),
        )
        edit_button = tk.Button(
            card,
            text="✏️  Edit book",
            bg=c["button"],
            fg=c["text"],
            activebackground=c["button_hover"],
            activeforeground=c["accent"],
            font=(
                self.font,
                9,
                "bold",
            ),
            relief="flat",
            bd=0,
            cursor="hand2",
            padx=12,
            pady=6,
            command=lambda current_side=side:
                self.edit_current_book(current_side),
        )

        edit_button.grid(
            row=2,
            column=1,
            sticky="e",
            padx=(4, 14),
            pady=(0, 6),
        )
        search_button.grid(
            row=2,
            column=0,
            sticky="ew",
            padx=14,
            pady=(0, 6),
        )

        tk.Label(
            card,
            textvariable=author_var,
            bg=c["card"],
            fg=c["muted"],
            font=(
                self.font,
                10,
            ),
            wraplength=450,
        ).grid(
            row=3,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=14,
        )

        tk.Label(
            card,
            textvariable=meta_var,
            bg=c["card"],
            fg=c["muted"],
            font=(
                self.font,
                8,
            ),
            wraplength=450,
        ).grid(
            row=4,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=10,
            pady=4,
        )

        details = tk.Text(
            card,
            height=13,
            wrap="word",
            bg=c["card"],
            fg=c["text"],
            insertbackground=c["text"],
            selectbackground=c["card2"],
            selectforeground=c["text"],
            relief="flat",
            bd=0,
            font=(
                self.font,
                9,
            ),
            padx=4,
            pady=4,
        )

        details.grid(
            row=5,
            column=0,
            sticky="nsew",
            padx=(10, 0),
            pady=(2, 0),
        )

        scrollbar = ttk.Scrollbar(
            card,
            orient="vertical",
            command=details.yview,
        )

        scrollbar.grid(
            row=5,
            column=1,
            sticky="ns",
            padx=(0, 8),
            pady=(2, 0),
        )

        details.configure(
            state="disabled",
            yscrollcommand=scrollbar.set,
        )

        setattr(
            self,
            f"{side}_details",
            details,
        )

        # Keep the old attribute name as an alias so older state/UI
        # code cannot break if it still refers to *_description.
        setattr(
            self,
            f"{side}_description",
            details,
        )

        lifecycle = tk.Frame(
            card,
            bg=c["card"],
        )

        lifecycle.grid(
            row=6,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=8,
            pady=7,
        )

        for index in range(4):
            lifecycle.grid_columnconfigure(
                index,
                weight=1,
            )

        actions = [
            (
                "📖 Reading",
                "currently-reading",
            ),
            (
                "✓ Read",
                "read",
            ),
            (
                "🚫 Ignore",
                "ignore",
            ),
            (
                "↺ To-read",
                "to-read",
            ),
        ]

        for index, (
            text,
            status,
        ) in enumerate(actions):
            ttk.Button(
                lifecycle,
                text=text,
                command=lambda
                s=status,
                current_side=side:
                    self.set_lifecycle(
                        current_side,
                        s,
                    ),
            ).grid(
                row=0,
                column=index,
                sticky="ew",
                padx=2,
            )

        self.bind_card_clicks(
            card,
            side,
        )

        return card

    def bind_card_clicks(
        self,
        widget,
        side,
    ):
        """Bind lifecycle menu to the card background only."""

        widget.bind(
            "<Button-1>",
            lambda event:
                self.lifecycle_menu(
                    event,
                    side,
                ),
        )

    # --------------------------------------------------------
    # Theme
    # --------------------------------------------------------

    def toggle_theme(self):
        self.theme_name = (
            "light"
            if self.theme_name == "dark"
            else "dark"
        )

        self.colors = (
            LIGHT
            if self.theme_name == "light"
            else DARK
        )

        self.apply_theme()

    def apply_theme(self):
        c = self.colors

        self.root.configure(
            bg=c["bg"]
        )

        self.configure_styles()

        self.theme_button.configure(
            text=(
                "☀ Light mode"
                if self.theme_name == "dark"
                else "🌙 Dark mode"
            )
        )

        self.rebuild_ui_theme(
            self.root
        )

        if self.engine:
            self.refresh()

    def rebuild_ui_theme(
        self,
        widget,
    ):
        c = self.colors

        try:
            if isinstance(
                widget,
                tk.Text,
            ):
                widget.configure(
                    bg=c["card"],
                    fg=c["text"],
                    insertbackground=c["text"],
                )

            elif isinstance(
                widget,
                tk.Frame,
            ):
                current_bg = widget.cget(
                    "bg"
                )

                if current_bg in {
                    DARK["bg"],
                    LIGHT["bg"],
                }:
                    widget.configure(
                        bg=c["bg"]
                    )

                elif current_bg in {
                    DARK["panel"],
                    LIGHT["panel"],
                }:
                    widget.configure(
                        bg=c["panel"]
                    )

                elif current_bg in {
                    DARK["card"],
                    LIGHT["card"],
                }:
                    widget.configure(
                        bg=c["card"]
                    )

                elif current_bg in {
                    DARK["card2"],
                    LIGHT["card2"],
                }:
                    widget.configure(
                        bg=c["card2"]
                    )

            elif isinstance(
                widget,
                tk.Label,
            ):
                current_bg = widget.cget(
                    "bg"
                )

                if current_bg in {
                    DARK["bg"],
                    LIGHT["bg"],
                }:
                    widget.configure(
                        bg=c["bg"]
                    )

                elif current_bg in {
                    DARK["panel"],
                    LIGHT["panel"],
                }:
                    widget.configure(
                        bg=c["panel"]
                    )

                elif current_bg in {
                    DARK["card"],
                    LIGHT["card"],
                }:
                    widget.configure(
                        bg=c["card"]
                    )

                elif current_bg in {
                    DARK["card2"],
                    LIGHT["card2"],
                }:
                    widget.configure(
                        bg=c["card2"]
                    )

        except tk.TclError:
            pass

        for child in widget.winfo_children():
            self.rebuild_ui_theme(
                child
            )

    # --------------------------------------------------------
    # Keyboard shortcuts
    # --------------------------------------------------------

    def bind_shortcuts(self):
        root = self.root

        root.bind_all(
            "<KeyPress-1>",
            lambda event:
                self.choose("left"),
        )

        root.bind_all(
            "<KeyPress-2>",
            lambda event:
                self.choose("right"),
        )

        root.bind_all(
            "<KeyPress-3>",
            lambda event:
                self.choose("tie"),
        )

        root.bind_all(
            "<KeyPress-t>",
            lambda event:
                self.choose("tie"),
        )

        root.bind_all(
            "<KeyPress-T>",
            lambda event:
                self.choose("tie"),
        )

        root.bind_all(
            "<KeyPress-4>",
            lambda event:
                self.skip(),
        )

        root.bind_all(
            "<KeyPress-s>",
            lambda event:
                self.skip(),
        )

        root.bind_all(
            "<KeyPress-S>",
            lambda event:
                self.skip(),
        )

        root.bind_all(
            "<Left>",
            lambda event:
                self.choose("left"),
        )

        root.bind_all(
            "<Right>",
            lambda event:
                self.choose("right"),
        )

        root.bind_all(
            "<KeyPress-u>",
            lambda event:
                self.undo(),
        )

        root.bind_all(
            "<KeyPress-U>",
            lambda event:
                self.undo(),
        )

        root.bind_all(
            "<KeyPress-i>",
            lambda event:
                self.set_lifecycle(
                    "left",
                    "ignore",
                ),
        )

        root.bind_all(
            "<KeyPress-r>",
            lambda event:
                self.set_lifecycle(
                    "left",
                    "read",
                ),
        )

        root.bind_all(
            "<KeyPress-c>",
            lambda event:
                self.set_lifecycle(
                    "left",
                    "currently-reading",
                ),
        )

        root.bind_all(
            "<Shift-I>",
            lambda event:
                self.set_lifecycle(
                    "right",
                    "ignore",
                ),
        )

        root.bind_all(
            "<Shift-R>",
            lambda event:
                self.set_lifecycle(
                    "right",
                    "read",
                ),
        )

        root.bind_all(
            "<Shift-C>",
            lambda event:
                self.set_lifecycle(
                    "right",
                    "currently-reading",
                ),
        )

        root.bind_all(
            "<Control-o>",
            lambda event:
                self.open_excel(),
        )

        root.bind_all(
            "<Control-e>",
            lambda event:
                self.export(),
        )

    # --------------------------------------------------------
    # Shortcut panel
    # --------------------------------------------------------

    def build_shortcut_panel(self):
        c = self.colors

        tk.Label(
            self.shortcut_panel,
            text=(
                "⌨  QUICK KEYBOARD REFERENCE"
            ),
            bg=c["panel"],
            fg=c["accent"],
            font=(
                self.font,
                9,
                "bold",
            ),
        ).pack(
            anchor="w",
            padx=9,
            pady=(5, 1),
        )

        shortcuts = [
            ("1", "Choose LEFT"),
            ("2", "Choose RIGHT"),
            ("3 / T", "Tie"),
            ("4 / S", "Skip"),
            ("← / →", "Choose left / right"),
            ("U", "Undo"),
            ("I", "Mark LEFT Ignore"),
            ("R", "Mark LEFT Read"),
            ("C", "Mark LEFT Currently Reading"),
            ("Shift+I", "Mark RIGHT Ignore"),
            ("Shift+R", "Mark RIGHT Read"),
            ("Shift+C", "Mark RIGHT Currently Reading"),
            ("Ctrl+O", "Open Goodreads Excel"),
            ("Ctrl+E", "Export Ranking"),
            ("Click book card", "Open lifecycle menu"),
        ]

        grid = tk.Frame(
            self.shortcut_panel,
            bg=c["panel"],
        )

        grid.pack(
            fill="x",
            padx=8,
            pady=(0, 6),
        )

        for index, (
            key,
            description,
        ) in enumerate(shortcuts):

            row = index // 3
            column = index % 3

            cell = tk.Frame(
                grid,
                bg=c["panel"],
            )

            cell.grid(
                row=row,
                column=column,
                sticky="ew",
                padx=3,
                pady=1,
            )

            grid.grid_columnconfigure(
                column,
                weight=1,
            )

            tk.Label(
                cell,
                text=key,
                bg=c["panel"],
                fg=c["accent"],
                font=(
                    "Consolas",
                    8,
                    "bold",
                ),
                width=12,
                anchor="w",
            ).pack(
                side="left"
            )

            tk.Label(
                cell,
                text=description,
                bg=c["panel"],
                fg=c["text"],
                font=(
                    self.font,
                    8,
                ),
                anchor="w",
            ).pack(
                side="left",
                fill="x",
                expand=True,
            )

    # --------------------------------------------------------
    # File opening
    # --------------------------------------------------------

    def open_excel(self):
        path = filedialog.askopenfilename(
            title="Open Goodreads Excel",
            filetypes=[
                (
                    "Excel workbooks",
                    "*.xlsx *.xlsm",
                ),
                (
                    "All files",
                    "*.*",
                ),
            ],
        )

        if not path:
            return

        try:
            rows = import_goodreads(
                path
            )

            if not rows:
                raise ValueError(
                    "No Goodreads books were found "
                    "in this workbook."
                )

            self.source_file = Path(
                path
            )

            self.state_store = StateStore(
                self.source_file
            )

            raw_state = (
                self.state_store.load_raw()
            )

            books = [
                book_from_row(row)
                for row in rows
                if isinstance(row, dict)
            ]

            if not books:
                raise ValueError(
                    "The workbook contained rows, but "
                    "none could be converted into books."
                )

            if raw_state:
                saved_books = raw_state.get(
                    "books",
                    {},
                )

                if not isinstance(
                    saved_books,
                    dict,
                ):
                    saved_books = {}

                # Restore manually managed lifecycle states.
                for book in books:
                    saved = saved_books.get(
                        book.id
                    )

                    if not isinstance(
                        saved,
                        dict,
                    ):
                        continue

                    saved_status = saved.get(
                        "status"
                    )

                    shelf = book.shelf

                    if (
                        saved_status
                        in RankingEngine.ALL_STATUSES
                    ):
                        if (
                            saved_status == "ignore"
                            and shelf == "to-read"
                        ):
                            book.status = "ignore"

                        elif (
                            saved_status
                            == "currently-reading"
                            and shelf
                            in {
                                "to-read",
                                "currently-reading",
                            }
                        ):
                            book.status = (
                                "currently-reading"
                            )

                saved_statuses = {
                    key: value.get(
                        "status",
                        "to-read",
                    )
                    for key, value
                    in saved_books.items()
                    if isinstance(
                        value,
                        dict,
                    )
                }

                self.engine = RankingEngine(
                    books,
                    mode=raw_state.get(
                        "mode",
                        "TOP_25_FOCUS",
                    ),
                    seed=raw_state.get(
                        "seed"
                    ),
                    comparisons=raw_state.get(
                        "comparisons",
                        [],
                    ),
                    statuses=saved_statuses,
                    skips=raw_state.get(
                        "skips",
                        {},
                    ),
                )

                self.engine.sync_goodreads(
                    rows
                )

                # Make the mode selector agree with loaded state.
                self.mode_var.set(
                    self.engine.mode
                )

            else:
                self.engine = RankingEngine(
                    books,
                    mode=self.mode_var.get(),
                )

            self.state_store.save(
                self.engine
            )

            self.current_pair = None

            self.refresh()

        except Exception as exc:
            messagebox.showerror(
                "Could not open Goodreads workbook",
                (
                    f"{exc}\n\n"
                    "If this is a Goodreads export, make "
                    "sure the workbook contains a 'Title' "
                    "column."
                ),
            )

    # --------------------------------------------------------
    # Mode
    # --------------------------------------------------------

    def change_mode(
        self,
        _event=None,
    ):
        if not self.engine:
            return

        mode = self.mode_var.get()

        if mode not in PRESETS:
            return

        self.engine.mode = mode

        (
            self.engine.target,
            self.engine.top_focus,
        ) = PRESETS[mode]

        self.save_state()

        self.current_pair = None

        self.refresh()

    # --------------------------------------------------------
    # Refresh
    # --------------------------------------------------------

    def refresh(self):
        if not self.engine:
            return

        self.engine.tick_skips()

        self.progress["value"] = (
            self.engine.progress()
        )

        self.current_pair = (
            self.engine.choose_pair()
        )

        self.show_current_pair()

        metrics = (
            self.engine.confidence_metrics()
        )

        self.info_var.set(
            f"{len(self.engine.books)} active books"
            f"  •  {len(self.engine.comparisons)} decisions"
            f"  •  Top-10 "
            f"{metrics['top10_confidence'] * 100:.0f}%"
            f"  •  Top-25 "
            f"{metrics['top25_confidence'] * 100:.0f}%"
            f"  •  stability "
            f"{metrics['top25_stability'] * 100:.0f}%"
            f"  •  "
            f"{metrics['unresolved_boundary']} "
            f"boundary unresolved"
        )

    def show_current_pair(self):
        if not self.current_pair:
            self.left_button.configure(
                state="disabled"
            )

            self.tie_button.configure(
                state="disabled"
            )

            self.right_button.configure(
                state="disabled"
            )

            self.left_title.set(
                "🏆 Ranking ready"
            )

            self.right_title.set(
                "Continue whenever you want"
            )

            self.left_status.set(
                "ANALYSIS"
            )

            self.right_status.set(
                "HUMAN CHOICE REQUIRED"
            )

            self.set_description(
                self.left_description,
                (
                    "The adaptive engine does not invent "
                    "preferences. Export the ranking or "
                    "continue refining it."
                ),
            )

            self.set_description(
                self.right_description,
                (
                    "Use the lifecycle buttons if a book "
                    "becomes Read, Currently Reading or Ignore."
                ),
            )

            return

        self.left_button.configure(
            state="normal"
        )

        self.tie_button.configure(
            state="normal"
        )

        self.right_button.configure(
            state="normal"
        )

        left = self.engine.library[
            self.current_pair[0]
        ]

        right = self.engine.library[
            self.current_pair[1]
        ]

        self.display_book(
            left,
            "left",
        )

        self.display_book(
            right,
            "right",
        )

    def search_duckduckgo(
        self,
        title,
        author="",
        isbn="",
        doi="",
    ):
        """Open a DuckDuckGo search using title, author, and ISBN or DOI."""
        title = str(title or "").strip()
        author = str(author or "").strip()
        isbn = str(isbn or "").strip()
        doi = str(doi or "").strip()

        identifier = isbn or doi

        parts = [
            value
            for value in (
                title,
                author,
                identifier,
            )
            if value
        ]

        query = " ".join(parts)

        if query:
            webbrowser.open_new_tab(
                "https://duckduckgo.com/?q="
                + urllib.parse.quote_plus(query)
            )

    def search_current_book(self, side):
        """Search the book currently displayed on the selected card."""

        if not self.engine or not self.current_pair:
            return

        if side == "left":
            book_id = self.current_pair[0]
        else:
            book_id = self.current_pair[1]

        book = self.engine.library.get(book_id)

        if not book:
            return

        fields = (
            getattr(
                book,
                "goodreads_fields",
                {},
            )
            or {}
        )

        isbn = (
            fields.get("isbn")
            or book.isbn
        )

        doi = fields.get("doi", "")

        self.search_duckduckgo(
            book.title,
            book.author,
            isbn,
            doi,
        )
        
    def edit_current_book(self, side):
        if not self.engine or not self.current_pair:
            return

        book_id = (
            self.current_pair[0]
            if side == "left"
            else self.current_pair[1]
        )

        book = self.engine.library.get(book_id)

        if not book:
            return

        fields = dict(
            getattr(book, "goodreads_fields", {}) or {}
        )

        editor = tk.Toplevel(self.root)
        editor.title(f"Edit Book — {book.title}")
        editor.geometry("700x750")
        editor.configure(bg=self.colors["bg"])
        editor.transient(self.root)
        editor.grab_set()

        outer = tk.Frame(
            editor,
            bg=self.colors["bg"],
            padx=14,
            pady=14,
        )
        outer.pack(
            fill="both",
            expand=True,
        )

        canvas = tk.Canvas(
            outer,
            bg=self.colors["bg"],
            highlightthickness=0,
        )

        scrollbar = ttk.Scrollbar(
            outer,
            orient="vertical",
            command=canvas.yview,
        )

        form = tk.Frame(
            canvas,
            bg=self.colors["bg"],
        )

        canvas_window = canvas.create_window(
            (0, 0),
            window=form,
            anchor="nw",
        )

        def configure_form(_event=None):
            canvas.configure(
                scrollregion=canvas.bbox("all")
            )

        form.bind(
            "<Configure>",
            configure_form,
        )

        canvas.bind(
            "<Configure>",
            lambda event: canvas.itemconfigure(
                canvas_window,
                width=event.width,
            ),
        )

        canvas.configure(
            yscrollcommand=scrollbar.set
        )

        canvas.pack(
            side="left",
            fill="both",
            expand=True,
        )

        scrollbar.pack(
            side="right",
            fill="y",
        )

        fields_to_edit = [
            ("Pages", "pages"),
            ("Title", "title"),
            ("Author l-f", "author l-f"),
            ("cat", "cat"),
            ("Private Notes", "private notes"),
            ("Exclusive Shelf", "exclusive shelf"),
            ("Date Read", "date read"),
            ("My Rating", "my rating"),
            ("Read Count", "read count"),
            ("Year Published", "year published"),
            (
                "Original Publication Year",
                "original publication year",
            ),
            ("Publisher", "publisher"),
            ("ISBN", "isbn"),
            ("DOI", "doi"),
            ("Series", "series"),
            ("Series-#", "series-#"),
            ("Description", "description"),
        ]

        variables = {}
        widgets = {}

        for row_index, (label, key) in enumerate(fields_to_edit):

            value = fields.get(key, "")

            # Fall back to the Book model.
            if not value:
                fallback = {
                    "pages": book.pages,
                    "title": book.title,
                    "author l-f": book.author,
                    "exclusive shelf": book.shelf,
                    "my rating": book.my_rating,
                    "year published": book.year,
                    "publisher": book.publisher,
                    "isbn": book.isbn,
                    "book id - goodreads": book.goodreads_id,
                    "series": series_name(book),
                    "series-#": (
                        "" if series_number(book) is None
                        else str(series_number(book))
                    ),
                    "description": book.description,
                }

                value = fallback.get(key, "")

            tk.Label(
                form,
                text=label,
                bg=self.colors["bg"],
                fg=self.colors["text"],
                font=(
                    self.font,
                    9,
                    "bold",
                ),
                anchor="w",
            ).grid(
                row=row_index,
                column=0,
                sticky="nw",
                padx=(0, 12),
                pady=6,
            )

            if key == "description":
                widget = tk.Text(
                    form,
                    height=8,
                    wrap="word",
                    bg=self.colors["card"],
                    fg=self.colors["text"],
                    insertbackground=self.colors["text"],
                    relief="flat",
                    bd=1,
                    font=(
                        self.font,
                        9,
                    ),
                )

                widget.insert(
                    "1.0",
                    value,
                )

            else:
                variable = tk.StringVar(
                    value=value
                )

                variables[key] = variable

                widget = tk.Entry(
                    form,
                    textvariable=variable,
                    bg=self.colors["card"],
                    fg=self.colors["text"],
                    insertbackground=self.colors["text"],
                    relief="flat",
                    bd=1,
                    font=(
                        self.font,
                        9,
                    ),
                )

            # Goodreads ID is deliberately not editable.
            if key == "book id - goodreads":
                widget.configure(
                    state="disabled"
                )

            widget.grid(
                row=row_index,
                column=1,
                sticky="ew",
                pady=6,
            )

            widgets[key] = widget

        form.grid_columnconfigure(
            1,
            weight=1,
        )

        # --------------------------------------------------------
        # Buttons
        # --------------------------------------------------------

        buttons = tk.Frame(
            editor,
            bg=self.colors["bg"],
        )

        buttons.pack(
            fill="x",
            padx=14,
            pady=(0, 14),
        )

        def save_changes():
            # Update all Goodreads fields.
            for key, variable in variables.items():
                fields[key] = normalize(
                    variable.get()
                )

            description_widget = widgets.get(
                "description"
            )

            if description_widget:
                fields["description"] = normalize(
                    description_widget.get(
                        "1.0",
                        "end-1c",
                    )
                )

            # Update the full source dictionary.
            book.goodreads_fields = fields

            # Keep the Book model synchronized with the editable fields.
            book.title = fields.get(
                "title",
                book.title,
            )

            book.author = fields.get(
                "author l-f",
                book.author,
            )

            book.pages = fields.get(
                "pages",
                book.pages,
            )

            book.my_rating = fields.get(
                "my rating",
                book.my_rating,
            )

            book.shelf = fields.get(
                "exclusive shelf",
                book.shelf,
            )

            book.description = fields.get(
                "description",
                book.description,
            )

            book.isbn = clean_isbn(
                fields.get(
                    "isbn",
                    book.isbn,
                )
            )

            book.year = fields.get(
                "year published",
                book.year,
            )

            book.publisher = fields.get(
                "publisher",
                book.publisher,
            )

            book.goodreads_id = fields.get(
                "book id - goodreads",
                book.goodreads_id,
            )

            # Keep status/ranking lifecycle synchronized.
            shelf = normalize(
                fields.get(
                    "exclusive shelf",
                    book.status,
                )
            ).lower()

            if shelf in {
                "currently-reading",
                "currently reading",
            }:
                shelf = "currently-reading"
            elif shelf == "read":
                shelf = "read"
            elif shelf == "to-read":
                shelf = "to-read"

            if shelf in self.engine.ALL_STATUSES:
                book.shelf = shelf

                # Do not accidentally resurrect an ignored book
                # simply because its Goodreads shelf says to-read.
                if book.status != "ignore":
                    book.status = shelf

                self.engine.statuses[
                    book.id
                ] = book.status

            self.engine._sync_books()
            self.engine._replay()

            self.save_state()

            editor.destroy()

            # Redraw the current cards.
            self.display_book(
                book,
                side,
            )

            self.info_var.set(
                f"Saved changes to: {book.title}"
            )

        ttk.Button(
            buttons,
            text="Cancel",
            command=editor.destroy,
        ).pack(
            side="right",
            padx=4,
        )

        ttk.Button(
            buttons,
            text="Save Changes",
            style="Accent.TButton",
            command=save_changes,
        ).pack(
            side="right",
            padx=4,
        )

        editor.bind(
            "<Control-s>",
            lambda event: save_changes(),
        )

        # Put keyboard focus in the first field.
        if fields_to_edit:
            first_key = fields_to_edit[0][1]
            first_widget = widgets.get(first_key)

            if first_widget:
                first_widget.focus_set()

    def display_book(
        self,
        book,
        side,
    ):
        rating = self.engine.ratings[
            book.id
        ]

        statistics = (
            self.engine.statistics()
        )

        item = next(
            item
            for item in statistics
            if item["book"].id
            == book.id
        )

        status_labels = {
            "currently-reading":
                "📖 CURRENTLY READING",
            "to-read":
                "📚 TO-READ",
            "read":
                "✓ READ",
            "ignore":
                "🚫 IGNORED",
        }

        getattr(
            self,
            f"{side}_title",
        ).set(
            book.title
        )

        getattr(
            self,
            f"{side}_author",
        ).set(
            book.author
            or "Unknown author"
        )

        getattr(
            self,
            f"{side}_status",
        ).set(
            status_labels.get(
                book.status,
                book.status.upper(),
            )
        )

        meta = [
            f"Model #{item['rank']}",
            (
                f"Likely "
                f"{item['rank_low']}"
                f"–"
                f"{item['rank_high']}"
            ),
            (
                f"Top-25 "
                f"{item['top25_probability'] * 100:.0f}%"
            ),
            (
                f"Rating "
                f"{rating.rating:.0f}"
                f" ±{rating.rd:.0f}"
            ),
            (
                f"{rating.comparisons}"
                f" decisions"
            ),
        ]

        if book.pages:
            meta.insert(
                0,
                f"{book.pages} pages",
            )

        getattr(
            self,
            f"{side}_meta",
        ).set(
            "  •  ".join(meta)
        )

        self.set_card_details(
            getattr(
                self,
                f"{side}_details",
            ),
            book,
        )

    def format_book_details(
        self,
        book,
    ):
        """Return every useful Goodreads source column for a book."""

        fields = dict(
            getattr(
                book,
                "goodreads_fields",
                {},
            )
            or {}
        )

        # These are the columns the user specifically wants to see,
        # in the same logical order as the Goodreads export.
        ordered = [
            ("Pages", "pages"),
            ("Title", "title"),
            ("Author l-f", "author l-f"),
            ("cat", "cat"),
            ("Private Notes", "private notes"),
            ("Exclusive Shelf", "exclusive shelf"),
            ("Date Read", "date read"),
            ("My Rating", "my rating"),
            ("Read Count", "read count"),
            ("Year Published", "year published"),
            (
                "Original Publication Year",
                "original publication year",
            ),
            ("Publisher", "publisher"),
            ("ISBN", "isbn"),
            ("DOI", "doi"),
            ("Book Id - Goodreads", "book id - goodreads"),
            ("Description", "description"),
        ]

        lines = []
        used = set()

        for label, key in ordered:
            value = fields.get(key, "")

            # Fall back to the Book model for older state files.
            if not value:
                fallback = {
                    "number of pages": book.pages,
                    "title": book.title,
                    "author l-f": book.author,
                    "exclusive shelf": book.shelf,
                    "my rating": book.my_rating,
                    "year published": book.year,
                    "publisher": book.publisher,
                    "isbn": book.isbn,
                    "book id - goodreads": book.goodreads_id,
                    "series": series_name(book),
                    "series-#": (
                        "" if series_number(book) is None
                        else str(series_number(book))
                    ),
                    "description": book.description,
                }
                value = fallback.get(key, "")

            if key == "description" and value:
                # Keep the full description available via the scrollable
                # card rather than silently cutting it off.
                pass

            lines.append(
                f"{label}: {value or '—'}"
            )
            used.add(key)

        # Also show any additional columns present in a particular
        # Goodreads export, so the card genuinely exposes all source
        # columns rather than only a fixed subset.
        extras = [
            (key, value)
            for key, value in fields.items()
            if key not in used
        ]

        if extras:
            lines.append("")
            lines.append("Other Goodreads Columns:")
            for key, value in extras:
                label = key.title()
                lines.append(
                    f"{label}: {value or '—'}"
                )

        return "\n".join(lines)

    def set_card_details(
        self,
        widget,
        book,
    ):
        widget.configure(
            state="normal"
        )

        widget.delete(
            "1.0",
            "end",
        )

        widget.insert(
            "1.0",
            self.format_book_details(book),
        )

        widget.configure(
            state="disabled"
        )
        widget.yview_moveto(0.0)

    # Backward-compatible helper retained for older callers.
    def set_description(
        self,
        widget,
        text,
    ):
        widget.configure(
            state="normal"
        )
        widget.delete(
            "1.0",
            "end",
        )
        widget.insert(
            "1.0",
            text,
        )
        widget.configure(
            state="disabled"
        )

    # --------------------------------------------------------
    # Human actions
    # --------------------------------------------------------

    def choose(
        self,
        result,
    ):
        if (
            not self.engine
            or not self.current_pair
        ):
            return

        left, right = (
            self.current_pair
        )

        try:
            self.engine.apply_match(
                left,
                right,
                result,
            )

            self.save_state()

            self.current_pair = None

            self.refresh()

        except Exception as exc:
            messagebox.showerror(
                "Could not record choice",
                str(exc),
            )

    def skip(self):
        if (
            not self.engine
            or not self.current_pair
        ):
            return

        self.engine.skip_pair(
            *self.current_pair
        )

        self.current_pair = None

        self.refresh()

    # --------------------------------------------------------
    # Lifecycle
    # --------------------------------------------------------

    def set_lifecycle(
        self,
        side,
        status,
    ):
        if (
            not self.engine
            or not self.current_pair
        ):
            return

        book_id = (
            self.current_pair[0]
            if side == "left"
            else self.current_pair[1]
        )

        book = self.engine.library[book_id]

        old_status = book.status
        old_shelf = book.shelf
        old_exclusive_shelf = (
            book.goodreads_fields.get(
                "exclusive shelf",
                "",
            )
            if isinstance(book.goodreads_fields, dict)
            else ""
        )

        try:
            # Update application state.
            self.engine.set_lifecycle(
                book_id,
                status,
            )

            # Immediately update the original Goodreads workbook.
            self.save_source_workbook()

            # Also persist the ranker's internal JSON state.
            self.save_state()

        except Exception as exc:
            # Roll back the in-memory change if the source workbook
            # could not be updated.
            book.status = old_status
            book.shelf = old_shelf

            if not isinstance(book.goodreads_fields, dict):
                book.goodreads_fields = {}

            book.goodreads_fields[
                "exclusive shelf"
            ] = old_exclusive_shelf

            self.engine.statuses[
                book_id
            ] = old_status

            self.engine._sync_books()
            self.engine._replay()

            messagebox.showerror(
                "Could not update Goodreads shelf",
                str(exc),
            )
            return

        self.current_pair = None

        self.refresh()

        self.info_var.set(
            f"{book.title}  →  "
            f"{status.replace('-', ' ').title()}"
        )

    def lifecycle_menu(
        self,
        event,
        side,
    ):
        if (
            not self.engine
            or not self.current_pair
        ):
            return

        menu = tk.Menu(
            self.root,
            tearoff=False,
            bg=self.colors["panel"],
            fg=self.colors["text"],
            activebackground=self.colors["purple"],
            activeforeground="#FFFFFF",
        )

        menu.add_command(
            label="📖 Currently Reading",
            command=lambda:
                self.set_lifecycle(
                    side,
                    "currently-reading",
                ),
        )

        menu.add_command(
            label="✓ Read",
            command=lambda:
                self.set_lifecycle(
                    side,
                    "read",
                ),
        )

        menu.add_command(
            label="🚫 Ignore",
            command=lambda:
                self.set_lifecycle(
                    side,
                    "ignore",
                ),
        )

        menu.add_command(
            label="↺ To-read",
            command=lambda:
                self.set_lifecycle(
                    side,
                    "to-read",
                ),
        )

        try:
            menu.tk_popup(
                event.x_root,
                event.y_root,
            )
        finally:
            menu.grab_release()

    # --------------------------------------------------------
    # Undo
    # --------------------------------------------------------

    def undo(self):
        if not self.engine:
            return

        if self.engine.undo():
            self.save_state()

            self.current_pair = None

            self.refresh()

    def save_source_workbook(self):
        """Persist current shelf/status changes to the original Goodreads file."""

        if not self.engine or not self.source_file:
            return

        source = Path(self.source_file)

        keep_vba = (
            source.suffix.lower() == ".xlsm"
        )

        workbook = load_workbook(
            source,
            keep_vba=keep_vba,
        )

        temporary = source.with_name(
            source.stem + ".lifecycle.tmp" + source.suffix
        )

        try:
            source_sheet = workbook.active

            update_goodreads_source_sheet(
                source_sheet,
                self.engine,
            )

            workbook.save(temporary)

        finally:
            workbook.close()

        os.replace(
            temporary,
            source,
        )
        
    def save_state(self):
        if (
            self.engine
            and self.state_store
        ):
            try:
                self.state_store.save(
                    self.engine
                )
            except Exception as exc:
                print(
                    "Warning: could not save state:",
                    exc,
                )

    # --------------------------------------------------------
    # Ranking window
    # --------------------------------------------------------

    def show_ranking(self):
        if not self.engine:
            messagebox.showinfo(
                "No ranking",
                "Open a Goodreads workbook first.",
            )
            return

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Ranking • uncertainty aware"
        )

        window.geometry(
            "1450x760"
        )

        window.configure(
            bg=self.colors["bg"]
        )

        metrics = (
            self.engine.confidence_metrics()
        )

        tk.Label(
            window,
            text=(
                "🏆  CURRENT READING + TOP-25"
            ),
            bg=self.colors["bg"],
            fg=self.colors["text"],
            font=(
                self.font,
                18,
                "bold",
            ),
        ).pack(
            anchor="w",
            padx=14,
            pady=(12, 2),
        )

        tk.Label(
            window,
            text=(
                f"Top-10 confidence "
                f"{metrics['top10_confidence'] * 100:.0f}%"
                f"  •  "
                f"Top-25 confidence "
                f"{metrics['top25_confidence'] * 100:.0f}%"
                f"  •  "
                f"stability "
                f"{metrics['top25_stability'] * 100:.0f}%"
                f"  •  "
                f"{metrics['unresolved_boundary']}"
                f" unresolved boundary"
            ),
            bg=self.colors["bg"],
            fg=self.colors["muted"],
            font=(
                self.font,
                9,
            ),
        ).pack(
            anchor="w",
            padx=14,
            pady=(0, 8),
        )

        columns = (
            "priority",
            "rank",
            "title",
            "author",
            "status",
            "rating",
            "rd",
            "interval",
            "top10",
            "top25",
            "decisions",
        )

        tree = ttk.Treeview(
            window,
            columns=columns,
            show="headings",
        )

        headings = {
            "priority": "Priority",
            "rank": "Rank",
            "title": "Title",
            "author": "Author",
            "status": "Status",
            "rating": "Rating",
            "rd": "RD",
            "interval": "Likely rank",
            "top10": "Top 10 %",
            "top25": "Top 25 %",
            "decisions": "Decisions",
        }

        widths = {
            "priority": 110,
            "rank": 55,
            "title": 340,
            "author": 200,
            "status": 170,
            "rating": 75,
            "rd": 65,
            "interval": 90,
            "top10": 85,
            "top25": 85,
            "decisions": 85,
        }

        for column in columns:
            tree.heading(
                column,
                text=headings[column],
            )

            tree.column(
                column,
                width=widths[column],
                anchor=(
                    "w"
                    if column
                    in {
                        "title",
                        "author",
                        "status",
                    }
                    else "center"
                ),
            )

        scrollbar = ttk.Scrollbar(
            window,
            orient="vertical",
            command=tree.yview,
        )

        tree.configure(
            yscrollcommand=scrollbar.set
        )

        tree.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(14, 0),
            pady=(0, 14),
        )

        scrollbar.pack(
            side="right",
            fill="y",
            padx=(0, 14),
            pady=(0, 14),
        )

        statistics = (
            self.engine.statistics()
        )

        current = [
            item
            for item in statistics
            if item["book"].status
            == "currently-reading"
        ]

        rest = [
            item
            for item in statistics
            if item["book"].status
            != "currently-reading"
        ]

        for item in current + rest:
            book = item["book"]
            rating = item["rating"]

            if (
                book.status
                == "currently-reading"
            ):
                priority = "📖 NOW"

            elif item["rank"] <= TOP_K:
                priority = "🏆 TOP 25"

            else:
                priority = "📚 QUEUE"

            tree.insert(
                "",
                "end",
                values=(
                    priority,
                    item["rank"],
                    book.title,
                    book.author,
                    book.status,
                    f"{rating.rating:.0f}",
                    f"±{rating.rd:.0f}",
                    (
                        f"{item['rank_low']}"
                        f"–"
                        f"{item['rank_high']}"
                    ),
                    (
                        f"{item['top10_probability'] * 100:.0f}%"
                    ),
                    (
                        f"{item['top25_probability'] * 100:.0f}%"
                    ),
                    rating.comparisons,
                ),
            )

    # --------------------------------------------------------
    # Shortcut dialog
    # --------------------------------------------------------

    def show_shortcuts(self):
        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Keyboard shortcuts"
        )

        window.geometry(
            "760x650"
        )

        window.configure(
            bg=self.colors["bg"]
        )

        tk.Label(
            window,
            text="⌨  FULL KEYBOARD REFERENCE",
            bg=self.colors["bg"],
            fg=self.colors["text"],
            font=(
                self.font,
                18,
                "bold",
            ),
        ).pack(
            anchor="w",
            padx=20,
            pady=16,
        )

        shortcuts = [
            ("1", "Choose LEFT"),
            ("2", "Choose RIGHT"),
            ("3 / T", "Tie"),
            ("4 / S", "Skip"),
            ("← / →", "Choose left / right"),
            ("U", "Undo"),
            ("I", "Mark LEFT Ignore"),
            ("R", "Mark LEFT Read"),
            ("C", "Mark LEFT Currently Reading"),
            ("Shift+I", "Mark RIGHT Ignore"),
            ("Shift+R", "Mark RIGHT Read"),
            ("Shift+C", "Mark RIGHT Currently Reading"),
            ("Ctrl+O", "Open Goodreads Excel"),
            ("Ctrl+E", "Export Ranking"),
            ("Click book card", "Open lifecycle menu"),
        ]

        for key, description in shortcuts:
            row = tk.Frame(
                window,
                bg=self.colors["panel"],
            )

            row.pack(
                fill="x",
                padx=20,
                pady=2,
            )

            tk.Label(
                row,
                text=key,
                bg=self.colors["panel"],
                fg=self.colors["accent"],
                font=(
                    "Consolas",
                    10,
                    "bold",
                ),
                width=20,
                anchor="w",
                padx=10,
                pady=6,
            ).pack(
                side="left"
            )

            tk.Label(
                row,
                text=description,
                bg=self.colors["panel"],
                fg=self.colors["text"],
                font=(
                    self.font,
                    9,
                ),
                anchor="w",
                padx=10,
                pady=6,
            ).pack(
                side="left",
                fill="x",
                expand=True,
            )

    # --------------------------------------------------------
    # Export
    # --------------------------------------------------------

    def export(self):
        if (
            not self.engine
            or not self.source_file
        ):
            messagebox.showinfo(
                "Nothing to export",
                "Open a Goodreads workbook first.",
            )
            return

        try:
            output = export_results(
                self.source_file,
                self.engine,
            )

            self.save_state()

            if messagebox.askyesno(
                "Export complete",
                (
                    f"Saved:\n\n"
                    f"{output}\n\n"
                    f"Open it now?"
                ),
            ):
                self.open_external(
                    output
                )

        except PermissionError:
            messagebox.showerror(
                "Export failed",
                (
                    "Windows could not save the workbook. "
                    "Close it in Excel and try again."
                ),
            )

        except Exception as exc:
            messagebox.showerror(
                "Export failed",
                str(exc),
            )

    def overwrite_source(self):
        if (
            not self.engine
            or not self.source_file
        ):
            return

        if not messagebox.askyesno(
            "Overwrite Goodreads source?",
            (
                "This will replace the original "
                "Goodreads workbook.\n\n"
                "A timestamped backup will be created first.\n\n"
                "Continue?"
            ),
        ):
            return

        timestamp = (
            datetime.now()
            .strftime(
                "%Y%m%d-%H%M%S"
            )
        )

        backup = (
            self.source_file.with_name(
                self.source_file.stem
                + ".backup-"
                + timestamp
                + self.source_file.suffix
            )
        )

        try:
            shutil.copy2(
                self.source_file,
                backup,
            )

            export_results(
                self.source_file,
                self.engine,
                output=self.source_file,
            )

            self.save_state()

            messagebox.showinfo(
                "Source updated",
                (
                    "The Goodreads workbook was updated.\n\n"
                    f"Backup:\n{backup}"
                ),
            )

        except PermissionError:
            messagebox.showerror(
                "Overwrite failed",
                (
                    "Close the workbook in Excel "
                    "and try again."
                ),
            )

        except Exception as exc:
            messagebox.showerror(
                "Overwrite failed",
                str(exc),
            )

    # --------------------------------------------------------
    # External file
    # --------------------------------------------------------

    def open_external(
        self,
        path,
    ):
        try:
            if sys.platform.startswith(
                "win"
            ):
                os.startfile(
                    str(path)
                )

            elif sys.platform == "darwin":
                subprocess.Popen(
                    [
                        "open",
                        str(path),
                    ]
                )

            else:
                subprocess.Popen(
                    [
                        "xdg-open",
                        str(path),
                    ]
                )

        except Exception:
            pass

    # --------------------------------------------------------
    # Close
    # --------------------------------------------------------

    def close(self):
        self.save_state()
        self.root.destroy()


# ============================================================
# SELF TEST
# ============================================================

def run_self_test():
    started = datetime.now()

    # --------------------------------------------------------
    # Glicko direction
    # --------------------------------------------------------

    winner = Rating()
    loser = Rating()

    winner_after = glicko_update(
        winner,
        loser,
        1.0,
    )

    loser_after = glicko_update(
        loser,
        winner,
        0.0,
    )

    assert (
        winner_after.rating
        > winner.rating
    )

    assert (
        loser_after.rating
        < loser.rating
    )

    # --------------------------------------------------------
    # Goodreads normalized import simulation
    # --------------------------------------------------------

    fake_row = {
        "book id - goodreads": "123456",
        "isbn": "9781234567890",
        "title": "Test Book",
        "author l-f": "Test Author",
        "description": "A test book.",
        "number of pages": "300",
        "my rating": "0",
        "exclusive shelf": "to-read",
        "year published": "2025",
        "publisher": "Test Publisher",
    }

    fake_book = book_from_row(
        fake_row
    )

    assert (
        fake_book.id
        == "gr:123456"
    )

    assert (
        fake_book.title
        == "Test Book"
    )

    assert (
        fake_book.author
        == "Test Author"
    )

    # --------------------------------------------------------
    # Small library
    # --------------------------------------------------------

    books = [
        Book(
            id=str(index),
            title=f"Book {index}",
            author=f"Author {index % 5}",
        )
        for index in range(30)
    ]

    engine = RankingEngine(
        books,
        mode="TOP_25_FOCUS",
        seed=12345,
    )

    seen = set()

    for index in range(25):
        pair = engine.choose_pair()

        assert pair is not None

        pair_key = engine.pair_key(
            *pair
        )

        assert pair_key not in seen

        seen.add(pair_key)

        result = (
            "left"
            if index % 3
            else "right"
        )

        engine.apply_match(
            pair[0],
            pair[1],
            result,
        )

    assert (
        len(engine.comparisons)
        == 25
    )

    # --------------------------------------------------------
    # Undo
    # --------------------------------------------------------

    before = len(
        engine.comparisons
    )

    assert engine.undo()

    assert (
        len(engine.comparisons)
        == before - 1
    )

    # --------------------------------------------------------
    # Shelf-segmented comparison
    # --------------------------------------------------------
    segmented_books = [
        Book(
            id="cr-1",
            title="Currently Reading 1",
            shelf="currently-reading",
            status="currently-reading",
        ),
        Book(
            id="cr-2",
            title="Currently Reading 2",
            shelf="currently-reading",
            status="currently-reading",
        ),
        Book(
            id="tr-1",
            title="To Read 1",
            shelf="to-read",
            status="to-read",
        ),
        Book(
            id="tr-2",
            title="To Read 2",
            shelf="to-read",
            status="to-read",
        ),
    ]

    segmented_engine = RankingEngine(
        segmented_books,
        seed=99,
    )

    # When there are at least two currently-reading books, only that
    # cohort is eligible for comparison.
    segmented_pair = segmented_engine.choose_pair()
    assert segmented_pair is not None
    assert segmented_engine._same_pairing_cohort(
        *segmented_pair
    )
    assert {
        segmented_engine.library[book_id].status
        for book_id in segmented_pair
    } == {"currently-reading"}

    segmented_engine.apply_match(
        segmented_pair[0],
        segmented_pair[1],
        "left",
    )

    # The only currently-reading pair has now been played, so the selector
    # must not cross over to to-read merely to find another pair.
    assert segmented_engine.choose_pair() is None

    # A separate to-read cohort still works normally.
    to_read_engine = RankingEngine(
        [
            segmented_books[2],
            segmented_books[3],
        ],
        seed=100,
    )
    to_read_pair = to_read_engine.choose_pair()
    assert to_read_pair is not None
    assert {
        to_read_engine.library[book_id].status
        for book_id in to_read_pair
    } == {"to-read"}

    # A single currently-reading book is never paired with a to-read book.
    single_current_engine = RankingEngine(
        [
            segmented_books[0],
            segmented_books[2],
            segmented_books[3],
        ],
        seed=101,
    )
    single_current_pair = single_current_engine.choose_pair()
    assert single_current_pair is not None
    assert {
        single_current_engine.library[book_id].status
        for book_id in single_current_pair
    } == {"to-read"}

    # Direct application of a cross-shelf comparison is rejected too.
    try:
        single_current_engine.apply_match(
            "cr-1",
            "tr-1",
            "left",
        )
    except ValueError:
        pass
    else:
        raise AssertionError(
            "Cross-shelf comparisons must be rejected."
        )

    # --------------------------------------------------------
    # Lifecycle
    # --------------------------------------------------------

    engine.set_lifecycle(
        "0",
        "currently-reading",
    )

    assert any(
        book.id == "0"
        for book in engine.books
    )

    engine.set_lifecycle(
        "1",
        "ignore",
    )

    assert all(
        book.id != "1"
        for book in engine.books
    )

    engine.set_lifecycle(
        "1",
        "to-read",
    )

    assert any(
        book.id == "1"
        for book in engine.books
    )

    # --------------------------------------------------------
    # Large library bounded selection
    # --------------------------------------------------------

    large_books = [
        Book(
            id=f"large-{index}",
            title=f"Large Book {index}",
        )
        for index in range(5000)
    ]

    large_engine = RankingEngine(
        large_books,
        seed=7,
        mode="TOP_25_FOCUS",
    )

    pair = large_engine.choose_pair()

    assert pair is not None

    # --------------------------------------------------------
    # State persistence
    # --------------------------------------------------------

    with tempfile.TemporaryDirectory() as directory:
        source = (
            Path(directory)
            / "library.xlsx"
        )

        temporary_workbook = Workbook()

        worksheet = (
            temporary_workbook.active
        )

        worksheet["A1"] = "Title"
        worksheet["A2"] = "Test Book"

        temporary_workbook.save(
            source
        )

        temporary_workbook.close()

        store = StateStore(
            source
        )

        store.save(
            engine
        )

        loaded_state = (
            store.load_raw()
        )

        assert (
            loaded_state
            is not None
        )

        assert (
            loaded_state[
                "state_version"
            ]
            == STATE_VERSION
        )

        assert (
            len(
                loaded_state[
                    "comparisons"
                ]
            )
            == len(
                engine.comparisons
            )
        )

    # --------------------------------------------------------
    # Series ordering test
    # --------------------------------------------------------

    series_books = [
        Book(
            id="s3",
            title="Series 3",
            shelf="to-read",
            status="to-read",
            goodreads_fields={"series": "Saga", "series-#": "3"},
        ),
        Book(
            id="s1",
            title="Series 1",
            shelf="to-read",
            status="to-read",
            goodreads_fields={"series": "Saga", "series-#": "1"},
        ),
        Book(
            id="s2",
            title="Series 2",
            shelf="to-read",
            status="to-read",
            goodreads_fields={"series": "Saga", "series-#": "2"},
        ),
    ]

    series_engine = RankingEngine(
        series_books,
        seed=7,
    )
    series_pair = series_engine.choose_pair()
    assert series_pair == ("s1", "s2"), series_pair

    series_stats = series_engine.statistics()
    saga_order = [
        item["book"].id
        for item in series_stats
        if series_name(item["book"]).casefold() == "saga"
    ]
    assert saga_order == ["s1", "s2", "s3"], saga_order

    # --------------------------------------------------------
    # Actual import pipeline test
    # --------------------------------------------------------

    with tempfile.TemporaryDirectory() as directory:
        source = (
            Path(directory)
            / "goodreads.xlsx"
        )

        workbook = Workbook()

        worksheet = (
            workbook.active
        )

        worksheet.append(
            [
                "Book Id - Goodreads",
                "Series",
                "Series-#",
                "Title",
                "Author l-f",
                "Exclusive Shelf",
                "ISBN",
                "Number of Pages",
                "Description",
            ]
        )

        worksheet.append(
            [
                "1001",
                "Example Series",
                "1",
                "Alpha",
                "Author, A",
                "to-read",
                "9781111111111",
                "200",
                "Alpha description",
            ]
        )

        worksheet.append(
            [
                "1002",
                "Example Series",
                "2",
                "Beta",
                "Author, B",
                "currently-reading",
                "9782222222222",
                "250",
                "Beta description",
            ]
        )

        workbook.save(
            source
        )

        workbook.close()

        imported_rows = import_goodreads(
            source
        )

        assert isinstance(
            imported_rows,
            list,
        )

        assert len(
            imported_rows
        ) == 2

        assert all(
            isinstance(row, dict)
            for row in imported_rows
        )

        imported_books = [
            book_from_row(row)
            for row in imported_rows
        ]

        assert (
            len(imported_books)
            == 2
        )

        assert (
            imported_books[0].id
            == "gr:1001"
        )

        assert (
            imported_books[1].id
            == "gr:1002"
        )

        assert (
            imported_books[0].goodreads_fields[
                "description"
            ]
            == "Alpha description"
        )

        imported_engine = RankingEngine(
            imported_books,
            seed=42,
        )

        result = (
            imported_engine.sync_goodreads(
                imported_rows
            )
        )

        assert (
            result["active"]
            == 2
        )

    _test_series_import_creates_real_comparisons()

    elapsed = (
        datetime.now()
        - started
    ).total_seconds()

    print(
        f"SELF-TEST PASSED in {elapsed:.2f}s"
    )

    return 0


# ============================================================
# MAIN
# ============================================================

def main():
    if "--self-test" in sys.argv:
        return run_self_test()

    root = tk.Tk()

    RankerApp(
        root
    )

    root.mainloop()

    return 0


def _test_series_import_creates_real_comparisons():
    books = [
        Book(id="s3", title="Series 3", shelf="to-read", status="to-read",
             goodreads_fields={"series": "Saga", "series-#": "3"}),
        Book(id="s1", title="Series 1", shelf="to-read", status="to-read",
             goodreads_fields={"series": "Saga", "series-#": "1"}),
        Book(id="s2", title="Series 2", shelf="to-read", status="to-read",
             goodreads_fields={"series": "Saga", "series-#": "2"}),
    ]
    engine = RankingEngine(books, seed=7)
    created = engine._apply_series_order_comparisons()
    assert created == 2
    pairs = {
        engine.pair_key(item["left"], item["right"]): item
        for item in engine.comparisons
    }
    assert engine.pair_key("s1", "s2") in pairs
    assert engine.pair_key("s2", "s3") in pairs
    assert pairs[engine.pair_key("s1", "s2")]["result"] == "left"
    assert pairs[engine.pair_key("s2", "s3")]["result"] == "left"
    assert all(
        item.get("source") == "series-order"
        for item in engine.comparisons
    )
    assert engine._apply_series_order_comparisons() == 0

    # Different shelves are never automatically mixed.
    books[2].status = "currently-reading"
    books[2].shelf = "currently-reading"
    engine = RankingEngine(books, seed=7)
    assert engine._apply_series_order_comparisons() == 1

if __name__ == "__main__":
    try:
        raise SystemExit(
            main()
        )

    except Exception:
        error = traceback.format_exc()

        print(error)

        try:
            messagebox.showerror(
                "Unexpected error",
                (
                    "The application encountered "
                    "an unexpected error:\n\n"
                    + error
                ),
            )

        except Exception:
            pass