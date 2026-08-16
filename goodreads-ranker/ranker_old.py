from __future__ import annotations

import copy
import json
import math
import os
import random
import re
import subprocess
import sys
import traceback
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

APP_NAME = "Goodreads To-Read Ranker"
APP_VERSION = "1.1"

DEFAULT_RATING = 1500.0
DEFAULT_RD = 350.0
DEFAULT_VOLATILITY = 0.06

# Glicko-2 constants
GLICKO_SCALE = 173.7178
TAU = 0.5

DEFAULT_TARGET_COMPARISONS = 12
MIN_TARGET_COMPARISONS = 5
MAX_TARGET_COMPARISONS = 30

STATE_DIRECTORY_NAME = ".ranker_state"


# ============================================================
# GENERAL UTILITIES
# ============================================================

def normalize(value) -> str:
    """Convert an Excel value to a clean string."""
    if value is None:
        return ""

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))

    return str(value).strip()


def normalize_header(value) -> str:
    """Normalize an Excel column header for comparison."""
    return (
        normalize(value)
        .replace("\ufeff", "")
        .strip()
        .lower()
    )


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def safe_filename(value: str) -> str:
    value = re.sub(
        r'[<>:"/\\|?*]',
        "_",
        value,
    )

    value = re.sub(
        r"\s+",
        " ",
        value,
    ).strip()

    return value or "ranking"


def truncate(text: str, length: int) -> str:
    text = normalize(text)

    if len(text) <= length:
        return text

    return text[: length - 1].rstrip() + "…"


# ============================================================
# GLICKO-2
# ============================================================

@dataclass
class Rating:
    rating: float = DEFAULT_RATING
    rd: float = DEFAULT_RD
    volatility: float = DEFAULT_VOLATILITY

    comparisons: int = 0
    wins: int = 0
    losses: int = 0
    ties: int = 0


def to_glicko(rating: Rating):
    mu = (rating.rating - 1500.0) / GLICKO_SCALE
    phi = rating.rd / GLICKO_SCALE

    return mu, phi


def from_glicko(
    mu: float,
    phi: float,
    volatility: float,
) -> Rating:
    return Rating(
        rating=1500.0 + GLICKO_SCALE * mu,
        rd=max(1.0, GLICKO_SCALE * phi),
        volatility=volatility,
    )


def g_function(phi: float) -> float:
    return 1.0 / math.sqrt(
        1.0
        + 3.0 * phi * phi / (math.pi * math.pi)
    )


def expected_score(
    mu: float,
    opponent_mu: float,
    opponent_phi: float,
) -> float:
    exponent = (
        -g_function(opponent_phi)
        * (mu - opponent_mu)
    )

    if exponent > 700:
        return 0.0

    if exponent < -700:
        return 1.0

    return 1.0 / (1.0 + math.exp(exponent))


def volatility_function(
    x: float,
    delta: float,
    phi: float,
    variance: float,
    a: float,
) -> float:
    exp_x = math.exp(x)

    denominator = (
        2.0
        * (
            phi * phi
            + variance
            + exp_x
        )
        * (
            phi * phi
            + variance
            + exp_x
        )
    )

    numerator = (
        exp_x
        * (
            delta * delta
            - phi * phi
            - variance
            - exp_x
        )
    )

    return (
        numerator / denominator
        - (x - a) / (TAU * TAU)
    )


def calculate_volatility(
    phi: float,
    volatility: float,
    delta: float,
    variance: float,
) -> float:
    """
    Glicko-2 volatility iteration.
    """

    a = math.log(
        volatility * volatility
    )

    A = a

    if delta * delta > (
        phi * phi + variance
    ):
        B = math.log(
            delta * delta
            - phi * phi
            - variance
        )
    else:
        k = 1

        while True:
            B = a - k * TAU

            value = volatility_function(
                B,
                delta,
                phi,
                variance,
                a,
            )

            if value >= 0:
                break

            k += 1

            if k > 100:
                break

    f_a = volatility_function(
        A,
        delta,
        phi,
        variance,
        a,
    )

    f_b = volatility_function(
        B,
        delta,
        phi,
        variance,
        a,
    )

    for _ in range(100):
        if abs(B - A) < 1e-8:
            break

        denominator = f_b - f_a

        if abs(denominator) < 1e-15:
            break

        C = (
            A
            + (A - B)
            * f_a
            / denominator
        )

        f_c = volatility_function(
            C,
            delta,
            phi,
            variance,
            a,
        )

        if f_c * f_b < 0:
            A = B
            f_a = f_b
        else:
            f_a /= 2.0

        B = C
        f_b = f_c

    return math.exp(A / 2.0)


def glicko_update(
    player: Rating,
    opponents: list[Rating],
    scores: list[float],
) -> Rating:
    """
    Perform a Glicko-2 update.

    score:
        1.0 = win
        0.5 = tie
        0.0 = loss
    """

    if not opponents:
        result = copy.deepcopy(player)

        phi = (
            player.rd
            / GLICKO_SCALE
        )

        phi_star = math.sqrt(
            phi * phi
            + player.volatility
            * player.volatility
        )

        result.rd = min(
            GLICKO_SCALE * phi_star,
            350.0,
        )

        return result

    mu, phi = to_glicko(player)

    variance_inverse = 0.0
    score_sum = 0.0

    for opponent, score in zip(
        opponents,
        scores,
    ):
        opponent_mu, opponent_phi = (
            to_glicko(opponent)
        )

        g_value = g_function(
            opponent_phi
        )

        expected = expected_score(
            mu,
            opponent_mu,
            opponent_phi,
        )

        variance_inverse += (
            g_value
            * g_value
            * expected
            * (1.0 - expected)
        )

        score_sum += (
            g_value
            * (score - expected)
        )

    if variance_inverse <= 0:
        return copy.deepcopy(player)

    variance = (
        1.0 / variance_inverse
    )

    delta = (
        variance * score_sum
    )

    new_volatility = calculate_volatility(
        phi,
        player.volatility,
        delta,
        variance,
    )

    phi_star = math.sqrt(
        phi * phi
        + new_volatility
        * new_volatility
    )

    new_phi = 1.0 / math.sqrt(
        1.0 / (
            phi_star * phi_star
        )
        + 1.0 / variance
    )

    new_mu = (
        mu
        + new_phi
        * new_phi
        * score_sum
    )

    result = from_glicko(
        new_mu,
        new_phi,
        new_volatility,
    )

    return result


# ============================================================
# BOOK MODEL
# ============================================================

@dataclass
class Book:
    id: str
    title: str
    author: str

    pages: str = ""
    year: str = ""
    publication_year: str = ""
    description: str = ""
    my_rating: str = ""
    isbn: str = ""
    book_id: str = ""

    original_row: int = 0

    original: dict = None

    def __post_init__(self):
        if self.original is None:
            self.original = {}


# ============================================================
# GOODREADS IMPORTER
# ============================================================

def find_header(
    headers: list,
    wanted: str,
):
    """
    Find a header by normalized text.
    """

    wanted_normalized = normalize_header(
        wanted
    )

    for header in headers:
        if (
            normalize_header(header)
            == wanted_normalized
        ):
            return header

    return None


def find_first_header(
    headers: list,
    candidates: list[str],
):
    """
    Return the first matching header from
    a list of possible Goodreads column names.
    """

    for candidate in candidates:
        result = find_header(
            headers,
            candidate,
        )

        if result is not None:
            return result

    return None


def load_goodreads(
    path: Path,
):
    """
    Load a Goodreads XLSX export.

    IMPORTANT:
    Only rows where:

        Exclusive Shelf == to-read

    are imported.
    """

    if not path.exists():
        raise FileNotFoundError(
            f"File not found:\n{path}"
        )

    if path.suffix.lower() != ".xlsx":
        raise ValueError(
            "Please select a Goodreads .xlsx export."
        )

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook.active

        rows = worksheet.iter_rows(
            values_only=True
        )

        try:
            header_values = next(rows)
        except StopIteration:
            raise ValueError(
                "The Excel file is empty."
            )

        headers = list(
            header_values
        )

        # ----------------------------------------------------
        # Required columns
        # ----------------------------------------------------

        shelf_header = find_header(
            headers,
            "Exclusive Shelf",
        )

        title_header = find_header(
            headers,
            "Title",
        )

        if shelf_header is None:
            raise ValueError(
                'Could not find the "Exclusive Shelf" '
                "column in this Goodreads export."
            )

        if title_header is None:
            raise ValueError(
                'Could not find the "Title" '
                "column in this Goodreads export."
            )

        # ----------------------------------------------------
        # AUTHOR FIX
        #
        # Your Goodreads export uses:
        #
        #     Author l-f
        #
        # rather than:
        #
        #     Author
        #
        # We explicitly support both.
        # ----------------------------------------------------

        author_header = find_first_header(
            headers,
            [
                "Author l-f",
                "Author",
                "Authors",
                "Author Name",
            ],
        )

        # ----------------------------------------------------
        # Build column index
        # ----------------------------------------------------

        header_index = {}

        for index, header in enumerate(
            headers
        ):
            if header is not None:
                header_index[header] = index

        shelf_index = header_index[
            shelf_header
        ]

        title_index = header_index[
            title_header
        ]

        author_index = None

        if author_header is not None:
            author_index = header_index[
                author_header
            ]

        # Optional columns.
        pages_header = find_header(
            headers,
            "Pages",
        )

        year_header = find_header(
            headers,
            "Year Published",
        )

        original_year_header = find_header(
            headers,
            "Original Publication Year",
        )

        description_header = find_header(
            headers,
            "Description",
        )

        my_rating_header = find_header(
            headers,
            "My Rating",
        )

        isbn_header = find_header(
            headers,
            "ISBN",
        )

        book_id_header = find_header(
            headers,
            "Book Id - Goodreads",
        )

        # ----------------------------------------------------
        # Helper for optional values
        # ----------------------------------------------------

        def get_value(
            values,
            header,
        ):
            if header is None:
                return ""

            index = header_index.get(
                header
            )

            if index is None:
                return ""

            if index >= len(values):
                return ""

            return normalize(
                values[index]
            )

        # ----------------------------------------------------
        # Import rows
        # ----------------------------------------------------

        books = []

        excel_row_number = 2

        for raw_values in rows:
            values = list(
                raw_values
            )

            # Make sure every row has enough
            # cells for all headers.
            while len(values) < len(
                headers
            ):
                values.append(None)

            shelf = get_value(
                values,
                shelf_header,
            ).lower()

            # ------------------------------------------------
            # THIS IS THE FILTER YOU ASKED FOR
            # ------------------------------------------------

            if shelf != "to-read":
                excel_row_number += 1
                continue

            title = get_value(
                values,
                title_header,
            )

            if not title:
                excel_row_number += 1
                continue

            # ------------------------------------------------
            # AUTHOR
            # ------------------------------------------------

            author = ""

            if author_index is not None:
                author = normalize(
                    values[author_index]
                )

            # Support Goodreads exports which might
            # instead contain separate author fields.
            if not author:
                first_name_header = (
                    find_header(
                        headers,
                        "Author First Name",
                    )
                )

                last_name_header = (
                    find_header(
                        headers,
                        "Author Last Name",
                    )
                )

                first_name = get_value(
                    values,
                    first_name_header,
                )

                last_name = get_value(
                    values,
                    last_name_header,
                )

                author = " ".join(
                    part
                    for part in [
                        first_name,
                        last_name,
                    ]
                    if part
                )

            # ------------------------------------------------
            # Goodreads ID
            # ------------------------------------------------

            goodreads_id = get_value(
                values,
                book_id_header,
            )

            if goodreads_id:
                book_id = (
                    "goodreads:"
                    + goodreads_id
                )
            else:
                # Fallback to source row.
                book_id = (
                    "row:"
                    + str(excel_row_number)
                )

            # ------------------------------------------------
            # Preserve original row
            # ------------------------------------------------

            original = {}

            for index, header in enumerate(
                headers
            ):
                if header is None:
                    continue

                value = (
                    values[index]
                    if index < len(values)
                    else None
                )

                # Convert dates/datetimes into
                # JSON-safe strings.
                if isinstance(
                    value,
                    datetime,
                ):
                    value = value.isoformat()

                original[
                    str(header)
                ] = value

            book = Book(
                id=book_id,
                title=title,
                author=author,
                pages=get_value(
                    values,
                    pages_header,
                ),
                year=get_value(
                    values,
                    year_header,
                ),
                publication_year=get_value(
                    values,
                    original_year_header,
                ),
                description=get_value(
                    values,
                    description_header,
                ),
                my_rating=get_value(
                    values,
                    my_rating_header,
                ),
                isbn=get_value(
                    values,
                    isbn_header,
                ),
                book_id=goodreads_id,
                original_row=excel_row_number,
                original=original,
            )

            books.append(book)

            excel_row_number += 1

        return headers, books

    finally:
        workbook.close()


# ============================================================
# RANKING ENGINE
# ============================================================

class RankingEngine:
    def __init__(
        self,
        books: list[Book],
        target_comparisons: int = (
            DEFAULT_TARGET_COMPARISONS
        ),
    ):
        self.books = books

        self.target_comparisons = max(
            MIN_TARGET_COMPARISONS,
            min(
                MAX_TARGET_COMPARISONS,
                int(target_comparisons),
            ),
        )

        self.ratings = {
            book.id: Rating()
            for book in books
        }

        self.comparisons = []

    # --------------------------------------------------------
    # Helpers
    # --------------------------------------------------------

    def book_map(self):
        return {
            book.id: book
            for book in self.books
        }

    @staticmethod
    def pair_key(
        first: str,
        second: str,
    ) -> str:
        return "|".join(
            sorted(
                [first, second]
            )
        )

    def played_pairs(self):
        return {
            self.pair_key(
                match["left"],
                match["right"],
            )
            for match in self.comparisons
        }

    def comparison_counts(self):
        counts = {
            book.id: 0
            for book in self.books
        }

        for match in self.comparisons:
            left = match["left"]
            right = match["right"]

            if left in counts:
                counts[left] += 1

            if right in counts:
                counts[right] += 1

        return counts

    # --------------------------------------------------------
    # Apply a comparison
    # --------------------------------------------------------

    def apply_match(
        self,
        left_id: str,
        right_id: str,
        result: str,
    ):
        if left_id not in self.ratings:
            raise ValueError(
                "Unknown left book."
            )

        if right_id not in self.ratings:
            raise ValueError(
                "Unknown right book."
            )

        if result not in (
            "left",
            "right",
            "tie",
        ):
            raise ValueError(
                "Invalid comparison result."
            )

        old_left = copy.deepcopy(
            self.ratings[left_id]
        )

        old_right = copy.deepcopy(
            self.ratings[right_id]
        )

        if result == "left":
            left_score = 1.0
            right_score = 0.0

        elif result == "right":
            left_score = 0.0
            right_score = 1.0

        else:
            left_score = 0.5
            right_score = 0.5

        new_left = glicko_update(
            old_left,
            [old_right],
            [left_score],
        )

        new_right = glicko_update(
            old_right,
            [old_left],
            [right_score],
        )

        # Preserve match statistics.
        new_left.comparisons = (
            old_left.comparisons + 1
        )

        new_right.comparisons = (
            old_right.comparisons + 1
        )

        new_left.wins = old_left.wins
        new_left.losses = old_left.losses
        new_left.ties = old_left.ties

        new_right.wins = old_right.wins
        new_right.losses = old_right.losses
        new_right.ties = old_right.ties

        if result == "left":
            new_left.wins += 1
            new_right.losses += 1

        elif result == "right":
            new_right.wins += 1
            new_left.losses += 1

        else:
            new_left.ties += 1
            new_right.ties += 1

        self.ratings[left_id] = (
            new_left
        )

        self.ratings[right_id] = (
            new_right
        )

        self.comparisons.append({
            "left": left_id,
            "right": right_id,
            "result": result,
            "timestamp": now_iso(),
        })

    # --------------------------------------------------------
    # Rebuild
    # --------------------------------------------------------

    def rebuild_from_history(self):
        history = copy.deepcopy(
            self.comparisons
        )

        self.ratings = {
            book.id: Rating()
            for book in self.books
        }

        self.comparisons = []

        for match in history:
            try:
                self.apply_match(
                    match["left"],
                    match["right"],
                    match["result"],
                )
            except Exception:
                # Ignore corrupted historical matches
                # rather than destroying the entire session.
                continue

    # --------------------------------------------------------
    # Undo
    # --------------------------------------------------------

    def undo(self) -> bool:
        if not self.comparisons:
            return False

        self.comparisons.pop()

        self.rebuild_from_history()

        return True

    # --------------------------------------------------------
    # Choose next pair
    # --------------------------------------------------------

    def choose_pair(self):
        if len(self.books) < 2:
            return None

        played = self.played_pairs()
        counts = self.comparison_counts()

        # Find books which still need evidence.
        needs = []

        for book in self.books:
            rating = self.ratings[
                book.id
            ]

            remaining = max(
                0,
                self.target_comparisons
                - counts[book.id],
            )

            # RD is deliberately included because a book
            # with high uncertainty deserves more comparisons.
            weight = (
                1.0
                + remaining * 5.0
                + rating.rd * 0.03
            )

            needs.append(
                (
                    book,
                    weight,
                )
            )

        # Random weighted primary book.
        primary = random.choices(
            [
                item[0]
                for item in needs
            ],
            weights=[
                item[1]
                for item in needs
            ],
            k=1,
        )[0]

        primary_rating = self.ratings[
            primary.id
        ]

        candidates = []

        for book in self.books:
            if book.id == primary.id:
                continue

            if (
                self.pair_key(
                    primary.id,
                    book.id,
                )
                in played
            ):
                continue

            rating = self.ratings[
                book.id
            ]

            rating_distance = abs(
                primary_rating.rating
                - rating.rating
            )

            # Prefer reasonably close books.
            similarity = math.exp(
                -rating_distance
                / 250.0
            )

            uncertainty = (
                primary_rating.rd
                + rating.rd
            ) / 2.0

            remaining = max(
                0,
                self.target_comparisons
                - counts[book.id],
            )

            score = (
                similarity * 100.0
                + uncertainty * 0.5
                + remaining * 8.0
            )

            candidates.append(
                (
                    score,
                    book,
                )
            )

        if candidates:
            candidates.sort(
                key=lambda item: item[0],
                reverse=True,
            )

            shortlist = candidates[
                : min(
                    5,
                    len(candidates),
                )
            ]

            opponent = random.choice(
                shortlist
            )[1]

            return (
                primary.id,
                opponent.id,
            )

        # Final fallback: highest combined uncertainty.
        fallback = []

        for i in range(
            len(self.books)
        ):
            for j in range(
                i + 1,
                len(self.books),
            ):
                left = self.books[i]
                right = self.books[j]

                if (
                    self.pair_key(
                        left.id,
                        right.id,
                    )
                    in played
                ):
                    continue

                score = (
                    self.ratings[
                        left.id
                    ].rd
                    + self.ratings[
                        right.id
                    ].rd
                )

                fallback.append(
                    (
                        score,
                        left.id,
                        right.id,
                    )
                )

        if fallback:
            fallback.sort(
                reverse=True
            )

            _, left, right = (
                fallback[0]
            )

            return left, right

        return None

    # --------------------------------------------------------
    # Ranking
    # --------------------------------------------------------

    def ranking(self):
        result = []

        for book in self.books:
            result.append({
                "book": book,
                "rating": self.ratings[
                    book.id
                ],
            })

        result.sort(
            key=lambda item: (
                item["rating"].rating,
                -item["rating"].rd,
            ),
            reverse=True,
        )

        return result

    # --------------------------------------------------------
    # Progress
    # --------------------------------------------------------

    def progress(self):
        if not self.books:
            return 0.0

        counts = (
            self.comparison_counts()
        )

        total = 0.0

        for book in self.books:
            total += min(
                counts[book.id]
                / self.target_comparisons,
                1.0,
            )

        return total / len(
            self.books
        )

    def is_finished(self):
        counts = (
            self.comparison_counts()
        )

        return all(
            counts[book.id]
            >= self.target_comparisons
            for book in self.books
        )


# ============================================================
# SAVED STATE
# ============================================================

class StateStore:
    def __init__(
        self,
        source_file: Path,
    ):
        self.source_file = (
            source_file
        )

        self.directory = (
            source_file.parent
            / STATE_DIRECTORY_NAME
        )

        self.directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        self.path = (
            self.directory
            / (
                safe_filename(
                    source_file.stem
                )
                + ".json"
            )
        )

    def exists(self):
        return self.path.exists()

    def save(
        self,
        engine: RankingEngine,
    ):
        data = {
            "version": 2,
            "source_file": str(
                self.source_file.resolve()
            ),
            "target_comparisons": (
                engine.target_comparisons
            ),
            "books": [
                asdict(book)
                for book in engine.books
            ],
            "comparisons": (
                engine.comparisons
            ),
            "saved_at": now_iso(),
        }

        temporary = self.path.with_suffix(
            ".tmp"
        )

        with temporary.open(
            "w",
            encoding="utf-8",
        ) as file:
            json.dump(
                data,
                file,
                indent=2,
                ensure_ascii=False,
            )

        temporary.replace(
            self.path
        )

    def load(
        self,
        books: list[Book],
    ) -> Optional[RankingEngine]:
        if not self.path.exists():
            return None

        try:
            with self.path.open(
                "r",
                encoding="utf-8",
            ) as file:
                data = json.load(file)

            saved_books = data.get(
                "books",
                [],
            )

            saved_ids = {
                book.get("id")
                for book in saved_books
            }

            current_ids = {
                book.id
                for book in books
            }

            # The saved ranking must belong to the same
            # to-read list.
            if saved_ids != current_ids:
                return None

            target = int(
                data.get(
                    "target_comparisons",
                    DEFAULT_TARGET_COMPARISONS,
                )
            )

            engine = RankingEngine(
                books,
                target,
            )

            history = data.get(
                "comparisons",
                [],
            )

            engine.comparisons = (
                copy.deepcopy(history)
            )

            engine.rebuild_from_history()

            return engine

        except Exception:
            return None


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_results(
    source_file: Path,
    engine: RankingEngine,
):
    """
    Create:

        originalname_ranked.xlsx

    The original Goodreads workbook is not modified.
    """

    workbook = load_workbook(
        filename=source_file
    )

    try:
        # ----------------------------------------------------
        # Ranking sheet
        # ----------------------------------------------------

        if "Ranking" in workbook.sheetnames:
            del workbook["Ranking"]

        ranking_sheet = (
            workbook.create_sheet(
                "Ranking",
                0,
            )
        )

        headers = [
            "Rank",
            "Title",
            "Author",
            "Rating",
            "Uncertainty",
            "Comparisons",
            "Wins",
            "Losses",
            "Ties",
            "Goodreads Book Id",
        ]

        ranking_sheet.append(
            headers
        )

        for rank, item in enumerate(
            engine.ranking(),
            start=1,
        ):
            book = item["book"]
            rating = item["rating"]

            ranking_sheet.append([
                rank,
                book.title,
                book.author,
                round(
                    rating.rating,
                    2,
                ),
                round(
                    rating.rd,
                    2,
                ),
                rating.comparisons,
                rating.wins,
                rating.losses,
                rating.ties,
                book.book_id,
            ])

        # ----------------------------------------------------
        # Styling
        # ----------------------------------------------------

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="2F5597",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        for cell in ranking_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = (
                Alignment(
                    vertical="center"
                )
            )

        ranking_sheet.freeze_panes = "A2"

        ranking_sheet.auto_filter.ref = (
            ranking_sheet.dimensions
        )

        widths = [
            8,
            45,
            30,
            14,
            16,
            15,
            10,
            10,
            10,
            22,
        ]

        for index, width in enumerate(
            widths,
            start=1,
        ):
            ranking_sheet.column_dimensions[
                get_column_letter(index)
            ].width = width

        # ----------------------------------------------------
        # Comparisons sheet
        # ----------------------------------------------------

        if (
            "Comparisons"
            in workbook.sheetnames
        ):
            del workbook["Comparisons"]

        comparison_sheet = (
            workbook.create_sheet(
                "Comparisons"
            )
        )

        comparison_sheet.append([
            "Left Book",
            "Left Author",
            "Right Book",
            "Right Author",
            "Result",
            "Timestamp",
        ])

        book_map = (
            engine.book_map()
        )

        for match in engine.comparisons:
            left = book_map.get(
                match["left"]
            )

            right = book_map.get(
                match["right"]
            )

            if left is None or right is None:
                continue

            if match["result"] == "left":
                result = left.title

            elif match["result"] == "right":
                result = right.title

            else:
                result = "Tie"

            comparison_sheet.append([
                left.title,
                left.author,
                right.title,
                right.author,
                result,
                match["timestamp"],
            ])

        for cell in comparison_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        comparison_sheet.freeze_panes = "A2"

        comparison_widths = {
            "A": 45,
            "B": 30,
            "C": 45,
            "D": 30,
            "E": 45,
            "F": 22,
        }

        for column, width in (
            comparison_widths.items()
        ):
            comparison_sheet.column_dimensions[
                column
            ].width = width

        # ----------------------------------------------------
        # Add ranking columns to original sheet
        # ----------------------------------------------------

        original_sheet = (
            workbook[
                workbook.sheetnames[
                    -1
                ]
            ]
        )

        # Find the original Goodreads sheet.
        # Usually it is the first non-generated sheet.
        for sheet in workbook.worksheets:
            if sheet.title not in (
                "Ranking",
                "Comparisons",
            ):
                original_sheet = sheet
                break

        existing_headers = [
            cell.value
            for cell in original_sheet[1]
        ]

        extra_headers = [
            "Rank",
            "Rank Rating",
            "Rating Uncertainty",
            "Pairwise Comparisons",
            "Wins",
            "Losses",
            "Ties",
        ]

        start_column = (
            len(existing_headers)
            + 1
        )

        for offset, header in enumerate(
            extra_headers
        ):
            cell = (
                original_sheet.cell(
                    row=1,
                    column=(
                        start_column
                        + offset
                    ),
                )
            )

            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Map book IDs to ranking information.
        ranked = {}

        for rank, item in enumerate(
            engine.ranking(),
            start=1,
        ):
            ranked[
                item["book"].id
            ] = (
                rank,
                item["rating"],
            )

        # Locate Goodreads Book Id column.
        book_id_column = None

        for index, header in enumerate(
            existing_headers,
            start=1,
        ):
            if (
                normalize_header(
                    header
                )
                == normalize_header(
                    "Book Id - Goodreads"
                )
            ):
                book_id_column = index
                break

        # Map original source rows.
        books_by_row = {
            book.original_row: book
            for book in engine.books
        }

        # ----------------------------------------------------
        # Populate ranking data
        # ----------------------------------------------------

        for row in range(
            2,
            original_sheet.max_row + 1,
        ):
            book = None

            # Best match: Goodreads ID.
            if book_id_column is not None:
                goodreads_id = normalize(
                    original_sheet.cell(
                        row=row,
                        column=book_id_column,
                    ).value
                )

                if goodreads_id:
                    candidate_id = (
                        "goodreads:"
                        + goodreads_id
                    )

                    for candidate in (
                        engine.books
                    ):
                        if (
                            candidate.id
                            == candidate_id
                        ):
                            book = candidate
                            break

            # Fallback: source row.
            if book is None:
                book = books_by_row.get(
                    row
                )

            if book is None:
                continue

            ranking_data = ranked.get(
                book.id
            )

            if ranking_data is None:
                continue

            rank, rating = (
                ranking_data
            )

            values = [
                rank,
                round(
                    rating.rating,
                    2,
                ),
                round(
                    rating.rd,
                    2,
                ),
                rating.comparisons,
                rating.wins,
                rating.losses,
                rating.ties,
            ]

            for offset, value in enumerate(
                values
            ):
                original_sheet.cell(
                    row=row,
                    column=(
                        start_column
                        + offset
                    ),
                    value=value,
                )

        # ----------------------------------------------------
        # Save
        # ----------------------------------------------------

        output_path = (
            source_file.parent
            / (
                safe_filename(
                    source_file.stem
                )
                + "_ranked.xlsx"
            )
        )

        workbook.save(
            output_path
        )

        return output_path

    finally:
        workbook.close()


# ============================================================
# GUI APPLICATION
# ============================================================

class RankerApp:
    def __init__(
        self,
        root: tk.Tk,
    ):
        self.root = root

        self.root.title(
            APP_NAME
        )

        self.root.geometry(
            "1180x760"
        )

        self.root.minsize(
            950,
            650,
        )

        self.engine = None
        self.source_file = None
        self.state_store = None
        self.current_pair = None

        self.target_var = tk.IntVar(
            value=DEFAULT_TARGET_COMPARISONS
        )

        self.status_var = tk.StringVar(
            value=(
                "Open your Goodreads "
                "Excel export to begin."
            )
        )

        self.progress_var = (
            tk.DoubleVar(
                value=0.0
            )
        )

        self.stats_var = (
            tk.StringVar(
                value=""
            )
        )

        self.left_title_var = (
            tk.StringVar()
        )

        self.left_author_var = (
            tk.StringVar()
        )

        self.left_meta_var = (
            tk.StringVar()
        )

        self.right_title_var = (
            tk.StringVar()
        )

        self.right_author_var = (
            tk.StringVar()
        )

        self.right_meta_var = (
            tk.StringVar()
        )

        self.build_styles()
        self.build_menu()
        self.build_ui()
        self.bind_keys()

        self.root.protocol(
            "WM_DELETE_WINDOW",
            self.on_close,
        )

    # --------------------------------------------------------
    # Styles
    # --------------------------------------------------------

    def build_styles(self):
        style = ttk.Style()

        try:
            style.theme_use(
                "vista"
            )
        except tk.TclError:
            pass

        style.configure(
            "Title.TLabel",
            font=(
                "Segoe UI",
                23,
                "bold",
            ),
        )

        style.configure(
            "BookTitle.TLabel",
            font=(
                "Segoe UI",
                18,
                "bold",
            ),
        )

        style.configure(
            "BookAuthor.TLabel",
            font=(
                "Segoe UI",
                12,
            ),
        )

        style.configure(
            "BookMeta.TLabel",
            font=(
                "Segoe UI",
                10,
            ),
            foreground="#666666",
        )

        style.configure(
            "Choice.TButton",
            font=(
                "Segoe UI",
                12,
                "bold",
            ),
            padding=14,
        )

        style.configure(
            "Secondary.TButton",
            font=(
                "Segoe UI",
                10,
            ),
            padding=8,
        )

    # --------------------------------------------------------
    # Menu
    # --------------------------------------------------------

    def build_menu(self):
        menu = tk.Menu(
            self.root
        )

        file_menu = tk.Menu(
            menu,
            tearoff=False,
        )

        file_menu.add_command(
            label="Open Goodreads Excel…",
            command=self.open_file,
            accelerator="Ctrl+O",
        )

        file_menu.add_command(
            label="Export Ranking",
            command=self.export,
            accelerator="Ctrl+E",
        )

        file_menu.add_separator()

        file_menu.add_command(
            label="Exit",
            command=self.on_close,
        )

        menu.add_cascade(
            label="File",
            menu=file_menu,
        )

        ranking_menu = tk.Menu(
            menu,
            tearoff=False,
        )

        ranking_menu.add_command(
            label="Undo Last Choice",
            command=self.undo,
            accelerator="U",
        )

        ranking_menu.add_command(
            label="View Ranking",
            command=self.show_ranking,
        )

        menu.add_cascade(
            label="Ranking",
            menu=ranking_menu,
        )

        help_menu = tk.Menu(
            menu,
            tearoff=False,
        )

        help_menu.add_command(
            label="Keyboard Shortcuts",
            command=self.show_shortcuts,
        )

        help_menu.add_command(
            label="About",
            command=self.show_about,
        )

        menu.add_cascade(
            label="Help",
            menu=help_menu,
        )

        self.root.config(
            menu=menu
        )

    # --------------------------------------------------------
    # Main UI
    # --------------------------------------------------------

    def build_ui(self):
        outer = ttk.Frame(
            self.root,
            padding=22,
        )

        outer.pack(
            fill="both",
            expand=True,
        )

        # Header
        header = ttk.Frame(
            outer
        )

        header.pack(
            fill="x",
            pady=(0, 15),
        )

        ttk.Label(
            header,
            text=(
                "📚 Goodreads "
                "To-Read Ranker"
            ),
            style="Title.TLabel",
        ).pack(
            side="left"
        )

        controls = ttk.Frame(
            header
        )

        controls.pack(
            side="right"
        )

        ttk.Label(
            controls,
            text="Comparisons/book:",
        ).pack(
            side="left",
            padx=(0, 7),
        )

        spinbox = ttk.Spinbox(
            controls,
            from_=MIN_TARGET_COMPARISONS,
            to=MAX_TARGET_COMPARISONS,
            textvariable=self.target_var,
            width=5,
        )

        spinbox.pack(
            side="left"
        )

        spinbox.bind(
            "<Return>",
            self.target_changed,
        )

        ttk.Button(
            controls,
            text="Open Excel",
            command=self.open_file,
        ).pack(
            side="left",
            padx=(15, 0),
        )

        # Status
        status = ttk.Frame(
            outer
        )

        status.pack(
            fill="x",
            pady=(0, 15),
        )

        ttk.Label(
            status,
            textvariable=self.status_var,
        ).pack(
            anchor="w"
        )

        ttk.Progressbar(
            status,
            variable=self.progress_var,
            maximum=100,
        ).pack(
            fill="x",
            pady=(8, 0),
        )

        ttk.Label(
            status,
            textvariable=self.stats_var,
        ).pack(
            anchor="w",
            pady=(5, 0),
        )

        # Cards
        choices = ttk.Frame(
            outer
        )

        choices.pack(
            fill="both",
            expand=True,
        )

        self.left_card = (
            self.make_book_card(
                choices,
                "left",
            )
        )

        self.left_card.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(0, 8),
        )

        middle = ttk.Frame(
            choices,
            width=70,
        )

        middle.pack(
            side="left",
            fill="y",
        )

        middle.pack_propagate(
            False
        )

        ttk.Label(
            middle,
            text="VS",
            font=(
                "Segoe UI",
                16,
                "bold",
            ),
        ).pack(
            pady=(155, 10)
        )

        self.right_card = (
            self.make_book_card(
                choices,
                "right",
            )
        )

        self.right_card.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(8, 0),
        )

        # Main buttons
        buttons = ttk.Frame(
            outer
        )

        buttons.pack(
            fill="x",
            pady=(15, 0),
        )

        self.left_button = ttk.Button(
            buttons,
            text="←  Choose Left",
            command=lambda: self.choose(
                "left"
            ),
            style="Choice.TButton",
        )

        self.left_button.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(0, 5),
        )

        self.tie_button = ttk.Button(
            buttons,
            text="≈  Tie",
            command=lambda: self.choose(
                "tie"
            ),
            style="Choice.TButton",
        )

        self.tie_button.pack(
            side="left",
            fill="x",
            expand=True,
            padx=5,
        )

        self.right_button = ttk.Button(
            buttons,
            text="Choose Right  →",
            command=lambda: self.choose(
                "right"
            ),
            style="Choice.TButton",
        )

        self.right_button.pack(
            side="left",
            fill="x",
            expand=True,
            padx=(5, 0),
        )

        # Bottom controls
        bottom = ttk.Frame(
            outer
        )

        bottom.pack(
            fill="x",
            pady=(12, 0),
        )

        ttk.Button(
            bottom,
            text="Undo",
            command=self.undo,
            style="Secondary.TButton",
        ).pack(
            side="left"
        )

        ttk.Button(
            bottom,
            text="Skip",
            command=self.skip,
            style="Secondary.TButton",
        ).pack(
            side="left",
            padx=7,
        )

        ttk.Button(
            bottom,
            text="View Ranking",
            command=self.show_ranking,
            style="Secondary.TButton",
        ).pack(
            side="right"
        )

        ttk.Button(
            bottom,
            text="Export Excel",
            command=self.export,
            style="Secondary.TButton",
        ).pack(
            side="right",
            padx=7,
        )

    # --------------------------------------------------------
    # Book card
    # --------------------------------------------------------

    def make_book_card(
        self,
        parent,
        side,
    ):
        frame = tk.Frame(
            parent,
            bg="#F6F7F9",
            highlightbackground="#D8DCE2",
            highlightthickness=1,
        )

        content = ttk.Frame(
            frame,
            padding=22,
        )

        content.pack(
            fill="both",
            expand=True,
        )

        if side == "left":
            title_var = (
                self.left_title_var
            )

            author_var = (
                self.left_author_var
            )

            meta_var = (
                self.left_meta_var
            )

        else:
            title_var = (
                self.right_title_var
            )

            author_var = (
                self.right_author_var
            )

            meta_var = (
                self.right_meta_var
            )

        ttk.Label(
            content,
            textvariable=title_var,
            style="BookTitle.TLabel",
            wraplength=410,
            justify="center",
        ).pack(
            pady=(25, 10)
        )

        ttk.Label(
            content,
            textvariable=author_var,
            style="BookAuthor.TLabel",
            wraplength=410,
            justify="center",
        ).pack(
            pady=(0, 18)
        )

        ttk.Label(
            content,
            textvariable=meta_var,
            style="BookMeta.TLabel",
            wraplength=410,
            justify="center",
        ).pack(
            pady=(0, 15)
        )

        description = tk.Text(
            content,
            height=9,
            width=45,
            wrap="word",
            relief="flat",
            borderwidth=0,
            bg="#F6F7F9",
            fg="#444444",
            font=(
                "Segoe UI",
                10,
            ),
            padx=5,
            pady=5,
        )

        description.pack(
            fill="both",
            expand=True,
        )

        description.configure(
            state="disabled"
        )

        if side == "left":
            self.left_description = (
                description
            )
        else:
            self.right_description = (
                description
            )

        return frame

    # --------------------------------------------------------
    # Keyboard
    # --------------------------------------------------------

    def bind_keys(self):
        self.root.bind(
            "<Left>",
            lambda event: self.choose(
                "left"
            ),
        )

        self.root.bind(
            "<Right>",
            lambda event: self.choose(
                "right"
            ),
        )

        self.root.bind(
            "<KeyPress-t>",
            lambda event: self.choose(
                "tie"
            ),
        )

        self.root.bind(
            "<KeyPress-T>",
            lambda event: self.choose(
                "tie"
            ),
        )

        self.root.bind(
            "<KeyPress-s>",
            lambda event: self.skip(),
        )

        self.root.bind(
            "<KeyPress-S>",
            lambda event: self.skip(),
        )

        self.root.bind(
            "<KeyPress-u>",
            lambda event: self.undo(),
        )

        self.root.bind(
            "<KeyPress-U>",
            lambda event: self.undo(),
        )

        self.root.bind(
            "<Control-o>",
            lambda event: self.open_file(),
        )

        self.root.bind(
            "<Control-e>",
            lambda event: self.export(),
        )

    # --------------------------------------------------------
    # Open file
    # --------------------------------------------------------

    def open_file(self):
        path = (
            filedialog.askopenfilename(
                title=(
                    "Open Goodreads Excel export"
                ),
                filetypes=[
                    (
                        "Excel files",
                        "*.xlsx",
                    ),
                    (
                        "All files",
                        "*.*",
                    ),
                ],
            )
        )

        if not path:
            return

        self.load_file(
            Path(path)
        )

    def load_file(
        self,
        path: Path,
    ):
        try:
            _, books = load_goodreads(
                path
            )

        except Exception as exc:
            messagebox.showerror(
                "Could not open Goodreads file",
                str(exc),
            )
            return

        if len(books) < 2:
            messagebox.showwarning(
                "Not enough books",
                (
                    "I found fewer than two books "
                    'where "Exclusive Shelf" is '
                    '"to-read".'
                ),
            )
            return

        self.source_file = path

        self.state_store = (
            StateStore(path)
        )

        existing = (
            self.state_store.load(
                books
            )
        )

        if existing is not None:
            resume = messagebox.askyesno(
                "Resume previous ranking?",
                (
                    f"I found a saved ranking for:\n\n"
                    f"{path.name}\n\n"
                    f"{len(existing.comparisons)} "
                    f"comparisons have already been "
                    f"made.\n\n"
                    f"Resume it?"
                ),
            )

            if resume:
                self.engine = existing

            else:
                self.engine = (
                    RankingEngine(
                        books,
                        self.target_var.get(),
                    )
                )

        else:
            self.engine = (
                RankingEngine(
                    books,
                    self.target_var.get(),
                )
            )

        self.target_var.set(
            self.engine.target_comparisons
        )

        self.current_pair = None

        self.status_var.set(
            (
                f"{len(books)} books on your "
                f"to-read shelf · {path.name}"
            )
        )

        self.refresh()

    # --------------------------------------------------------
    # Target comparisons
    # --------------------------------------------------------

    def target_changed(
        self,
        event=None,
    ):
        if self.engine is None:
            return

        try:
            value = int(
                self.target_var.get()
            )

        except ValueError:
            value = (
                DEFAULT_TARGET_COMPARISONS
            )

        value = max(
            MIN_TARGET_COMPARISONS,
            min(
                MAX_TARGET_COMPARISONS,
                value,
            ),
        )

        self.target_var.set(
            value
        )

        self.engine.target_comparisons = (
            value
        )

        self.save_state()

        self.current_pair = None

        self.refresh()

    # --------------------------------------------------------
    # Refresh
    # --------------------------------------------------------

    def refresh(self):
        if self.engine is None:
            return

        if (
            self.current_pair is None
            or not self.current_pair_valid()
        ):
            self.current_pair = (
                self.engine.choose_pair()
            )

        if self.current_pair is None:
            self.disable_choices()

            self.left_title_var.set(
                "Ranking complete"
            )

            self.left_author_var.set(
                ""
            )

            self.left_meta_var.set(
                ""
            )

            self.right_title_var.set(
                "No more comparisons"
            )

            self.right_author_var.set(
                ""
            )

            self.right_meta_var.set(
                ""
            )

            self.set_description(
                self.left_description,
                "",
            )

            self.set_description(
                self.right_description,
                "",
            )

            return

        self.enable_choices()

        left_id, right_id = (
            self.current_pair
        )

        book_map = (
            self.engine.book_map()
        )

        left = book_map[
            left_id
        ]

        right = book_map[
            right_id
        ]

        self.show_book(
            left,
            self.engine.ratings[
                left.id
            ],
            self.left_title_var,
            self.left_author_var,
            self.left_meta_var,
            self.left_description,
        )

        self.show_book(
            right,
            self.engine.ratings[
                right.id
            ],
            self.right_title_var,
            self.right_author_var,
            self.right_meta_var,
            self.right_description,
        )

        progress = (
            self.engine.progress()
            * 100.0
        )

        self.progress_var.set(
            progress
        )

        comparisons = len(
            self.engine.comparisons
        )

        if self.engine.books:
            average = (
                sum(
                    self.engine.ratings[
                        book.id
                    ].comparisons
                    for book
                    in self.engine.books
                )
                / len(
                    self.engine.books
                )
            )
        else:
            average = 0

        self.stats_var.set(
            (
                f"{comparisons} decisions · "
                f"{average:.1f} average comparisons/book · "
                f"{progress:.0f}% evidence target"
            )
        )

        if self.engine.is_finished():
            self.status_var.set(
                (
                    "Target reached for every book. "
                    "You can keep ranking or export."
                )
            )
        else:
            self.status_var.set(
                "Which book would you rather read?"
            )

    def current_pair_valid(self):
        if (
            self.engine is None
            or self.current_pair is None
        ):
            return False

        left, right = (
            self.current_pair
        )

        return (
            left in self.engine.ratings
            and right in self.engine.ratings
        )

    # --------------------------------------------------------
    # Display book
    # --------------------------------------------------------

    def show_book(
        self,
        book,
        rating,
        title_var,
        author_var,
        meta_var,
        description_widget,
    ):
        title_var.set(
            book.title
        )

        # This is the important author fix.
        author_var.set(
            book.author
            if book.author
            else "Unknown author"
        )

        meta = []

        if book.pages:
            meta.append(
                f"{book.pages} pages"
            )

        if book.year:
            meta.append(
                f"Published {book.year}"
            )

        meta.append(
            f"Rating {rating.rating:.0f}"
        )

        meta.append(
            f"±{rating.rd:.0f}"
        )

        meta.append(
            f"{rating.comparisons} comparisons"
        )

        meta_var.set(
            "  ·  ".join(meta)
        )

        description = (
            book.description
            if book.description
            else "No description available."
        )

        self.set_description(
            description_widget,
            description,
        )

    def set_description(
        self,
        widget,
        text,
    ):
        widget.configure(
            state="normal"
        )

        widget.delete(
            "1.0",
            "end",
        )

        widget.insert(
            "1.0",
            truncate(
                text,
                1800,
            ),
        )

        widget.configure(
            state="disabled"
        )

    # --------------------------------------------------------
    # Choices
    # --------------------------------------------------------

    def choose(
        self,
        result: str,
    ):
        if (
            self.engine is None
            or self.current_pair is None
        ):
            return

        left_id, right_id = (
            self.current_pair
        )

        try:
            self.engine.apply_match(
                left_id,
                right_id,
                result,
            )

            self.save_state()

        except Exception as exc:
            messagebox.showerror(
                "Could not record choice",
                str(exc),
            )

            return

        self.current_pair = None

        self.refresh()

    def skip(self):
        if self.engine is None:
            return

        self.current_pair = None

        self.refresh()

    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------

    def save_state(self):
        if (
            self.engine is None
            or self.state_store is None
        ):
            return

        try:
            self.state_store.save(
                self.engine
            )
        except Exception as exc:
            # Saving should never stop the ranking UI.
            print(
                "Warning: could not save state:",
                exc,
            )

    # --------------------------------------------------------
    # Undo
    # --------------------------------------------------------

    def undo(self):
        if (
            self.engine is None
            or not self.engine.comparisons
        ):
            return

        if self.engine.undo():
            self.save_state()

            self.current_pair = None

            self.refresh()

    # --------------------------------------------------------
    # Button states
    # --------------------------------------------------------

    def enable_choices(self):
        self.left_button.configure(
            state="normal"
        )

        self.tie_button.configure(
            state="normal"
        )

        self.right_button.configure(
            state="normal"
        )

    def disable_choices(self):
        self.left_button.configure(
            state="disabled"
        )

        self.tie_button.configure(
            state="disabled"
        )

        self.right_button.configure(
            state="disabled"
        )

    # --------------------------------------------------------
    # Ranking window
    # --------------------------------------------------------

    def show_ranking(self):
        if self.engine is None:
            messagebox.showinfo(
                "No ranking",
                (
                    "Open a Goodreads file first."
                ),
            )
            return

        window = tk.Toplevel(
            self.root
        )

        window.title(
            "Current Ranking"
        )

        window.geometry(
            "1100x650"
        )

        frame = ttk.Frame(
            window,
            padding=15,
        )

        frame.pack(
            fill="both",
            expand=True,
        )

        ttk.Label(
            frame,
            text="Current Ranking",
            style="Title.TLabel",
        ).pack(
            anchor="w",
            pady=(0, 10),
        )

        columns = (
            "rank",
            "title",
            "author",
            "rating",
            "rd",
            "comparisons",
            "wins",
            "losses",
            "ties",
        )

        tree = ttk.Treeview(
            frame,
            columns=columns,
            show="headings",
        )

        headings = {
            "rank": "Rank",
            "title": "Title",
            "author": "Author",
            "rating": "Rating",
            "rd": "Uncertainty",
            "comparisons": "Comparisons",
            "wins": "Wins",
            "losses": "Losses",
            "ties": "Ties",
        }

        widths = {
            "rank": 55,
            "title": 350,
            "author": 230,
            "rating": 80,
            "rd": 95,
            "comparisons": 100,
            "wins": 70,
            "losses": 70,
            "ties": 60,
        }

        for column in columns:
            tree.heading(
                column,
                text=headings[column],
            )

            tree.column(
                column,
                width=widths[column],
                anchor=(
                    "w"
                    if column
                    in (
                        "title",
                        "author",
                    )
                    else "center"
                ),
            )

        scrollbar = ttk.Scrollbar(
            frame,
            orient="vertical",
            command=tree.yview,
        )

        tree.configure(
            yscrollcommand=scrollbar.set
        )

        tree.pack(
            side="left",
            fill="both",
            expand=True,
        )

        scrollbar.pack(
            side="right",
            fill="y",
        )

        for rank, item in enumerate(
            self.engine.ranking(),
            start=1,
        ):
            book = item["book"]
            rating = item["rating"]

            tree.insert(
                "",
                "end",
                values=(
                    rank,
                    book.title,
                    book.author,
                    f"{rating.rating:.0f}",
                    f"±{rating.rd:.0f}",
                    rating.comparisons,
                    rating.wins,
                    rating.losses,
                    rating.ties,
                ),
            )

    # --------------------------------------------------------
    # Export
    # --------------------------------------------------------

    def export(self):
        if (
            self.engine is None
            or self.source_file is None
        ):
            messagebox.showinfo(
                "Nothing to export",
                (
                    "Open a Goodreads file first."
                ),
            )
            return

        try:
            output_path = export_results(
                self.source_file,
                self.engine,
            )

        except PermissionError:
            messagebox.showerror(
                "Could not save",
                (
                    "Windows could not save the "
                    "ranked workbook.\n\n"
                    "If the output file is open in "
                    "Excel, close it and try again."
                ),
            )
            return

        except Exception as exc:
            messagebox.showerror(
                "Export failed",
                str(exc),
            )
            return

        open_file = messagebox.askyesno(
            "Ranking exported",
            (
                "Ranking exported successfully.\n\n"
                f"{output_path}\n\n"
                "Open it now?"
            ),
        )

        if open_file:
            self.open_external(
                output_path
            )

    @staticmethod
    def open_external(
        path: Path,
    ):
        try:
            if sys.platform.startswith(
                "win"
            ):
                os.startfile(
                    str(path)
                )

            elif sys.platform == "darwin":
                subprocess.Popen([
                    "open",
                    str(path),
                ])

            else:
                subprocess.Popen([
                    "xdg-open",
                    str(path),
                ])

        except Exception:
            pass

    # --------------------------------------------------------
    # Help
    # --------------------------------------------------------

    def show_shortcuts(self):
        messagebox.showinfo(
            "Keyboard shortcuts",
            (
                "←   Choose left book\n"
                "→   Choose right book\n"
                "T   Tie / equal\n"
                "S   Skip pair\n"
                "U   Undo\n"
                "Ctrl+O   Open Goodreads Excel\n"
                "Ctrl+E   Export ranking"
            ),
        )

    def show_about(self):
        messagebox.showinfo(
            "About",
            (
                f"{APP_NAME}\n"
                f"Version {APP_VERSION}\n\n"
                "Ranks Goodreads to-read books using "
                "pairwise comparisons and Glicko-2.\n\n"
                'Only rows where "Exclusive Shelf" '
                'equals "to-read" are included.\n\n'
                "Your original Goodreads workbook is "
                "never modified."
            ),
        )

    # --------------------------------------------------------
    # Close
    # --------------------------------------------------------

    def on_close(self):
        self.save_state()

        self.root.destroy()


# ============================================================
# MAIN
# ============================================================

def main():
    root = tk.Tk()

    RankerApp(
        root
    )

    root.mainloop()


if __name__ == "__main__":
    try:
        main()

    except Exception:
        error = traceback.format_exc()

        print(error)

        try:
            messagebox.showerror(
                "Unexpected error",
                (
                    "The application encountered an "
                    "unexpected error:\n\n"
                    + error
                ),
            )
        except Exception:
            pass